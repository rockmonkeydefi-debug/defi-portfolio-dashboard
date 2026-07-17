"""Flask web app for DeFi Portfolio Visualization."""

import sys
import os
import math
import time
import requests
import warnings
import re
import json
import threading
import traceback
from datetime import datetime
from dotenv import load_dotenv, set_key
from web3 import Web3
from flask import Flask, render_template, jsonify, request, session, redirect, url_for
import bcrypt
from functools import wraps

# Suppress urllib3 SSL warning
warnings.filterwarnings('ignore', message='.*OpenSSL.*')

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from src.connectors.uniswap_v3 import UniswapV3Connector, POOL_ABI, ERC20_ABI
from src.connectors.aerodrome_slipstream import (
    AerodromeSlipstreamConnector,
    POOL_ABI as AERO_POOL_ABI,
    POOL_FACTORY_ABI as AERO_FACTORY_ABI,
    POOL_FACTORY_ADDRESS as AERO_FACTORY_ADDRESS,
)
from src.connectors.aave_v3 import get_aave_positions
from src.connectors.zerion import (
    ZerionConnector,
    categorize_zerion_positions,
    map_zerion_token_to_app,
    map_zerion_lending_to_app,
    map_zerion_lp_to_app,
)
from src.models import Chain
from src.engines.telegram_service import (
    load_telegram_config, save_telegram_config, validate_telegram_config,
    mask_bot_token, send_telegram_message, build_notification_content,
)
from src.engines.range_optimizer import (
    discover_pools, run_optimization, load_regime_probabilities,
)

# ── Hyperliquid coin-name resolution ────────────────────────────────────
# Crypto perps use bare names ('BTC'). TradFi perps are HIP-3 builder-deployed
# on dex 'xyz' and require the namespaced form ('xyz:NVDA'). We cache both
# universe name-sets so candleSnapshot calls hit the correct string without a
# metadata fetch per ticker. Single gunicorn worker → one shared cache.
_HL_UNIVERSE_CACHE = {
    'crypto': None,      # set of bare names, e.g. {'BTC','ETH',...}
    'xyz': None,         # dict: bare ticker -> full name, e.g. {'NVDA':'xyz:NVDA'}
    'fetched_at': 0.0,
}
_HL_CACHE_TTL = 3600     # seconds; refresh universe sets hourly

# ── Global Hyperliquid rate limiter (token bucket) ──────────────────────────
# HL shares an aggregated 1200 weight/min budget. candleSnapshot is ~20-23
# weight each → ~50 fetches/min ceiling. ALL HL requests go through _hl_post,
# which acquires a token first — so the cap is shared across the whole process
# (scanner worker pool, market snapshot, universe refresh, etc.).
# 55/min sits just above HL's nominal candleSnapshot ceiling: safe because the
# token bucket paces evenly (no bursts) and retries absorb the rare 429. This
# is the main tuning knob — dial back toward 45 if 429s reappear.
_HL_RATE_MAX_PER_MIN = 55
_HL_RATE_LOCK = threading.Lock()
_HL_RATE_STATE = {
    'tokens': float(_HL_RATE_MAX_PER_MIN),
    'last_refill': time.time(),
    'cap_hit_count': 0,        # cumulative times we had to block (for logging)
}

# Only one scan (manual or scheduled) runs at a time. They share the HL rate
# budget and write the same table, so serialize them. Neither blocks-waits
# (avoids the gunicorn timeout) — the loser is told (409) / skips.
_SCAN_LOCK = threading.Lock()

# Live progress + cooperative-cancel state for the in-flight scan (manual or
# scheduled). Read by GET /scanner/status; cancel sets _SCAN_CANCEL, which the
# scan checks between tickers (never kills a thread mid-write).
_SCAN_STATE_LOCK = threading.Lock()
_SCAN_STATE = {
    'running': False,
    'kind': None,          # 'manual' | 'scheduled'
    'started_at': None,    # epoch seconds
    'total': 0,            # total ticker×pair units to scan
    'done': 0,             # units completed
    'current': None,       # current ticker symbol
    'cancelled': False,    # set when a cancel was requested
    'last_result': None,   # short summary of the last finished scan
}
_SCAN_CANCEL = threading.Event()


def _scan_state_update(**kwargs):
    with _SCAN_STATE_LOCK:
        _SCAN_STATE.update(kwargs)


def _scan_state_snapshot():
    with _SCAN_STATE_LOCK:
        return dict(_SCAN_STATE)


def _hl_rate_acquire():
    """Block until a token is available. The bucket refills continuously at
    _HL_RATE_MAX_PER_MIN per 60s."""
    while True:
        with _HL_RATE_LOCK:
            now = time.time()
            elapsed = now - _HL_RATE_STATE['last_refill']
            refill = elapsed * (_HL_RATE_MAX_PER_MIN / 60.0)
            if refill > 0:
                _HL_RATE_STATE['tokens'] = min(
                    float(_HL_RATE_MAX_PER_MIN),
                    _HL_RATE_STATE['tokens'] + refill)
                _HL_RATE_STATE['last_refill'] = now
            if _HL_RATE_STATE['tokens'] >= 1.0:
                _HL_RATE_STATE['tokens'] -= 1.0
                return
            # No token available — record the cap hit (for any future logging)
            # and compute the wait. Hitting the cap is expected during wide
            # scans; the scan still completes, so we don't notify on it.
            _HL_RATE_STATE['cap_hit_count'] += 1
            deficit = 1.0 - _HL_RATE_STATE['tokens']
            wait = deficit / (_HL_RATE_MAX_PER_MIN / 60.0)
        # Wait outside the lock, then re-check.
        time.sleep(min(wait, 2.0))


def _hl_post(payload, timeout=10, retries=4):
    """The ONE place HL requests go through. Acquires a rate token, then POSTs
    with retry-on-429/5xx backoff. Returns parsed JSON. Raises on exhausted
    retries / real 4xx."""
    _RETRY_STATUS = {429, 500, 502, 503, 504}
    _BACKOFF = [0.5, 1.5, 3.0]
    _hl_rate_acquire()   # one token per LOGICAL fetch — retries don't re-acquire
    for _attempt in range(retries):
        try:
            resp = requests.post(
                'https://api.hyperliquid.xyz/info',
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=timeout)
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError):
            if _attempt >= retries - 1:
                raise
            time.sleep(_BACKOFF[min(_attempt, len(_BACKOFF) - 1)])
            continue
        if resp.status_code in _RETRY_STATUS:
            if _attempt >= retries - 1:
                resp.raise_for_status()
            _wait = _BACKOFF[min(_attempt, len(_BACKOFF) - 1)]
            if resp.status_code == 429:
                _ra = resp.headers.get('Retry-After')
                if _ra:
                    try:
                        _wait = min(float(_ra), 10.0)
                    except (TypeError, ValueError):
                        pass
            time.sleep(_wait)
            continue
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError("HL post exhausted retries")


def _hl_refresh_universes(force=False):
    """Populate/refresh the HL universe caches. Best-effort: on failure,
    leaves any existing cache in place and returns silently."""
    now = time.time()
    if (not force and _HL_UNIVERSE_CACHE['crypto'] is not None
            and (now - _HL_UNIVERSE_CACHE['fetched_at']) < _HL_CACHE_TTL):
        return
    # Default (crypto) universe
    try:
        meta = _hl_post({'type': 'metaAndAssetCtxs'})
        uni = (meta[0].get('universe') if meta and isinstance(meta, list) else None) or []
        _HL_UNIVERSE_CACHE['crypto'] = {
            u['name'].upper() for u in uni if u.get('name')
        }
    except Exception as e:
        print(f"[HL] crypto universe refresh failed: {type(e).__name__}: {e}", flush=True)
    # TradFi 'xyz' dex universe
    try:
        meta = _hl_post({'type': 'meta', 'dex': 'xyz'})
        uni = (meta.get('universe') if isinstance(meta, dict) else None) or []
        xyz_map = {}
        for u in uni:
            full = u.get('name')          # e.g. 'xyz:NVDA'
            if not full:
                continue
            bare = full.split(':', 1)[1] if ':' in full else full
            xyz_map[bare.upper()] = full
        _HL_UNIVERSE_CACHE['xyz'] = xyz_map
    except Exception as e:
        print(f"[HL] xyz universe refresh failed: {type(e).__name__}: {e}", flush=True)
    _HL_UNIVERSE_CACHE['fetched_at'] = now

def _hl_resolve_coin(ticker):
    """Map a bare ticker to the coin string candleSnapshot accepts.
       1) crypto universe → bare name unchanged
       2) xyz (TradFi) universe → 'xyz:TICKER'
       3) unknown → bare name (will fail as before; genuinely unavailable)."""
    if not ticker:
        return ticker
    _hl_refresh_universes()
    up = ticker.upper()
    crypto = _HL_UNIVERSE_CACHE.get('crypto') or set()
    xyz    = _HL_UNIVERSE_CACHE.get('xyz') or {}
    if up in crypto:
        return ticker            # crypto: unchanged
    if up in xyz:
        return xyz[up]           # TradFi: namespaced 'xyz:NVDA'
    return ticker                # unknown: leave bare

# Auto-create config files from examples on first run (local dev only;
# in Docker the entrypoint handles this via the config volume).
_env_path = os.getenv("DOTENV_PATH", ".env")
if not os.path.exists(_env_path) and os.path.exists(".env.example"):
    import shutil
    shutil.copy(".env.example", _env_path)
    print(f"Created {_env_path} from .env.example — edit it with your API keys")

_wc_path = os.getenv("WALLET_CONFIG_PATH", "wallet_config.json")
if not os.path.exists(_wc_path):
    with open(_wc_path, "w") as f:
        f.write("{}")

# Load environment variables from persistent config path (Docker) or local .env
ENV_FILE = os.getenv("DOTENV_PATH", ".env")
load_dotenv(ENV_FILE)

# Startup validation — warn about missing optional keys
_optional_keys = {
    "ZERION_API_KEY": "Zerion portfolio data",
    "ETHEREUM_RPC_URL": "Ethereum on-chain data",
    "ARBITRUM_RPC_URL": "Arbitrum on-chain data",
    "BASE_RPC_URL": "Base on-chain data",
    "ETHERSCAN_API_KEY": "Position age & collected fees",
    "FRED_API_KEY": "FRED macro indicators (US10Y, DXY, M2, Fed Funds)",
}
for _key, _desc in _optional_keys.items():
    if not os.getenv(_key):
        print(f"[Startup] Warning: {_key} not set — {_desc} will be unavailable")

app = Flask(__name__)
# Persistent secret key — generate once and save to .env if missing
_flask_secret = os.getenv("FLASK_SECRET_KEY")
if not _flask_secret:
    import secrets as _secrets
    _flask_secret = _secrets.token_hex(32)
    set_key(ENV_FILE, "FLASK_SECRET_KEY", _flask_secret)
app.secret_key = _flask_secret
app.permanent_session_lifetime = __import__('datetime').timedelta(hours=24)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB upload limit

# Inject version string into all templates for static asset cache-busting.
# Prefer git short hash; fall back to a startup timestamp so the version
# always changes on redeploy even when git is not available at runtime.
try:
    import subprocess as _sp
    _git_hash = _sp.check_output(['git', 'rev-parse', '--short', 'HEAD'], stderr=_sp.DEVNULL).decode().strip()
    if not _git_hash:
        raise ValueError('empty')
except Exception:
    import time as _time
    _git_hash = str(int(_time.time()))

@app.context_processor
def inject_static_version():
    return {'static_version': _git_hash}


@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    return response


@app.errorhandler(Exception)
def handle_exception(e):
    """Return JSON errors for API routes, HTML for others."""
    if request.path.startswith('/api/'):
        print(f"API error on {request.path}: {e}")
        return jsonify({"error": str(e)}), 500
    # Let non-API errors propagate to default Flask handler
    raise e


# --- Authentication ---
# Rate limiting for login
_login_attempts = {}  # ip -> (count, first_attempt_time)
MAX_LOGIN_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300  # 5 minutes


def get_password_hash():
    """Get the stored bcrypt password hash from .env."""
    return os.getenv("APP_PASSWORD_HASH", "")


def check_rate_limit(ip):
    """Check if IP has exceeded login attempt limit."""
    now = datetime.now().timestamp()
    if ip in _login_attempts:
        count, first_time = _login_attempts[ip]
        if now - first_time > LOGIN_WINDOW_SECONDS:
            _login_attempts[ip] = (1, now)
            return True
        if count >= MAX_LOGIN_ATTEMPTS:
            return False
        _login_attempts[ip] = (count + 1, first_time)
        return True
    _login_attempts[ip] = (1, now)
    return True


def login_required(f):
    """Decorator to require authentication."""
    @wraps(f)
    def decorated(*args, **kwargs):
        pw_hash = get_password_hash()
        # If no password is set, skip auth (local dev convenience)
        if not pw_hash:
            return f(*args, **kwargs)
        if not session.get('authenticated'):
            # For API calls, return 401; for pages, redirect to login
            if request.path.startswith('/api/'):
                return jsonify({"error": "Authentication required"}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


@app.before_request
def require_auth():
    """Global auth check — all routes except login/logout/static."""
    # Skip auth for login, logout, and static files
    if request.path in ('/login', '/logout') or request.path.startswith('/static/'):
        return
    pw_hash = get_password_hash()
    if not pw_hash:
        # No password set — redirect to setup (which is the login page)
        if request.path != '/login':
            return redirect(url_for('login_page'))
        return
    if not session.get('authenticated'):
        if request.path.startswith('/api/'):
            return jsonify({"error": "Authentication required"}), 401
        return redirect(url_for('login_page'))

# Configuration
WALLET_CONFIG_FILE = os.getenv("WALLET_CONFIG_PATH", "wallet_config.json")

# Cache for portfolio data
_portfolio_cache = None

def load_wallet_config():
    """Load wallet configuration from JSON file."""
    if os.path.exists(WALLET_CONFIG_FILE):
        try:
            with open(WALLET_CONFIG_FILE, 'r') as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except (json.JSONDecodeError, ValueError):
            return {}
    return {}

def save_wallet_config(config):
    """Save wallet configuration to JSON file."""
    with open(WALLET_CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)

def get_wallet_addresses():
    """Get wallet addresses from config file, excluding hidden wallets."""
    config = load_wallet_config()
    return [addr for addr, info in config.items() if not info.get("hidden", False)]

def get_wallet_label(address):
    """Get label for a wallet address."""
    config = load_wallet_config()
    return config.get(address, {}).get('label', address[:10] + '...')

def save_wallet_addresses(addresses):
    """Save wallet addresses to .env file (for backward compatibility)."""
    wallet_str = ",".join(addresses)
    set_key(ENV_FILE, "WALLET_ADDRESS", wallet_str)
    
    # Reload environment variables
    load_dotenv(ENV_FILE)

def is_valid_address(address):
    """Validate Ethereum address format."""
    if not address:
        return False
    if not address.startswith("0x"):
        return False
    if len(address) != 42:
        return False
    try:
        int(address[2:], 16)
        return True
    except ValueError:
        return False

FACTORY_ADDRESSES = {
    Chain.ETHEREUM: "0x1F98431c8aD98523631AE4a59f267346ea31F984",
    Chain.ARBITRUM: "0x1F98431c8aD98523631AE4a59f267346ea31F984",
    Chain.BASE: "0x33128a8fC17869897dcE68Ed026d694621f6FDfD"
}

FACTORY_ABI = [
    {
        "inputs": [
            {"name": "tokenA", "type": "address"},
            {"name": "tokenB", "type": "address"},
            {"name": "fee", "type": "uint24"}
        ],
        "name": "getPool",
        "outputs": [{"name": "pool", "type": "address"}],
        "stateMutability": "view",
        "type": "function"
    }
]

_price_cache = {}


def get_token_price_usd(symbol: str, address: str = None, chain: str = "ethereum") -> float:
    """Get token price in USD using DeFiLlama API."""
    cache_key = f"{chain}:{address}" if address else symbol
    
    if cache_key in _price_cache:
        return _price_cache[cache_key]
    
    try:
        if symbol == "ETH" and not address:
            address = "0x0000000000000000000000000000000000000000"
        
        # Bitcoin — use CoinGecko ID via DeFiLlama
        if symbol == "BTC" and chain == "bitcoin":
            for attempt in range(3):
                try:
                    url = "https://coins.llama.fi/prices/current/coingecko:bitcoin"
                    response = requests.get(url, timeout=10)
                    if response.status_code == 200:
                        price = response.json().get("coins", {}).get("coingecko:bitcoin", {}).get("price", 0.0)
                        if price > 0:
                            _price_cache[cache_key] = price
                            return price
                    if response.status_code == 429 and attempt < 2:
                        import time
                        time.sleep(2)
                        continue
                except Exception as e:
                    if attempt < 2:
                        import time
                        time.sleep(1)
                    else:
                        print(f"Price fetch failed for BTC: {e}")
                    pass
                break
        
        if address:
            try:
                url = f"https://coins.llama.fi/prices/current/{chain}:{address}"
                response = requests.get(url, timeout=5)
                
                if response.status_code == 200:
                    data = response.json()
                    price = data.get("coins", {}).get(f"{chain}:{address}", {}).get("price", 0.0)
                    if price > 0:
                        _price_cache[cache_key] = price
                        return price
            except requests.exceptions.RequestException as e:
                print(f"Price fetch failed for {symbol} ({chain}:{address}): {e}")
        _price_cache[cache_key] = 0.0
        return 0.0
        
    except Exception as e:
        print(f"Unexpected error in get_token_price_usd({symbol}): {e}")
        _price_cache[cache_key] = 0.0
        return 0.0




def calculate_token_amounts(liquidity: int, tick_lower: int, tick_upper: int, current_tick: int,
                           sqrt_price_x96: int, token0_decimals: int, token1_decimals: int):
    """Calculate the actual token amounts from liquidity."""
    try:
        sqrt_price_lower = math.sqrt(1.0001 ** tick_lower)
        sqrt_price_upper = math.sqrt(1.0001 ** tick_upper)
        sqrt_price_current = sqrt_price_x96 / (2 ** 96)
        
        if current_tick < tick_lower:
            amount0 = liquidity * (1/sqrt_price_lower - 1/sqrt_price_upper)
            amount1 = 0
        elif current_tick >= tick_upper:
            amount0 = 0
            amount1 = liquidity * (sqrt_price_upper - sqrt_price_lower)
        else:
            amount0 = liquidity * (1/sqrt_price_current - 1/sqrt_price_upper)
            amount1 = liquidity * (sqrt_price_current - sqrt_price_lower)
        
        amount0_adjusted = amount0 / (10 ** token0_decimals)
        amount1_adjusted = amount1 / (10 ** token1_decimals)
        
        return amount0_adjusted, amount1_adjusted
    except (ZeroDivisionError, OverflowError, ValueError) as e:
        print(f"Token amount calculation error: {e}")
        return 0, 0


def subIn256(x: int, y: int) -> int:
    """Handle underflow for 256-bit subtraction (wraps around)."""
    Q256 = 2 ** 256
    difference = x - y
    if difference < 0:
        return Q256 + difference
    return difference


def calculate_uncollected_fees(
    liquidity: int,
    tick_lower: int,
    tick_upper: int,
    tick_current: int,
    fee_growth_global_0: int,
    fee_growth_global_1: int,
    fee_growth_outside_lower_0: int,
    fee_growth_outside_lower_1: int,
    fee_growth_outside_upper_0: int,
    fee_growth_outside_upper_1: int,
    fee_growth_inside_last_0: int,
    fee_growth_inside_last_1: int
) -> tuple[int, int]:
    """Calculate uncollected fees for a Uniswap V3 position."""
    # Calculate fee growth below lower tick
    if tick_current >= tick_lower:
        fee_growth_below_0 = fee_growth_outside_lower_0
        fee_growth_below_1 = fee_growth_outside_lower_1
    else:
        fee_growth_below_0 = subIn256(fee_growth_global_0, fee_growth_outside_lower_0)
        fee_growth_below_1 = subIn256(fee_growth_global_1, fee_growth_outside_lower_1)
    
    # Calculate fee growth above upper tick
    if tick_current >= tick_upper:
        fee_growth_above_0 = subIn256(fee_growth_global_0, fee_growth_outside_upper_0)
        fee_growth_above_1 = subIn256(fee_growth_global_1, fee_growth_outside_upper_1)
    else:
        fee_growth_above_0 = fee_growth_outside_upper_0
        fee_growth_above_1 = fee_growth_outside_upper_1
    
    # Calculate fee growth inside the position range
    fee_growth_inside_0 = subIn256(
        subIn256(fee_growth_global_0, fee_growth_below_0),
        fee_growth_above_0
    )
    fee_growth_inside_1 = subIn256(
        subIn256(fee_growth_global_1, fee_growth_below_1),
        fee_growth_above_1
    )
    
    # Calculate uncollected fees
    Q128 = 2 ** 128
    uncollected_fees_0 = (liquidity * subIn256(fee_growth_inside_0, fee_growth_inside_last_0)) // Q128
    uncollected_fees_1 = (liquidity * subIn256(fee_growth_inside_1, fee_growth_inside_last_1)) // Q128
    
    return uncollected_fees_0, uncollected_fees_1


def get_position_creation_time(
    w3: Web3,
    position_manager_address: str,
    token_id: int,
    pool_address: str = None,
    tick_lower: int = None,
    tick_upper: int = None,
    from_block: int = 0
) -> int:
    """Get position creation timestamp using Etherscan V2 API or RPC fallback."""
    try:
        chain_id = w3.eth.chain_id
        api_key = os.getenv("ETHERSCAN_API_KEY")

        api_url = "https://api.etherscan.io/v2/api"
        
        # Base chain: no free Etherscan API, Alchemy RPC rejects getLogs
        if chain_id == 8453:
            return None

        # Try Etherscan V2 API first (Ethereum, Arbitrum)
        if api_key and chain_id in [1, 42161]:
            try:
                print(f"Using Etherscan V2 API (chain {chain_id}) for position {token_id} creation time")

                # Use getLogs to find the Transfer(address(0), to, tokenId) mint event
                transfer_topic = Web3.keccak(text="Transfer(address,address,uint256)").hex()
                if not transfer_topic.startswith('0x'):
                    transfer_topic = '0x' + transfer_topic
                # from = zero address (mint)
                zero_topic = '0x' + '0' * 64
                token_id_topic = '0x' + hex(token_id)[2:].zfill(64)

                params = {
                    "chainid": chain_id,
                    "module": "logs",
                    "action": "getLogs",
                    "address": position_manager_address,
                    "topic0": transfer_topic,
                    "topic1": zero_topic,
                    "topic3": token_id_topic,
                    "topic0_1_opr": "and",
                    "topic1_3_opr": "and",
                    "startblock": 1,
                    "endblock": 99999999999,
                    "sort": "asc",
                    "apikey": api_key
                }

                response = requests.get(api_url, params=params, timeout=30)

                if response.status_code == 200:
                    data = response.json()

                    if data.get("status") == "1" and data.get("result"):
                        log = data["result"][0]
                        timestamp = int(log["timeStamp"], 16) if log["timeStamp"].startswith("0x") else int(log["timeStamp"])
                        block_number = int(log["blockNumber"], 16) if log["blockNumber"].startswith("0x") else int(log["blockNumber"])
                        print(f"Position {token_id} minted at block {block_number}, timestamp {timestamp}")
                        return timestamp
                    else:
                        msg = data.get("message", "")
                        if "No records found" in msg or "No records found" in str(data.get("result", "")):
                            print(f"No mint event found for token {token_id}")
                        else:
                            print(f"Etherscan API: {msg} | {data.get('result', '')}")
            except Exception as e:
                print(f"Etherscan V2 API failed for token {token_id}: {e}")
                print("Falling back to RPC method...")

        # Fallback: RPC with chunked scanning
        current_block = w3.eth.block_number

        # Chain-aware lookback and chunk sizes
        if chain_id == 42161:  # Arbitrum
            lookback = 700_000  # ~2 days
            chunk_size = 5_000
        elif chain_id == 8453:  # Base
            lookback = 130_000  # ~3 days
            chunk_size = 2_000
        else:  # Ethereum
            lookback = 50_000  # ~7 days
            chunk_size = 50_000

        scan_from = max(1, current_block - lookback)

        print(f"Using RPC to search for position {token_id} creation from block {scan_from} to {current_block}")

        transfer_event_sig = w3.keccak(text="Transfer(address,address,uint256)").hex()
        token_id_topic = '0x' + hex(token_id)[2:].zfill(64)

        earliest_block = None

        while scan_from <= current_block:
            scan_to = min(scan_from + chunk_size - 1, current_block)

            try:
                logs = w3.eth.get_logs({
                    'fromBlock': scan_from,
                    'toBlock': scan_to,
                    'address': Web3.to_checksum_address(position_manager_address),
                    'topics': [
                        transfer_event_sig,
                        None,  # from
                        None,  # to
                        token_id_topic
                    ]
                })

                if logs:
                    block_num = min(log['blockNumber'] for log in logs)
                    if earliest_block is None or block_num < earliest_block:
                        earliest_block = block_num
                    break  # Found it, no need to scan further

            except Exception as e:
                if chunk_size > 500:
                    chunk_size = chunk_size // 2
                    continue
                else:
                    print(f"RPC scan failed for blocks {scan_from}-{scan_to}: {e}")
                    scan_from = scan_to + 1
                    continue

            scan_from = scan_to + 1

        if earliest_block:
            block = w3.eth.get_block(earliest_block)
            print(f"Position {token_id} minted at block {earliest_block}, timestamp {block['timestamp']}")
            return block['timestamp']

        print(f"Could not find creation event for position {token_id}")
        return 0

    except Exception as e:
        print(f"Error fetching position creation time for token {token_id}: {e}")
        return 0


def get_collected_fees_from_events(
    w3: Web3,
    position_manager_address: str,
    token_id: int,
    token0_decimals: int,
    token1_decimals: int
) -> tuple[float, float]:
    """Query NonfungiblePositionManager Collect events for a specific tokenId."""
    try:
        chain_id = w3.eth.chain_id
        api_key = os.getenv("ETHERSCAN_API_KEY")

        # Base chain: no free API support, Alchemy RPC rejects getLogs — skip entirely
        # Arbitrum: Alchemy free tier rejects getLogs for large ranges — skip RPC fallback
        if chain_id == 8453:
            return 0.0, 0.0

        # Use Etherscan V2 API for supported chains
        if api_key and chain_id in [1, 42161]:  # Ethereum, Arbitrum
            result = _get_collected_fees_etherscan(
                chain_id, api_key, position_manager_address, token_id,
                token0_decimals, token1_decimals
            )
            if result is not None:
                return result
            print("Etherscan API failed, falling back to RPC...")

        # Fallback: RPC log scanning (Ethereum only — Arbitrum/Base RPCs reject large getLogs)
        if chain_id == 42161:
            print(f"Skipping RPC fallback for Arbitrum (getLogs not supported on free tier)")
            return 0.0, 0.0

        return _get_collected_fees_rpc(
            w3, position_manager_address, token_id,
            token0_decimals, token1_decimals
        )

    except Exception as e:
        print(f"Warning: Could not fetch collected fees from events: {e}")
        return 0.0, 0.0


def _get_collected_fees_etherscan(
    chain_id: int,
    api_key: str,
    position_manager_address: str,
    token_id: int,
    token0_decimals: int,
    token1_decimals: int
) -> tuple[float, float] | None:
    """Fetch collected fees using Etherscan V2 API getLogs on NonfungiblePositionManager.

    NonfungiblePositionManager Collect event:
      Collect(uint256 indexed tokenId, address recipient, uint256 amount0Collect, uint256 amount1Collect)
    Topics: [event_sig, tokenId]
    Data: [recipient, amount0Collect, amount1Collect]
    """
    try:
        api_url = "https://api.etherscan.io/v2/api"
        collect_topic = Web3.keccak(text="Collect(uint256,address,uint256,uint256)").hex()
        if not collect_topic.startswith('0x'):
            collect_topic = '0x' + collect_topic
        token_id_topic = '0x' + hex(token_id)[2:].zfill(64)

        params = {
            "chainid": chain_id,
            "module": "logs",
            "action": "getLogs",
            "address": position_manager_address,
            "topic0": collect_topic,
            "topic1": token_id_topic,
            "topic0_1_opr": "and",
            "startblock": 1,
            "endblock": 99999999999,
            "sort": "asc",
            "apikey": api_key
        }

        response = requests.get(api_url, params=params, timeout=30)

        if response.status_code != 200:
            print(f"Etherscan API HTTP error: {response.status_code}")
            return None

        data = response.json()

        if data.get("status") != "1":
            msg = data.get("message", "")
            result_str = data.get("result", "")
            if "No records found" in msg or "No records found" in str(result_str):
                return 0.0, 0.0
            print(f"Etherscan API error: {msg} | result: {result_str}")
            return None

        total_collected_0 = 0
        total_collected_1 = 0

        for log in data["result"]:
            # Data layout: recipient (32 bytes) + amount0Collect (32 bytes) + amount1Collect (32 bytes)
            log_data = log["data"]
            if log_data.startswith("0x"):
                log_data = log_data[2:]

            # Skip recipient (first 64 hex chars = 32 bytes)
            amount0 = int(log_data[64:128], 16)
            amount1 = int(log_data[128:192], 16)

            total_collected_0 += amount0
            total_collected_1 += amount1

        collected_0 = total_collected_0 / (10 ** token0_decimals)
        collected_1 = total_collected_1 / (10 ** token1_decimals)

        print(f"Etherscan: found {len(data['result'])} Collect events for token {token_id}, fees {collected_0:.6f} token0, {collected_1:.6f} token1")
        return collected_0, collected_1

    except Exception as e:
        print(f"Etherscan collected fees lookup failed: {e}")
        return None


def _get_collected_fees_rpc(
    w3: Web3,
    position_manager_address: str,
    token_id: int,
    token0_decimals: int,
    token1_decimals: int
) -> tuple[float, float]:
    """Fallback: fetch collected fees via RPC getLogs on NonfungiblePositionManager.

    NonfungiblePositionManager Collect event:
      Collect(uint256 indexed tokenId, address recipient, uint256 amount0Collect, uint256 amount1Collect)
    Topics: [event_sig, tokenId]
    Data: [recipient, amount0Collect, amount1Collect]
    """
    try:
        collect_event_signature = w3.keccak(text="Collect(uint256,address,uint256,uint256)").hex()
        token_id_topic = '0x' + hex(token_id)[2:].zfill(64)
        current_block = w3.eth.block_number
        chain_id = w3.eth.chain_id

        # Conservative RPC lookback
        if chain_id == 42161:  # Arbitrum
            default_lookback = 700_000
            chunk_size = 5_000
        elif chain_id == 8453:  # Base
            default_lookback = 130_000
            chunk_size = 2_000
        else:  # Ethereum
            default_lookback = 50_000
            chunk_size = 50_000

        from_block = max(1, current_block - default_lookback)

        total_collected_0 = 0
        total_collected_1 = 0

        scan_from = from_block
        while scan_from <= current_block:
            scan_to = min(scan_from + chunk_size - 1, current_block)

            try:
                logs = w3.eth.get_logs({
                    'fromBlock': scan_from,
                    'toBlock': scan_to,
                    'address': Web3.to_checksum_address(position_manager_address),
                    'topics': [
                        collect_event_signature,
                        token_id_topic
                    ]
                })
            except Exception as chunk_err:
                if chunk_size > 500:
                    chunk_size = chunk_size // 2
                    continue
                else:
                    print(f"Warning: Could not fetch logs for blocks {scan_from}-{scan_to}: {chunk_err}")
                    scan_from = scan_to + 1
                    continue

            for log in logs:
                # Data: recipient (32 bytes) + amount0Collect (32 bytes) + amount1Collect (32 bytes)
                data = log['data'].hex() if isinstance(log['data'], bytes) else log['data']
                if data.startswith('0x'):
                    data = data[2:]

                # Skip recipient (first 64 hex chars)
                amount0 = int(data[64:128], 16)
                amount1 = int(data[128:192], 16)

                total_collected_0 += amount0
                total_collected_1 += amount1

            scan_from = scan_to + 1

        collected_0 = total_collected_0 / (10 ** token0_decimals)
        collected_1 = total_collected_1 / (10 ** token1_decimals)

        return collected_0, collected_1

    except Exception as e:
        print(f"Warning: RPC collected fees scan failed: {e}")
        return 0.0, 0.0


def check_lp_position(connector: UniswapV3Connector, token_id: int, chain: Chain) -> dict:
    """Check a single LP position and return data."""
    try:
        position_data = connector.position_manager.functions.positions(token_id).call()
        
        token0 = position_data[2]
        token1 = position_data[3]
        fee = position_data[4]
        tick_lower = position_data[5]
        tick_upper = position_data[6]
        liquidity = position_data[7]
        fee_growth_inside0_last = position_data[8]
        fee_growth_inside1_last = position_data[9]
        tokens_owed0 = position_data[10]
        tokens_owed1 = position_data[11]
        
        if liquidity == 0:
            return None
        
        token0_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(token0),
            abi=ERC20_ABI
        )
        token1_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(token1),
            abi=ERC20_ABI
        )
        
        token0_symbol = token0_contract.functions.symbol().call()
        token1_symbol = token1_contract.functions.symbol().call()
        token0_decimals = token0_contract.functions.decimals().call()
        token1_decimals = token1_contract.functions.decimals().call()
        
        factory_address = FACTORY_ADDRESSES[chain]
        factory = connector.w3.eth.contract(
            address=Web3.to_checksum_address(factory_address),
            abi=FACTORY_ABI
        )
        pool_address = factory.functions.getPool(token0, token1, fee).call()
        
        pool_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(pool_address),
            abi=POOL_ABI
        )
        
        slot0 = pool_contract.functions.slot0().call()
        sqrt_price_x96 = slot0[0]
        current_tick = slot0[1]
        
        # Get global fee growth and tick data
        fee_growth_global_0 = pool_contract.functions.feeGrowthGlobal0X128().call()
        fee_growth_global_1 = pool_contract.functions.feeGrowthGlobal1X128().call()
        
        tick_lower_data = pool_contract.functions.ticks(tick_lower).call()
        tick_upper_data = pool_contract.functions.ticks(tick_upper).call()
        
        fee_growth_outside_lower_0 = tick_lower_data[2]
        fee_growth_outside_lower_1 = tick_lower_data[3]
        fee_growth_outside_upper_0 = tick_upper_data[2]
        fee_growth_outside_upper_1 = tick_upper_data[3]
        
        # Calculate uncollected fees
        uncollected_fees_0, uncollected_fees_1 = calculate_uncollected_fees(
            liquidity,
            tick_lower,
            tick_upper,
            current_tick,
            fee_growth_global_0,
            fee_growth_global_1,
            fee_growth_outside_lower_0,
            fee_growth_outside_lower_1,
            fee_growth_outside_upper_0,
            fee_growth_outside_upper_1,
            fee_growth_inside0_last,
            fee_growth_inside1_last
        )
        
        # Get historical collected fees
        collected_fees_0, collected_fees_1 = get_collected_fees_from_events(
            connector.w3,
            connector.position_manager_address,
            token_id,
            token0_decimals,
            token1_decimals
        )
        
        # Get position creation time
        creation_timestamp = get_position_creation_time(
            connector.w3,
            connector.position_manager_address,
            token_id,
            pool_address,
            tick_lower,
            tick_upper
        )
        
        sqrt_price = sqrt_price_x96 / (2 ** 96)
        current_price_raw = sqrt_price ** 2
        decimal_adjustment = 10 ** (token0_decimals - token1_decimals)
        current_price_adjusted = current_price_raw * decimal_adjustment
        
        price_lower_raw = 1.0001 ** tick_lower
        price_upper_raw = 1.0001 ** tick_upper
        price_lower = price_lower_raw * decimal_adjustment
        price_upper = price_upper_raw * decimal_adjustment
        
        amount0, amount1 = calculate_token_amounts(
            liquidity, tick_lower, tick_upper, current_tick,
            sqrt_price_x96, token0_decimals, token1_decimals
        )
        
        price0_usd = get_token_price_usd(token0_symbol, token0, chain.value)
        price1_usd = get_token_price_usd(token1_symbol, token1, chain.value)
        
        value0_usd = amount0 * price0_usd
        value1_usd = amount1 * price1_usd
        total_value_usd = value0_usd + value1_usd
        
        # Calculate fee values
        total_fees_0 = uncollected_fees_0 + tokens_owed0
        total_fees_1 = uncollected_fees_1 + tokens_owed1
        
        fees_owed0 = total_fees_0 / (10 ** token0_decimals)
        fees_owed1 = total_fees_1 / (10 ** token1_decimals)
        
        fees0_usd = fees_owed0 * price0_usd
        fees1_usd = fees_owed1 * price1_usd
        total_fees_usd = fees0_usd + fees1_usd
        
        collected_fees_0_usd = collected_fees_0 * price0_usd
        collected_fees_1_usd = collected_fees_1 * price1_usd
        total_collected_fees_usd = collected_fees_0_usd + collected_fees_1_usd
        
        total_earned_fees_0 = collected_fees_0 + fees_owed0
        total_earned_fees_1 = collected_fees_1 + fees_owed1
        total_earned_fees_usd = total_collected_fees_usd + total_fees_usd
        
        # Calculate position age (after we have fee values for fallback estimation)
        current_time = datetime.now().timestamp()
        
        # Handle None (N/A) for Base positions
        if creation_timestamp is None:
            # Base chain - no age data available
            age_days = None
            age_hours = None
            age_display = "N/A"
            daily_apr = None
            monthly_apr = None
            daily_earnings = None
            print(f"Position {token_id}: Age not available (Base chain)")
        elif creation_timestamp > 0:
            age_seconds = current_time - creation_timestamp
            age_days = int(age_seconds // 86400)
            age_hours = int((age_seconds % 86400) // 3600)
            age_display = f"{age_days}d {age_hours}h"
            print(f"Position {token_id} created at {datetime.fromtimestamp(creation_timestamp)}")
            
            # Calculate APR
            daily_apr = 0.0
            monthly_apr = 0.0
            daily_earnings = 0.0
            
            fractional_days = age_days + age_hours / 24.0
            if fractional_days > 0.04 and total_value_usd > 0:
                # Daily earnings = total earned fees / fractional days
                daily_earnings = total_earned_fees_usd / fractional_days
                # Daily APR = (daily earnings / position value) * 100
                daily_apr = (daily_earnings / total_value_usd) * 100
                # Monthly APR = daily APR * 30
                monthly_apr = daily_apr * 30
        else:
            # Fallback: Use total earned fees to estimate age
            # Typical Uniswap V3 0.3% pool earns ~0.1-0.5% daily on active positions
            # Conservative estimate: if we have $100 in fees, position is at least 30-60 days old
            if total_earned_fees_usd > 10:
                # Estimate based on fees: assume 0.2% daily return
                estimated_days = (total_earned_fees_usd / total_value_usd) / 0.002 if total_value_usd > 0 else 30
                estimated_days = max(7, min(estimated_days, 365))  # Clamp between 7-365 days
                age_seconds = estimated_days * 86400
                print(f"Position {token_id}: Estimated age {int(estimated_days)} days based on ${total_earned_fees_usd:.2f} fees")
            elif collected_fees_0 > 0 or collected_fees_1 > 0:
                age_seconds = 30 * 86400  # 30 days if has collected fees
                print(f"Position {token_id}: Using 30 day estimate (has collected fees)")
            else:
                age_seconds = 7 * 86400  # 7 days for new positions
                print(f"Position {token_id}: Using 7 day estimate (new position)")
            
            age_days = int(age_seconds // 86400)
            age_hours = int((age_seconds % 86400) // 3600)
            age_display = f"{age_days}d {age_hours}h (est)"
            
            # Calculate APR
            daily_apr = 0.0
            monthly_apr = 0.0
            daily_earnings = 0.0
            
            fractional_days = age_days + age_hours / 24.0
            if fractional_days > 0.04 and total_value_usd > 0:
                # Daily earnings = total earned fees / fractional days
                daily_earnings = total_earned_fees_usd / fractional_days
                # Daily APR = (daily earnings / position value) * 100
                daily_apr = (daily_earnings / total_value_usd) * 100
                # Monthly APR = daily APR * 30
                monthly_apr = daily_apr * 30
        
        in_range = tick_lower <= current_tick <= tick_upper
        
        return {
            "token_id": token_id,
            "chain": chain.value,
            "pair": f"{token0_symbol}/{token1_symbol}",
            "token0_symbol": token0_symbol,
            "token1_symbol": token1_symbol,
            "amount0": amount0,
            "amount1": amount1,
            "price0_usd": price0_usd,
            "price1_usd": price1_usd,
            "value0_usd": value0_usd,
            "value1_usd": value1_usd,
            "total_value_usd": total_value_usd,
            "fees_owed0": fees_owed0,
            "fees_owed1": fees_owed1,
            "fees0_usd": fees0_usd,
            "fees1_usd": fees1_usd,
            "total_fees_usd": total_fees_usd,
            "collected_fees_0": collected_fees_0,
            "collected_fees_1": collected_fees_1,
            "collected_fees_0_usd": collected_fees_0_usd,
            "collected_fees_1_usd": collected_fees_1_usd,
            "total_collected_fees_usd": total_collected_fees_usd,
            "total_earned_fees_0": total_earned_fees_0,
            "total_earned_fees_1": total_earned_fees_1,
            "total_earned_fees_usd": total_earned_fees_usd,
            "age_days": age_days,
            "age_hours": age_hours,
            "daily_apr": daily_apr,
            "monthly_apr": monthly_apr,
            "daily_earnings": daily_earnings,
            "in_range": in_range,
            "fee_tier": fee / 10000,
            "current_price": current_price_adjusted,
            "price_lower": price_lower,
            "price_upper": price_upper
        }
        
    except Exception:
        return None


def check_aerodrome_slipstream_lp_position(
    connector: AerodromeSlipstreamConnector,
    token_id: int,
) -> dict | None:
    """Check a single Aerodrome Slipstream LP position and return enriched data.

    Mirrors check_lp_position() for Uniswap V3, but adapted for Aerodrome Slipstream's
    NonfungiblePositionManager interface:
      - positions() returns tickSpacing at index 4 (not fee)
      - Pools are resolved via PoolFactory.getPool(token0, token1, tickSpacing)
      - Pool's fee() is dynamic, set by SwapFeeModule

    Returns None if liquidity is 0 or on any error.
    """
    try:
        pm = connector.position_manager
        position_data = pm.functions.positions(token_id).call()

        # position_data layout: (nonce, operator, token0, token1, tickSpacing,
        #                        tickLower, tickUpper, liquidity,
        #                        feeGrowthInside0LastX128, feeGrowthInside1LastX128,
        #                        tokensOwed0, tokensOwed1)
        token0 = position_data[2]
        token1 = position_data[3]
        tick_spacing = position_data[4]
        tick_lower = position_data[5]
        tick_upper = position_data[6]
        liquidity = position_data[7]
        fee_growth_inside0_last = position_data[8]
        fee_growth_inside1_last = position_data[9]
        tokens_owed0 = position_data[10]
        tokens_owed1 = position_data[11]

        if liquidity == 0:
            return None

        # Token metadata
        token0_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(token0), abi=ERC20_ABI
        )
        token1_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(token1), abi=ERC20_ABI
        )
        token0_symbol = token0_contract.functions.symbol().call()
        token1_symbol = token1_contract.functions.symbol().call()
        token0_decimals = token0_contract.functions.decimals().call()
        token1_decimals = token1_contract.functions.decimals().call()

        # Resolve pool via Aerodrome factory (takes tickSpacing, not fee)
        pool_address = connector.pool_factory.functions.getPool(
            token0, token1, tick_spacing
        ).call()
        pool_contract = connector.w3.eth.contract(
            address=Web3.to_checksum_address(pool_address), abi=AERO_POOL_ABI
        )

        # Current pool state
        slot0 = pool_contract.functions.slot0().call()
        sqrt_price_x96 = slot0[0]
        current_tick = slot0[1]

        # Dynamic fee (may differ from Uniswap V3's fixed tiers)
        try:
            fee_raw = pool_contract.functions.fee().call()  # in hundredths of bps
        except Exception:
            fee_raw = 0

        # Fee growth globals
        fee_growth_global_0 = pool_contract.functions.feeGrowthGlobal0X128().call()
        fee_growth_global_1 = pool_contract.functions.feeGrowthGlobal1X128().call()

        # Tick data — Aerodrome returns 10 fields vs Uniswap's 8.
        # Fee growth outside is at indices 3,4 instead of 2,3.
        tick_lower_data = pool_contract.functions.ticks(tick_lower).call()
        tick_upper_data = pool_contract.functions.ticks(tick_upper).call()
        fee_growth_outside_lower_0 = tick_lower_data[3]
        fee_growth_outside_lower_1 = tick_lower_data[4]
        fee_growth_outside_upper_0 = tick_upper_data[3]
        fee_growth_outside_upper_1 = tick_upper_data[4]

        # Uncollected fees — same math as Uniswap V3 (shared helper)
        uncollected_fees_0, uncollected_fees_1 = calculate_uncollected_fees(
            liquidity, tick_lower, tick_upper, current_tick,
            fee_growth_global_0, fee_growth_global_1,
            fee_growth_outside_lower_0, fee_growth_outside_lower_1,
            fee_growth_outside_upper_0, fee_growth_outside_upper_1,
            fee_growth_inside0_last, fee_growth_inside1_last
        )

        # Collected fees from on-chain Collect events
        collected_fees_0, collected_fees_1 = get_collected_fees_from_events(
            connector.w3,
            connector.position_manager_address,
            token_id,
            token0_decimals,
            token1_decimals,
        )

        # Position creation time
        creation_timestamp = get_position_creation_time(
            connector.w3,
            connector.position_manager_address,
            token_id,
            pool_address,
            tick_lower,
            tick_upper,
        )

        # Price math — identical to Uniswap V3
        sqrt_price = sqrt_price_x96 / (2 ** 96)
        current_price_raw = sqrt_price ** 2
        decimal_adjustment = 10 ** (token0_decimals - token1_decimals)
        current_price_adjusted = current_price_raw * decimal_adjustment

        price_lower = (1.0001 ** tick_lower) * decimal_adjustment
        price_upper = (1.0001 ** tick_upper) * decimal_adjustment

        amount0, amount1 = calculate_token_amounts(
            liquidity, tick_lower, tick_upper, current_tick,
            sqrt_price_x96, token0_decimals, token1_decimals,
        )

        price0_usd = get_token_price_usd(token0_symbol, token0, "base")
        price1_usd = get_token_price_usd(token1_symbol, token1, "base")
        value0_usd = amount0 * price0_usd
        value1_usd = amount1 * price1_usd
        total_value_usd = value0_usd + value1_usd

        # Fee values (include any tokensOwed carried on the NFT)
        total_fees_0 = uncollected_fees_0 + tokens_owed0
        total_fees_1 = uncollected_fees_1 + tokens_owed1
        fees_owed0 = total_fees_0 / (10 ** token0_decimals)
        fees_owed1 = total_fees_1 / (10 ** token1_decimals)
        fees0_usd = fees_owed0 * price0_usd
        fees1_usd = fees_owed1 * price1_usd
        total_fees_usd = fees0_usd + fees1_usd

        collected_fees_0_usd = collected_fees_0 * price0_usd
        collected_fees_1_usd = collected_fees_1 * price1_usd
        total_collected_fees_usd = collected_fees_0_usd + collected_fees_1_usd

        # Resolve gauge for this pool — caller will use it to fetch pending AERO
        # for staked positions (it already knows which wallet owns the deposit
        # because it discovered this tokenId via get_staked_token_ids).
        gauge_address = connector.gauge_for_pool(pool_address)

        total_earned_fees_0 = collected_fees_0 + fees_owed0
        total_earned_fees_1 = collected_fees_1 + fees_owed1
        total_earned_fees_usd = total_collected_fees_usd + total_fees_usd

        # Age & APR (NOTE: APR recomputed by caller after AERO is folded in)
        current_time = datetime.now().timestamp()
        age_days = None
        age_hours = None
        daily_apr = None
        monthly_apr = None
        daily_earnings = None

        if creation_timestamp and creation_timestamp > 0:
            age_seconds = current_time - creation_timestamp
            age_days = int(age_seconds // 86400)
            age_hours = int((age_seconds % 86400) // 3600)
            fractional_days = age_days + age_hours / 24.0
            if fractional_days > 0.04 and total_value_usd > 0:
                daily_earnings = total_earned_fees_usd / fractional_days
                daily_apr = (daily_earnings / total_value_usd) * 100
                monthly_apr = daily_apr * 30

        in_range = tick_lower <= current_tick <= tick_upper

        # fee_raw is in 1/100 bps; convert to percent (e.g. 500 -> 0.05%)
        fee_tier_pct = (fee_raw / 10000) if fee_raw else 0

        return {
            "token_id": token_id,
            "chain": "base",
            "pair": f"{token0_symbol}/{token1_symbol}",
            "token0_symbol": token0_symbol,
            "token1_symbol": token1_symbol,
            "amount0": amount0,
            "amount1": amount1,
            "price0_usd": price0_usd,
            "price1_usd": price1_usd,
            "value0_usd": value0_usd,
            "value1_usd": value1_usd,
            "total_value_usd": total_value_usd,
            "fees_owed0": fees_owed0,
            "fees_owed1": fees_owed1,
            "fees0_usd": fees0_usd,
            "fees1_usd": fees1_usd,
            "total_fees_usd": total_fees_usd,
            "collected_fees_0": collected_fees_0,
            "collected_fees_1": collected_fees_1,
            "collected_fees_0_usd": collected_fees_0_usd,
            "collected_fees_1_usd": collected_fees_1_usd,
            "total_collected_fees_usd": total_collected_fees_usd,
            "total_earned_fees_0": total_earned_fees_0,
            "total_earned_fees_1": total_earned_fees_1,
            "total_earned_fees_usd": total_earned_fees_usd,
            "age_days": age_days,
            "age_hours": age_hours,
            "daily_apr": daily_apr,
            "monthly_apr": monthly_apr,
            "daily_earnings": daily_earnings,
            "in_range": in_range,
            "fee_tier": fee_tier_pct,
            "current_price": current_price_adjusted,
            "price_lower": price_lower,
            "price_upper": price_upper,
            "tick_spacing": tick_spacing,
            # Staking-reward placeholders — populated by caller when the position
            # is staked (caller already knows which wallet owns the staked deposit).
            "gauge_address": gauge_address,
            "reward_symbol": "AERO",
            "reward_decimals": 18,
            "reward_address": "",
            "reward_pending": 0.0,
            "reward_pending_usd": 0.0,
            "reward_price_usd": 0.0,
            "reward_claimed": 0.0,
            "reward_claimed_usd": 0.0,
            "reward_total_usd": 0.0,
        }

    except Exception as e:
        print(f"Aerodrome Slipstream position {token_id} fetch failed: {e}")
        return None


def apply_aerodrome_rewards(
    connector: AerodromeSlipstreamConnector,
    position_data: dict,
    wallet: str,
) -> None:
    """Fill in AERO staking-reward fields on an Aerodrome position dict (in place).

    Reads:
      - gauge.earned(wallet, tokenId) for currently-pending AERO
      - rewardToken() + ERC-20 metadata for the reward token (AERO on Base)
      - DB high-water-mark for claimed AERO (lp_aero_claimed_total field;
        snapshot_service uses drop-detection to grow this over time)

    Folds reward USD into total_earned_fees_usd / total_fees_usd so the existing
    APR math (later in the pipeline) treats AERO like swap fees.
    """
    gauge_address = position_data.get("gauge_address")
    token_id = position_data.get("token_id")
    if not gauge_address or token_id is None:
        return

    pending_raw = connector.get_pending_aero(gauge_address, wallet, token_id)
    if pending_raw <= 0 and not position_data.get("reward_claimed"):
        return  # not staked, nothing to add

    reward_token_addr = connector.get_reward_token(gauge_address) or ""
    reward_symbol = "AERO"
    reward_decimals = 18
    if reward_token_addr:
        try:
            erc20 = connector.w3.eth.contract(
                address=Web3.to_checksum_address(reward_token_addr), abi=ERC20_ABI
            )
            reward_symbol = erc20.functions.symbol().call() or "AERO"
            reward_decimals = int(erc20.functions.decimals().call() or 18)
        except Exception:
            pass

    pending = pending_raw / (10 ** reward_decimals)
    reward_price_usd = get_token_price_usd(reward_symbol, reward_token_addr, "base")
    pending_usd = pending * reward_price_usd

    # Claimed history: read high-water-mark for the AERO reward leg.
    # snapshot_service grows this when pending drops between snapshots.
    claimed = 0.0
    claimed_usd = 0.0
    try:
        from src.storage.portfolio_db import get_lp_aero_claimed
        claimed = get_lp_aero_claimed(1, str(token_id))
        claimed_usd = claimed * reward_price_usd
    except Exception:
        pass

    total_reward_usd = pending_usd + claimed_usd

    position_data["reward_symbol"] = reward_symbol
    position_data["reward_decimals"] = reward_decimals
    position_data["reward_address"] = reward_token_addr
    position_data["reward_pending"] = pending
    position_data["reward_pending_usd"] = pending_usd
    position_data["reward_price_usd"] = reward_price_usd
    position_data["reward_claimed"] = claimed
    position_data["reward_claimed_usd"] = claimed_usd
    position_data["reward_total_usd"] = total_reward_usd

    # Fold reward USD into the totals — APR math reads total_earned_fees_usd.
    position_data["total_fees_usd"] = (position_data.get("total_fees_usd") or 0) + pending_usd
    position_data["total_collected_fees_usd"] = (
        position_data.get("total_collected_fees_usd") or 0
    ) + claimed_usd
    position_data["total_earned_fees_usd"] = (
        position_data.get("total_earned_fees_usd") or 0
    ) + total_reward_usd


from src.models import STABLECOIN_SYMBOLS as STABLECOINS


def _calc_aave_liquidation_price(aave_data: dict) -> dict | None:
    """Calculate liquidation price for an AAVE position.
    
    Only meaningful when one side is volatile and the other is stable.
    Returns dict with price info or None if not calculable.
    """
    supplied = aave_data.get('supplied', [])
    borrowed = aave_data.get('borrowed', [])
    liq_threshold = aave_data.get('liquidation_threshold', 0) / 100  # Convert % to decimal
    
    if not supplied or not borrowed or liq_threshold == 0:
        return None
    
    # Classify assets as stable or volatile
    stable_supply = [s for s in supplied if s['symbol'] in STABLECOINS]
    volatile_supply = [s for s in supplied if s['symbol'] not in STABLECOINS]
    stable_borrow = [b for b in borrowed if b['symbol'] in STABLECOINS]
    volatile_borrow = [b for b in borrowed if b['symbol'] not in STABLECOINS]
    
    # Case 1: Volatile collateral, stable debt (e.g., ETH supplied, USDC borrowed)
    if volatile_supply and not volatile_borrow and stable_borrow:
        # Liquidation when: collateral_value * liq_threshold = debt_value
        # collateral_value = sum(amount * price), debt is stable (~fixed USD)
        total_debt_usd = sum(b.get('value_usd', 0) for b in stable_borrow)
        # If single volatile collateral, we can give a specific price
        if len(volatile_supply) == 1:
            s = volatile_supply[0]
            # liq_price * amount * liq_threshold = total_debt
            if s['balance'] > 0 and liq_threshold > 0:
                liq_price = total_debt_usd / (s['balance'] * liq_threshold)
                return {
                    "type": "collateral_drop",
                    "asset": s['symbol'],
                    "current_price": s.get('price_usd', 0),
                    "liquidation_price": liq_price,
                    "description": f"{s['symbol']} price must drop to"
                }
        # Multiple volatile collaterals — show as % drop needed
        total_volatile_collateral = sum(s.get('value_usd', 0) for s in volatile_supply)
        if total_volatile_collateral > 0:
            liq_ratio = total_debt_usd / (total_volatile_collateral * liq_threshold)
            pct_drop = (1 - liq_ratio) * 100
            return {
                "type": "collateral_drop_pct",
                "pct_drop": pct_drop,
                "description": "Collateral must drop by"
            }
    
    # Case 2: Stable collateral, volatile debt (e.g., USDC supplied, ETH borrowed)
    if stable_supply and not stable_borrow and volatile_borrow:
        total_collateral_usd = sum(s.get('value_usd', 0) for s in stable_supply)
        max_debt_usd = total_collateral_usd * liq_threshold
        if len(volatile_borrow) == 1:
            b = volatile_borrow[0]
            if b['balance'] > 0:
                liq_price = max_debt_usd / b['balance']
                return {
                    "type": "debt_rise",
                    "asset": b['symbol'],
                    "current_price": b.get('price_usd', 0),
                    "liquidation_price": liq_price,
                    "description": f"{b['symbol']} price must rise to"
                }
    
    # Both stable or both volatile — no simple liquidation price
    return None


def get_portfolio_data(force_refresh=False):
    """Fetch complete portfolio data with caching."""
    global _portfolio_cache
    
    # Return cached data if available and not forcing refresh
    if _portfolio_cache is not None and not force_refresh:
        return _portfolio_cache
    
    WALLET_ADDRESSES = get_wallet_addresses()
    
    # Load wallet config once for the entire function
    _wallet_config = load_wallet_config()
    
    def _label(addr):
        return _wallet_config.get(addr, {}).get('label', addr[:10] + '...')
    
    def _is_xpub(addr):
        return _wallet_config.get(addr, {}).get("type") == "bitcoin_xpub"
    
    evm_wallets = [w for w in WALLET_ADDRESSES if not _is_xpub(w)]
    xpub_wallets = [w for w in WALLET_ADDRESSES if _is_xpub(w)]
    
    def _btc_price_from_tokens(tokens):
        """Fallback: derive BTC price from WBTC/cbBTC in token list."""
        for t in tokens:
            if t.get("symbol") in ("WBTC", "cbBTC") and t.get("price_usd", 0) > 0:
                return t["price_usd"]
        return 0.0
    
    all_tokens = []
    all_lp_positions = []
    all_lending_positions = []
    all_gmx_positions = []
    total_tokens_value = 0.0
    total_lp_value = 0.0
    api_failures = []
    zerion_lp_groups = {}  # group_id -> {positions, wallet, wallet_label}
    
    zerion = ZerionConnector()
    zerion_configured = zerion.is_configured()
    zerion_not_configured_logged = False
    
    for wallet_index, wallet in enumerate(evm_wallets):
        wallet_label = _label(wallet)
        
        if zerion_configured:
            try:
                # Add inter-call delay for free tier rate limiting (skip for first wallet)
                if wallet_index > 0:
                    time.sleep(1)
                
                positions = zerion.get_wallet_positions(wallet)
                categorized = categorize_zerion_positions(positions)
                
                # Map tokens
                for pos in categorized['tokens']:
                    token = map_zerion_token_to_app(pos, wallet, wallet_label)
                    all_tokens.append(token)
                
                # Group lending positions by (chain, protocol) using group_id
                lending_groups = {}
                for pos in categorized['lending']:
                    attrs = pos.get("attributes", {})
                    group_id = attrs.get("group_id")
                    if group_id:
                        key = group_id
                    else:
                        chain_id = (
                            pos.get("relationships", {})
                            .get("chain", {})
                            .get("data", {})
                            .get("id", "unknown")
                        )
                        protocol = attrs.get("protocol", "unknown")
                        key = f"{chain_id}:{protocol}"
                    lending_groups.setdefault(key, []).append(pos)
                
                for group_key, group_positions in lending_groups.items():
                    lending = map_zerion_lending_to_app(group_positions, wallet, wallet_label)
                    all_lending_positions.append(lending)
                
                # Collect LP positions grouped by group_id for later processing
                for pos in categorized['lp_basic']:
                    attrs = pos.get("attributes", {})
                    group_id = attrs.get("group_id") or pos.get("id", f"unknown_{len(zerion_lp_groups)}")
                    group_key = (group_id, wallet)
                    if group_key not in zerion_lp_groups:
                        zerion_lp_groups[group_key] = {
                            "positions": [],
                            "wallet": wallet,
                            "wallet_label": wallet_label,
                        }
                    zerion_lp_groups[group_key]["positions"].append(pos)
                
            except Exception as e:
                print(f"Error fetching Zerion data for {wallet}: {e}")
                api_failures.append(f"zerion:{wallet}")
        else:
            if not zerion_not_configured_logged:
                api_failures.append("zerion:not_configured")
                zerion_not_configured_logged = True
    
    # Enrich Aave lending positions with on-chain data (health factor, APYs, liquidation)
    AAVE_NAMES = {"aave v3", "aave", "aave_v3", "aave-v3"}
    LENDING_CHAIN_TO_AAVE = {"ethereum": "ethereum", "arbitrum": "arbitrum", "base": "base"}
    enriched_lending = []
    aave_fetched = set()  # track (wallet, chain) already fetched

    for pos in all_lending_positions:
        protocol = (pos.get("protocol_name") or "").lower()
        chain_id = pos.get("chain", "")
        wallet = pos.get("wallet", "")
        aave_chain = LENDING_CHAIN_TO_AAVE.get(chain_id)

        if protocol in AAVE_NAMES and aave_chain and wallet:
            fetch_key = (wallet, aave_chain)
            if fetch_key in aave_fetched:
                continue  # already enriched for this wallet+chain
            aave_fetched.add(fetch_key)
            try:
                from src.models import get_web3, CHAIN_DISPLAY_NAMES
                w3 = get_web3(aave_chain)
                if w3:
                    aave_data = get_aave_positions(w3, wallet, aave_chain)
                    if aave_data:
                        # Add price info for liquidation calculation
                        for s in aave_data.get('supplied', []):
                            s['price_usd'] = get_token_price_usd(s['symbol'], s.get('token_address'), aave_chain)
                            s['value_usd'] = s['balance'] * s['price_usd']
                        for b in aave_data.get('borrowed', []):
                            b['price_usd'] = get_token_price_usd(b['symbol'], b.get('token_address'), aave_chain)
                            b['value_usd'] = b['balance'] * b['price_usd']

                        chain_name = CHAIN_DISPLAY_NAMES.get(aave_chain, aave_chain)
                        liq_price = _calc_aave_liquidation_price(aave_data)
                        enriched_lending.append({
                            "chain": aave_chain,
                            "chain_name": chain_name,
                            "protocol_name": "Aave V3",
                            "total_collateral_usd": aave_data["total_collateral_usd"],
                            "total_debt_usd": aave_data["total_debt_usd"],
                            "available_borrows_usd": aave_data["available_borrows_usd"],
                            "ltv": aave_data["ltv"],
                            "max_ltv": aave_data.get("max_ltv", 0),
                            "liquidation_threshold": aave_data["liquidation_threshold"],
                            "health_factor": aave_data["health_factor"],
                            "supplied": aave_data["supplied"],
                            "borrowed": aave_data["borrowed"],
                            "wallet": wallet,
                            "wallet_label": _label(wallet),
                            "liquidation_price": liq_price,
                        })
                        continue
            except Exception as e:
                print(f"Error enriching Aave data for {wallet} on {aave_chain}: {e}")
                api_failures.append(f"aave_enrich:{aave_chain}:{wallet[:8]}")
        # Keep Zerion data as fallback for non-Aave or failed enrichment
        enriched_lending.append(pos)

    all_lending_positions = enriched_lending

    # --- Check for previously-seen Aave positions that Zerion no longer reports ---
    # If a (wallet, chain) had lending snapshots before but Zerion didn't mention it
    # this time, do an on-chain check. If confirmed closed (collateral ≈ 0), write a
    # zero-balance snapshot so the DB reflects the closure.
    try:
        from src.storage.portfolio_db import get_connection as _get_conn
        _conn = _get_conn()
        prev_lending = _conn.execute("""
            SELECT DISTINCT wallet, chain FROM lending_account_snapshots
            WHERE total_collateral_usd > 1
        """).fetchall()
        _conn.close()

        for row in prev_lending:
            pw, pc = row['wallet'], row['chain']
            if (pw, pc) in aave_fetched:
                continue  # already checked via Zerion path
            # Zerion didn't report this — verify on-chain
            try:
                from src.models import get_web3
                w3 = get_web3(pc)
                if w3:
                    aave_data = get_aave_positions(w3, pw, pc)
                    coll = aave_data.get('total_collateral_usd', 0) if aave_data else 0
                    if coll < 1:
                        print(f"[Lending] Position {pw[:10]}... on {pc} confirmed closed (collateral ${coll:.2f})")
                    else:
                        # Still active — add to results
                        from src.models import CHAIN_DISPLAY_NAMES
                        for s in aave_data.get('supplied', []):
                            s['price_usd'] = get_token_price_usd(s['symbol'], s.get('token_address'), pc)
                            s['value_usd'] = s['balance'] * s['price_usd']
                        for b in aave_data.get('borrowed', []):
                            b['price_usd'] = get_token_price_usd(b['symbol'], b.get('token_address'), pc)
                            b['value_usd'] = b['balance'] * b['price_usd']
                        liq_price = _calc_aave_liquidation_price(aave_data)
                        all_lending_positions.append({
                            "chain": pc,
                            "chain_name": CHAIN_DISPLAY_NAMES.get(pc, pc),
                            "protocol_name": "Aave V3",
                            "total_collateral_usd": aave_data["total_collateral_usd"],
                            "total_debt_usd": aave_data["total_debt_usd"],
                            "available_borrows_usd": aave_data.get("available_borrows_usd", 0),
                            "ltv": aave_data["ltv"],
                            "max_ltv": aave_data.get("max_ltv", 0),
                            "liquidation_threshold": aave_data["liquidation_threshold"],
                            "health_factor": aave_data["health_factor"],
                            "supplied": aave_data["supplied"],
                            "borrowed": aave_data["borrowed"],
                            "wallet": pw,
                            "wallet_label": _label(pw),
                            "liquidation_price": liq_price,
                        })
                        print(f"[Lending] Position {pw[:10]}... on {pc} still active (${coll:,.0f}), added from on-chain")
                    aave_fetched.add((pw, pc))
            except Exception as e:
                print(f"[Lending] Could not verify {pw[:10]}... on {pc}: {e}")
    except Exception as e:
        print(f"[Lending] Stale position check failed: {e}")
    
    # Process LP positions discovered via Zerion
    # For Uniswap V3 / Aerodrome Slipstream: enrich with on-chain data (fees, range, APR)
    # For other protocols: use Zerion data as-is
    UNISWAP_V3_NAMES = {"uniswap v3", "uniswap_v3", "uniswap-v3"}
    AERODROME_SLIPSTREAM_NAMES = {
        "aerodrome v3", "aerodrome_v3", "aerodrome-v3",
        "aerodrome slipstream", "aerodrome cl",
    }
    CHAIN_TO_ENUM = {"ethereum": Chain.ETHEREUM, "arbitrum": Chain.ARBITRUM, "base": Chain.BASE}
    uniswap_v3_fetched = set()  # track (wallet, chain) pairs already fetched on-chain
    uniswap_v3_onchain_count = {}  # (wallet, chain) -> number of positions found on-chain
    uniswap_v3_zerion_seen = {}  # (wallet, chain) -> number of Zerion groups processed
    aero_fetched = set()
    aero_onchain_count = {}
    aero_zerion_seen = {}
    aero_seen_tids: dict = {}      # (wallet, "base") -> set of tokenIds already enriched
    aero_connector_cache: dict = {}  # (wallet, "base") -> AerodromeSlipstreamConnector

    for group_id, group_data in zerion_lp_groups.items():
        group_positions = group_data["positions"]
        wallet = group_data["wallet"]
        wallet_label = group_data["wallet_label"]

        first_attrs = group_positions[0].get("attributes", {})
        app_meta = first_attrs.get("application_metadata", {})
        protocol_name = (app_meta.get("name") or first_attrs.get("protocol", "")).lower()
        chain_id = (
            group_positions[0].get("relationships", {})
            .get("chain", {})
            .get("data", {})
            .get("id", "unknown")
        )
        chain_enum = CHAIN_TO_ENUM.get(chain_id)

        # Try on-chain enrichment for Uniswap V3 positions
        if protocol_name in UNISWAP_V3_NAMES and chain_enum is not None:
            fetch_key = (wallet, chain_id)
            if fetch_key in uniswap_v3_fetched:
                # Already fetched on-chain for this wallet+chain.
                # If Zerion reports more groups than on-chain positions found,
                # the extras are managed by contracts (e.g. Krystal) — use Zerion data.
                uniswap_v3_zerion_seen[fetch_key] = uniswap_v3_zerion_seen.get(fetch_key, 1) + 1
                if uniswap_v3_zerion_seen[fetch_key] > uniswap_v3_onchain_count.get(fetch_key, 0):
                    lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                    all_lp_positions.append(lp_data)
                    total_lp_value += lp_data["total_value_usd"]
                continue
            uniswap_v3_fetched.add(fetch_key)
            uniswap_v3_zerion_seen[fetch_key] = 1  # this is the first group
            try:
                connector = UniswapV3Connector(chain=chain_enum)
                raw_positions = connector.get_positions(wallet, chain_enum.value)
                uniswap_v3_onchain_count[fetch_key] = len(raw_positions)
                for raw_pos in raw_positions:
                    try:
                        position_data = check_lp_position(connector, raw_pos.token_id, chain_enum)
                        if position_data:
                            position_data['wallet'] = wallet
                            position_data['wallet_label'] = wallet_label
                            position_data['protocol'] = 'uniswap_v3'
                            all_lp_positions.append(position_data)
                            total_lp_value += position_data["total_value_usd"]
                    except Exception as e:
                        print(f"Error fetching LP position {raw_pos.token_id} on {chain_enum.value}: {e}")
                        api_failures.append(f"lp:{raw_pos.token_id}")
            except Exception as e:
                # Fallback to Zerion data if on-chain fetch fails
                print(f"On-chain Uniswap V3 fetch failed for {wallet} on {chain_id}, using Zerion data: {e}")
                lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                all_lp_positions.append(lp_data)
                total_lp_value += lp_data["total_value_usd"]
        elif protocol_name in AERODROME_SLIPSTREAM_NAMES and chain_id == "base":
            # Aerodrome Slipstream is Base-only; Zerion calls it "Aerodrome V3"
            # Staked positions are owned by the gauge, so we combine
            #   wallet-owned NFTs (positionManager.balanceOf)
            # + gauge-staked NFTs discovered via Zerion's pool_address per group.
            #
            # Discovery is keyed by (wallet, "base") so we only enumerate once per wallet,
            # but we still walk each group to collect pool addresses for gauge lookup.
            fetch_key = (wallet, "base")

            # Collect pool addresses Zerion reported for this group
            group_pools = {
                (p.get("attributes", {}).get("pool_address") or "").lower()
                for p in group_positions
                if p.get("attributes", {}).get("pool_address")
            }
            group_pools.discard("")

            if fetch_key in aero_fetched:
                # Already enriched this wallet — but this is a NEW group for the same
                # wallet (different pool). Look up its staked tokenIds and enrich them.
                try:
                    a_connector = aero_connector_cache.get(fetch_key)
                    if a_connector is None:
                        raise RuntimeError("connector missing from cache")
                    new_tids = []
                    for pa in group_pools:
                        new_tids.extend(a_connector.get_staked_token_ids(wallet, pa))
                    new_tids = [t for t in new_tids if t not in aero_seen_tids[fetch_key]]
                    aero_seen_tids[fetch_key].update(new_tids)
                    enriched_any = False
                    for tid in new_tids:
                        position_data = check_aerodrome_slipstream_lp_position(a_connector, tid)
                        if position_data:
                            position_data['wallet'] = wallet
                            position_data['wallet_label'] = wallet_label
                            position_data['protocol'] = 'aerodrome_slipstream'
                            position_data['protocol_display'] = 'Aerodrome Slipstream'
                            apply_aerodrome_rewards(a_connector, position_data, wallet)
                            all_lp_positions.append(position_data)
                            total_lp_value += position_data["total_value_usd"]
                            enriched_any = True
                    if not enriched_any:
                        # Fall back to Zerion data so the position still appears
                        lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                        all_lp_positions.append(lp_data)
                        total_lp_value += lp_data["total_value_usd"]
                except Exception as e:
                    print(f"Aerodrome staked enrichment failed for additional group: {e}")
                    lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                    all_lp_positions.append(lp_data)
                    total_lp_value += lp_data["total_value_usd"]
                continue

            aero_fetched.add(fetch_key)
            aero_seen_tids[fetch_key] = set()
            try:
                a_connector = AerodromeSlipstreamConnector()
                aero_connector_cache[fetch_key] = a_connector

                # Wallet-owned NFTs (unstaked)
                raw_positions = a_connector.get_positions(wallet, "base")
                token_ids = [r.token_id for r in raw_positions]

                # Gauge-staked NFTs for this group's pool(s)
                for pa in group_pools:
                    token_ids.extend(a_connector.get_staked_token_ids(wallet, pa))

                # Dedupe while preserving order
                seen = set()
                unique_ids = []
                for t in token_ids:
                    if t not in seen:
                        seen.add(t)
                        unique_ids.append(t)
                aero_seen_tids[fetch_key].update(unique_ids)
                aero_onchain_count[fetch_key] = len(unique_ids)

                enriched_any = False
                for tid in unique_ids:
                    try:
                        position_data = check_aerodrome_slipstream_lp_position(a_connector, tid)
                        if position_data:
                            position_data['wallet'] = wallet
                            position_data['wallet_label'] = wallet_label
                            position_data['protocol'] = 'aerodrome_slipstream'
                            position_data['protocol_display'] = 'Aerodrome Slipstream'
                            apply_aerodrome_rewards(a_connector, position_data, wallet)
                            all_lp_positions.append(position_data)
                            total_lp_value += position_data["total_value_usd"]
                            enriched_any = True
                    except Exception as e:
                        print(f"Error fetching Aerodrome LP position {tid}: {e}")
                        api_failures.append(f"lp:{tid}")

                if not enriched_any:
                    # No on-chain match (e.g. RPC failures) — fall back to Zerion data
                    lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                    all_lp_positions.append(lp_data)
                    total_lp_value += lp_data["total_value_usd"]
            except Exception as e:
                print(f"On-chain Aerodrome Slipstream fetch failed for {wallet}, using Zerion data: {e}")
                lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
                all_lp_positions.append(lp_data)
                total_lp_value += lp_data["total_value_usd"]
        else:
            # Other protocols — use Zerion data directly
            lp_data = map_zerion_lp_to_app(group_positions, wallet, wallet_label)
            all_lp_positions.append(lp_data)
            total_lp_value += lp_data["total_value_usd"]
    
    # Enrich LP positions missing age data using Zerion transactions API
    # For on-chain positions we have token_id — match by NFT token_id in transaction transfers
    # Group by (wallet, chain) for batched API calls
    if zerion_configured:
        age_needed: dict = {}  # (wallet, chain) -> list of (index, token_id)
        for idx, lp in enumerate(all_lp_positions):
            if lp.get("age_days") is None and lp.get("wallet") and lp.get("chain"):
                tid = lp.get("token_id")
                if tid is not None:
                    key = (lp["wallet"], lp["chain"])
                    age_needed.setdefault(key, []).append((idx, str(tid)))

        for (wallet, chain), positions_info in age_needed.items():
            if not positions_info:
                continue
            try:
                time.sleep(1)  # Rate limit
                txns = zerion.get_wallet_transactions(
                    wallet,
                    chain_ids=chain,
                    operation_types="deposit,mint",
                    sort="mined_at",  # oldest first
                    limit=100,
                )

                # Build a token_id -> {mined_at, entry_value} map from transaction transfers
                tid_to_mint: dict = {}
                for tx in txns:
                    tx_attrs = tx.get("attributes", {})
                    mined_at = tx_attrs.get("mined_at", "")
                    if not mined_at:
                        continue
                    transfers = tx_attrs.get("transfers", [])
                    # Check if this transaction involves an NFT we care about
                    tx_nft_tid = None
                    for tr in transfers:
                        nft = tr.get("nft_info", {})
                        nft_tid = str(nft.get("token_id", ""))
                        if nft_tid:
                            tx_nft_tid = nft_tid
                            break
                    if tx_nft_tid and tx_nft_tid not in tid_to_mint:
                        # Sum the USD value of fungible transfers (the deposited tokens)
                        entry_value = 0.0
                        for tr in transfers:
                            if tr.get("nft_info"):
                                continue  # skip the NFT transfer itself
                            tr_value = abs(tr.get("value") or 0.0)
                            entry_value += tr_value
                        tid_to_mint[tx_nft_tid] = {"mined_at": mined_at, "entry_value": entry_value}

                # Apply to positions
                for idx, tid in positions_info:
                    mint_data = tid_to_mint.get(tid)
                    if not mint_data:
                        continue
                    earliest_ts = mint_data["mined_at"]
                    try:
                        from datetime import datetime as dt
                        ts = dt.fromisoformat(earliest_ts.replace("Z", "+00:00"))
                        creation_epoch = ts.timestamp()
                        age_seconds = datetime.now().timestamp() - creation_epoch
                        age_days = int(age_seconds // 86400)
                        age_hours = int((age_seconds % 86400) // 3600)

                        lp = all_lp_positions[idx]
                        lp["age_days"] = age_days
                        lp["age_hours"] = age_hours

                        # Entry value from the mint transaction
                        if mint_data["entry_value"] > 0:
                            lp["entry_value_usd"] = mint_data["entry_value"]

                        print(f"Zerion txns: Position #{tid} on {chain} created {age_days}d ago")
                    except Exception as e:
                        print(f"Error parsing Zerion timestamp for #{tid}: {e}")

            except Exception as e:
                print(f"Error fetching Zerion transactions for {wallet[:8]} on {chain}: {e}")

    # Fetch GMX V2 perpetual positions (Arbitrum only)
    from src.models import get_web3 as _get_web3
    arb_w3 = _get_web3("arbitrum")
    if arb_w3:
        for wallet in evm_wallets:
            try:
                from src.connectors.gmx_v2 import get_gmx_positions
                gmx_pos = get_gmx_positions(arb_w3, wallet)
                for p in gmx_pos:
                    p['wallet'] = wallet
                    p['wallet_label'] = _label(wallet)
                    current_price = get_token_price_usd(p['index_symbol'], None, "arbitrum")
                    if current_price == 0:
                        cg_map = {"WETH": "ETH", "WBTC": "BTC", "BTC": "BTC", "ETH": "ETH"}
                        mapped = cg_map.get(p['index_symbol'], p['index_symbol'])
                        if mapped == "BTC":
                            current_price = get_token_price_usd("BTC", None, "bitcoin")
                        else:
                            current_price = get_token_price_usd(mapped, None, "ethereum")
                    # Fallback: derive from Zerion token prices already fetched
                    if current_price == 0:
                        if p['index_symbol'] in ("BTC", "WBTC"):
                            current_price = _btc_price_from_tokens(all_tokens)
                        else:
                            for t in all_tokens:
                                if t.get("symbol") == p['index_symbol'] and t.get("price_usd", 0) > 0:
                                    current_price = t["price_usd"]
                                    break
                    p['current_price'] = current_price
                    if p['entry_price'] > 0 and current_price > 0:
                        if p['is_long']:
                            p['pnl_usd'] = p['size_usd'] * (current_price - p['entry_price']) / p['entry_price']
                        else:
                            p['pnl_usd'] = p['size_usd'] * (p['entry_price'] - current_price) / p['entry_price']
                        p['pnl_pct'] = (p['pnl_usd'] / p['collateral_amount']) * 100 if p['collateral_amount'] > 0 else 0
                    else:
                        p['pnl_usd'] = 0
                        p['pnl_pct'] = 0
                    # Estimate liquidation price (simplified: when losses = collateral)
                    if p['entry_price'] > 0 and p['size_usd'] > 0:
                        liq_move = p['collateral_amount'] / p['size_usd'] * p['entry_price']
                        if p['is_long']:
                            p['liquidation_price'] = p['entry_price'] - liq_move
                        else:
                            p['liquidation_price'] = p['entry_price'] + liq_move
                    else:
                        p['liquidation_price'] = 0
                    all_gmx_positions.append(p)
                if gmx_pos:
                    print(f"GMX: Found {len(gmx_pos)} positions for {_label(wallet)}")
            except Exception as e:
                print(f"Error fetching GMX positions: {e}")

    # Fetch Bitcoin balances from xpub wallets
    for wallet_key in xpub_wallets:
        try:
            from src.connectors.bitcoin import get_btc_balance_for_xpub
            print(f"Fetching BTC balance for xpub wallet: {_label(wallet_key)}")
            btc_data = get_btc_balance_for_xpub(wallet_key)
            btc_balance = btc_data["total_btc"]
            if btc_balance > 0:
                btc_price = get_token_price_usd("BTC", None, "bitcoin")
                # Fallback: derive from WBTC/cbBTC price already fetched via Zerion
                if btc_price == 0:
                    btc_price = _btc_price_from_tokens(all_tokens)
                btc_value = btc_balance * btc_price
                all_tokens.append({
                    "chain": "Bitcoin",
                    "symbol": "BTC",
                    "balance": btc_balance,
                    "value_usd": btc_value,
                    "price_usd": btc_price,
                    "wallet": wallet_key,
                    "wallet_label": _label(wallet_key)
                })
                total_tokens_value += btc_value
                print(f"BTC balance: {btc_balance:.8f} BTC (${btc_value:.2f}), {btc_data['addresses_used']} addresses used")
        except Exception as e:
            print(f"Error fetching BTC balance: {e}")
    
    # Sort tokens by value descending
    all_tokens.sort(key=lambda x: x["value_usd"], reverse=True)
    
    # Enrich LP positions with DB high water mark for collected fees
    # The high water mark tracks the maximum total_earned_fees_usd ever recorded.
    # collected = high_water_mark - current_uncollected
    try:
        from src.storage.portfolio_db import get_lp_fee_high_water_mark
        for lp in all_lp_positions:
            if lp.get('total_collected_fees_usd', 0) > 0:
                continue  # already has on-chain collected fees
            position_id = str(lp.get('token_id') or '')
            if not position_id:
                continue
            hwm = get_lp_fee_high_water_mark(1, position_id)
            if hwm <= 0:
                continue
            uncollected = lp.get('total_fees_usd', 0)
            total_earned = max(hwm, uncollected)
            collected = max(0, total_earned - uncollected)
            if collected > 0 or total_earned > uncollected:
                lp['total_earned_fees_usd'] = total_earned
                lp['total_collected_fees_usd'] = collected
                lp['collected_fees_0_usd'] = collected / 2
                lp['collected_fees_1_usd'] = collected / 2
    except Exception as e:
        print(f"Error enriching LP fees from DB: {e}")

    # Recalculate APR for all positions using consistent fractional-days formula
    # daily_apr = (total_earned_fees / fractional_days) / position_value * 100
    for lp in all_lp_positions:
        age_d = lp.get('age_days')
        age_h = lp.get('age_hours', 0) or 0
        total_earned = lp.get('total_earned_fees_usd', 0)
        total_value = lp.get('total_value_usd', 0)
        if age_d is not None and total_value > 0 and total_earned > 0:
            fractional_days = age_d + age_h / 24.0
            if fractional_days > 0.04:  # at least ~1 hour
                daily_earnings = total_earned / fractional_days
                lp['daily_earnings'] = daily_earnings
                lp['daily_apr'] = (daily_earnings / total_value) * 100
                lp['monthly_apr'] = lp['daily_apr'] * 30

    # Filter out Zerion LP positions the user has archived
    try:
        from src.storage.portfolio_db import get_db_path
        import sqlite3 as _sqlite3
        _db = _sqlite3.connect(get_db_path())
        hidden_keys = set(r[0] for r in _db.execute(
            "SELECT position_key FROM zerion_lp_hidden"
        ).fetchall())
        _db.close()
        if hidden_keys:
            def _make_zerion_lp_key(pos):
                return 'zerion-' + '-'.join([
                    pos.get('wallet', ''),
                    pos.get('protocol', ''),
                    pos.get('chain', ''),
                    pos.get('pair', ''),
                ])
            all_lp_positions = [p for p in all_lp_positions
                                 if _make_zerion_lp_key(p) not in hidden_keys]
    except Exception as _e:
        print(f"Warning: could not filter zerion_lp_hidden: {_e}")

    # --- Post-process: reconstruct borrowed data from variableDebt/stableDebt tokens,
    #     and remove receipt/yield tokens that aren't spendable holdings.
    try:
        _DEBT_PREFIXES = ('variableDebt', 'stableDebt')
        _RECEIPT_PREFIXES = ('aBas', 'aEth', 'aArb', 'aOpt', 'steak', 'variableDebt', 'stableDebt')
        _CHAIN_PREFIXES = ['Bas', 'Eth', 'Arb', 'Opt', 'Pol', 'Ava', 'Bnb', 'Gno']

        def _parse_debt_symbol(debt_sym):
            sym = debt_sym
            for p in _DEBT_PREFIXES:
                if sym.startswith(p):
                    sym = sym[len(p):]
                    break
            for cp in _CHAIN_PREFIXES:
                if sym.startswith(cp):
                    sym = sym[len(cp):]
                    break
            return sym or debt_sym

        def _find_token_price(symbol, tokens):
            for t in tokens:
                if t.get('symbol') == symbol and (t.get('price_usd') or 0) > 0:
                    return t['price_usd']
            return 0.0

        # Build wallet -> lending positions lookup
        _wallet_to_lending = {}
        for _lpos in all_lending_positions:
            _w = _lpos.get('wallet', '')
            _wallet_to_lending.setdefault(_w, []).append(_lpos)

        # Group debt tokens by wallet (index preserved for later removal)
        _wallet_debt_tokens = {}
        for _i, _tok in enumerate(all_tokens):
            if _tok.get('symbol', '').startswith(_DEBT_PREFIXES):
                _wallet_debt_tokens.setdefault(_tok.get('wallet', ''), []).append((_i, _tok))

        # Reconstruct borrowed arrays for positions that have none
        for _wallet, _debt_list in _wallet_debt_tokens.items():
            for _lpos in _wallet_to_lending.get(_wallet, []):
                if _lpos.get('borrowed'):
                    continue  # on-chain data already present
                for _idx, _tok in _debt_list:
                    _underlying = _parse_debt_symbol(_tok.get('symbol', ''))
                    _price = _find_token_price(_underlying, all_tokens)
                    _balance = _tok.get('balance', 0)
                    _value = _balance * _price
                    _lpos.setdefault('borrowed', []).append({
                        'symbol': _underlying,
                        'balance': _balance,
                        'value_usd': _value,
                        'borrow_apy': 0,
                        'variable_borrow_apy': 0,
                        'token_address': '',
                    })
                    _lpos['total_debt_usd'] = (_lpos.get('total_debt_usd') or 0) + _value
                # Recalculate health factor and available borrows after all debt added
                _total_coll = _lpos.get('total_collateral_usd', 0)
                _total_debt = _lpos.get('total_debt_usd', 0)
                if _total_debt > 0:
                    _liq_thresh = _lpos.get('liquidation_threshold') or 0.825
                    _ltv = _lpos.get('ltv') or 0.75
                    _lpos['health_factor'] = (_total_coll * _liq_thresh) / _total_debt
                    _lpos['available_borrows_usd'] = max(0, (_total_coll * _ltv) - _total_debt)

        # Remove receipt/debt tokens from the visible token list
        all_tokens = [
            t for t in all_tokens
            if not t.get('symbol', '').startswith(_RECEIPT_PREFIXES)
        ]
    except Exception as _e:
        print(f"Warning: debt token reconstruction failed: {_e}")

    # Calculate totals
    total_tokens_value = sum(t["value_usd"] for t in all_tokens)
    total_uncollected_fees = sum(pos.get('total_fees_usd', 0) for pos in all_lp_positions)
    total_hedge_value = sum(p.get('collateral_amount', 0) for p in all_gmx_positions)
    
    # Get unique wallet labels for filtering
    wallet_labels = {}
    for wallet in WALLET_ADDRESSES:
        wallet_labels[wallet] = _label(wallet)
    
    result = {
        "tokens": all_tokens,
        "lp_positions": all_lp_positions,
        "aave_positions": all_lending_positions,
        "gmx_positions": all_gmx_positions,
        "total_tokens_value": total_tokens_value,
        "total_lp_value": total_lp_value,
        "total_uncollected_fees": total_uncollected_fees,
        "total_value": total_tokens_value + total_lp_value + total_uncollected_fees + total_hedge_value,
        "wallet_count": len(WALLET_ADDRESSES),
        "wallet_labels": wallet_labels,
        "api_failures": api_failures,
        "fetched_at": datetime.now().isoformat(),
    }
    
    # Cache the result
    _portfolio_cache = result
    
    return result


@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """Login page, or first-time setup if no password exists."""
    import secrets as _secrets
    pw_hash = get_password_hash()

    def _check_csrf():
        return request.form.get('csrf_token') == session.get('_csrf_token')

    def _set_csrf():
        token = _secrets.token_hex(32)
        session['_csrf_token'] = token
        return token

    # No password set — show one-time setup page
    if not pw_hash:
        error = None
        if request.method == 'POST':
            if not _check_csrf():
                error = "Invalid request. Please try again."
            else:
                password = request.form.get('password', '')
                confirm = request.form.get('confirm', '')
                if len(password) < 6:
                    error = "Password must be at least 6 characters"
                elif password != confirm:
                    error = "Passwords don't match"
                else:
                    new_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    set_key(ENV_FILE, "APP_PASSWORD_HASH", new_hash)
                    session['authenticated'] = True
                    session.permanent = True
                    return redirect(url_for('index'))
        return render_template('setup.html', error=error, csrf_token=_set_csrf())

    # Already authenticated
    if session.get('authenticated'):
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        if not _check_csrf():
            error = "Invalid request. Please try again."
        else:
            ip = request.remote_addr
            if not check_rate_limit(ip):
                error = "Too many attempts. Try again in 5 minutes."
            else:
                password = request.form.get('password', '')
                if bcrypt.checkpw(password.encode('utf-8'), pw_hash.encode('utf-8')):
                    session['authenticated'] = True
                    session.permanent = True
                    _login_attempts.pop(ip, None)
                    return redirect(url_for('index'))
                else:
                    error = "Invalid password"

    return render_template('login.html', error=error, csrf_token=_set_csrf())


@app.route('/logout')
def logout():
    """Log out and clear session."""
    session.clear()
    return redirect(url_for('login_page'))


@app.route('/api/change-password', methods=['POST'])
def api_change_password():
    """Change the app password."""
    data = request.json
    current = data.get('current', '')
    new_pw = data.get('new', '')

    if not new_pw or len(new_pw) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    pw_hash = get_password_hash()
    # If a password is set, verify current password
    if pw_hash and not bcrypt.checkpw(current.encode('utf-8'), pw_hash.encode('utf-8')):
        return jsonify({"error": "Current password is incorrect"}), 403

    # Generate new hash and save to .env
    new_hash = bcrypt.hashpw(new_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    set_key(ENV_FILE, "APP_PASSWORD_HASH", new_hash)

    return jsonify({"status": "success", "message": "Password updated"})


@app.route('/')
def index():
    """Render the unified dashboard."""
    return render_template('index.html')


@app.route('/settings')
def settings():
    """Render the settings page."""
    return render_template('settings.html')


@app.route('/api/market/lending-rates')
def api_lending_rates():
    """Get current Aave V3 lending/borrow rates from on-chain."""
    from src.connectors.aave_v3 import get_aave_market_rates
    from src.models import get_web3 as _gw3
    results = {}
    for chain_name in ['arbitrum', 'base', 'ethereum']:
        w3 = _gw3(chain_name)
        if w3:
            try:
                rates = get_aave_market_rates(w3, chain_name)
                for r in rates:
                    results[f"{chain_name}_{r['asset'].lower()}"] = {
                        'chain': chain_name, 'asset': r['asset'],
                        'supply': r['supply_apy'], 'borrow': r['borrow_apy'],
                    }
            except Exception:
                pass
    return jsonify(results)


@app.route('/api/market-data')
def api_market_data():
    """Return latest market data from DB — single source of truth for frontend + AI."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()

    # Latest snapshot with prices — single source, no merging from older snapshots
    snap = conn.execute(
        "SELECT * FROM market_snapshots WHERE btc_price IS NOT NULL ORDER BY timestamp DESC LIMIT 1"
    ).fetchone()
    if not snap:
        snap = conn.execute("SELECT * FROM market_snapshots ORDER BY timestamp DESC LIMIT 1").fetchone()

    data = dict(snap) if snap else {}

    # Calculate 24h price changes from DB history
    if data.get('btc_price'):
        prev = conn.execute(
            "SELECT btc_price, eth_price FROM market_snapshots WHERE btc_price IS NOT NULL "
            "AND datetime(timestamp) BETWEEN datetime('now', '-28 hours') AND datetime('now', '-20 hours') "
            "ORDER BY ABS(julianday('now') - julianday(timestamp) - 1.0) ASC LIMIT 1"
        ).fetchone()
        if prev and prev['btc_price']:
            data['btc_24h_change'] = ((data['btc_price'] - prev['btc_price']) / prev['btc_price']) * 100
        if prev and prev['eth_price'] and data.get('eth_price'):
            data['eth_24h_change'] = ((data['eth_price'] - prev['eth_price']) / prev['eth_price']) * 100

    # Latest defi rates (lending + LP pools)
    latest_rate_ts = conn.execute("SELECT MAX(timestamp) as ts FROM defi_rates").fetchone()
    lending = {}
    lp_pools = {}
    if latest_rate_ts and latest_rate_ts['ts']:
        for r in conn.execute("SELECT * FROM defi_rates WHERE timestamp=? AND rate_type='lending'", (latest_rate_ts['ts'],)).fetchall():
            lending[f"{r['chain']}_{r['asset']}"] = {
                'chain': r['chain'], 'asset': r['asset'],
                'supply': r['supply_apy'], 'borrow': r['borrow_apy'],
            }
        for r in conn.execute("SELECT * FROM defi_rates WHERE timestamp=? AND rate_type='lp' AND fee_apr > 0 ORDER BY tvl DESC", (latest_rate_ts['ts'],)).fetchall():
            r = dict(r)
            lp_pools[f"{r['chain']}_{r['asset']}"] = {
                'chain': r['chain'], 'asset': r['asset'],
                'apyBase': r['fee_apr'], 'apyReward': r.get('reward_apr') or 0,
                'tvl': r['tvl'] or 0, 'vol1d': r.get('volume_1d') or 0,
            }

    # Stablecoin 7d change
    sc_current = conn.execute("SELECT stablecoin_supply FROM market_snapshots WHERE stablecoin_supply IS NOT NULL ORDER BY timestamp DESC LIMIT 1").fetchone()
    sc_prior = conn.execute("SELECT stablecoin_supply FROM market_snapshots WHERE stablecoin_supply IS NOT NULL AND timestamp <= datetime('now', '-5 days') ORDER BY timestamp DESC LIMIT 1").fetchone()
    sc_7d_change = None
    if sc_current and sc_prior and sc_prior['stablecoin_supply']:
        sc_7d_change = ((sc_current['stablecoin_supply'] - sc_prior['stablecoin_supply']) / sc_prior['stablecoin_supply']) * 100

    # Fear & Greed history (from recent snapshots)
    fg_rows = conn.execute(
        "SELECT fear_greed_index FROM market_snapshots WHERE fear_greed_index IS NOT NULL ORDER BY timestamp DESC LIMIT 30"
    ).fetchall()
    fg_history = [r['fear_greed_index'] for r in fg_rows]

    # Macro indicators
    macro_rows = []
    try:
        macro_result = fetch_fred_macro()
        if macro_result and macro_result.get('rows'):
            macro_rows = macro_result['rows']
    except Exception:
        pass

    conn.close()

    return jsonify({
        'snapshot': data,
        'lending': lending,
        'lp_pools': lp_pools,
        'stablecoin_7d_change': sc_7d_change,
        'fg_history': fg_history,
        'macro': macro_rows,
    })


@app.route('/api/market-data/refresh', methods=['POST'])
def api_market_data_refresh():
    """Trigger a fresh market snapshot, then return updated data."""
    try:
        from src.engines.snapshot_service import take_market_snapshot
        take_market_snapshot('manual')
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return api_market_data()


@app.route('/api/market/snapshot', methods=['POST'])
def api_market_snapshot():
    """Trigger a market data snapshot to DB."""
    import threading
    def _bg():
        try:
            from src.engines.snapshot_service import take_market_snapshot
            take_market_snapshot('manual')
        except Exception as e:
            print(f"Manual market snapshot error: {e}")
    threading.Thread(target=_bg, daemon=True, name='manual-market-snapshot').start()
    return jsonify({"status": "started"})


@app.route('/api/market/stablecoin-7d')
def api_stablecoin_7d():
    """Get stablecoin supply 7d change from DB."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    current = conn.execute("SELECT stablecoin_supply FROM market_snapshots WHERE stablecoin_supply IS NOT NULL ORDER BY timestamp DESC LIMIT 1").fetchone()
    prior = conn.execute("SELECT stablecoin_supply FROM market_snapshots WHERE stablecoin_supply IS NOT NULL AND timestamp <= datetime('now', '-5 days') ORDER BY timestamp DESC LIMIT 1").fetchone()
    conn.close()
    if current and prior and prior['stablecoin_supply'] > 0:
        change = ((current['stablecoin_supply'] - prior['stablecoin_supply']) / prior['stablecoin_supply']) * 100
        return jsonify({"change_pct": change, "current": current['stablecoin_supply'], "prior": prior['stablecoin_supply']})
    return jsonify({"change_pct": None})


@app.route('/api/market/stablecoin-supply')
def api_stablecoin_supply():
    """Proxy DefiLlama stablecoin chains endpoint (avoids CORS)."""
    try:
        resp = requests.get('https://stablecoins.llama.fi/stablecoinchains', timeout=15)
        resp.raise_for_status()
        return jsonify(resp.json())
    except Exception as e:
        return jsonify({"error": str(e)}), 502


# --- FRED Macro Data (24h in-memory cache) ---
_fred_cache = {"data": None, "fetched_at": None}
FRED_CACHE_TTL = 86400  # 24 hours

FRED_BASE_URL = "https://api.stlouisfed.org/fred/series/observations"
FOMC_2026 = [
    __import__('datetime').date(2026, 1, 28), __import__('datetime').date(2026, 3, 18),
    __import__('datetime').date(2026, 4, 29), __import__('datetime').date(2026, 6, 17),
    __import__('datetime').date(2026, 7, 29), __import__('datetime').date(2026, 9, 16),
    __import__('datetime').date(2026, 10, 28), __import__('datetime').date(2026, 12, 9),
]


def _fred_fetch(api_key, series_id, start=None):
    """Fetch FRED observations. Returns list of {date, value}."""
    params = {"series_id": series_id, "api_key": api_key, "file_type": "json", "sort_order": "desc"}
    if start:
        params["observation_start"] = start
    else:
        params["limit"] = 1
    try:
        resp = requests.get(FRED_BASE_URL, params=params, timeout=15)
        resp.raise_for_status()
        return [{"date": o["date"], "value": float(o["value"])}
                for o in resp.json().get("observations", []) if o.get("value") not in (".", "")]
    except Exception as e:
        print(f"FRED fetch error ({series_id}): {e}")
        return []


def fetch_fred_macro(force=False):
    """Fetch all FRED macro indicators. Returns dict with rows for display + LLM text."""
    global _fred_cache
    now_ts = datetime.now().timestamp()
    if not force and _fred_cache["data"] and _fred_cache["fetched_at"] and (now_ts - _fred_cache["fetched_at"]) < FRED_CACHE_TTL:
        return _fred_cache["data"]

    api_key = os.getenv("FRED_API_KEY")
    if not api_key:
        return None

    from datetime import timedelta
    rows = []
    llm_lines = []
    today = datetime.now().date()

    # US10Y
    start_14d = (today - timedelta(days=14)).isoformat()
    obs = _fred_fetch(api_key, "DGS10", start=start_14d)
    if obs:
        lat = obs[0]
        target = today - timedelta(days=7)
        pri = next((o for o in obs[1:] if __import__('datetime').date.fromisoformat(o["date"]) <= target), obs[-1] if len(obs) > 1 else None)
        trend = "rising" if pri and lat["value"] > pri["value"] else "declining" if pri and lat["value"] < pri["value"] else "flat"
        prior_str = f"7d prior: {pri['value']:.2f}%" if pri else ""
        rows.append({"metric": "US 10Y Yield", "value": f"{lat['value']:.2f}%", "comment": f"{prior_str}, {trend}"})
        llm_lines.append(f"US10Y: {lat['value']:.2f}% ({prior_str}, trend: {trend})")

    # DXY — real US Dollar Index from Yahoo Finance (DX-Y.NYB)
    try:
        dxy_resp = requests.get(
            "https://query1.finance.yahoo.com/v8/finance/chart/DX-Y.NYB",
            params={"interval": "1d", "range": "7d"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10,
        )
        if dxy_resp.status_code == 200:
            dxy_data = dxy_resp.json()
            closes = dxy_data.get("chart", {}).get("result", [{}])[0].get("indicators", {}).get("quote", [{}])[0].get("close", [])
            closes = [c for c in closes if c is not None]
            if closes:
                dxy_now = closes[-1]
                dxy_7d = closes[0] if len(closes) > 1 else None
                trend = "strengthening" if dxy_7d and dxy_now > dxy_7d else "weakening" if dxy_7d and dxy_now < dxy_7d else "flat"
                prior_str = f"7d prior: {dxy_7d:.2f}" if dxy_7d else ""
                rows.append({"metric": "DXY (USD Index)", "value": f"{dxy_now:.2f}", "comment": f"{prior_str}, {trend}"})
                llm_lines.append(f"DXY: {dxy_now:.2f} ({prior_str}, trend: {trend})")
    except Exception as e:
        print(f"DXY fetch from Yahoo failed: {e}")

    # M2 YoY
    start_500d = (today - timedelta(days=500)).isoformat()
    m2_obs = _fred_fetch(api_key, "WM2NS", start=start_500d)
    if m2_obs:
        m2_lat = m2_obs[0]
        lat_dt = __import__('datetime').date.fromisoformat(m2_lat["date"])
        m2_yoy = next((o for o in m2_obs if (lat_dt - __import__('datetime').date.fromisoformat(o["date"])).days >= 340), None)
        if m2_yoy:
            yoy_pct = (m2_lat["value"] - m2_yoy["value"]) / m2_yoy["value"] * 100
            trend = "expanding" if yoy_pct > 0 else "contracting"
            rows.append({"metric": "M2 Money Supply", "value": f"${m2_lat['value']:,.0f}B", "comment": f"{yoy_pct:+.1f}% YoY, {trend}"})
            llm_lines.append(f"M2: ${m2_lat['value']:,.0f}B, {yoy_pct:+.1f}% YoY (trend: {trend})")

    # Fed Funds + cycle
    obs_u = _fred_fetch(api_key, "DFEDTARU")
    obs_u2 = _fred_fetch(api_key, "DFEDTARU", start=(today - timedelta(days=365)).isoformat())
    obs_l = _fred_fetch(api_key, "DFEDTARL")
    if obs_u and obs_l:
        upper = obs_u[0]["value"]
        lower = obs_l[0]["value"]
        rate_range = f"{lower:.2f}%–{upper:.2f}%"
        # Cycle from last 3 decisions
        cycle = "paused"
        last_action = "N/A"
        last_date = obs_u[0]["date"]
        if len(obs_u2) >= 2:
            change_bp = round((obs_u2[0]["value"] - obs_u2[1]["value"]) * 100)
            last_action = f"+{change_bp}bp" if change_bp > 0 else f"{change_bp}bp" if change_bp < 0 else "held"
            if len(obs_u2) >= 3:
                moves = [round((obs_u2[i]["value"] - obs_u2[i+1]["value"]) * 100) for i in range(min(3, len(obs_u2)-1))]
                cuts = sum(1 for m in moves if m < 0)
                hikes = sum(1 for m in moves if m > 0)
                cycle = "cutting" if cuts > hikes else "hiking" if hikes > cuts else "paused"
        # Next FOMC
        next_fomc = "N/A"
        fomc_days = 0
        for d in FOMC_2026:
            if d >= today:
                next_fomc = d.strftime("%b %d, %Y")
                fomc_days = (d - today).days
                break
        rows.append({"metric": "Fed Funds Rate", "value": rate_range, "comment": f"cycle: {cycle}, last ({last_date}): {last_action}"})
        rows.append({"metric": "Next FOMC", "value": next_fomc, "comment": f"{fomc_days} days"})
        llm_lines.append(f"Fed Funds: {rate_range}, cycle: {cycle}, last FOMC ({last_date}): {last_action}, next FOMC: {next_fomc} ({fomc_days}d)")

    result = {
        "rows": rows,
        "llm_text": "\n".join(llm_lines),
        "fetched_at": datetime.now().isoformat(),
    }
    _fred_cache["data"] = result
    _fred_cache["fetched_at"] = now_ts
    return result


@app.route('/api/market/macro')
def api_market_macro():
    """Get FRED macro indicators (24h cached)."""
    force = request.args.get('refresh', 'false').lower() == 'true'
    data = fetch_fred_macro(force=force)
    if not data:
        return jsonify({"error": "FRED_API_KEY not configured", "rows": []})
    return jsonify(data)


@app.route('/api/portfolio')
def api_portfolio():
    """API endpoint for portfolio data."""
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    data = get_portfolio_data(force_refresh=force_refresh)
    # Auto-snapshot: when fresh data is pulled, save to DB in background
    if force_refresh:
        import threading
        def _bg_snapshot():
            try:
                from src.engines.snapshot_service import take_portfolio_snapshot
                wallets = get_wallet_addresses()
                if wallets:
                    # Pass a lambda that returns cached data (no re-fetch)
                    take_portfolio_snapshot(lambda **_: data, wallets)
            except Exception as e:
                print(f"Auto-snapshot error: {e}")
        threading.Thread(target=_bg_snapshot, daemon=True, name='auto-snapshot').start()
    return jsonify(data)


@app.route('/api/wallets', methods=['GET'])
def api_get_wallets():
    """Get list of wallet addresses with labels and roles."""
    config = load_wallet_config()
    wallets = [
        {
            "address": addr,
            "label": info.get("label", addr[:10] + "..."),
            "role": info.get("role", "active"),
            "hidden": bool(info.get("hidden", False)),
        }
        for addr, info in config.items()
    ]
    return jsonify({"wallets": wallets})


@app.route('/api/wallets', methods=['POST'])
def api_add_wallet():
    """Add a new wallet address or Bitcoin xpub with optional label."""
    global _portfolio_cache
    data = request.json
    address = data.get('address', '').strip()
    label = data.get('label', '').strip()

    # Detect if this is a Bitcoin xpub/ypub/zpub
    is_xpub = address.startswith(('xpub', 'ypub', 'zpub'))

    if not is_xpub and not is_valid_address(address):
        return jsonify({"error": "Invalid address format. Use an Ethereum address (0x...) or Bitcoin xpub/ypub/zpub."}), 400

    if is_xpub and len(address) < 100:
        return jsonify({"error": "Invalid xpub key — too short"}), 400

    config = load_wallet_config()

    # Check for duplicates
    if address in config:
        return jsonify({"error": "Wallet already exists"}), 400

    # Add wallet with label and type
    wallet_entry = {
        "label": label if label else ("BTC Ledger" if is_xpub else f"Wallet {len(config) + 1}"),
        "added_at": datetime.now().isoformat()
    }
    if is_xpub:
        wallet_entry["type"] = "bitcoin_xpub"

    config[address] = wallet_entry
    save_wallet_config(config)
    save_wallet_addresses(list(config.keys()))

    _portfolio_cache = None

    wallets = [
        {"address": addr, "label": info.get("label", addr[:10] + "...")}
        for addr, info in config.items()
    ]
    return jsonify({"success": True, "wallets": wallets})


@app.route('/api/wallets/<address>', methods=['PUT'])
def api_update_wallet(address):
    """Update wallet label and/or role."""
    global _portfolio_cache
    data = request.json
    label = data.get('label', '').strip()
    role = data.get('role', '').strip()
    hidden = data.get('hidden', None)

    if not label and not role and hidden is None:
        return jsonify({"error": "Nothing to update"}), 400

    config = load_wallet_config()

    wallet_key = None
    for key in config.keys():
        if key.lower() == address.lower():
            wallet_key = key
            break

    if not wallet_key:
        return jsonify({"error": "Wallet not found"}), 404

    if label:
        config[wallet_key]["label"] = label
    if role and role in ('active', 'treasury'):
        config[wallet_key]["role"] = role
    if hidden is not None:
        config[wallet_key]["hidden"] = bool(hidden)
    save_wallet_config(config)

    _portfolio_cache = None
    return jsonify({"success": True})


@app.route('/api/wallets/<address>', methods=['DELETE'])
def api_remove_wallet(address):
    """Remove a wallet address."""
    global _portfolio_cache
    config = load_wallet_config()
    
    # Find and remove the address (case-insensitive)
    wallet_key = None
    for key in config.keys():
        if key.lower() == address.lower():
            wallet_key = key
            break
    
    if wallet_key:
        del config[wallet_key]
        save_wallet_config(config)
        save_wallet_addresses(list(config.keys()))
        
        # Clear cache when wallets change
        _portfolio_cache = None
    
    wallets = [
        {"address": addr, "label": info.get("label", addr[:10] + "...")}
        for addr, info in config.items()
    ]
    return jsonify({"success": True, "wallets": wallets})



@app.route('/api/config', methods=['GET'])
def api_get_config():
    """Get API keys + RPC endpoints (masked — only first/last 4 chars visible)."""
    def _mask(val):
        if not val or len(val) < 8:
            return "****" if val else ""
        return val[:4] + "•" * (len(val) - 8) + val[-4:]

    def _describe_rpc(url: str) -> dict:
        """Detect provider + key from a known RPC URL pattern, or fall back to custom.

        Used by the Settings UI to round-trip the user's choice without storing
        provider info separately. Detection patterns (the prefix is identical
        for every Alchemy / Infura customer; only the trailing key segment varies):
          - https://eth-mainnet.g.alchemy.com/v2/<KEY>
          - https://arb-mainnet.g.alchemy.com/v2/<KEY>
          - https://base-mainnet.g.alchemy.com/v2/<KEY>
          - https://mainnet.infura.io/v3/<KEY>
          - https://arbitrum-mainnet.infura.io/v3/<KEY>
          - https://base-mainnet.infura.io/v3/<KEY>
        Anything else is reported as 'custom' so the UI shows a plain URL field.
        """
        if not url:
            return {"provider": "", "key_masked": "", "url_masked": ""}
        import re
        m = re.match(r"https://[a-z0-9-]+\.g\.alchemy\.com/v2/(.+)$", url)
        if m:
            return {"provider": "alchemy", "key_masked": _mask(m.group(1)), "url_masked": ""}
        m = re.match(r"https://[a-z0-9-]+\.infura\.io/v3/(.+)$", url)
        if m:
            return {"provider": "infura", "key_masked": _mask(m.group(1)), "url_masked": ""}
        # Unknown provider — mask the whole URL by treating it as the secret
        return {"provider": "custom", "key_masked": "", "url_masked": _mask(url)}

    return jsonify({
        "etherscan_api_key": _mask(os.getenv("ETHERSCAN_API_KEY", "")),
        "openai_api_key": _mask(os.getenv("OPENAI_API_KEY", "")),
        "anthropic_api_key": _mask(os.getenv("ANTHROPIC_API_KEY", "")),
        "aws_bearer_token": _mask(os.getenv("AWS_BEARER_TOKEN_BEDROCK", "")),
        "zerion_api_key": _mask(os.getenv("ZERION_API_KEY", "")),
        "fred_api_key": _mask(os.getenv("FRED_API_KEY", "")),
        "rpc_ethereum": _describe_rpc(os.getenv("ETHEREUM_RPC_URL", "")),
        "rpc_arbitrum": _describe_rpc(os.getenv("ARBITRUM_RPC_URL", "")),
        "rpc_base": _describe_rpc(os.getenv("BASE_RPC_URL", "")),
    })


@app.route('/api/config', methods=['POST'])
def api_update_config():
    """Update API key in .env file."""
    try:
        data = request.json
        key_name = data.get('key')
        key_value = data.get('value', '').strip()
        
        if not key_name or not key_value:
            return jsonify({"error": "Missing key or value"}), 400
        
        # Validate key name
        valid_keys = ['ETHERSCAN_API_KEY',
                      'OPENAI_API_KEY', 'ANTHROPIC_API_KEY',
                      'AWS_BEARER_TOKEN_BEDROCK', 'AWS_REGION',
                      'ZERION_API_KEY', 'FRED_API_KEY',
                      'ETHEREUM_RPC_URL', 'ARBITRUM_RPC_URL', 'BASE_RPC_URL']
        if key_name not in valid_keys:
            return jsonify({"error": "Invalid key name"}), 400
        
        # Read current .env file
        env_path = ENV_FILE
        env_lines = []
        key_found = False
        
        if os.path.exists(env_path):
            with open(env_path, 'r') as f:
                env_lines = f.readlines()
        
        # Update or add the key
        new_lines = []
        for line in env_lines:
            if line.strip().startswith(f'{key_name}='):
                new_lines.append(f'{key_name}={key_value}\n')
                key_found = True
            else:
                new_lines.append(line)
        
        # If key wasn't found, add it
        if not key_found:
            new_lines.append(f'\n{key_name}={key_value}\n')
        
        # Write back to .env
        with open(env_path, 'w') as f:
            f.writelines(new_lines)
        
        # Update environment variable for current session
        os.environ[key_name] = key_value
        
        return jsonify({"success": True, "message": f"{key_name} updated successfully"})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/manual-positions', methods=['GET'])
def api_get_manual_positions():
    """Get all active manual LP positions."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM lp_positions WHERE is_active=1 AND (is_permanently_hidden=0 OR is_permanently_hidden IS NULL) ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/manual-positions', methods=['POST'])
def api_add_manual_position():
    """Add a new manual LP position."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    
    # Validate required fields
    required = ['chain', 'protocol', 'token0', 'token1', 'amount0', 'amount1', 'range_lower', 'range_upper', 'fee_tier']
    missing = [f for f in required if not data.get(f) and data.get(f) != 0]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    
    # Validate and fetch token prices (use overrides if provided)
    price0 = float(data['price0_override']) if data.get('price0_override') else _get_coingecko_price(data['token0'])
    price1 = float(data['price1_override']) if data.get('price1_override') else _get_coingecko_price(data['token1'])
    
    # Retry once if rate limited
    import time as _time
    if price0 is None and not data.get('price0_override'):
        _time.sleep(1.5)
        price0 = _get_coingecko_price(data['token0'])
    if price1 is None and not data.get('price1_override'):
        _time.sleep(1.5)
        price1 = _get_coingecko_price(data['token1'])
    
    token0_valid = price0 is not None
    token1_valid = price1 is not None
    
    # If price not available and no override, reject
    price_errors = []
    if not token0_valid:
        price_errors.append(f"{data['token0']} price not found — enter Price Token0 manually")
    if not token1_valid:
        price_errors.append(f"{data['token1']} price not found — enter Price Token1 manually")
    if price_errors:
        return jsonify({"error": "; ".join(price_errors), "need_prices": True}), 400
    
    price0 = price0 or 0
    price1 = price1 or 0
    
    amount0 = float(data['amount0'])
    amount1 = float(data['amount1'])
    value_usd = amount0 * price0 + amount1 * price1
    
    # Calculate current price (token1 per token0)
    current_price = price0 / price1 if price1 > 0 else 0
    in_range = float(data['range_lower']) <= current_price <= float(data['range_upper'])
    
    conn = get_connection()
    c = conn.cursor()
    c.execute("""INSERT INTO lp_positions
        (user_id, source, chain, protocol, position_id, token0, token1, fee_tier,
         value_usd, range_lower, range_upper, current_price, in_range,
         fees_uncollected_usd, fees_collected_usd, notes, is_active,
         amount0, amount1, price0_usd, price1_usd, entry_value_usd)
        VALUES (1, 'manual', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, ?, 1, ?, ?, ?, ?, ?)""",
        (data['chain'], data['protocol'], data.get('position_id', ''),
         data['token0'], data['token1'], float(data['fee_tier']),
         value_usd, float(data['range_lower']), float(data['range_upper']),
         current_price, in_range, data.get('notes', ''),
         amount0, amount1, price0, price1, value_usd))
    pos_id = c.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({
        "success": True, "id": pos_id,
        "value_usd": value_usd,
        "current_price": current_price,
        "in_range": in_range,
        "token0_valid": token0_valid,
        "token1_valid": token1_valid,
        "warnings": [] if (token0_valid and token1_valid) else
            [f"Price not found for: {', '.join([data['token0']] * (not token0_valid) + [data['token1']] * (not token1_valid))}"]
    })


@app.route('/api/manual-positions/<int:pos_id>', methods=['PUT'])
def api_update_manual_position(pos_id):
    """Update a manual position (fees, notes, or close it)."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    conn = get_connection()
    
    if data.get('action') == 'close':
        # Fetch current prices for exit value
        row = conn.execute("SELECT * FROM lp_positions WHERE id=?", (pos_id,)).fetchone()
        if row:
            price0 = _get_coingecko_price(row['token0']) or 0
            price1 = _get_coingecko_price(row['token1']) or 0
            exit_value = (row['amount0'] or 0) * price0 + (row['amount1'] or 0) * price1
            conn.execute(
                "UPDATE lp_positions SET is_active=0, updated_at=CURRENT_TIMESTAMP, value_usd=?, current_price=? WHERE id=?",
                (exit_value, price0 / price1 if price1 > 0 else 0, pos_id))
        else:
            conn.execute("UPDATE lp_positions SET is_active=0, updated_at=CURRENT_TIMESTAMP WHERE id=?", (pos_id,))
    elif data.get('action') == 'edit':
        import time as _time
        price0 = float(data['price0_override']) if data.get('price0_override') else _get_coingecko_price(data['token0'])
        price1 = float(data['price1_override']) if data.get('price1_override') else _get_coingecko_price(data['token1'])
        if price0 is None and not data.get('price0_override'):
            _time.sleep(1.5)
            price0 = _get_coingecko_price(data['token0'])
        if price1 is None and not data.get('price1_override'):
            _time.sleep(1.5)
            price1 = _get_coingecko_price(data['token1'])
        price_errors = []
        if price0 is None:
            price_errors.append(f"{data['token0']} price not found — enter Price manually")
        if price1 is None:
            price_errors.append(f"{data['token1']} price not found — enter Price manually")
        if price_errors:
            conn.close()
            return jsonify({"error": "; ".join(price_errors), "need_prices": True}), 400
        price0 = price0 or 0
        price1 = price1 or 0
        amount0 = float(data['amount0'])
        amount1 = float(data['amount1'])
        value_usd = amount0 * price0 + amount1 * price1
        current_price = price0 / price1 if price1 > 0 else 0
        in_range = float(data['range_lower']) <= current_price <= float(data['range_upper'])
        conn.execute("""
            UPDATE lp_positions SET
              chain=?, protocol=?, token0=?, token1=?, fee_tier=?,
              amount0=?, amount1=?, range_lower=?, range_upper=?, notes=?,
              value_usd=?, entry_value_usd=?, current_price=?, in_range=?,
              price0_usd=?, price1_usd=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (
            data.get('chain', ''), data.get('protocol', ''),
            data['token0'], data['token1'], float(data['fee_tier']),
            amount0, amount1, float(data['range_lower']), float(data['range_upper']),
            data.get('notes', ''), value_usd, value_usd,
            current_price, in_range, price0, price1, pos_id
        ))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "value_usd": value_usd, "current_price": current_price, "in_range": in_range})
    else:
        updates = []
        params = []
        for field in ['fees_uncollected_usd', 'fees_collected_usd', 'notes', 'amount0', 'amount1']:
            if field in data:
                updates.append(f"{field}=?")
                params.append(data[field])
        if updates:
            updates.append("updated_at=CURRENT_TIMESTAMP")
            params.append(pos_id)
            conn.execute(f"UPDATE lp_positions SET {', '.join(updates)} WHERE id=?", params)
            
            # Recalculate value if amounts changed
            if 'amount0' in data or 'amount1' in data:
                row = conn.execute("SELECT * FROM lp_positions WHERE id=?", (pos_id,)).fetchone()
                if row:
                    price0 = _get_coingecko_price(row['token0']) or 0
                    price1 = _get_coingecko_price(row['token1']) or 0
                    value_usd = (row['amount0'] or 0) * price0 + (row['amount1'] or 0) * price1
                    current_price = price0 / price1 if price1 > 0 else 0
                    in_range = row['range_lower'] <= current_price <= row['range_upper']
                    conn.execute("UPDATE lp_positions SET value_usd=?, current_price=?, in_range=? WHERE id=?",
                        (value_usd, current_price, in_range, pos_id))
    
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route('/api/manual-positions/<int:pos_id>', methods=['DELETE'])
def api_delete_manual_position(pos_id):
    """Permanently hide a manual LP position (soft delete)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    conn.execute("UPDATE lp_positions SET is_permanently_hidden=1, is_active=0 WHERE id=?", (pos_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route('/api/validate-token/<symbol>')
def api_validate_token(symbol):
    """Check if a token symbol is recognized and return its price."""
    # First check if it's in our known mapping
    cg_map = {
        'BTC': 'bitcoin', 'ETH': 'ethereum', 'WETH': 'ethereum', 'WBTC': 'bitcoin',
        'SOL': 'solana', 'USDC': 'usd-coin', 'USDT': 'tether', 'DAI': 'dai',
        'ARB': 'arbitrum', 'OP': 'optimism', 'LINK': 'chainlink', 'UNI': 'uniswap',
        'AAVE': 'aave', 'SUI': 'sui', 'TAO': 'bittensor', 'DOGE': 'dogecoin',
        'AVAX': 'avalanche-2', 'MATIC': 'matic-network', 'NEAR': 'near',
        'DOT': 'polkadot', 'ATOM': 'cosmos', 'XRP': 'ripple',
        'stETH': 'staked-ether', 'wstETH': 'wrapped-steth', 'cbBTC': 'bitcoin',
        'USDe': 'ethena-usde', 'GHO': 'gho', 'CRV': 'curve-dao-token',
        'PEPE': 'pepe', 'INJ': 'injective-protocol', 'TIA': 'celestia',
    }
    sym_upper = symbol.upper()
    known = sym_upper in cg_map or symbol in cg_map
    price = _get_coingecko_price(symbol)
    return jsonify({
        "symbol": symbol,
        "valid": price is not None,
        "known": known,
        "price_usd": price or 0,
        "message": "" if price else ("Rate limited — try again" if known else "Token not recognized")
    })


@app.route('/api/manual-hedges', methods=['GET'])
def api_get_manual_hedges():
    """Get all active manual hedge positions with live prices."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    rows = conn.execute("SELECT * FROM hedge_positions WHERE is_active=1 ORDER BY created_at DESC").fetchall()
    conn.close()
    results = []
    for row in rows:
        h = dict(row)
        # Refresh current price and recalculate PnL
        symbol = h['market'].split('/')[0] if '/' in h['market'] else h['market']
        price = _get_coingecko_price(symbol)
        if price:
            h['current_price'] = price
            if h['entry_price'] and h['size_usd']:
                if h['direction'] == 'long':
                    h['pnl_usd'] = h['size_usd'] * (price - h['entry_price']) / h['entry_price']
                else:
                    h['pnl_usd'] = h['size_usd'] * (h['entry_price'] - price) / h['entry_price']
                h['pnl_pct'] = (h['pnl_usd'] / h['margin_usd'] * 100) if h['margin_usd'] > 0 else 0
        results.append(h)
    return jsonify(results)


@app.route('/api/manual-hedges', methods=['POST'])
def api_add_manual_hedge():
    """Add a new manual hedge position."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    
    required = ['market', 'direction', 'margin_usd', 'leverage']
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing: {', '.join(missing)}"}), 400
    
    margin = float(data['margin_usd'])
    leverage = float(data['leverage'])
    size_usd = margin * leverage
    direction = data['direction']
    
    # Get entry price (user-provided or fetched)
    symbol = data['market'].split('/')[0] if '/' in data['market'] else data['market']
    entry_price = float(data['entry_price']) if data.get('entry_price') else None
    if not entry_price:
        entry_price = _get_coingecko_price(symbol) or 0
    
    current_price = _get_coingecko_price(symbol) or entry_price
    
    # Calculate liquidation price
    if entry_price > 0 and size_usd > 0:
        liq_move = margin / size_usd * entry_price
        liq_price = (entry_price - liq_move) if direction == 'long' else (entry_price + liq_move)
    else:
        liq_price = 0
    
    # Calculate PnL
    pnl_usd = 0
    if entry_price > 0 and current_price > 0:
        if direction == 'long':
            pnl_usd = size_usd * (current_price - entry_price) / entry_price
        else:
            pnl_usd = size_usd * (entry_price - current_price) / entry_price
    pnl_pct = (pnl_usd / margin * 100) if margin > 0 else 0
    
    conn = get_connection()
    c = conn.cursor()
    c.execute("""INSERT INTO hedge_positions
        (user_id, source, exchange, market, direction, margin_usd, leverage, size_usd,
         entry_price, current_price, liquidation_price, pnl_usd, pnl_pct,
         stop_loss_price, take_profit_price, notes, is_active)
        VALUES (1, 'manual', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)""",
        (data.get('exchange', ''), data['market'], direction, margin, leverage, size_usd,
         entry_price, current_price, liq_price, pnl_usd, pnl_pct,
         float(data['stop_loss_price']) if data.get('stop_loss_price') else None,
         float(data['take_profit_price']) if data.get('take_profit_price') else None,
         data.get('notes', '')))
    hedge_id = c.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({"success": True, "id": hedge_id, "entry_price": entry_price,
                    "size_usd": size_usd, "liquidation_price": liq_price})


@app.route('/api/manual-hedges/<int:hedge_id>', methods=['PUT'])
def api_update_manual_hedge(hedge_id):
    """Update or close a manual hedge."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    conn = get_connection()
    
    if data.get('action') == 'close':
        row = conn.execute("SELECT * FROM hedge_positions WHERE id=?", (hedge_id,)).fetchone()
        if row:
            symbol = row['market'].split('/')[0] if '/' in row['market'] else row['market']
            price = _get_coingecko_price(symbol) or row['current_price']
            if row['direction'] == 'long':
                pnl = row['size_usd'] * (price - row['entry_price']) / row['entry_price'] if row['entry_price'] else 0
            else:
                pnl = row['size_usd'] * (row['entry_price'] - price) / row['entry_price'] if row['entry_price'] else 0
            conn.execute(
                "UPDATE hedge_positions SET is_active=0, updated_at=CURRENT_TIMESTAMP, current_price=?, pnl_usd=? WHERE id=?",
                (price, pnl, hedge_id))
        else:
            conn.execute("UPDATE hedge_positions SET is_active=0, updated_at=CURRENT_TIMESTAMP WHERE id=?", (hedge_id,))
    else:
        updates = []
        params = []
        for field in ['stop_loss_price', 'take_profit_price', 'notes', 'margin_usd']:
            if field in data:
                updates.append(f"{field}=?")
                params.append(data[field])
        if updates:
            updates.append("updated_at=CURRENT_TIMESTAMP")
            params.append(hedge_id)
            conn.execute(f"UPDATE hedge_positions SET {', '.join(updates)} WHERE id=?", params)
    
    conn.commit()
    conn.close()
    return jsonify({"success": True})


_cg_price_cache = {}  # symbol -> (price, timestamp)

def _get_coingecko_price(symbol: str) -> float | None:
    """Get token price from CoinGecko by symbol. Caches for 60 seconds."""
    import time as _t
    cache_key = symbol.upper()
    if cache_key in _cg_price_cache:
        price, ts = _cg_price_cache[cache_key]
        if _t.time() - ts < 60:  # Cache for 60 seconds
            return price

    # Check spot_token_config for a CoinGecko ID override
    _override_cg_id = None
    try:
        from src.storage.portfolio_db import get_connection as _gc
        _cfg_conn = _gc()
        _cfg_row = _cfg_conn.execute(
            "SELECT cg_id FROM spot_token_config WHERE symbol=?", (symbol.upper(),)
        ).fetchone()
        _cfg_conn.close()
        if _cfg_row and _cfg_row['cg_id']:
            _override_cg_id = _cfg_row['cg_id']
    except Exception:
        pass

    cg_map = {
        'BTC': 'bitcoin', 'ETH': 'ethereum', 'WETH': 'ethereum', 'WBTC': 'bitcoin',
        'SOL': 'solana', 'USDC': 'usd-coin', 'USDT': 'tether', 'DAI': 'dai',
        'ARB': 'arbitrum', 'OP': 'optimism', 'MATIC': 'matic-network', 'AVAX': 'avalanche-2',
        'LINK': 'chainlink', 'UNI': 'uniswap', 'AAVE': 'aave', 'CRV': 'curve-dao-token',
        'SUI': 'sui', 'TAO': 'bittensor', 'DOGE': 'dogecoin', 'XRP': 'ripple',
        'DOT': 'polkadot', 'ATOM': 'cosmos', 'NEAR': 'near', 'FTM': 'fantom',
        'APT': 'aptos', 'INJ': 'injective-protocol', 'TIA': 'celestia',
        'STX': 'blockstack', 'SEI': 'sei-network', 'PEPE': 'pepe',
        'cbBTC': 'bitcoin', 'stETH': 'staked-ether', 'wstETH': 'wrapped-steth',
        'rETH': 'rocket-pool-eth', 'USDe': 'ethena-usde', 'GHO': 'gho',
    }
    if _override_cg_id:
        cg_id = _override_cg_id
    else:
        cg_id = cg_map.get(symbol.upper(), cg_map.get(symbol))
        if not cg_id:
            # Try lowercase as CoinGecko ID directly
            cg_id = symbol.lower()
    try:
        r = requests.get(f'https://api.coingecko.com/api/v3/simple/price?ids={cg_id}&vs_currencies=usd', timeout=5)
        if r.ok:
            price = r.json().get(cg_id, {}).get('usd')
            if price and price > 0:
                _cg_price_cache[cache_key] = (float(price), _t.time())
                return float(price)
        elif r.status_code == 429:
            # Rate limited — retry once after delay
            import time
            time.sleep(1)
            r = requests.get(f'https://api.coingecko.com/api/v3/simple/price?ids={cg_id}&vs_currencies=usd', timeout=5)
            if r.ok:
                price = r.json().get(cg_id, {}).get('usd')
                if price and price > 0:
                    _cg_price_cache[cache_key] = (float(price), _t.time())
                    return float(price)
    except Exception:
        pass
    return None


def _get_dexscreener_price(contract_address: str) -> float | None:
    """Get token price via DexScreener by contract address. Caches for 60 seconds."""
    import time as _t
    cache_key = contract_address.lower()
    if cache_key in _cg_price_cache:
        price, ts = _cg_price_cache[cache_key]
        if _t.time() - ts < 60:
            return price
    try:
        r = requests.get(
            f'https://api.dexscreener.com/latest/dex/tokens/{contract_address}',
            timeout=5
        )
        if not r.ok:
            return None
        pairs = r.json().get('pairs') or []
        # Only consider pairs where our token is the base token
        base_pairs = [
            p for p in pairs
            if (p.get('baseToken') or {}).get('address', '').lower() == cache_key
            and p.get('priceUsd')
        ]
        if not base_pairs:
            return None
        best = max(base_pairs, key=lambda p: float((p.get('liquidity') or {}).get('usd') or 0))
        price = float(best['priceUsd'])
        _cg_price_cache[cache_key] = (price, _t.time())
        return price
    except Exception:
        pass
    return None


def _get_spot_price(symbol: str) -> float | None:
    """Price lookup respecting price_source field from spot_token_config.
    dexscreener -> requires contract_address
    coingecko -> uses cg_id override or symbol map
    manual -> returns None (no live price)
    Falls back to coingecko if price_source not set."""
    try:
        from src.storage.portfolio_db import get_connection as _gc
        _conn = _gc()
        _row = _conn.execute(
            "SELECT contract_address, price_source FROM spot_token_config WHERE symbol=?", (symbol.upper(),)
        ).fetchone()
        _conn.close()
        if _row:
            source = (_row['price_source'] or 'coingecko').lower()
            if source == 'dexscreener':
                if _row['contract_address']:
                    return _get_dexscreener_price(_row['contract_address'])
                return None  # dexscreener selected but no contract address
            elif source == 'manual':
                return None  # manual price, no live lookup
            # else coingecko (default)
    except Exception:
        pass
    return _get_coingecko_price(symbol)


# --- AI Advisor Routes ---

@app.route('/api/ai/config', methods=['GET'])
def api_ai_config_get():
    """Get AI advisor configuration."""
    from src.engines.ai_advisor import load_ai_config
    return jsonify(load_ai_config())


@app.route('/api/ai/config', methods=['POST'])
def api_ai_config_save():
    """Save AI advisor configuration."""
    from src.engines.ai_advisor import save_ai_config
    data = request.json
    save_ai_config(data)
    return jsonify({"status": "success"})


@app.route('/api/ai/prompt', methods=['GET'])
def api_ai_prompt_get():
    """Get the custom strategy prompt."""
    from src.engines.ai_advisor import load_ai_config
    config = load_ai_config()
    return jsonify({"prompt": config.get("custom_system_prompt", "")})


@app.route('/api/ai/prompt', methods=['POST'])
def api_ai_prompt_save():
    """Save the custom strategy prompt."""
    from src.engines.ai_advisor import load_ai_config, save_ai_config
    data = request.json or {}
    config = load_ai_config()
    config["custom_system_prompt"] = data.get("prompt", "")
    save_ai_config(config)
    return jsonify({"success": True})


@app.route('/api/ai/models/<provider>')
def api_ai_models(provider):
    """Fetch available models for a given AI provider, with hardcoded fallbacks."""
    _ANTHROPIC_FALLBACK = [
        'claude-opus-4-8',
        'claude-opus-4-7-20260416',
        'claude-sonnet-4-6',
        'claude-haiku-4-5-20251001',
    ]
    _OPENAI_FALLBACK = [
        'gpt-4o',
        'gpt-4o-mini',
        'gpt-4-turbo',
        'gpt-4',
        'gpt-3.5-turbo',
    ]

    if provider == 'anthropic':
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        if api_key:
            try:
                resp = requests.get(
                    'https://api.anthropic.com/v1/models',
                    headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01'},
                    timeout=8,
                )
                if resp.ok:
                    data = resp.json()
                    models = [m['id'] for m in data.get('data', []) if m.get('id', '').startswith('claude-')]
                    if models:
                        return jsonify({'models': models})
            except Exception:
                pass
        return jsonify({'models': _ANTHROPIC_FALLBACK})

    elif provider == 'openai':
        api_key = os.environ.get('OPENAI_API_KEY', '')
        if api_key:
            try:
                resp = requests.get(
                    'https://api.openai.com/v1/models',
                    headers={'Authorization': f'Bearer {api_key}'},
                    timeout=8,
                )
                if resp.ok:
                    data = resp.json()
                    models = sorted(
                        [m['id'] for m in data.get('data', []) if m.get('id', '').startswith('gpt-')],
                        reverse=True,
                    )
                    if models:
                        return jsonify({'models': models})
            except Exception:
                pass
        return jsonify({'models': _OPENAI_FALLBACK})

    else:
        return jsonify({'models': []})


# --- Telegram Settings Routes ---

@app.route('/api/settings/telegram', methods=['GET'])
def api_telegram_config_get():
    """Get Telegram configuration with masked bot token."""
    config = load_telegram_config()
    config["bot_token"] = mask_bot_token(config.get("bot_token", ""))
    return jsonify(config)


@app.route('/api/settings/telegram', methods=['POST'])
def api_telegram_config_save():
    """Validate and save Telegram configuration."""
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400
    # If the submitted token looks masked (starts with ****), preserve the existing one
    if data.get("bot_token", "").startswith("****"):
        existing = load_telegram_config()
        data["bot_token"] = existing.get("bot_token", "")
    ok, err = validate_telegram_config(data)
    if not ok:
        return jsonify({"error": err}), 400
    save_telegram_config(data)
    return jsonify({"status": "success"})


@app.route('/api/settings/telegram/test', methods=['POST'])
def api_telegram_test():
    """Send a test message with real content via Telegram."""
    data = request.json or {}
    chat_id = data.get("chat_id", "").strip()
    # Always read the real token from disk — the form only has the masked version
    config = load_telegram_config()
    bot_token = config.get("bot_token", "")
    # Allow chat_id override from form, fall back to saved config
    if not chat_id:
        chat_id = config.get("chat_id", "")
    if not bot_token:
        return jsonify({"error": "No bot token configured — save your config first"}), 400
    if not chat_id:
        return jsonify({"error": "No chat ID configured"}), 400
    try:
        messages = build_notification_content(config)
        if not messages:
            messages = ["DeFi Portfolio Dashboard — test notification OK\n\n(No digest or regime data available yet)"]
        for msg in messages:
            send_telegram_message(bot_token, chat_id, msg, timeout=10)
        return jsonify({"status": "success", "message": f"Test sent ({len(messages)} message(s))"})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400


# Triage-digest Telegram settings (file-KV; distinct from the daily-notification
# /api/settings/telegram routes above, which use load/save_telegram_config).
@app.route('/api/settings/telegram-digest', methods=['GET'])
@login_required
def api_settings_telegram_digest_get():
    s = _telegram_settings()
    # never return the full token to the client
    token = s.get('bot_token', '')
    masked = ('*' * (len(token) - 4) + token[-4:]) if len(token) > 4 else (
        '*' * len(token))
    return jsonify({**s, 'bot_token': masked})


@app.route('/api/settings/telegram-digest', methods=['PUT'])
@login_required
def api_settings_telegram_digest_put():
    data = request.get_json() or {}
    updates = {}
    if 'bot_token' in data and data['bot_token'] and '*' not in data['bot_token']:
        updates['bot_token'] = data['bot_token']
    if 'chat_id' in data:
        updates['chat_id'] = str(data['chat_id']).strip()
    if 'enabled' in data:
        updates['enabled'] = bool(data['enabled'])
    if 'reminders' in data:
        raw = data['reminders']
        if not isinstance(raw, list):
            return jsonify({'error': 'reminders must be a list'}), 400
        cleaned = []
        for r in raw:
            if not isinstance(r, dict):
                continue
            if not r.get('id') or not r.get('label'):
                continue
            times = []
            for t in (r.get('times_utc') or []):
                import re
                if re.match(r'^\d{2}:\d{2}$', str(t)):
                    times.append(str(t))
            cleaned.append({
                'id': str(r['id']),
                'label': str(r['label']),
                'message': str(r.get('message', '')),
                'enabled': bool(r.get('enabled', False)),
                'times_utc': times[:3]
            })
        updates['reminders'] = cleaned
    _telegram_settings(updates)
    return jsonify({'ok': True})


DISPLAY_PREFS_PATH = os.path.join("data", "display_prefs.json")
DISPLAY_PREFS_DEFAULTS = {"dust_threshold": 0.01, "lending_threshold": 1.0}


@app.route('/api/settings/display', methods=['GET'])
def api_display_prefs_get():
    """Return display preferences, falling back to defaults."""
    prefs = dict(DISPLAY_PREFS_DEFAULTS)
    if os.path.exists(DISPLAY_PREFS_PATH):
        try:
            with open(DISPLAY_PREFS_PATH, "r") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                prefs.update(saved)
        except (json.JSONDecodeError, IOError):
            pass
    return jsonify(prefs)


@app.route('/api/settings/display', methods=['POST'])
def api_display_prefs_save():
    """Persist display preferences to disk."""
    data = request.json
    if not isinstance(data, dict):
        return jsonify({"error": "Invalid payload"}), 400
    if "dust_threshold" in data:
        try:
            data["dust_threshold"] = float(data["dust_threshold"])
            if data["dust_threshold"] < 0:
                return jsonify({"error": "dust_threshold must be >= 0"}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "dust_threshold must be a number"}), 400
    if "lending_threshold" in data:
        try:
            data["lending_threshold"] = float(data["lending_threshold"])
            if data["lending_threshold"] < 0:
                return jsonify({"error": "lending_threshold must be >= 0"}), 400
        except (TypeError, ValueError):
            return jsonify({"error": "lending_threshold must be a number"}), 400
    os.makedirs(os.path.dirname(DISPLAY_PREFS_PATH), exist_ok=True)
    existing = dict(DISPLAY_PREFS_DEFAULTS)
    if os.path.exists(DISPLAY_PREFS_PATH):
        try:
            with open(DISPLAY_PREFS_PATH, "r") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                existing.update(saved)
        except (json.JSONDecodeError, IOError):
            pass
    existing.update(data)
    import tempfile
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(DISPLAY_PREFS_PATH), suffix=".tmp")
    try:
        with os.fdopen(fd, "w") as f:
            json.dump(existing, f, indent=2)
        os.replace(tmp, DISPLAY_PREFS_PATH)
    except Exception as e:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        return jsonify({"error": str(e)}), 500
    return jsonify({"status": "success"})


@app.route('/api/ai/generate', methods=['POST'])
def api_ai_generate():
    """Generate an AI advisor report."""
    try:
        from src.engines.ai_advisor import generate_report
        result = generate_report(get_portfolio_data, get_wallet_addresses)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/ai/digest', methods=['POST'])
def api_ai_digest_generate():
    """Generate a daily digest (pure DB, no LLM)."""
    try:
        from src.engines.ai_advisor import generate_daily_digest
        digest = generate_daily_digest()
        return jsonify(digest)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/ai/digest/latest')
def api_ai_digest_latest():
    """Get the most recent daily digest. Auto-generates if stale."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    
    # Check if latest digest is older than latest snapshot
    latest_snap = conn.execute(
        "SELECT MAX(timestamp) as ts FROM portfolio_snapshots WHERE status='completed'"
    ).fetchone()
    latest_digest = conn.execute(
        "SELECT MAX(timestamp) as ts FROM daily_digests"
    ).fetchone()
    
    snap_ts = latest_snap['ts'] if latest_snap else None
    digest_ts = latest_digest['ts'] if latest_digest else None
    
    # Auto-generate if no digest exists or if snapshot is newer
    if snap_ts and (not digest_ts or snap_ts > digest_ts):
        conn.close()
        try:
            from src.engines.ai_advisor import generate_daily_digest
            generate_daily_digest()
        except Exception as e:
            print(f"Auto-digest error: {e}")
        conn = get_connection()
    
    row = conn.execute("SELECT * FROM daily_digests ORDER BY timestamp DESC LIMIT 1").fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "No digests yet"}), 404
    result = dict(row)
    for field in ['positions_opened', 'positions_closed', 'positions_out_of_range', 'hedge_health_json', 'digest_json']:
        if result.get(field):
            try: result[field] = json.loads(result[field])
            except: pass
    return jsonify(result)


@app.route('/api/ai/daily-brief', methods=['POST'])
def api_ai_daily_brief_generate():
    """Generate an LLM-powered daily brief."""
    try:
        from src.engines.ai_advisor import load_ai_config, load_system_prompt, _get_market_data_smart
        from src.engines.llm_providers import get_provider
        from src.storage.portfolio_db import get_connection

        config = load_ai_config()
        system_prompt = load_system_prompt()
        freshness = {}
        market_text = _get_market_data_smart(freshness)

        conn = get_connection()
        snap = conn.execute(
            "SELECT btc_price, fear_greed_index FROM market_snapshots WHERE btc_price IS NOT NULL ORDER BY timestamp DESC LIMIT 1"
        ).fetchone()
        btc_price = snap['btc_price'] if snap else None
        fg = snap['fear_greed_index'] if snap else None

        fg_rows = conn.execute(
            "SELECT fear_greed_index FROM market_snapshots WHERE fear_greed_index IS NOT NULL ORDER BY timestamp DESC LIMIT 7"
        ).fetchall()
        fg_7d = sum(r['fear_greed_index'] for r in fg_rows) / len(fg_rows) if fg_rows else None

        pv_row = conn.execute(
            "SELECT total_value_usd, value_change_24h_pct FROM daily_digests ORDER BY timestamp DESC LIMIT 1"
        ).fetchone()
        portfolio_value = pv_row['total_value_usd'] if pv_row else None
        change_24h = pv_row['value_change_24h_pct'] if pv_row else None

        oor_row = conn.execute(
            "SELECT positions_out_of_range FROM daily_digests ORDER BY timestamp DESC LIMIT 1"
        ).fetchone()
        oor = []
        if oor_row and oor_row['positions_out_of_range']:
            try:
                oor = json.loads(oor_row['positions_out_of_range'])
            except Exception:
                pass

        risk_row = conn.execute(
            "SELECT full_report_json FROM ai_reports ORDER BY timestamp DESC LIMIT 1"
        ).fetchone()
        risk_alerts = []
        if risk_row and risk_row['full_report_json']:
            try:
                rj = json.loads(risk_row['full_report_json'])
                risk_alerts = [
                    a.get('message', '')
                    for a in (rj.get('risk_alerts') or [])
                    if (a.get('severity') or '').lower() in ('critical', 'warning', 'warn')
                ]
            except Exception:
                pass
        conn.close()

        context_parts = [market_text]
        if portfolio_value:
            context_parts.append(
                f"Portfolio value: ${portfolio_value:,.0f}" +
                (f" ({change_24h:+.1f}% 24h)" if change_24h is not None else "")
            )
        if fg is not None:
            context_parts.append(
                f"Fear & Greed: {fg}" +
                (f" (7d avg: {fg_7d:.0f})" if fg_7d is not None else "")
            )
        if oor:
            context_parts.append(f"LP positions out of range: {', '.join(oor)}")
        if risk_alerts:
            context_parts.append("Open risk alerts: " + "; ".join(risk_alerts[:3]))

        user_message = (
            "\n".join(context_parts) +
            "\n\nGenerate a concise daily brief. Include: current market regime assessment in one "
            "sentence, BTC price and key signal, portfolio value and 24h change, any urgent action items. "
            "Be direct and specific. No preamble. Keep your response under 120 words total."
        )

        provider = get_provider(config)
        # Daily brief is plain text — override system prompt to suppress JSON requirement
        brief_system = "You are a concise DeFi portfolio advisor. Respond in plain text only. Do not use JSON. Do not use markdown code blocks."
        result = provider.complete(brief_system, user_message)
        response = result.get('response', '')

        # Handle truncated responses gracefully — use partial_text if available
        if isinstance(result, dict) and result.get('error') == 'LLM response was truncated':
            brief_text = result.get('partial_text', '')
        elif isinstance(response, dict):
            brief_text = (
                response.get('daily_brief') or
                response.get('brief') or
                response.get('summary') or
                (response.get('market_analysis') or {}).get('summary') or
                str(response)
            )
        else:
            brief_text = str(response) if response else ''

        # Strip the error wrapper if brief_text itself is a stringified error dict
        if brief_text.startswith("{'error':") or brief_text.startswith('{"error":'):
            try:
                import ast
                parsed = ast.literal_eval(brief_text) if brief_text.startswith("{'") else json.loads(brief_text)
                brief_text = parsed.get('partial_text') or parsed.get('brief_text') or brief_text
            except Exception:
                pass

        conn = get_connection()
        c = conn.execute(
            "INSERT INTO daily_briefs (user_id, timestamp, brief_text, btc_price, portfolio_value) VALUES (1, ?, ?, ?, ?)",
            (datetime.utcnow().isoformat(), brief_text, btc_price, portfolio_value)
        )
        conn.commit()
        row_id = c.lastrowid
        row = conn.execute("SELECT * FROM daily_briefs WHERE id=?", (row_id,)).fetchone()
        conn.close()
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/ai/daily-brief/latest')
def api_ai_daily_brief_latest():
    """Get the most recent daily brief."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM daily_briefs WHERE user_id=1 ORDER BY timestamp DESC LIMIT 1"
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"brief_text": None})
    return jsonify(dict(row))


@app.route('/api/ai/reports', methods=['GET'])
def api_ai_reports():
    """Get AI report history."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    limit = request.args.get('limit', 10, type=int)
    rows = conn.execute(
        "SELECT id, timestamp, provider, model, market_regime_json, portfolio_alignment, summary FROM ai_reports ORDER BY timestamp DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/ai/reports/<int:report_id>')
def api_ai_report_detail(report_id):
    """Get a specific AI report."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    row = conn.execute("SELECT * FROM ai_reports WHERE id=?", (report_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Report not found"}), 404
    result = dict(row)
    # Parse JSON fields
    for field in ['full_report_json', 'market_regime_json', 'previous_recs_review_json', 'data_freshness_json']:
        if result.get(field):
            try:
                result[field] = json.loads(result[field])
            except Exception:
                pass
    return jsonify(result)


# --- History / Snapshot Routes ---

from src.storage.portfolio_db import (
    init_db,
    get_db_path,
    get_portfolio_timeseries,
    get_market_timeseries,
    get_token_price_history,
)

# Initialize DB on startup
init_db()

_DEFAULT_SCANNER_PROMPT = """# Scanner Evaluation Rules

## HTF Bias Rules
- Bullish bias: most recent decisive Market Structure Break (MSB/BOS) is to the upside — a prior swing high taken out with follow-through
- Bearish bias: most recent decisive MSB is to the downside — a prior swing low taken out with follow-through
- Ranging: no clear MSB in either direction; price oscillating between equal highs and lows

## Dealing Range Rules
- Draw from most recent significant swing high to most recent significant swing low
- "Significant" = visually obvious pivot; if debatable, it is not significant
- Equilibrium (EQ) = 50% midpoint of the range
- Discount zone = below EQ (look for longs here)
- Premium zone = above EQ (look for shorts here)
- A lick sweep (price briefly exceeds the high/low then immediately snaps back) does NOT reset the dealing range

## Valid Entry Triggers
- Order Block (OB) tap: price returns to the last opposing candle before an impulsive displacement
- Fair Value Gap (FVG) return: price retraces into a 3-candle imbalance zone created by a large displacement candle
- Liquidity Sweep + CHoCH: sell-side or buy-side liquidity swept followed by a Change of Character on the LTF
- Combined confluence: OB + FVG overlap in the same zone (highest probability)

## Setup Validity Requirements (ALL must be true for 'active' status)
- HTF bias is clear and unambiguous
- Price is in the correct zone (discount for longs, premium for shorts)
- At least one liquidity pool has been swept (internal or external)
- A valid entry trigger is present or printing on the LTF
- Risk:Reward minimum 2:1; prefer 3:1

## Invalidation Rules
- HTF bias shifts against the trade (new MSB in opposite direction)
- Price closes through the full order block (not just a wick)
- Stop level taken out

## Status Definitions
- active: HTF bias clear + price at POI + LTF entry trigger confirmed
- forming: HTF bias clear + price approaching POI + entry trigger not yet confirmed
- watching: HTF setup developing but missing 1-2 criteria (e.g. liquidity not yet swept)
- quiet: no meaningful setup; price in no-trade zone or bias unclear

## R:R Guidelines
- Minimum: 2:1
- Target: 3:1 or better
- Entry at POI, stop beyond the structure that invalidates, target at next external liquidity
"""

# Migrate strategy_documents CHECK constraint to include trading/trading_strategy
try:
    from src.storage.portfolio_db import get_connection as _gc
    _mc = _gc()
    _mc.execute("""
        CREATE TABLE IF NOT EXISTS strategy_documents_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            category TEXT NOT NULL CHECK(category IN ('bear','bull','stablecoin','cashflow_other','trading','trading_strategy')),
            extracted_text TEXT,
            file_size_bytes INTEGER,
            notes TEXT DEFAULT '',
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _mc.execute("""
        INSERT OR IGNORE INTO strategy_documents_new
        SELECT id, filename, category, extracted_text, file_size_bytes, notes, uploaded_at
        FROM strategy_documents
    """)
    _mc.execute("DROP TABLE IF EXISTS strategy_documents")
    _mc.execute("ALTER TABLE strategy_documents_new RENAME TO strategy_documents")
    _mc.commit()
    _mc.close()
    print("[startup] strategy_documents CHECK constraint updated", flush=True)
except Exception as _mc_err:
    print(f"[startup] strategy_documents migration skipped: {_mc_err}", flush=True)


def _normalize_symbol(sym):
    """Canonical watchlist symbol — the ONE form shared by manual-add, HL-import
    and the dedup migration. Strips dashes/spaces, uppercases, and gives a bare
    base coin a USDT quote so the same asset can never enter twice under two
    spellings:
        BTC->BTCUSDT   WBTC->WBTCUSDT   SOL->SOLUSDT   BTC-USDT->BTCUSDT
        BTCUSDT->BTCUSDT   WOWUSDC->WOWUSDC (kept)   xyz:CL->XYZ:CLUSDT
    A namespace prefix ("xyz:CL") is preserved. This replaces the older logic
    where manual-adds left BTC/ETH/BNB bare (they matched a quote-suffix list),
    which created "BTC" + "BTCUSDT" duplicate rows.

    KNOWN LIMITATION (out of scope): an HL spot import as -USDC (WOW->WOWUSDC)
    vs a manual add of "WOW" (->WOWUSDT) keep different quotes and are treated as
    different assets. Rare; not the bug being fixed here.
    """
    if not sym:
        return sym
    s = str(sym).strip().upper().replace('-', '').replace(' ', '')
    # preserve a namespace prefix if present (e.g. "XYZ:CL")
    prefix = ''
    if ':' in s:
        prefix, s = s.split(':', 1)
        prefix += ':'
    # if it already ends in a quote suffix, keep as-is
    if s.endswith('USDT') or s.endswith('USDC'):
        return prefix + s
    # otherwise it's a bare base coin -> append USDT (canonical)
    return prefix + s + 'USDT'


# Add new columns to scanner_signals and scanner_watchlist for ICT evaluation
try:
    from src.storage.portfolio_db import get_connection as _gc2
    _mc2 = _gc2()
    for _col, _ctype in [
        ("htf_label", "TEXT"), ("ltf_label", "TEXT"),
        ("recent_closes_htf", "TEXT"), ("recent_closes_ltf", "TEXT"),
        ("concepts_triggered", "TEXT"), ("why_flagged", "TEXT"),
        ("proposed_entry", "REAL"), ("proposed_stop", "REAL"),
        ("proposed_target", "REAL"), ("rr_ratio", "REAL"),
        ("confidence_score", "INTEGER"), ("status", "TEXT"),
        ("signal_text", "TEXT"), ("htf_timeframe", "TEXT"),
        ("ltf_timeframe", "TEXT"), ("raw_indicators_json", "TEXT"),
        ("current_price", "REAL"),
    ]:
        try:
            _mc2.execute(f"ALTER TABLE scanner_signals ADD COLUMN {_col} {_ctype}")
        except Exception:
            pass
    for _col, _ctype in [
        ("htf_timeframe", "TEXT DEFAULT '4h'"),
        ("ltf_timeframe", "TEXT DEFAULT '15m'"),
        ("contract_address", "TEXT DEFAULT ''"),
        # Constant-default columns that the canonical schema relies on. These used
        # to appear only via the old table-rebuild; add them here (ALTER-safe — all
        # constant defaults) so they exist on fresh/legacy DBs too. created_at is
        # handled by the dedup+rebuild below (its CURRENT_TIMESTAMP default cannot
        # be added via ALTER).
        ("asset_type", "TEXT DEFAULT 'crypto'"),
        ("is_active", "INTEGER DEFAULT 1"),
        ("display_order", "INTEGER DEFAULT 0"),
    ]:
        try:
            _mc2.execute(f"ALTER TABLE scanner_watchlist ADD COLUMN {_col} {_ctype}")
        except Exception:
            pass
    _mc2.commit()
    _mc2.close()
    print("[startup] scanner column migrations applied", flush=True)
except Exception as _mc2_err:
    print(f"[startup] scanner column migration skipped: {_mc2_err}", flush=True)

# ── One-time DEDUP + rebuild scanner_watchlist to UNIQUE(symbol) ──────────────
# Supersedes the prior UNIQUE(symbol, htf_timeframe, ltf_timeframe) design. The
# scanner drives timeframes from V3_PAIRS, so htf/ltf on a watchlist row are
# vestigial and a symbol must never legitimately appear twice. Manual-add and
# HL-import used to normalize symbols differently (e.g. "BTC" vs "BTCUSDT"),
# producing duplicate rows; this collapses them via the shared _normalize_symbol
# and enforces UNIQUE(symbol) so duplicates become structurally impossible.
#
# Conservative by construction: single explicit transaction, fully logged,
# idempotent, and it ABORTS (rolls back, leaves the table untouched) on anything
# ambiguous — specifically when two rows that collapse to one canonical symbol
# carry DIFFERENT non-empty contract_addresses (i.e. genuinely different assets).
# The rebuild reproduces the table's CURRENT columns from PRAGMA table_info (no
# hardcoded list), so any ALTER-added column and its type/default survive.
#
# NOTE: scanner_signals is intentionally left untouched — it is regenerated on
# every scan and stale rows age out, so it needs no dedup here.
try:
    from src.storage.portfolio_db import get_connection as _gc3
    _mc3 = _gc3()
    _mc3.isolation_level = None   # explicit BEGIN/COMMIT/ROLLBACK; no auto-txn around DDL
    _tbl_row = _mc3.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='scanner_watchlist'"
    ).fetchone()
    _tbl_sql = (_tbl_row[0] or '') if _tbl_row else ''
    # Idempotency guard: detect the *single-column* UNIQUE(symbol). The old
    # UNIQUE(symbol, htf_timeframe, ltf_timeframe) compacts to "unique(symbol,"
    # which does NOT contain the substring "unique(symbol)".
    _compact = ''.join(_tbl_sql.lower().split())
    if not _tbl_sql:
        print("[startup] watchlist dedup+rebuild skipped: table absent", flush=True)
    elif 'unique(symbol)' in _compact:
        print("[startup] watchlist dedup+rebuild skipped: already migrated", flush=True)
    else:
        # Capture the CURRENT schema so the rebuilt table preserves every column
        # (and its type/default/PK) — we only swap the UNIQUE(...) constraint.
        _info = _mc3.execute("PRAGMA table_info(scanner_watchlist)").fetchall()
        _cols = [_ci[1] for _ci in _info]          # column names, in order
        _has = set(_cols)
        # Explicit indexes (sql IS NULL = auto-index from a constraint; skip those).
        _index_sqls = [_ir[0] for _ir in _mc3.execute(
            "SELECT sql FROM sqlite_master WHERE type='index' AND tbl_name='scanner_watchlist' "
            "AND sql IS NOT NULL"
        ).fetchall()]

        _rows = [dict(zip(_cols, r)) for r in
                 _mc3.execute("SELECT * FROM scanner_watchlist").fetchall()]

        def _nonempty(_v):
            return _v is not None and str(_v).strip() != ''

        # Group rows by canonical symbol (single source of truth: _normalize_symbol).
        _groups = {}
        for _r in _rows:
            _c = _normalize_symbol(_r.get('symbol'))
            _groups.setdefault(_c, []).append(_r)

        # Abort guard: a canonical group whose rows carry two *distinct* non-empty
        # contract_addresses is almost certainly two different assets that merely
        # normalize alike — collapsing would silently drop a real asset. Bail.
        _ambiguous = []
        for _c, _grp in _groups.items():
            _addrs = {str(_r.get('contract_address')).strip().lower()
                      for _r in _grp if _nonempty(_r.get('contract_address'))}
            if len(_addrs) > 1:
                _ambiguous.append((_c, sorted(_addrs)))

        if _ambiguous:
            for _c, _ad in _ambiguous:
                print(f"[startup] watchlist dedup+rebuild ABORTED — ambiguous '{_c}': "
                      f"distinct contract_addresses {_ad}", flush=True)
            print("[startup] watchlist dedup+rebuild left table UNTOUCHED "
                  "(resolve duplicates manually, then restart)", flush=True)
        else:
            # Deterministic survivor per canonical group:
            #   1. a row whose stored symbol already equals the canonical form
            #   2. then a row with a non-empty contract_address (on-chain info)
            #   3. then a row with a non-null last_scanned_at (more scan state)
            #   4. tiebreak: lowest id (oldest / most stable)
            def _rank(_r, _canon_sym):
                return (
                    1 if str(_r.get('symbol') or '').strip().upper() == _canon_sym else 0,
                    1 if _nonempty(_r.get('contract_address')) else 0,
                    1 if _nonempty(_r.get('last_scanned_at')) else 0,
                    -(_r.get('id') or 0),
                )

            # Reconstruct a column definition from PRAGMA table_info so the rebuilt
            # table matches the live schema exactly, minus the constraint swap.
            def _coldef(_ci):
                _name, _ctype, _nn, _dflt, _pk = _ci[1], (_ci[2] or 'TEXT'), _ci[3], _ci[4], _ci[5]
                _parts = ['"%s"' % _name, _ctype]
                if _pk:
                    _parts.append('PRIMARY KEY')
                    if _ctype.upper() == 'INTEGER':
                        _parts.append('AUTOINCREMENT')   # only valid on INTEGER PK
                elif _nn:
                    _parts.append('NOT NULL')
                if _dflt is not None:
                    _parts.append('DEFAULT %s' % _dflt)   # table_info returns the SQL literal verbatim
                return ' '.join(_parts)

            _dup_groups = sum(1 for _grp in _groups.values() if len(_grp) > 1)

            _mc3.execute("DROP TABLE IF EXISTS scanner_watchlist_new")   # clear any stale temp
            _mc3.execute("BEGIN")
            try:
                # (a) DEDUP in place: collapse each canonical group onto one survivor.
                for _c, _grp in _groups.items():
                    _surv = sorted(_grp, key=lambda _r: _rank(_r, _c), reverse=True)[0]
                    _drops = [_r for _r in _grp if _r.get('id') != _surv.get('id')]
                    if _drops:
                        # Carry the richest data onto the survivor before dropping the rest.
                        _set = {}
                        if 'last_scanned_at' in _has:
                            _ls = [_r.get('last_scanned_at') for _r in _grp
                                   if _nonempty(_r.get('last_scanned_at'))]
                            if _ls:
                                _set['last_scanned_at'] = max(_ls)   # most recent (ISO 'Z' sorts lexicographically)
                        for _f in ('contract_address', 'asset_type', 'notes'):
                            # survivor's value wins; else first non-empty from the dropped rows.
                            if _f in _has and not _nonempty(_surv.get(_f)):
                                _fb = next((_r.get(_f) for _r in _drops if _nonempty(_r.get(_f))), None)
                                if _fb is not None:
                                    _set[_f] = _fb
                        if _set:
                            _assign = ', '.join('"%s"=?' % _k for _k in _set)
                            _mc3.execute('UPDATE scanner_watchlist SET %s WHERE id=?' % _assign,
                                         tuple(_set.values()) + (_surv.get('id'),))
                        for _d in _drops:
                            _mc3.execute("DELETE FROM scanner_watchlist WHERE id=?", (_d.get('id'),))
                            print("[startup] watchlist dedup: merged %s — kept id=%s[%s], "
                                  "dropped id=%s[%s]" % (_c, _surv.get('id'), _surv.get('symbol'),
                                                         _d.get('id'), _d.get('symbol')), flush=True)
                    # (b) NORMALIZE: rewrite the survivor's symbol to canonical if it differs.
                    #     Post-dedup each canonical maps to one row, so no collision is possible.
                    if str(_surv.get('symbol') or '') != _c:
                        _mc3.execute("UPDATE scanner_watchlist SET symbol=? WHERE id=?",
                                     (_c, _surv.get('id')))

                # (c) REBUILD with UNIQUE(symbol), reproducing ALL current columns.
                _defs = [_coldef(_ci) for _ci in _info]
                _ins_cols = list(_cols)        # target columns in the new table
                _sel_exprs = ['"%s"' % _n for _n in _cols]   # how to source each from the old table
                # Guarantee created_at: downstream code does ORDER BY created_at, but its
                # CURRENT_TIMESTAMP default can't be ALTER-added, so seed it here if absent
                # (fall back to a legacy added_at column, else now).
                if 'created_at' not in _has:
                    _defs.append('"created_at" TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
                    _ins_cols.append('created_at')
                    _sel_exprs.append('COALESCE("added_at", CURRENT_TIMESTAMP)'
                                      if 'added_at' in _has else 'CURRENT_TIMESTAMP')
                _mc3.execute(
                    "CREATE TABLE scanner_watchlist_new (\n                        %s,\n"
                    "                        UNIQUE(symbol)\n                    )"
                    % ',\n                        '.join(_defs)
                )
                _mc3.execute(
                    "INSERT INTO scanner_watchlist_new (%s) SELECT %s FROM scanner_watchlist"
                    % (', '.join('"%s"' % _n for _n in _ins_cols), ', '.join(_sel_exprs))
                )
                _mc3.execute("DROP TABLE scanner_watchlist")
                _mc3.execute("ALTER TABLE scanner_watchlist_new RENAME TO scanner_watchlist")
                # Recreate any explicit indexes that lived on the old table (best-effort).
                for _isql in _index_sqls:
                    try:
                        _mc3.execute(_isql)
                    except Exception as _ie:
                        print(f"[startup] watchlist dedup+rebuild: index recreate skipped ({_ie})", flush=True)
                _mc3.execute("COMMIT")
                print(f"[startup] watchlist dedup+rebuild done: {len(_rows)} row(s) -> "
                      f"{len(_groups)} canonical ({_dup_groups} group(s) collapsed); "
                      f"UNIQUE(symbol) enforced", flush=True)
            except Exception as _tx_err:
                _mc3.execute("ROLLBACK")
                try:
                    _mc3.execute("DROP TABLE IF EXISTS scanner_watchlist_new")
                except Exception:
                    pass
                print(f"[startup] watchlist dedup+rebuild ROLLED BACK (table untouched): "
                      f"{_tx_err}", flush=True)
    _mc3.close()
except Exception as _mc3_err:
    print(f"[startup] watchlist dedup+rebuild skipped: {_mc3_err}", flush=True)

# Add last_scanned_at to scanner_watchlist (idempotent). The dedup+rebuild above
# preserves it when present; on a brand-new DB it doesn't exist yet at rebuild time,
# so this ALTER adds it afterward (nullable TIMESTAMP — ALTER-safe). Stamps when a
# ticker was last scanned, regardless of drop/survivor outcome.
try:
    from src.storage.portfolio_db import get_connection as _gc_lsa
    _mc_lsa = _gc_lsa()
    _wl_cols = [r[1] for r in _mc_lsa.execute("PRAGMA table_info(scanner_watchlist)").fetchall()]
    if 'last_scanned_at' not in _wl_cols:
        _mc_lsa.execute("ALTER TABLE scanner_watchlist ADD COLUMN last_scanned_at TIMESTAMP")
        _mc_lsa.commit()
        print("[startup] scanner_watchlist: added last_scanned_at column", flush=True)
    _mc_lsa.close()
except Exception as _lsa_err:
    print(f"[startup] scanner_watchlist last_scanned_at migration skipped: {_lsa_err}", flush=True)

# Rebuild scanner_signals with UNIQUE(symbol, htf_timeframe, ltf_timeframe) if not already done
try:
    from src.storage.portfolio_db import get_connection as _gc4
    _mc4 = _gc4()
    _sig_row = _mc4.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='scanner_signals'"
    ).fetchone()
    _sig_sql = (_sig_row[0] or '') if _sig_row else ''
    if 'unique(symbol, htf_timeframe, ltf_timeframe)' not in _sig_sql.lower():
        _mc4.execute("""
            CREATE TABLE IF NOT EXISTS scanner_signals_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                htf_timeframe TEXT,
                ltf_timeframe TEXT,
                interval TEXT,
                signal_type TEXT,
                price REAL,
                signal_data_json TEXT,
                status TEXT DEFAULT 'quiet',
                signal_text TEXT,
                confidence_score INTEGER DEFAULT 0,
                why_flagged TEXT,
                proposed_entry REAL,
                proposed_stop REAL,
                proposed_target REAL,
                rr_ratio REAL,
                concepts_triggered TEXT,
                raw_indicators_json TEXT,
                htf_label TEXT,
                ltf_label TEXT,
                recent_closes_htf TEXT,
                recent_closes_ltf TEXT,
                current_price REAL,
                strategy_id INTEGER,
                detected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(symbol, htf_timeframe, ltf_timeframe)
            )
        """)
        _mc4.execute("""
            INSERT OR IGNORE INTO scanner_signals_new
            SELECT id, htf_timeframe, ltf_timeframe, interval, signal_type, price,
                   signal_data_json, status, signal_text, confidence_score, why_flagged,
                   proposed_entry, proposed_stop, proposed_target, rr_ratio,
                   concepts_triggered, raw_indicators_json, htf_label, ltf_label,
                   recent_closes_htf, recent_closes_ltf, current_price, NULL, detected_at
            FROM scanner_signals
        """)
        _mc4.execute("DROP TABLE IF EXISTS scanner_signals")
        _mc4.execute("ALTER TABLE scanner_signals_new RENAME TO scanner_signals")
        _mc4.commit()
        print("[startup] scanner_signals UNIQUE constraint updated to symbol+htf+ltf", flush=True)
    else:
        print("[startup] scanner_signals UNIQUE constraint already correct, skipping", flush=True)
    _mc4.close()
except Exception as _mc4_err:
    print(f"[startup] scanner_signals migration skipped: {_mc4_err}", flush=True)

# Scanner v2 migration: add pair_key column, switch to UNIQUE(symbol, pair_key),
# and clear stale pre-v2 signals (NULL pair_key) so the grouped UI starts clean.
try:
    from src.storage.portfolio_db import get_connection as _gc4b
    _mc4b = _gc4b()
    try:
        _mc4b.execute("ALTER TABLE scanner_signals ADD COLUMN pair_key TEXT")
    except Exception:
        pass  # column already exists
    try:
        _mc4b.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_signal_pair ON scanner_signals(symbol, pair_key)")
    except Exception:
        pass
    try:
        _mc4b.execute("DELETE FROM scanner_signals WHERE pair_key IS NULL")
    except Exception:
        pass
    _mc4b.commit()
    _mc4b.close()
    print("[startup] scanner_signals v2 pair_key migration applied", flush=True)
except Exception as _mc4b_err:
    print(f"[startup] scanner_signals v2 migration skipped: {_mc4b_err}", flush=True)

# Scanner v3 migration: clear stale swing/intra/scalp pair_key rows so the
# grouped UI starts clean with HTF-name keys (1W, 1D, 12H, 4H, 1H).
try:
    from src.storage.portfolio_db import get_connection as _gc4c
    _mc4c = _gc4c()
    _mc4c.execute(
        "DELETE FROM scanner_signals WHERE pair_key IN ('swing', 'intra', 'scalp')"
    )
    _mc4c.commit()
    _mc4c.close()
    print("[startup] scanner_signals v3 migration: stale v2 pair_key rows cleared", flush=True)
except Exception as _mc4c_err:
    print(f"[startup] scanner_signals v3 migration skipped: {_mc4c_err}", flush=True)

# Add new columns to trading_journal for trade tracking
try:
    from src.storage.portfolio_db import get_connection as _gc5
    _mc5 = _gc5()
    for _jcol, _jtype in [
        ("ticker", "TEXT"),
        ("direction", "TEXT"),
        ("timeframe", "TEXT"),
        ("outcome", "TEXT DEFAULT 'open'"),
        ("r_multiple", "REAL"),
        ("ready_score", "TEXT"),
        ("execution_plan_json", "TEXT"),
        ("concepts_json", "TEXT"),
        ("validator_answers_json", "TEXT"),
    ]:
        try:
            _mc5.execute(f"ALTER TABLE trading_journal ADD COLUMN {_jcol} {_jtype}")
        except Exception:
            pass  # column already exists
    _mc5.commit()
    _mc5.close()
    print("[startup] trading_journal column migrations applied", flush=True)
except Exception as _mc5_err:
    print(f"[startup] trading_journal migration skipped: {_mc5_err}", flush=True)

# Create journal_screenshots table
try:
    from src.storage.portfolio_db import get_connection as _gc_js
    _mc_js = _gc_js()
    _mc_js.execute("""
        CREATE TABLE IF NOT EXISTS journal_screenshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_entry_id INTEGER NOT NULL,
            filename TEXT,
            mime_type TEXT,
            data TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (journal_entry_id) REFERENCES trading_journal(id)
        )
    """)
    _mc_js.commit()
    _mc_js.close()
    print("[startup] journal_screenshots table ready", flush=True)
except Exception as _mc_js_err:
    print(f"[startup] journal_screenshots migration skipped: {_mc_js_err}", flush=True)

# Create strategies table and seed Mayne OB/FVG strategy if empty
_MAYNE_FLOW_JSON = '{"id":"mayne_ob_fvg","name":"Mayne — OB/FVG System","version":"1.1","step_labels":["HTF Bias","Dealing Range","Context Filter","POI Quality","Prelim R:R","HTF POI Valid","LTF Confirm","Risk / Inv."],"base_flow":[{"id":"htf_bias","label":"HTF Bias","stepType":"HTF","concept":"market structure","question":"What is the HTF bias and market structure?","subtext":"Mark your weekly or daily dealing range. Identify the most recent decisive MSB/BOS direction.","options":[{"key":"A","title":"Bullish — last decisive MSB to upside","description":"Most recent MSB swept a prior swing low then drove up with follow-through. Looking for longs.","branches":3},{"key":"B","title":"Bearish — last decisive MSB to downside","description":"Most recent MSB swept a prior swing high then drove down with follow-through. Looking for shorts.","branches":3},{"key":"C","title":"Ranging — no clear MSB","description":"Price oscillating between equal highs and lows. No decisive break in either direction.","branches":2},{"key":"D","title":"Bias unclear — conflicting signals","description":"Multiple conflicting MSBs; structure is ambiguous.","endsFlow":true}]},{"id":"dealing_range","label":"Dealing Range","stepType":"HTF","concept":"dealing range","question":"Where is price within the HTF dealing range, and does it align with your bias?","subtext":"EQ = 50% of the weekly/daily range. Bullish bias requires price in discount (below EQ). Bearish requires premium (above EQ). OTE zone is 61.8–78.6%.","options":[{"key":"A","title":"Discount — below EQ, bias is bullish","description":"Price below 50% equilibrium. Aligned for longs. Ideal if inside 61.8–78.6% OTE.","branches":3},{"key":"B","title":"Premium — above EQ, bias is bearish","description":"Price above 50% equilibrium. Aligned for shorts. Ideal if inside 61.8–78.6% OTE.","branches":3},{"key":"C","title":"At equilibrium (near 50%)","description":"Price at the midpoint. Less ideal — higher risk, lower probability.","branches":2},{"key":"D","title":"Wrong zone for bias — structurally misaligned","description":"Bullish bias but price in premium, or bearish bias in discount. Low-probability. Do not force.","endsFlow":true}]},{"id":"context_filter","label":"Context Filter","stepType":"HTF","concept":"market context","question":"What is the macro context, and which POI does it favour?","subtext":"Aggressive trend = price unlikely to reach OB, prioritise FVG. Ranging = OB provides cleaner boundaries. Reversal at major high/low = OB or Breaker only.","options":[{"key":"A","title":"Aggressive trend — fast momentum, shallow pullbacks","description":"Price leaving levels quickly. OB likely stays untouched. Prioritise FVG.","branches":3},{"key":"B","title":"Ranging / sideways — choppy price action","description":"Consecutive FVGs failing. OB provides cleaner structural boundary. Prioritise OB.","branches":3},{"key":"C","title":"Reversal — approaching major HTF high or low","description":"Key macro turning point. OB and Breaker Blocks dominate. FVG entries low priority.","branches":3},{"key":"D","title":"Context unclear — cannot classify","description":"No clear read on market regime. Lower conviction — consider standing aside.","branches":1}]},{"id":"poi_quality","label":"POI Quality","stepType":"HTF","concept":"order block","question":"What POIs are present in the zone, and do they meet quality criteria?","subtext":"High-quality OB: swept prior liquidity + explosive MSB + created its own FVG. High-quality FVG: massive displacement candles after OB or MSB. Unicorn = OB + FVG + Breaker overlapping.","options":[{"key":"A","title":"OB + FVG + Breaker overlapping — unicorn setup","description":"All three POIs stacked in the same zone. Highest priority. Size up or prioritise accordingly.","branches":3},{"key":"B","title":"High-quality OB only — swept liq, MSB, created FVG","description":"OB meets all three quality criteria. Clean structural anchor.","branches":3},{"key":"C","title":"High-quality FVG only — massive displacement after MSB","description":"Large obvious FVG with clear displacement. No qualifying OB present.","branches":3},{"key":"D","title":"POI present but quality criteria not met — ambiguous","description":"Setup is not visually obvious. Per Mayne: if ambiguous, discard it.","endsFlow":true}]},{"id":"prelim_rr","label":"Prelim R:R","stepType":"HTF","concept":"risk management","question":"Does the HTF structure support a minimum 2R trade?","subtext":"Using POI boundaries as your rough entry and next external liquidity as target — does the structure allow at least 2R before hitting opposing liquidity or invalidation? This is a structural eyeball, not exact prices.","options":[{"key":"A","title":"Yes — structure clearly supports 2R or better","description":"Enough room between POI and target. Proceed to LTF confirmation.","branches":3},{"key":"B","title":"Marginal — structure might support 2R at OB but not FVG","description":"FVG entry too close to price for 2R. Skip FVG, wait for OB level only.","branches":2},{"key":"C","title":"No — insufficient room for 2R at any POI","description":"Structure does not support a minimum 2R trade. Pass on this setup entirely.","endsFlow":true}]},{"id":"htf_poi_valid","label":"HTF POI Valid","stepType":"HTF","concept":"order block","question":"Is the HTF Order Block / FVG still valid?","subtext":"Check the HTF POI that brought price into this zone. A body close past the 50% midpoint indicates structural weakness. A body close through the outer boundary, or price wicking fully through it on an open candle, invalidates the POI entirely. OTE invalidation: a candle close through the 78.6 fib level opposite to trend also invalidates the setup.","options":[{"key":"A","title":"Valid — price wicked in, no body close past 50% midpoint","description":"POI is holding. Structure intact. Proceed to LTF confirmation.","branches":0},{"key":"B","title":"Weakened — body closed past 50% midpoint but not through outer boundary","description":"Structural weakness present. POI is degraded but not invalidated. Proceed with caution — reduce size or wait for stronger LTF confirmation.","warn":true,"branches":0},{"key":"C","title":"Invalidated — body closed through the outer boundary of the POI","description":"POI is fully invalidated by a candle body close. Step aside. Reassess HTF structure.","endsFlow":true},{"key":"D","title":"Invalidated (live wick) — price traded through outer boundary on open candle","description":"Even without a candle close, price has fully breached the POI outer boundary. Treat as invalidated. Wait for candle close to confirm, then reassess.","endsFlow":true},{"key":"E","title":"OTE invalidated — candle closed through 78.6 fib level against trend","description":"The OTE zone is broken. Setup no longer has a valid entry framework. Step aside and reassess dealing range.","endsFlow":true}]},{"id":"ltf_confirm","label":"LTF Confirm","stepType":"LTF","concept":"market structure shift","question":"Has the LTF produced a Market Structure Shift (MSS) at your entry level?","subtext":"Drop to entry timeframe when price enters the HTF zone. You will be counter-trend. Wait for LTF MSS to flip back in alignment with HTF direction. Breaker is Mayne\'s preferred LTF entry tool.","options":[{"key":"A","title":"LTF MSS fired at FVG level","description":"Counter-trend LTF structure shifted back to HTF direction at the FVG. Enter using LTF Breaker, OB, or FVG on entry TF.","branches":3},{"key":"B","title":"LTF MSS did not fire at FVG — fired at OB level","description":"FVG passed through without LTF structure change. MSS confirmed at OB. Enter at OB using LTF tools.","branches":3},{"key":"C","title":"FVG aggressively breached without LTF MSS — FVG invalidated","description":"FVG traded through with no structure shift. No valid entry here. Re-run validator when price reaches OB level.","endsFlow":true},{"key":"D","title":"OB also breached without LTF MSS — entire thesis invalidated","description":"Both levels failed without structure confirmation. Higher timeframe reversal likely underway. Step aside completely. Reassess HTF thesis.","endsFlow":true}]}],"adaptive_steps":{},"risk_step":{"id":"risk","label":"Risk / Inv.","stepType":"LTF","concept":"risk management","isForm":true,"question":"Define your exact entry, stop, and targets.","subtext":"LTF entry is now confirmed. Enter precise price levels. Stop goes beyond the LTF structure that invalidates. Target is next external liquidity. R:R calculated automatically — must confirm 2R minimum established at Step 5."}}'

try:
    from src.storage.portfolio_db import get_connection as _gc6
    _mc6 = _gc6()
    _mc6.execute("""
        CREATE TABLE IF NOT EXISTS strategies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            ai_prompt TEXT,
            flow_json TEXT,
            is_default INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    _mc6.commit()
    # Add strategy_id to scanner_signals if missing
    try:
        _mc6.execute("ALTER TABLE scanner_signals ADD COLUMN strategy_id INTEGER")
        _mc6.commit()
    except Exception:
        pass  # column already exists
    # Seed Mayne strategy if table is empty
    _strat_count = _mc6.execute("SELECT COUNT(*) FROM strategies").fetchone()[0]
    if _strat_count == 0:
        try:
            with open(_get_scanner_prompt_path(), 'r') as _sp_f:
                _seed_prompt = _sp_f.read()
            print(f"[startup] Strategy seed: ai_prompt loaded from scanner_prompt.md ({len(_seed_prompt)} chars)", flush=True)
        except FileNotFoundError:
            _seed_prompt = ''
            print("[startup] Strategy seed: scanner_prompt.md not found, ai_prompt left empty", flush=True)
        except Exception as _seed_err:
            _seed_prompt = ''
            print(f"[startup] Strategy seed: could not read scanner_prompt.md: {_seed_err}", flush=True)
        _mc6.execute(
            "INSERT INTO strategies (name, description, ai_prompt, flow_json, is_default, is_active) VALUES (?,?,?,?,1,1)",
            ('Mayne — OB/FVG System', 'ICT/SMC approach using Order Blocks, Fair Value Gaps, and Breaker Blocks. Developed from Mayne methodology.', _seed_prompt, _MAYNE_FLOW_JSON)
        )
        _mc6.commit()
        print("[startup] strategies table seeded with Mayne OB/FVG strategy", flush=True)
    else:
        print(f"[startup] strategies table already has {_strat_count} row(s), skipping seed", flush=True)
    _mc6.close()
    print("[startup] strategies table ready", flush=True)
except Exception as _mc6_err:
    print(f"[startup] strategies migration failed: {_mc6_err}", flush=True)

# Migration: insert the "HTF POI Valid" step before "ltf_confirm" in the live
# strategies flow_json. Idempotent — skips any row that already has the step.
# (Defined as a Python dict here, not re-parsed from _MAYNE_FLOW_JSON.)
_HTF_POI_VALID_STEP = {
    "id": "htf_poi_valid",
    "label": "HTF POI Valid",
    "stepType": "HTF",
    "concept": "order block",
    "question": "Is the HTF Order Block / FVG still valid?",
    "subtext": "Check the HTF POI that brought price into this zone. A body close past the 50% midpoint indicates structural weakness. A body close through the outer boundary, or price wicking fully through it on an open candle, invalidates the POI entirely. OTE invalidation: a candle close through the 78.6 fib level opposite to trend also invalidates the setup.",
    "options": [
        {"key": "A", "title": "Valid — price wicked in, no body close past 50% midpoint", "description": "POI is holding. Structure intact. Proceed to LTF confirmation.", "branches": 0},
        {"key": "B", "title": "Weakened — body closed past 50% midpoint but not through outer boundary", "description": "Structural weakness present. POI is degraded but not invalidated. Proceed with caution — reduce size or wait for stronger LTF confirmation.", "warn": True, "branches": 0},
        {"key": "C", "title": "Invalidated — body closed through the outer boundary of the POI", "description": "POI is fully invalidated by a candle body close. Step aside. Reassess HTF structure.", "endsFlow": True},
        {"key": "D", "title": "Invalidated (live wick) — price traded through outer boundary on open candle", "description": "Even without a candle close, price has fully breached the POI outer boundary. Treat as invalidated. Wait for candle close to confirm, then reassess.", "endsFlow": True},
        {"key": "E", "title": "OTE invalidated — candle closed through 78.6 fib level against trend", "description": "The OTE zone is broken. Setup no longer has a valid entry framework. Step aside and reassess dealing range.", "endsFlow": True},
    ],
}
try:
    from src.storage.portfolio_db import get_connection as _gc_poi
    _mc_poi = _gc_poi()
    _poi_rows = _mc_poi.execute("SELECT id, flow_json FROM strategies WHERE is_active=1").fetchall()
    for _prow in _poi_rows:
        try:
            _pflow = json.loads(_prow['flow_json']) if _prow['flow_json'] else None
        except Exception:
            _pflow = None
        if not isinstance(_pflow, dict) or not isinstance(_pflow.get('base_flow'), list):
            continue
        _bf = _pflow['base_flow']
        _ids = [s.get('id') for s in _bf if isinstance(s, dict)]
        if 'htf_poi_valid' in _ids:
            print("[startup] validator: htf_poi_valid already present, skipping", flush=True)
            continue
        if 'ltf_confirm' not in _ids:
            continue  # no anchor in this flow — leave it untouched
        _bf.insert(_ids.index('ltf_confirm'), dict(_HTF_POI_VALID_STEP))
        # Keep step_labels in sync — it drives the frontend step counter/progress UI.
        _labels = _pflow.get('step_labels')
        if isinstance(_labels, list) and 'HTF POI Valid' not in _labels and 'LTF Confirm' in _labels:
            _labels.insert(_labels.index('LTF Confirm'), 'HTF POI Valid')
        _mc_poi.execute("UPDATE strategies SET flow_json=? WHERE id=?",
                        (json.dumps(_pflow), _prow['id']))
        _mc_poi.commit()
        print("[startup] validator: inserted htf_poi_valid step before ltf_confirm", flush=True)
    _mc_poi.close()
except Exception as _poi_err:
    print(f"[startup] validator htf_poi_valid migration skipped: {_poi_err}", flush=True)

# Create default scanner_prompt.md on persistent volume if absent
try:
    _prompt_dir = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', '').strip() or '/app/data'
    _prompt_path = os.path.join(_prompt_dir, 'scanner_prompt.md')
    if not os.path.exists(_prompt_path):
        os.makedirs(_prompt_dir, exist_ok=True)
        with open(_prompt_path, 'w') as _spf:
            _spf.write(_DEFAULT_SCANNER_PROMPT)
        print("[startup] scanner_prompt.md created", flush=True)
except Exception as _sp_err:
    print(f"[startup] scanner_prompt.md creation skipped: {_sp_err}", flush=True)

# Backfill ai_prompt for the single seeded strategy if it was left empty
# (seed runs before scanner_prompt.md is written, so the file is always missing at seed time)
try:
    from src.storage.portfolio_db import get_connection as _gc_bp
    _bp_conn = _gc_bp()
    _bp_row = _bp_conn.execute(
        "SELECT id, ai_prompt FROM strategies WHERE is_active=1"
    ).fetchall()
    if len(_bp_row) == 1 and not (_bp_row[0]['ai_prompt'] or '').strip():
        try:
            with open(_get_scanner_prompt_path(), 'r') as _bp_f:
                _bp_prompt = _bp_f.read()
            if _bp_prompt.strip():
                _bp_conn.execute(
                    "UPDATE strategies SET ai_prompt=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (_bp_prompt, _bp_row[0]['id'])
                )
                _bp_conn.commit()
                print(f"[startup] Strategy seed: ai_prompt loaded from scanner_prompt.md ({len(_bp_prompt)} chars)", flush=True)
            else:
                print("[startup] Strategy seed: scanner_prompt.md exists but is empty, ai_prompt left empty", flush=True)
        except FileNotFoundError:
            print("[startup] Strategy seed: scanner_prompt.md not found, ai_prompt left empty", flush=True)
        except Exception as _bp_read_err:
            print(f"[startup] Strategy seed: could not read scanner_prompt.md: {_bp_read_err}", flush=True)
    _bp_conn.close()
except Exception as _bp_err:
    print(f"[startup] Strategy ai_prompt backfill skipped: {_bp_err}", flush=True)

# Normalise spot_transactions.trade_date to D/M/YYYY (no leading zeros)
try:
    from src.storage.portfolio_db import get_connection as _gc_dt
    _dt_conn = _gc_dt()
    _dt_rows = _dt_conn.execute("SELECT id, trade_date FROM spot_transactions").fetchall()
    _dt_updated = 0
    for _dt_r in _dt_rows:
        _normalized = normalize_date(_dt_r['trade_date'])
        if _normalized and _normalized != _dt_r['trade_date']:
            _dt_conn.execute("UPDATE spot_transactions SET trade_date=? WHERE id=?",
                             (_normalized, _dt_r['id']))
            _dt_updated += 1
    if _dt_updated:
        _dt_conn.commit()
    _dt_conn.close()
    print(f"[startup] spot_transactions date normalisation: {_dt_updated} rows updated", flush=True)
except Exception as _dt_err:
    print(f"[startup] spot_transactions date normalisation skipped: {_dt_err}", flush=True)

# Startup diagnostics — module-level so gunicorn always runs them; flush=True bypasses buffering
_startup_db_path = get_db_path()
print(f"[startup] db path: {_startup_db_path}", flush=True)
try:
    import sqlite3 as _sq3
    _diag_conn = _sq3.connect(_startup_db_path)
    _spot_rows = _diag_conn.execute("SELECT COUNT(*) FROM spot_transactions").fetchone()[0]
    print(f"[startup] spot_transactions rows: {_spot_rows}", flush=True)
    _diag_conn.close()
except Exception as _diag_err:
    print(f"[startup] spot_transactions check failed: {_diag_err}", flush=True)
if os.path.exists('/app/data'):
    print(f"[startup] /app/data contents: {os.listdir('/app/data')}", flush=True)
else:
    print("[startup] /app/data does not exist", flush=True)
_zerion_key = os.environ.get("ZERION_API_KEY", "")
if _zerion_key:
    print(f"[startup] zerion key prefix: {_zerion_key[:4]}...", flush=True)
else:
    print("[startup] zerion key: NOT SET", flush=True)

# Background scheduler
_scheduler_started = False


def start_snapshot_scheduler():
    """Start background scheduler for portfolio (2h) and market (3x daily) snapshots."""
    global _scheduler_started
    if _scheduler_started:
        return
    _scheduler_started = True
    from src.engines.snapshot_service import start_scheduler
    start_scheduler(get_portfolio_data, get_wallet_addresses)


@app.route('/api/snapshot', methods=['POST'])
def api_take_snapshot():
    """Manually trigger a portfolio snapshot."""
    try:
        wallets = get_wallet_addresses()
        if not wallets:
            return jsonify({"error": "No wallets configured"}), 400
        from src.engines.snapshot_service import take_portfolio_snapshot
        take_portfolio_snapshot(get_portfolio_data, wallets)
        return jsonify({"status": "success", "message": f"Snapshot completed for {len(wallets)} wallet(s)"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/history/portfolio')
def api_history_portfolio():
    """Get portfolio value timeseries for charts."""
    days = request.args.get('days', 90, type=int)
    wallet = request.args.get('wallet')
    data = get_portfolio_timeseries(days=days, wallet=wallet)
    return jsonify(data)


@app.route('/api/history/portfolio-chart')
def api_history_portfolio_chart():
    """Get aggregated portfolio chart data — sums across all wallets per timestamp."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    
    days = request.args.get('days', 30, type=int)
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    
    if date_from:
        query = """SELECT ts,
            SUM(total_tokens_usd) as tokens_value,
            SUM(total_lp_usd) as lp_value,
            SUM(total_lending_usd) as lending_value,
            SUM(total_hedge_collateral_usd) as hedge_value,
            SUM(total_value_usd) as total_value
            FROM (
                SELECT strftime('%Y-%m-%dT%H:%M:00', timestamp) as ts, wallet,
                    total_tokens_usd, total_lp_usd, total_lending_usd, total_hedge_collateral_usd, total_value_usd,
                    ROW_NUMBER() OVER (PARTITION BY strftime('%Y-%m-%dT%H:%M:00', timestamp), wallet ORDER BY id DESC) as rn
                FROM portfolio_snapshots
                WHERE status='completed' AND timestamp >= ? AND timestamp <= ?
            ) WHERE rn=1
            GROUP BY ts ORDER BY ts ASC"""
        rows = conn.execute(query, (date_from, date_to + 'T23:59:59' if date_to else '9999-12-31')).fetchall()
    else:
        date_filter = f"AND timestamp >= datetime('now', '-{int(days)} days')" if days < 9999 else ""
        query = f"""SELECT ts,
            SUM(total_tokens_usd) as tokens_value,
            SUM(total_lp_usd) as lp_value,
            SUM(total_lending_usd) as lending_value,
            SUM(total_hedge_collateral_usd) as hedge_value,
            SUM(total_value_usd) as total_value
            FROM (
                SELECT strftime('%Y-%m-%dT%H:%M:00', timestamp) as ts, wallet,
                    total_tokens_usd, total_lp_usd, total_lending_usd, total_hedge_collateral_usd, total_value_usd,
                    ROW_NUMBER() OVER (PARTITION BY strftime('%Y-%m-%dT%H:%M:00', timestamp), wallet ORDER BY id DESC) as rn
                FROM portfolio_snapshots
                WHERE status='completed' {date_filter}
            ) WHERE rn=1
            GROUP BY ts ORDER BY ts ASC"""
        rows = conn.execute(query).fetchall()
    
    # Get total fees — for each portfolio timestamp, use the closest LP snapshot
    lp_fee_map = {}
    lp_fee_rows = conn.execute(
        """SELECT ts, COALESCE(SUM(fees_uncollected_usd), 0) as total_fees FROM (
            SELECT strftime('%Y-%m-%dT%H:%M:00', timestamp) as ts, position_id, fees_uncollected_usd,
                ROW_NUMBER() OVER (PARTITION BY strftime('%Y-%m-%dT%H:%M:00', timestamp), position_id ORDER BY id DESC) as rn
            FROM lp_snapshots
        ) WHERE rn=1 GROUP BY ts ORDER BY ts ASC"""
    ).fetchall()
    for fr in lp_fee_rows:
        lp_fee_map[fr['ts']] = fr['total_fees']
    lp_timestamps = sorted(lp_fee_map.keys())
    
    result = []
    for r in rows:
        ts = r['ts']
        # Find the latest LP timestamp <= this portfolio timestamp
        fees = 0
        if lp_timestamps:
            import bisect
            idx = bisect.bisect_right(lp_timestamps, ts)
            if idx > 0:
                fees = lp_fee_map[lp_timestamps[idx - 1]]
        
        result.append({
            'timestamp': ts,
            'tokens_value': r['tokens_value'] or 0,
            'lp_value': r['lp_value'] or 0,
            'lending_value': r['lending_value'] or 0,
            'hedge_value': r['hedge_value'] or 0,
            'total_value': r['total_value'] or 0,
            'total_fees': fees,
        })
    
    conn.close()
    return jsonify(result)


@app.route('/api/history/token/<symbol>')
def api_history_token(symbol):
    """Get token price history."""
    days = request.args.get('days', 90, type=int)
    data = get_token_price_history(symbol, days=days)
    return jsonify(data)


@app.route('/api/history/closed-positions')
def api_history_closed_positions():
    """Get closed LP and hedge positions by comparing snapshots."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()

    days = request.args.get('days', 9999, type=int)
    date_from = request.args.get('from')
    date_to = request.args.get('to')

    # Build a parameterized date filter — `?` placeholders only, no string
    # interpolation of request data. The integer `days` branch is safe because
    # Flask coerces the query param via type=int, but we re-cast defensively.
    if date_from:
        if date_to:
            date_filter = "AND timestamp >= ? AND timestamp <= ?"
            date_params = (date_from, date_to + 'T23:59:59')
        else:
            date_filter = "AND timestamp >= ?"
            date_params = (date_from,)
    elif days < 9999:
        date_filter = f"AND timestamp >= datetime('now', '-{int(days)} days')"
        date_params = ()
    else:
        date_filter = ""
        date_params = ()

    # Same fragment with the `timestamp` column rewritten to `updated_at`
    # for queries against lp_positions / hedge_positions which use that column.
    date_filter_updated_at = date_filter.replace('timestamp', 'updated_at')
    
    # Get the latest snapshot timestamp (rounded to minute)
    latest = conn.execute(
        "SELECT MAX(strftime('%Y-%m-%dT%H:%M:00', timestamp)) as ts FROM portfolio_snapshots WHERE status='completed'"
    ).fetchone()
    if not latest or not latest['ts']:
        # No snapshots yet, but check manual positions
        closed_lps = []
        closed_hedges = []
        manual_lps_early = conn.execute(
            f"SELECT * FROM lp_positions WHERE is_active=0 {date_filter_updated_at}",
            date_params,
        ).fetchall()
        for mp in manual_lps_early:
            mp = dict(mp)
            entry_value = mp.get('entry_value_usd') or mp.get('value_usd') or 0
            exit_value = mp.get('value_usd') or 0
            total_fees = (mp.get('fees_uncollected_usd') or 0) + (mp.get('fees_collected_usd') or 0)
            try:
                from datetime import datetime as dt
                t0 = dt.fromisoformat(str(mp.get('created_at', '')).replace('Z', ''))
                t1 = dt.fromisoformat(str(mp.get('updated_at', '')).replace('Z', ''))
                days_held = max((t1 - t0).total_seconds() / 86400, 1)
            except Exception:
                days_held = 1
            total_return = ((exit_value - entry_value + total_fees) / entry_value * 100) if entry_value > 0 else 0
            annual_apr = total_return * (365 / days_held) if days_held > 0 else 0
            closed_lps.append({
                'pair': (mp.get('token0') or '?') + '/' + (mp.get('token1') or '?'),
                'chain': mp.get('chain') or 'Manual',
                'entry_value': entry_value, 'exit_value': exit_value,
                'range_lower': mp.get('range_lower'), 'range_upper': mp.get('range_upper'),
                'total_fees': total_fees, 'total_return_pct': total_return,
                'annual_apr': annual_apr, 'days_held': int(days_held),
                'entry_date': str(mp.get('created_at', ''))[:10],
                'exit_date': str(mp.get('updated_at', ''))[:10],
            })
        manual_hedges_early = conn.execute(
            f"SELECT * FROM hedge_positions WHERE is_active=0 {date_filter_updated_at}",
            date_params,
        ).fetchall()
        for mh in manual_hedges_early:
            mh = dict(mh)
            closed_hedges.append({
                'market': mh.get('market') or '?', 'direction': mh.get('direction') or '?',
                'entry_price': mh.get('entry_price') or 0, 'exit_price': mh.get('current_price') or 0,
                'size_usd': mh.get('size_usd') or 0, 'pnl_usd': mh.get('pnl_usd') or 0,
                'entry_date': str(mh.get('created_at', ''))[:10],
                'exit_date': str(mh.get('updated_at', ''))[:10],
            })
        conn.close()
        return jsonify({'closed_lps': closed_lps, 'closed_hedges': closed_hedges})
    latest_ts = latest['ts']
    
    # Get position_ids in the latest snapshot
    active_lp_ids = set(r['position_id'] for r in conn.execute(
        "SELECT DISTINCT position_id FROM lp_snapshots WHERE strftime('%Y-%m-%dT%H:%M:00', timestamp)=?", (latest_ts,)
    ).fetchall())
    
    # Get all position_ids that existed within the date range
    all_lp_ids = set(r['position_id'] for r in conn.execute(
        f"SELECT DISTINCT position_id FROM lp_snapshots WHERE 1=1 {date_filter}",
        date_params,
    ).fetchall())
    
    closed_lp_ids = all_lp_ids - active_lp_ids
    
    closed_lps = []
    for pid in closed_lp_ids:
        # First snapshot = entry
        first = conn.execute(
            "SELECT * FROM lp_snapshots WHERE position_id=? ORDER BY timestamp ASC LIMIT 1", (pid,)
        ).fetchone()
        # Last snapshot = exit
        last = conn.execute(
            "SELECT * FROM lp_snapshots WHERE position_id=? ORDER BY timestamp DESC LIMIT 1", (pid,)
        ).fetchone()
        # Max fees ever
        max_fees = conn.execute(
            "SELECT MAX(total_earned_fees_usd) as fees FROM lp_snapshots WHERE position_id=?", (pid,)
        ).fetchone()
        
        if first and last:
            entry_value = (first['entry_value_usd'] if 'entry_value_usd' in first.keys() else None) or first['value_usd'] or 0
            exit_value = last['value_usd'] or 0
            total_fees = max_fees['fees'] if max_fees and max_fees['fees'] else 0
            # Calculate days held
            from datetime import datetime as dt
            try:
                t0 = dt.fromisoformat(first['timestamp'].replace('Z',''))
                t1 = dt.fromisoformat(last['timestamp'].replace('Z',''))
                days_held = max((t1 - t0).total_seconds() / 86400, 1)
            except Exception:
                days_held = 1
            # Total return = (exit - entry + fees) / entry
            total_return = ((exit_value - entry_value + total_fees) / entry_value * 100) if entry_value > 0 else 0
            # Annualized APR
            annual_apr = total_return * (365 / days_held) if days_held > 0 else 0
            
            closed_lps.append({
                'pair': first['token0'] + '/' + first['token1'],
                'chain': first['chain'],
                'entry_value': entry_value,
                'exit_value': exit_value,
                'range_lower': first['range_lower'],
                'range_upper': first['range_upper'],
                'total_fees': total_fees,
                'total_return_pct': total_return,
                'annual_apr': annual_apr,
                'days_held': int(days_held),
                'entry_date': first['timestamp'][:10],
                'exit_date': last['timestamp'][:10],
            })
    
    # Closed hedges
    active_hedge_keys = set()
    for r in conn.execute(
        "SELECT DISTINCT market, direction, wallet FROM hedge_snapshots WHERE strftime('%Y-%m-%dT%H:%M:00', timestamp)=?", (latest_ts,)
    ).fetchall():
        active_hedge_keys.add((r['market'], r['direction'], r['wallet']))
    
    all_hedge_keys = set()
    for r in conn.execute(
        f"SELECT DISTINCT market, direction, wallet FROM hedge_snapshots WHERE 1=1 {date_filter}",
        date_params,
    ).fetchall():
        all_hedge_keys.add((r['market'], r['direction'], r['wallet']))
    
    closed_hedge_keys = all_hedge_keys - active_hedge_keys
    
    closed_hedges = []
    for market, direction, wallet in closed_hedge_keys:
        first = conn.execute(
            "SELECT * FROM hedge_snapshots WHERE market=? AND direction=? AND wallet=? ORDER BY timestamp ASC LIMIT 1",
            (market, direction, wallet)
        ).fetchone()
        last = conn.execute(
            "SELECT * FROM hedge_snapshots WHERE market=? AND direction=? AND wallet=? ORDER BY timestamp DESC LIMIT 1",
            (market, direction, wallet)
        ).fetchone()
        if first and last:
            closed_hedges.append({
                'market': market,
                'direction': direction,
                'entry_price': first['entry_price'],
                'exit_price': last['current_price'],
                'size_usd': first['size_usd'],
                'pnl_usd': last['pnl_usd'] or 0,
                'entry_date': first['timestamp'][:10],
                'exit_date': last['timestamp'][:10],
            })
    
    # Also include closed manual LP positions
    manual_lps = conn.execute(
        f"SELECT * FROM lp_positions WHERE is_active=0 {date_filter_updated_at}",
        date_params,
    ).fetchall()
    for mp in manual_lps:
        mp = dict(mp)
        entry_value = mp.get('entry_value_usd') or mp.get('value_usd') or 0
        exit_value = mp.get('value_usd') or 0
        total_fees = (mp.get('fees_uncollected_usd') or 0) + (mp.get('fees_collected_usd') or 0)
        try:
            from datetime import datetime as dt
            t0 = dt.fromisoformat(str(mp.get('created_at', '')).replace('Z', ''))
            t1 = dt.fromisoformat(str(mp.get('updated_at', '')).replace('Z', ''))
            days_held = max((t1 - t0).total_seconds() / 86400, 1)
        except Exception:
            days_held = 1
        total_return = ((exit_value - entry_value + total_fees) / entry_value * 100) if entry_value > 0 else 0
        annual_apr = total_return * (365 / days_held) if days_held > 0 else 0
        closed_lps.append({
            'pair': (mp.get('token0') or '?') + '/' + (mp.get('token1') or '?'),
            'chain': mp.get('chain') or 'Manual',
            'entry_value': entry_value,
            'exit_value': exit_value,
            'range_lower': mp.get('range_lower'),
            'range_upper': mp.get('range_upper'),
            'total_fees': total_fees,
            'total_return_pct': total_return,
            'annual_apr': annual_apr,
            'days_held': int(days_held),
            'entry_date': str(mp.get('created_at', ''))[:10],
            'exit_date': str(mp.get('updated_at', ''))[:10],
        })
    
    # Also include closed manual hedges
    manual_hedges = conn.execute(
        f"SELECT * FROM hedge_positions WHERE is_active=0 {date_filter_updated_at}",
        date_params,
    ).fetchall()
    for mh in manual_hedges:
        mh = dict(mh)
        closed_hedges.append({
            'market': mh.get('market') or '?',
            'direction': mh.get('direction') or '?',
            'entry_price': mh.get('entry_price') or 0,
            'exit_price': mh.get('current_price') or 0,
            'size_usd': mh.get('size_usd') or 0,
            'pnl_usd': mh.get('pnl_usd') or 0,
            'entry_date': str(mh.get('created_at', ''))[:10],
            'exit_date': str(mh.get('updated_at', ''))[:10],
        })
    
    conn.close()
    return jsonify({'closed_lps': closed_lps, 'closed_hedges': closed_hedges})


@app.route('/api/history/lp/<position_id>')
def api_history_lp(position_id):
    """Get fee claim history for an LP position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM lp_fee_claims WHERE position_id=? ORDER BY claimed_at DESC",
            (position_id,)
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/history/fees')
def api_history_fees():
    """Get all fee claims across all LP positions."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM lp_fee_claims ORDER BY claimed_at DESC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/history/runs')
def api_history_runs():
    """Get recent snapshot runs."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM portfolio_snapshots ORDER BY timestamp DESC LIMIT 20"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/history/positions')
def api_history_positions():
    """Get all tracked LP positions."""
    return jsonify([])


@app.route('/api/history/latest')
def api_history_latest():
    """Get the latest snapshot timestamp."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    row = conn.execute(
        "SELECT MAX(timestamp) as latest FROM portfolio_snapshots WHERE status='completed'"
    ).fetchone()
    conn.close()
    return jsonify({"latest_snapshot": row['latest'] if row else None})


@app.route('/api/history/wallets')
def api_history_wallets():
    """Get wallets that have snapshot data, with labels."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT wallet FROM portfolio_snapshots WHERE status='completed' ORDER BY wallet"
    ).fetchall()
    conn.close()
    
    wallets = []
    for r in rows:
        addr = r['wallet']
        label = get_wallet_label(addr)
        short = addr[:6] + '...' + addr[-4:]
        wallets.append({"address": addr, "label": label, "short": short})
    return jsonify(wallets)


# --- Backup / Export / Import ---

from flask import send_file
import shutil
import io
from src.storage.portfolio_db import get_db_path


@app.route('/api/backup/db', methods=['GET'])
def api_backup_db():
    """Download the SQLite database file."""
    db_path = get_db_path()
    if not os.path.exists(db_path):
        return jsonify({"error": "No database found"}), 404
    
    # Create a safe copy to avoid locking issues
    backup_path = db_path + '.backup'
    shutil.copy2(db_path, backup_path)
    try:
        return send_file(
            backup_path,
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name='portfolio_backup.db'
        )
    finally:
        # Clean up backup copy after a delay (Flask sends async)
        pass


def api_import_db():
    """Import/restore a SQLite database file."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    db_path = get_db_path()

    # Save uploaded file to a temp location first
    temp_path = db_path + '.import'
    file.save(temp_path)

    # Validate it's a valid SQLite file with at least some expected tables
    try:
        import sqlite3
        conn = sqlite3.connect(temp_path)
        tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        table_names = [t[0] for t in tables]
        conn.close()

        # Must have at least one of our core tables
        known_tables = ['portfolio_snapshots', 'token_snapshots', 'lp_snapshots',
                        'hedge_snapshots', 'lending_snapshots', 'lending_account_snapshots',
                        'market_snapshots', 'lp_positions', 'hedge_positions',
                        'manual_positions', 'manual_hedges',
                        'users', 'token_prices_daily']
        found = [t for t in known_tables if t in table_names]
        if not found:
            os.remove(temp_path)
            return jsonify({"error": "Not a valid portfolio database — no recognized tables found"}), 400
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": f"Invalid SQLite file: {e}"}), 400

    # Replace current DB
    if os.path.exists(db_path):
        shutil.copy2(db_path, db_path + '.pre_import')  # safety backup
    shutil.move(temp_path, db_path)

    # Re-initialize to add any missing tables from newer schema
    from src.storage.portfolio_db import init_db
    init_db()

    return jsonify({
        "success": True,
        "message": f"Database imported ({len(found)} tables recognized: {', '.join(found)})"
    })


@app.route('/api/backup/config', methods=['GET'])
def api_export_config():
    """Export all config: .env variables + wallet config + AI config as JSON."""
    env_data = {}
    if os.path.exists(ENV_FILE):
        with open(ENV_FILE, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, _, value = line.partition('=')
                    env_data[key.strip()] = value.strip().strip("'\"")

    wallet_config = load_wallet_config()

    # AI config
    ai_config = {}
    ai_config_path = os.path.join("data", "ai_config.json")
    if os.path.exists(ai_config_path):
        try:
            with open(ai_config_path, 'r') as f:
                ai_config = json.load(f)
        except Exception:
            pass

    # Telegram config
    telegram_config = {}
    telegram_config_path = os.path.join("data", "telegram_config.json")
    if os.path.exists(telegram_config_path):
        try:
            with open(telegram_config_path, 'r') as f:
                telegram_config = json.load(f)
        except Exception:
            pass

    export = {
        "env": env_data,
        "wallets": wallet_config,
        "ai_config": ai_config,
        "telegram_config": telegram_config,
        "exported_at": datetime.now().isoformat()
    }

    buf = io.BytesIO()
    buf.write(json.dumps(export, indent=2).encode('utf-8'))
    buf.seek(0)

    return send_file(buf, mimetype='application/json', as_attachment=True, download_name='portfolio_config.json')


@app.route('/api/backup/config', methods=['POST'])
def api_import_config():
    """Import config: restore .env variables + wallet config from JSON."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400
    
    try:
        data = json.loads(file.read().decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        return jsonify({"error": f"Invalid JSON file: {e}"}), 400
    
    restored = []
    
    # Restore .env
    if 'env' in data and isinstance(data['env'], dict):
        env_path = ENV_FILE
        # Read existing lines to preserve comments
        existing_lines = []
        existing_keys = set()
        if os.path.exists(env_path):
            # Backup first
            shutil.copy2(env_path, env_path + '.pre_import')
            with open(env_path, 'r') as f:
                existing_lines = f.readlines()
        
        # Update existing keys, track which ones we've seen
        new_lines = []
        for line in existing_lines:
            stripped = line.strip()
            if stripped and not stripped.startswith('#') and '=' in stripped:
                key = stripped.split('=', 1)[0].strip()
                if key in data['env']:
                    new_lines.append(f"{key}={data['env'][key]}\n")
                    existing_keys.add(key)
                else:
                    new_lines.append(line)
            else:
                new_lines.append(line)
        
        # Add new keys not in existing file
        for key, value in data['env'].items():
            if key not in existing_keys:
                new_lines.append(f"{key}={value}\n")
        
        with open(env_path, 'w') as f:
            f.writelines(new_lines)
        
        # Update current process env
        for key, value in data['env'].items():
            os.environ[key] = value
        
        restored.append('environment variables')
    
    # Restore wallet config
    if 'wallets' in data and isinstance(data['wallets'], dict):
        if os.path.exists(WALLET_CONFIG_FILE):
            shutil.copy2(WALLET_CONFIG_FILE, WALLET_CONFIG_FILE + '.pre_import')
        save_wallet_config(data['wallets'])
        restored.append('wallet configuration')
    
    # Restore AI config
    if 'ai_config' in data and isinstance(data['ai_config'], dict):
        ai_config_path = os.path.join("data", "ai_config.json")
        os.makedirs(os.path.dirname(ai_config_path), exist_ok=True)
        with open(ai_config_path, 'w') as f:
            json.dump(data['ai_config'], f, indent=2)
        restored.append('AI configuration')
    
    # Restore Telegram config
    if 'telegram_config' in data and isinstance(data['telegram_config'], dict):
        telegram_path = os.path.join("data", "telegram_config.json")
        os.makedirs(os.path.dirname(telegram_path), exist_ok=True)
        with open(telegram_path, 'w') as f:
            json.dump(data['telegram_config'], f, indent=2)
        restored.append('Telegram configuration')
    
    if not restored:
        return jsonify({"error": "No valid data found in file"}), 400
    
    return jsonify({
        "success": True,
        "message": f"Restored: {', '.join(restored)}"
    })


# ── LP Optimizer API ─────────────────────────────────────────────────

@app.route('/api/optimizer/pools', methods=['POST'])
def api_optimizer_pools():
    """Search DeFiLlama for Uniswap V3 pools matching filters."""
    data = request.get_json(silent=True) or {}
    chain = data.get("chain") or None
    symbol = data.get("symbol") or None
    min_tvl = data.get("min_tvl", 500_000)
    try:
        min_tvl = float(min_tvl)
    except (TypeError, ValueError):
        return jsonify({"error": "min_tvl must be a number"}), 400
    try:
        pools = discover_pools(chain=chain, symbol_filter=symbol, min_tvl=min_tvl)
        return jsonify(pools)
    except requests.RequestException as e:
        return jsonify({"error": f"Pool data unavailable: {e}"}), 502


@app.route('/api/optimizer/run', methods=['POST'])
def api_optimizer_run():
    """Run the full LP range optimization pipeline."""
    data = request.get_json(silent=True) or {}
    try:
        result = run_optimization(data)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except requests.RequestException as e:
        return jsonify({"error": f"External data unavailable: {e}"}), 502


@app.route('/api/optimizer/regimes', methods=['GET'])
def api_optimizer_regimes():
    """Load regime probabilities for a given horizon."""
    horizon_str = request.args.get("horizon", "")
    try:
        horizon = int(horizon_str)
    except (TypeError, ValueError):
        return jsonify({"error": "horizon must be 7, 14, or 30"}), 400
    try:
        regimes = load_regime_probabilities(horizon)
        return jsonify(regimes)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route('/api/optimizer/portfolio-positions', methods=['GET'])
def api_optimizer_portfolio_positions():
    """Return active V3 LP positions from the cached portfolio."""
    from src.engines.range_optimizer import resolve_coin_id
    data = get_portfolio_data()
    lp_positions = data.get("lp_positions", [])
    v3_positions = []
    for pos in lp_positions:
        protocol = (pos.get("protocol") or "").lower()
        if "v3" in protocol or "camelot" in protocol:
            pair = pos.get("pair") or ""
            t0 = pos.get("token0_symbol") or ""
            t1 = pos.get("token1_symbol") or ""
            symbol = pair or (t0 + "-" + t1 if t0 and t1 else "")
            raw_chain = pos.get("chain") or ""
            display_chain = raw_chain.capitalize() if raw_chain else raw_chain
            v3_positions.append({
                "chain": display_chain,
                "symbol": symbol,
                "token0": t0,
                "token1": t1,
                "fee_tier": pos.get("fee_tier"),
                "price_lower": pos.get("price_lower"),
                "price_upper": pos.get("price_upper"),
                "current_price": pos.get("current_price"),
                "value_usd": pos.get("total_value_usd"),
                "coin_id": resolve_coin_id(symbol),
            })
    return jsonify(v3_positions)


def _parse_trade_date(date_str):
    if not date_str:
        return None
    from datetime import datetime
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%-m/%-d/%Y', '%m-%d-%Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def normalize_date(value):
    """Parse any common date format and return M/D/YYYY (no leading zeros)."""
    from datetime import datetime
    if not value:
        return None
    v = str(value).strip()
    for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y'):
        try:
            d = datetime.strptime(v, fmt)
            return f"{d.month}/{d.day}/{d.year}"
        except ValueError:
            continue
    return None


def _calculate_spot_fifo(conn):
    """
    FIFO P&L across all spot_transactions rows.
    Returns (open_positions, closed_positions) dicts keyed by uppercase symbol.
    """
    from collections import defaultdict, deque

    rows = conn.execute(
        "SELECT * FROM spot_transactions ORDER BY id ASC"
    ).fetchall()
    rows = sorted(rows, key=lambda r: (_parse_trade_date(r['trade_date']) or 0, r['id']))

    lots            = defaultdict(deque)   # symbol -> deque of {units, price, date}
    realized_pnl    = defaultdict(float)
    total_invested  = defaultdict(float)
    total_proceeds  = defaultdict(float)
    last_sell_date  = defaultdict(str)
    all_symbols     = set()

    for row in rows:
        sym   = row['symbol'].upper()
        side  = row['side'].lower()
        units = float(row['units'])
        tx_amt = float(row['price_usd'])          # total transaction amount (incl. fees)
        price  = tx_amt / units if units > 1e-12 else 0.0  # derive per-unit cost for FIFO lots
        total  = tx_amt
        date  = row['trade_date']

        all_symbols.add(sym)

        if side == 'buy':
            lots[sym].append({'units': units, 'price': price, 'date': date})
            total_invested[sym] += total
        elif side == 'sell':
            remaining  = units
            cost_basis = 0.0
            while remaining > 1e-9 and lots[sym]:
                lot = lots[sym][0]
                if lot['units'] <= remaining + 1e-9:
                    cost_basis += lot['units'] * lot['price']
                    remaining  -= lot['units']
                    lots[sym].popleft()
                else:
                    cost_basis     += remaining * lot['price']
                    lot['units']   -= remaining
                    remaining       = 0.0
            total_proceeds[sym] += total
            realized_pnl[sym]   += total - cost_basis
            last_sell_date[sym]  = date

    open_positions = {}
    for sym, lot_queue in lots.items():
        remaining_units = sum(l['units'] for l in lot_queue)
        if remaining_units > 1e-6:
            total_cost = sum(l['units'] * l['price'] for l in lot_queue)
            open_positions[sym] = {
                'symbol':           sym,
                'units':            remaining_units,
                'avg_cost_usd':     total_cost / remaining_units,
                'total_cost_basis': total_cost,
                'oldest_lot_date':  lot_queue[0]['date'],
                'lot_count':        len(lot_queue),
                'lots':             list(lot_queue),
                'realized_pnl':     realized_pnl.get(sym, 0.0),
            }

    closed_positions = {}
    for sym in all_symbols:
        if total_proceeds.get(sym, 0.0) > 0:
            invested  = total_invested[sym]
            proceeds  = total_proceeds[sym]
            rpnl      = realized_pnl.get(sym, 0.0)
            # Cost basis of sold units = proceeds minus the profit on those units.
            # Using total_invested as denominator is wrong when only some units were sold
            # (it includes the cost of unsold lots, making profitable trades look negative).
            cost_basis_sold = proceeds - rpnl
            closed_positions[sym] = {
                'symbol':          sym,
                'realized_pnl':    rpnl,
                'total_invested':  invested,
                'total_proceeds':  proceeds,
                'last_sell_date':  last_sell_date.get(sym, ''),
                'roi_pct':         (rpnl / cost_basis_sold * 100) if cost_basis_sold > 0 else 0.0,
            }

    return open_positions, closed_positions


# ── Spot P&L Routes ──

@app.route('/api/spot/transactions', methods=['GET'])
def api_spot_transactions_list():
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        symbol = request.args.get('symbol', '').strip().upper()
        if symbol:
            rows = conn.execute(
                "SELECT * FROM spot_transactions WHERE symbol=?", (symbol,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM spot_transactions").fetchall()
        conn.close()
        # Sort by parsed date desc (handles D/M/YYYY and legacy ISO formats), then id desc
        from datetime import datetime as _dt
        def _sort_key(r):
            d = _parse_trade_date(r['trade_date'])
            return (d or _dt.min, r['id'])
        rows_sorted = sorted([dict(r) for r in rows], key=_sort_key, reverse=True)
        return jsonify(rows_sorted)
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/transactions', methods=['POST'])
def api_spot_transactions_create():
    try:
        from src.storage.portfolio_db import get_connection
        data = request.json or {}
        required = ['trade_date', 'symbol', 'side', 'units', 'price_usd']
        missing = [f for f in required if not data.get(f) and data.get(f) != 0]
        if missing:
            return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
        if data['side'] not in ('buy', 'sell'):
            return jsonify({"error": "side must be 'buy' or 'sell'"}), 400
        units = float(data['units'])
        price_usd = float(data['price_usd'])  # total tx amount (incl. fees)
        total_usd = price_usd                 # total_usd == tx amount; no per-unit multiplication
        conn = get_connection()
        c = conn.execute(
            """INSERT INTO spot_transactions (trade_date, symbol, side, units, price_usd, total_usd, platform, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (normalize_date(data['trade_date']) or data['trade_date'], data['symbol'].upper(), data['side'], units, price_usd, total_usd,
             data.get('platform', ''), data.get('notes', ''))
        )
        new_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": new_id, "total_usd": total_usd})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/transactions/<int:tx_id>', methods=['PUT'])
def api_spot_transactions_update(tx_id):
    try:
        from src.storage.portfolio_db import get_connection
        data = request.json or {}
        required = ['trade_date', 'symbol', 'side', 'units', 'price_usd']
        missing = [f for f in required if not data.get(f) and data.get(f) != 0]
        if missing:
            return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
        if data['side'] not in ('buy', 'sell'):
            return jsonify({"error": "side must be 'buy' or 'sell'"}), 400
        units = float(data['units'])
        price_usd = float(data['price_usd'])  # total tx amount (incl. fees)
        total_usd = price_usd                 # total_usd == tx amount; no per-unit multiplication
        conn = get_connection()
        conn.execute(
            """UPDATE spot_transactions SET trade_date=?, symbol=?, side=?, units=?, price_usd=?, total_usd=?,
               platform=?, notes=? WHERE id=?""",
            (normalize_date(data['trade_date']) or data['trade_date'], data['symbol'].upper(), data['side'], units, price_usd, total_usd,
             data.get('platform', ''), data.get('notes', ''), tx_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "total_usd": total_usd})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/transactions/<int:tx_id>', methods=['DELETE'])
def api_spot_transactions_delete(tx_id):
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        conn.execute("DELETE FROM spot_transactions WHERE id=?", (tx_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/transactions/all', methods=['DELETE'])
def api_spot_transactions_delete_all():
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        cur = conn.execute("DELETE FROM spot_transactions")
        deleted = cur.rowcount
        conn.commit()
        conn.close()
        return jsonify({"success": True, "deleted": deleted})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/token-config', methods=['GET'])
def api_spot_token_config_list():
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        rows = conn.execute("SELECT * FROM spot_token_config ORDER BY symbol ASC").fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/token-config', methods=['POST'])
def api_spot_token_config_upsert():
    try:
        from src.storage.portfolio_db import get_connection
        data = request.json or {}
        if not data.get('symbol'):
            return jsonify({"error": "symbol is required"}), 400
        conn = get_connection()
        c = conn.execute(
            """INSERT OR REPLACE INTO spot_token_config (symbol, cg_id, contract_address, chain, notes, price_source, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)""",
            (data['symbol'].upper(), data.get('cg_id', ''), data.get('contract_address', ''),
             data.get('chain', ''), data.get('notes', ''), data.get('price_source', 'coingecko'))
        )
        new_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/token-config/<symbol>', methods=['DELETE'])
def api_spot_token_config_delete(symbol):
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        conn.execute("DELETE FROM spot_token_config WHERE symbol=?", (symbol.upper(),))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/import-csv', methods=['POST'])
def api_spot_import_csv():
    try:
        import csv, io
        from src.storage.portfolio_db import get_connection
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        f = request.files['file']
        content = f.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(content))

        # Normalise header → lowercase, strip, collapse spaces to underscores
        def _norm(k):
            return k.lower().strip().replace(' ', '_').replace('$', '')

        # Build a normalised → original header map from the CSV
        raw_headers = reader.fieldnames or []
        norm_header_map = {_norm(h): h for h in raw_headers}
        norm_headers = set(norm_header_map.keys())

        # Detect which price column is present (priority order)
        TX_AMT_ALIASES    = {'tx_amount', 'tx_amt', 'total_usd', 'total'}
        UNIT_PRICE_ALIASES = {'unit_price', 'unit_price_usd', 'unitprice'}
        PRICE_USD_ALIASES  = {'price_usd', 'price'}

        if norm_headers & TX_AMT_ALIASES:
            _price_mode = 'tx_amt'
            _price_col  = norm_header_map[next(iter(norm_headers & TX_AMT_ALIASES))]
        elif norm_headers & UNIT_PRICE_ALIASES:
            _price_mode = 'unit_price'
            _price_col  = norm_header_map[next(iter(norm_headers & UNIT_PRICE_ALIASES))]
        else:
            _price_mode = 'price_usd'
            matched = norm_headers & PRICE_USD_ALIASES
            _price_col = norm_header_map[next(iter(matched))] if matched else None

        DATE_ALIASES    = {'date', 'trade_date', 'entry_date'}
        SYMBOL_ALIASES  = {'symbol', 'ticker', 'token'}
        SIDE_ALIASES    = {'side', 'type', 'buy_sell'}
        UNITS_ALIASES   = {'units', 'quantity', 'transacted_units', 'amount'}
        PLATFORM_ALIASES = {'platform'}
        NOTES_ALIASES   = {'notes'}

        def _find_col(aliases):
            match = norm_headers & aliases
            return norm_header_map[next(iter(match))] if match else None

        date_col     = _find_col(DATE_ALIASES)
        symbol_col   = _find_col(SYMBOL_ALIASES)
        side_col     = _find_col(SIDE_ALIASES)
        units_col    = _find_col(UNITS_ALIASES)
        platform_col = _find_col(PLATFORM_ALIASES)
        notes_col    = _find_col(NOTES_ALIASES)

        side_map = {
            'buy': 'buy', 'b': 'buy', 'bought': 'buy',
            'sell': 'sell', 's': 'sell', 'sold': 'sell',
        }

        def _clean_num(v):
            return v.strip().replace(',', '').replace('$', '') if v else ''

        def _parse_date(v):
            """Normalise to M/D/YYYY (no leading zeros) via normalize_date()."""
            return normalize_date(v)

        imported = 0
        skipped  = 0
        errors   = []
        conn = get_connection()

        for line_num, row in enumerate(reader, start=2):
            def _get(col):
                return (row.get(col) or '').strip()

            trade_date = _parse_date(_get(date_col)) if date_col else None
            symbol_raw = _get(symbol_col) if symbol_col else None
            side_raw   = _get(side_col)   if side_col   else None
            units_raw  = _clean_num(_get(units_col)) if units_col else None
            price_raw  = _clean_num(_get(_price_col)) if _price_col else None

            # Skip blank rows silently
            if not any([trade_date, symbol_raw, side_raw, units_raw, price_raw]):
                continue

            missing = []
            if not trade_date: missing.append('date')
            if not symbol_raw: missing.append('symbol')
            if not side_raw:   missing.append('type/side')
            if not units_raw:  missing.append('units')
            if not price_raw:  missing.append('price/tx_amount')
            if missing:
                errors.append(f"Row {line_num}: missing {', '.join(missing)}")
                skipped += 1; continue

            side = side_map.get(side_raw.lower().strip())
            if not side:
                errors.append(f"Row {line_num}: unrecognised side '{side_raw}'")
                skipped += 1; continue

            try:
                units = float(units_raw)
                raw_price = float(price_raw)
                if _price_mode == 'unit_price':
                    price_usd = raw_price * units   # derive total from per-unit × units
                else:
                    price_usd = raw_price           # tx_amt or price_usd already total
            except ValueError as exc:
                errors.append(f"Row {line_num}: number parse error — {exc}")
                skipped += 1; continue

            total_usd = price_usd
            platform  = _get(platform_col) if platform_col else ''
            notes     = _get(notes_col)    if notes_col    else ''

            conn.execute(
                """INSERT INTO spot_transactions (trade_date, symbol, side, units, price_usd, total_usd, platform, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (trade_date, symbol_raw.upper(), side, units, price_usd, total_usd, platform, notes)
            )
            imported += 1

        conn.commit()
        conn.close()
        return jsonify({"success": True, "imported": imported, "skipped": skipped, "errors": errors})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/pnl', methods=['GET'])
def api_spot_pnl():
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        open_positions, _ = _calculate_spot_fifo(conn)
        conn.close()

        results = []
        for sym, pos in open_positions.items():
            current_price = _get_spot_price(sym)
            current_value = (pos['units'] * current_price) if current_price is not None else None
            unrealized_pnl = (current_value - pos['total_cost_basis']) if current_value is not None else None
            unrealized_pct = (unrealized_pnl / pos['total_cost_basis'] * 100) if (unrealized_pnl is not None and pos['total_cost_basis'] > 0) else None
            results.append({
                'symbol':             sym,
                'units':              pos['units'],
                'avg_cost_usd':       pos['avg_cost_usd'],
                'total_cost_basis':   pos['total_cost_basis'],
                'current_price_usd':  current_price,
                'current_value_usd':  current_value,
                'unrealized_pnl_usd': unrealized_pnl,
                'unrealized_pct':     unrealized_pct,
                'realized_pnl_usd':   pos['realized_pnl'],
                'oldest_lot_date':    pos['oldest_lot_date'],
                'lot_count':          pos['lot_count'],
            })

        results.sort(key=lambda x: (x['current_value_usd'] is None, -(x['current_value_usd'] or 0)))
        return jsonify(results)
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/price-test/<symbol>', methods=['GET'])
def api_spot_price_test(symbol):
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        row = conn.execute(
            "SELECT contract_address, price_source FROM spot_token_config WHERE symbol=?",
            (symbol.upper(),)
        ).fetchone()
        conn.close()
        source = 'coingecko'
        contract_address = ''
        if row:
            source = (row['price_source'] or 'coingecko').lower()
            contract_address = row['contract_address'] or ''
        if source == 'manual':
            return jsonify({'symbol': symbol, 'price': None, 'source': 'manual', 'note': 'Manual price — no live lookup'})
        elif source == 'dexscreener':
            if not contract_address:
                return jsonify({'symbol': symbol, 'price': None, 'source': 'dexscreener', 'error': 'No contract address set'})
            price = _get_dexscreener_price(contract_address)
            return jsonify({'symbol': symbol, 'price': price, 'source': 'dexscreener'})
        else:
            price = _get_coingecko_price(symbol)
            return jsonify({'symbol': symbol, 'price': price, 'source': 'coingecko'})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/history', methods=['GET'])
def api_spot_history():
    try:
        from src.storage.portfolio_db import get_connection
        conn = get_connection()
        _, closed_positions = _calculate_spot_fifo(conn)
        conn.close()

        results = list(closed_positions.values())
        results.sort(key=lambda x: _parse_trade_date(x.get('last_sell_date', '')) or 0, reverse=True)
        return jsonify(results)
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/spot/stablecoins', methods=['GET'])
def api_spot_stablecoins():
    try:
        from src.storage.portfolio_db import get_connection
        STABLES = ('USDC', 'USDT', 'DAI', 'FRAX', 'LUSD', 'BUSD', 'TUSD', 'USDS', 'CRVUSD')
        conn = get_connection()
        # Latest completed snapshot id per wallet, then fetch matching stablecoin rows
        rows = conn.execute("""
            SELECT t.symbol, t.value_usd, t.wallet
            FROM token_snapshots t
            JOIN (
                SELECT wallet, MAX(id) AS snap_id
                FROM portfolio_snapshots
                WHERE status = 'completed'
                GROUP BY wallet
            ) latest ON t.snapshot_id = latest.snap_id
            WHERE UPPER(t.symbol) IN ({})
        """.format(','.join('?' * len(STABLES))), STABLES).fetchall()
        conn.close()
        breakdown = [{'symbol': r['symbol'], 'value_usd': r['value_usd'] or 0, 'wallet': r['wallet']} for r in rows]
        total_usd = sum(b['value_usd'] for b in breakdown)
        return jsonify({'total_usd': total_usd, 'breakdown': breakdown})
    except Exception as e:
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500


# --- DeFi Journal Routes ---

@app.route('/api/defi-journal/<position_type>/<path:position_id>')
def api_defi_journal_get(position_type, position_id):
    """Get journal entries for a DeFi position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM defi_journal WHERE position_type=? AND position_id=? ORDER BY created_at DESC",
            (position_type, position_id)
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-journal', methods=['POST'])
def api_defi_journal_post():
    """Add a journal entry for a DeFi position."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400
    position_type = data.get('position_type', '').strip()
    position_id = str(data.get('position_id', '')).strip()
    action = data.get('action', '').strip()
    if not position_type or not position_id or not action:
        return jsonify({"error": "position_type, position_id, and action are required"}), 400
    details = data.get('details', '')
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute(
            "INSERT INTO defi_journal (position_type, position_id, action, details) VALUES (?, ?, ?, ?)",
            (position_type, position_id, action, details)
        )
        entry_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": entry_id})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-journal/<int:entry_id>', methods=['PUT'])
def api_defi_journal_put(entry_id):
    """Update a journal entry."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        updates = []
        params = []
        if 'action' in data:
            updates.append("action=?")
            params.append(data['action'])
        if 'details' in data:
            updates.append("details=?")
            params.append(data['details'])
        if not updates:
            conn.close()
            return jsonify({"error": "No fields to update"}), 400
        params.append(entry_id)
        conn.execute(f"UPDATE defi_journal SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-journal/<int:entry_id>', methods=['DELETE'])
def api_defi_journal_delete(entry_id):
    """Delete a journal entry."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM defi_journal WHERE id=?", (entry_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- LP Fee Claims Routes ---

@app.route('/api/lp/fee-claim', methods=['POST'])
def api_lp_fee_claim_post():
    """Record a fee claim event for an LP position."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400
    position_id = data.get('position_id')
    claimed_at = data.get('claimed_at', '').strip()
    value_usd = data.get('value_usd')
    if not position_id or not claimed_at or value_usd is None:
        return jsonify({"error": "position_id, claimed_at, and value_usd are required"}), 400
    try:
        value_usd = float(value_usd)
    except (TypeError, ValueError):
        return jsonify({"error": "value_usd must be a number"}), 400
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute(
            """INSERT INTO lp_fee_claims
               (position_id, claimed_at, token0_amount, token1_amount,
                token0_price_usd, token1_price_usd, value_usd, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (position_id, claimed_at,
             float(data.get('token0_amount', 0) or 0),
             float(data.get('token1_amount', 0) or 0),
             float(data.get('token0_price_usd', 0) or 0),
             float(data.get('token1_price_usd', 0) or 0),
             value_usd,
             data.get('notes', ''))
        )
        claim_id = c.lastrowid
        # Increment fees_collected_usd on the lp_positions row
        conn.execute(
            "UPDATE lp_positions SET fees_collected_usd = COALESCE(fees_collected_usd, 0) + ? WHERE id=?",
            (value_usd, position_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": claim_id})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lp/fee-claims/<path:position_id>')
def api_lp_fee_claims_get(position_id):
    """Get fee claim history for an LP position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM lp_fee_claims WHERE position_id=? ORDER BY claimed_at DESC",
            (position_id,)
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lp/fee-claims/<int:claim_id>', methods=['DELETE'])
def api_lp_fee_claims_delete(claim_id):
    """Delete a fee claim record."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM lp_fee_claims WHERE id=?", (claim_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- DeFi Staking Routes ---

@app.route('/api/defi-staking')
def api_defi_staking_get():
    """Get active custom DeFi staking positions."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM defi_staking WHERE is_active=1 AND (is_archived IS NULL OR is_archived=0) ORDER BY created_at DESC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking', methods=['POST'])
def api_defi_staking_post():
    """Add a custom DeFi staking position."""
    from src.storage.portfolio_db import get_connection
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400
    app_name = data.get('app_name', '').strip()
    chain = data.get('chain', '').strip()
    token_symbol = data.get('token_symbol', '').strip()
    staked_amount = data.get('staked_amount')
    if not app_name or not chain or not token_symbol or staked_amount is None:
        return jsonify({"error": "app_name, chain, token_symbol, and staked_amount are required"}), 400
    try:
        staked_amount = float(staked_amount)
    except (TypeError, ValueError):
        return jsonify({"error": "staked_amount must be a number"}), 400
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute(
            """INSERT INTO defi_staking
               (app_name, chain, position_label, token_symbol, staked_amount,
                staked_value_usd, reward_token, reward_amount, reward_value_usd,
                entry_date, lock_end_date, notes, wallet_address)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (app_name, chain,
             data.get('position_label', ''),
             token_symbol, staked_amount,
             data.get('staked_value_usd'),
             data.get('reward_token', ''),
             float(data.get('reward_amount', 0) or 0),
             float(data.get('reward_value_usd', 0) or 0),
             data.get('entry_date'),
             data.get('lock_end_date'),
             data.get('notes', ''),
             data.get('wallet_address', ''))
        )
        pos_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": pos_id})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking/<int:pos_id>', methods=['PUT'])
def api_defi_staking_put(pos_id):
    """Update a staking position. action='close' sets is_active=0."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        if data.get('action') == 'close':
            conn.execute(
                "UPDATE defi_staking SET is_active=0, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (pos_id,)
            )
        else:
            allowed = ('app_name', 'chain', 'position_label', 'token_symbol', 'staked_amount',
                       'staked_value_usd', 'reward_token', 'reward_amount', 'reward_value_usd',
                       'entry_date', 'lock_end_date', 'notes', 'wallet_address', 'is_active')
            updates = [(k, data[k]) for k in allowed if k in data]
            if not updates:
                conn.close()
                return jsonify({"error": "No fields to update"}), 400
            set_clause = ', '.join(f"{k}=?" for k, _ in updates)
            params = [v for _, v in updates] + [pos_id]
            conn.execute(
                f"UPDATE defi_staking SET {set_clause}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                params
            )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking/<int:pos_id>', methods=['DELETE'])
def api_defi_staking_delete(pos_id):
    """Hard-delete a staking position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM defi_staking WHERE id=?", (pos_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- Archive & Hide Routes ---

@app.route('/api/archive/lp', methods=['GET'])
def api_archive_lp_list():
    """List archived LP positions: manual (is_active=0) and Zerion (zerion_lp_hidden)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        manual_rows = conn.execute(
            "SELECT * FROM lp_positions WHERE is_active=0 AND (is_permanently_hidden=0 OR is_permanently_hidden IS NULL) ORDER BY updated_at DESC"
        ).fetchall()
        zerion_rows = conn.execute(
            "SELECT * FROM zerion_lp_hidden ORDER BY hidden_at DESC"
        ).fetchall()
        conn.close()
        zerion_result = []
        for r in zerion_rows:
            d = dict(r)
            zerion_result.append({
                "id": d["id"],
                "source": "zerion",
                "position_key": d["position_key"],
                "pair": d.get("pair", ""),
                "chain": d.get("chain", ""),
                "wallet": d.get("wallet", ""),
                "protocol": d.get("protocol", ""),
                "archived_at": d.get("hidden_at"),
                "is_permanently_hidden": bool(d.get("is_permanently_deleted", 0)),
            })
        return jsonify({"manual": [dict(r) for r in manual_rows], "zerion": zerion_result})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/lending', methods=['GET'])
def api_archive_lending_list():
    """List archived lending positions (from lending_archive_overrides)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM lending_archive_overrides WHERE is_permanently_deleted=0 ORDER BY archived_at DESC"
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            item = dict(r)
            if item.get('snapshot_json'):
                try:
                    item['snapshot'] = json.loads(item['snapshot_json'])
                except Exception:
                    item['snapshot'] = {}
            result.append(item)
        return jsonify(result)
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/staking', methods=['GET'])
def api_archive_staking_list():
    """List archived staking positions (is_archived=1)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM defi_staking WHERE is_archived=1 ORDER BY updated_at DESC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/lending', methods=['POST'])
def api_archive_lending_create():
    """Archive a Zerion lending position by storing an override record."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    position_key = data.get('position_key', '').strip()
    if not position_key:
        return jsonify({"error": "position_key required"}), 400
    conn = get_connection()
    try:
        conn.execute(
            """INSERT INTO lending_archive_overrides
               (position_key, protocol, chain, wallet, snapshot_json)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(position_key) DO UPDATE SET
                 is_permanently_deleted=0,
                 archived_at=CURRENT_TIMESTAMP,
                 snapshot_json=excluded.snapshot_json""",
            (position_key,
             data.get('protocol', ''),
             data.get('chain', ''),
             data.get('wallet', ''),
             json.dumps(data.get('snapshot', {})))
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/zerion-lp', methods=['POST'])
def api_archive_zerion_lp():
    """Archive a Zerion LP position by storing a hidden override."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    position_key = data.get('position_key', '').strip()
    if not position_key:
        return jsonify({"error": "position_key required"}), 400
    conn = get_connection()
    try:
        conn.execute(
            """INSERT INTO zerion_lp_hidden (position_key, pair, chain, wallet, protocol)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(position_key) DO UPDATE SET
                 hidden_at=CURRENT_TIMESTAMP,
                 is_permanently_deleted=0""",
            (position_key, data.get('pair', ''), data.get('chain', ''),
             data.get('wallet', ''), data.get('protocol', ''))
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/zerion-lp/<int:row_id>', methods=['DELETE'])
def api_archive_zerion_lp_delete(row_id):
    """Restore a Zerion LP position by removing its hidden override."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM zerion_lp_hidden WHERE id=?", (row_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/zerion-lp/<int:row_id>/permanent', methods=['PUT'])
def api_archive_zerion_lp_permanent(row_id):
    """Permanently hide a Zerion LP position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE zerion_lp_hidden SET is_permanently_deleted=1 WHERE id=?", (row_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/restore/lp/<int:pos_id>', methods=['POST'])
def api_restore_lp(pos_id):
    """Restore an archived manual LP position (set is_active=1)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE lp_positions SET is_active=1, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (pos_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/restore/staking/<int:pos_id>', methods=['POST'])
def api_restore_staking(pos_id):
    """Restore an archived staking position (set is_archived=0)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE defi_staking SET is_archived=0, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (pos_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/restore/lending/<path:position_key>', methods=['POST'])
def api_restore_lending(position_key):
    """Restore an archived lending position (remove override record)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "DELETE FROM lending_archive_overrides WHERE position_key=?",
            (position_key,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/lp/<int:pos_id>', methods=['DELETE'])
def api_archive_lp_delete(pos_id):
    """Permanently hide an archived LP position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("UPDATE lp_positions SET is_permanently_hidden=1 WHERE id=? AND is_active=0", (pos_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/lending/<path:position_key>', methods=['DELETE'])
def api_archive_lending_delete(position_key):
    """Permanently delete an archived lending position override."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE lending_archive_overrides SET is_permanently_deleted=1 WHERE position_key=?",
            (position_key,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/staking/<int:pos_id>', methods=['DELETE'])
def api_archive_staking_delete(pos_id):
    """Permanently hide an archived staking position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("UPDATE defi_staking SET is_permanently_hidden=1 WHERE id=? AND is_archived=1", (pos_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/permanently-hidden', methods=['GET'])
def api_archive_permanently_hidden():
    """List all permanently hidden positions across all types."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        lp_rows = conn.execute(
            "SELECT * FROM lp_positions WHERE is_permanently_hidden=1 ORDER BY updated_at DESC"
        ).fetchall()
        staking_rows = conn.execute(
            "SELECT * FROM defi_staking WHERE is_permanently_hidden=1 ORDER BY updated_at DESC"
        ).fetchall()
        zerion_lp_rows = conn.execute(
            "SELECT * FROM zerion_lp_hidden WHERE is_permanently_deleted=1 ORDER BY hidden_at DESC"
        ).fetchall()
        lending_rows = conn.execute(
            "SELECT * FROM lending_archive_overrides WHERE is_permanently_deleted=1 ORDER BY archived_at DESC"
        ).fetchall()
        conn.close()
        zerion_lp_result = []
        for r in zerion_lp_rows:
            d = dict(r)
            zerion_lp_result.append({
                "id": d["id"],
                "source": "zerion",
                "position_key": d["position_key"],
                "pair": d.get("pair", ""),
                "chain": d.get("chain", ""),
                "wallet": d.get("wallet", ""),
                "protocol": d.get("protocol", ""),
                "archived_at": d.get("hidden_at"),
            })
        return jsonify({
            "lp": [dict(r) for r in lp_rows],
            "staking": [dict(r) for r in staking_rows],
            "zerion_lp": zerion_lp_result,
            "lending": [dict(r) for r in lending_rows],
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/archive/permanently-hidden/<string:type_>/<int:pos_id>/restore', methods=['POST'])
def api_archive_permanently_hidden_restore(type_, pos_id):
    """Restore a permanently hidden position back to archived state."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        if type_ == 'lp':
            conn.execute("UPDATE lp_positions SET is_permanently_hidden=0 WHERE id=?", (pos_id,))
        elif type_ == 'staking':
            conn.execute("UPDATE defi_staking SET is_permanently_hidden=0 WHERE id=?", (pos_id,))
        elif type_ == 'zerion_lp':
            conn.execute("UPDATE zerion_lp_hidden SET is_permanently_deleted=0 WHERE id=?", (pos_id,))
        elif type_ == 'lending':
            conn.execute("UPDATE lending_archive_overrides SET is_permanently_deleted=0 WHERE id=?", (pos_id,))
        else:
            conn.close()
            return jsonify({"error": f"Unknown type: {type_}"}), 400
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking/<int:pos_id>/hide', methods=['PUT'])
def api_defi_staking_hide(pos_id):
    """Hide a staking position from the main portfolio view."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE defi_staking SET is_hidden=1, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (pos_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking/<int:pos_id>/unhide', methods=['PUT'])
def api_defi_staking_unhide(pos_id):
    """Unhide a staking position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE defi_staking SET is_hidden=0, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (pos_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-staking/<int:pos_id>/archive', methods=['PUT'])
def api_defi_staking_archive(pos_id):
    """Archive a staking position (moves it to archive tab)."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE defi_staking SET is_archived=1, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (pos_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/defi-protocols', methods=['GET'])
def api_defi_protocols():
    """Return distinct app_name values from active, non-archived staking positions."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT DISTINCT app_name FROM defi_staking WHERE is_active=1 AND (is_archived IS NULL OR is_archived=0) ORDER BY app_name"
        ).fetchall()
        conn.close()
        return jsonify([r['app_name'] for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- Strategy Documents Routes ---

@app.route('/api/strategies')
def api_strategies_list():
    """List all strategy documents with a short text preview."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT id, filename, category, file_size_bytes, uploaded_at, notes, substr(extracted_text, 1, 200) as preview FROM strategy_documents ORDER BY uploaded_at DESC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/strategies/upload', methods=['POST'])
def api_strategies_upload():
    """Upload and extract text from a strategy document (.md, .pdf, .docx, .xlsx, .csv)."""
    from src.storage.portfolio_db import get_connection
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400
    category = request.form.get('category', '').strip()
    if category not in ('bear', 'bull', 'stablecoin', 'cashflow_other', 'trading', 'trading_strategy'):
        return jsonify({"error": "Invalid category"}), 400
    notes = request.form.get('notes', '').strip()
    filename = f.filename
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.md', '.pdf', '.docx', '.xlsx', '.csv'):
        return jsonify({"error": "Unsupported file type"}), 400
    raw = f.read()
    file_size = len(raw)
    try:
        if ext in ('.md', '.csv'):
            extracted_text = raw.decode('utf-8', errors='replace')
        elif ext == '.pdf':
            import pdfplumber, io as _io
            with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                extracted_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        elif ext == '.docx':
            import docx as _docx, io as _io
            doc = _docx.Document(_io.BytesIO(raw))
            extracted_text = '\n'.join(p.text for p in doc.paragraphs)
        elif ext == '.xlsx':
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append('\t'.join('' if v is None else str(v) for v in row))
            extracted_text = '\n'.join(parts)
    except Exception as e:
        return jsonify({"error": f"Failed to extract text: {str(e)}"}), 500
    conn = get_connection()
    try:
        c = conn.cursor()
        c.execute(
            "INSERT INTO strategy_documents (filename, category, extracted_text, file_size_bytes, notes) VALUES (?, ?, ?, ?, ?)",
            (filename, category, extracted_text, file_size, notes)
        )
        doc_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": doc_id, "filename": filename, "preview": extracted_text[:200]})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/strategies/<int:doc_id>', methods=['DELETE'])
def api_strategies_delete(doc_id):
    """Hard-delete a strategy document."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM strategy_documents WHERE id=?", (doc_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/strategies/for-ai/<regime>')
def api_strategies_for_ai(regime):
    """Return concatenated strategy doc text matching the current market regime."""
    from src.storage.portfolio_db import get_connection
    regime_map = {
        'bear':     ['bear',      'trading', 'trading_strategy'],
        'bull':     ['bull',      'trading', 'trading_strategy'],
        'sideways': ['stablecoin', 'cashflow_other', 'trading', 'trading_strategy'],
        'unknown':  ['cashflow_other', 'trading', 'trading_strategy'],
    }
    categories = regime_map.get(regime, ['cashflow_other'])
    conn = get_connection()
    try:
        placeholders = ','.join('?' for _ in categories)
        rows = conn.execute(
            f"SELECT filename, category, extracted_text FROM strategy_documents WHERE category IN ({placeholders}) ORDER BY uploaded_at DESC",
            categories
        ).fetchall()
        conn.close()
        parts = [f"=== {r['filename']} ({r['category']}) ===\n{r['extracted_text']}\n" for r in rows]
        return jsonify({
            "regime": regime,
            "categories_used": categories,
            "strategy_text": '\n'.join(parts),
            "doc_count": len(rows),
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lp/zerion-note/<path:position_key>', methods=['GET'])
def api_get_zerion_lp_note(position_key):
    """Get the persistent note for a Zerion LP position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        row = conn.execute("SELECT note FROM zerion_lp_notes WHERE position_key=?", (position_key,)).fetchone()
        conn.close()
        return jsonify({"note": row["note"] if row else ""})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lp/zerion-note', methods=['PUT'])
def api_put_zerion_lp_note():
    """Save the persistent note for a Zerion LP position."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    position_key = data.get("position_key", "")
    note = data.get("note", "")
    if not position_key:
        return jsonify({"error": "position_key required"}), 400
    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO zerion_lp_notes (position_key, note, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
            "ON CONFLICT(position_key) DO UPDATE SET note=excluded.note, updated_at=CURRENT_TIMESTAMP",
            (position_key, note)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lending/note/<path:position_key>', methods=['GET'])
def api_get_lending_note(position_key):
    """Get the persistent note for a lending position."""
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        row = conn.execute("SELECT note FROM zerion_lending_notes WHERE position_key=?", (position_key,)).fetchone()
        conn.close()
        return jsonify({"note": row["note"] if row else ""})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/lending/note', methods=['PUT'])
def api_put_lending_note():
    """Save the persistent note for a lending position."""
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    position_key = data.get("position_key", "")
    note = data.get("note", "")
    if not position_key:
        return jsonify({"error": "position_key required"}), 400
    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO zerion_lending_notes (position_key, note, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
            "ON CONFLICT(position_key) DO UPDATE SET note=excluded.note, updated_at=CURRENT_TIMESTAMP",
            (position_key, note)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# ===== TRADING TOOLS =====

def _ema_series(closes, period):
    """Return full EMA series for given period."""
    if len(closes) < period:
        return []
    k = 2.0 / (period + 1)
    ema = [sum(closes[:period]) / period]
    for price in closes[period:]:
        ema.append(price * k + ema[-1] * (1.0 - k))
    return ema


def _calc_rsi(closes, period=14):
    if len(closes) < period + 1:
        return None
    deltas = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
    gains = [max(d, 0.0) for d in deltas]
    losses = [max(-d, 0.0) for d in deltas]
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
    if avg_loss == 0:
        return 100.0
    return 100.0 - (100.0 / (1 + avg_gain / avg_loss))


def _calc_macd(closes):
    """Returns (macd, signal, histogram, prev_macd, prev_signal) for last candle."""
    ema12 = _ema_series(closes, 12)
    ema26 = _ema_series(closes, 26)
    if not ema12 or not ema26:
        return (None,) * 5
    offset = len(ema12) - len(ema26)
    macd_vals = [ema12[i + offset] - ema26[i] for i in range(len(ema26))]
    if len(macd_vals) < 9:
        return (macd_vals[-1] if macd_vals else None, None, None, None, None)
    signal_vals = _ema_series(macd_vals, 9)
    if len(signal_vals) < 2 or len(macd_vals) < 2:
        return (None,) * 5
    return (macd_vals[-1], signal_vals[-1], macd_vals[-1] - signal_vals[-1], macd_vals[-2], signal_vals[-2])


def _calc_bollinger(closes, period=20, mult=2.0):
    if len(closes) < period:
        return None, None, None
    recent = closes[-period:]
    sma = sum(recent) / period
    variance = sum((p - sma) ** 2 for p in recent) / period
    std = variance ** 0.5
    return sma + mult * std, sma, sma - mult * std


def _detect_signals(symbol, interval, candles):
    if len(candles) < 30:
        return []
    closes = [float(c[4]) for c in candles]
    current = closes[-1]
    signals = []

    rsi = _calc_rsi(closes)
    if rsi is not None:
        if rsi < 30:
            signals.append(('rsi_oversold', {'rsi': round(rsi, 2), 'price': current}))
        elif rsi > 70:
            signals.append(('rsi_overbought', {'rsi': round(rsi, 2), 'price': current}))

    ema20 = _ema_series(closes, 20)
    if len(ema20) >= 2:
        if current > ema20[-1] and closes[-2] <= ema20[-2]:
            signals.append(('ema20_cross_up', {'ema20': round(ema20[-1], 4), 'price': current}))
        elif current < ema20[-1] and closes[-2] >= ema20[-2]:
            signals.append(('ema20_cross_down', {'ema20': round(ema20[-1], 4), 'price': current}))

    macd, sig, hist, prev_macd, prev_sig = _calc_macd(closes)
    if all(v is not None for v in [macd, sig, prev_macd, prev_sig]):
        if macd > sig and prev_macd <= prev_sig:
            signals.append(('macd_bullish', {'macd': round(macd, 6), 'signal': round(sig, 6)}))
        elif macd < sig and prev_macd >= prev_sig:
            signals.append(('macd_bearish', {'macd': round(macd, 6), 'signal': round(sig, 6)}))

    bb_upper, bb_mid, bb_lower = _calc_bollinger(closes)
    if bb_upper is not None:
        if current <= bb_lower:
            signals.append(('bb_lower_touch', {'bb_lower': round(bb_lower, 4), 'price': round(current, 4)}))
        elif current >= bb_upper:
            signals.append(('bb_upper_touch', {'bb_upper': round(bb_upper, 4), 'price': round(current, 4)}))

    return signals


def _fetch_binance_ohlcv(symbol, interval='4h', limit=500):
    sym = symbol.upper().replace('-', '').replace('/', '')
    resp = requests.get(
        'https://api.binance.com/api/v3/klines',
        params={'symbol': sym, 'interval': interval, 'limit': limit},
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json()


def _get_scanner_prompt_path():
    d = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', '').strip() or '/app/data'
    return os.path.join(d, 'scanner_prompt.md')


def _parse_binance_candles(raw):
    return [{'open': float(c[1]), 'high': float(c[2]), 'low': float(c[3]),
             'close': float(c[4]), 'volume': float(c[5])} for c in raw]


def _ict_ema20(candles):
    closes = [c['close'] for c in candles]
    if len(closes) < 20:
        return None
    k = 2.0 / 21.0
    ema = sum(closes[:20]) / 20.0
    for price in closes[20:]:
        ema = price * k + ema * (1.0 - k)
    return ema


def _calc_stoch_rsi(candles, rsi_len=14, stoch_len=14, k_smooth=3, d_smooth=3):
    """
    Stochastic RSI matching TradingView defaults.
    RSI Length=14, Stoch Length=14, K=3, D=3, OB=80, OS=20.
    Returns (k, d) as the two lines, or (None, None) if insufficient data.
    """
    if len(candles) < rsi_len + stoch_len + k_smooth + d_smooth + 5:
        return None, None
    closes = [c['close'] for c in candles]
    deltas = [closes[i] - closes[i-1] for i in range(1, len(closes))]
    gains  = [d if d > 0 else 0.0 for d in deltas]
    losses = [-d if d < 0 else 0.0 for d in deltas]
    rsi_series = [None] * len(closes)
    avg_gain = sum(gains[:rsi_len]) / rsi_len
    avg_loss = sum(losses[:rsi_len]) / rsi_len
    if avg_loss == 0:
        rsi_series[rsi_len] = 100.0
    else:
        rs = avg_gain / avg_loss
        rsi_series[rsi_len] = 100 - (100 / (1 + rs))
    for i in range(rsi_len + 1, len(closes)):
        avg_gain = (avg_gain * (rsi_len - 1) + gains[i-1]) / rsi_len
        avg_loss = (avg_loss * (rsi_len - 1) + losses[i-1]) / rsi_len
        if avg_loss == 0:
            rsi_series[i] = 100.0
        else:
            rs = avg_gain / avg_loss
            rsi_series[i] = 100 - (100 / (1 + rs))
    stoch_series = [None] * len(rsi_series)
    for i in range(len(rsi_series)):
        window = [rsi_series[j] for j in range(max(0, i - stoch_len + 1), i + 1)
                  if rsi_series[j] is not None]
        if len(window) < stoch_len:
            continue
        lo = min(window)
        hi = max(window)
        if hi == lo:
            stoch_series[i] = 50.0
        else:
            stoch_series[i] = (rsi_series[i] - lo) / (hi - lo) * 100.0
    valid = [(i, v) for i, v in enumerate(stoch_series) if v is not None]
    if len(valid) < k_smooth + d_smooth:
        return None, None
    k_series = []
    for i in range(k_smooth - 1, len(valid)):
        window = [valid[j][1] for j in range(i - k_smooth + 1, i + 1)]
        k_series.append(sum(window) / k_smooth)
    d_series = []
    for i in range(d_smooth - 1, len(k_series)):
        window = k_series[i - d_smooth + 1: i + 1]
        d_series.append(sum(window) / d_smooth)
    if not k_series or not d_series:
        return None, None
    return round(k_series[-1], 2), round(d_series[-1], 2)


def _calc_rsi(candles, period=14):
    """Standard RSI, Wilder smoothing. Returns final value or None."""
    if len(candles) < period + 1:
        return None
    closes = [c['close'] for c in candles]
    deltas = [closes[i] - closes[i-1] for i in range(1, len(closes))]
    gains  = [d if d > 0 else 0.0 for d in deltas]
    losses = [-d if d < 0 else 0.0 for d in deltas]
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return round(100 - (100 / (1 + rs)), 2)


def _calc_rsi_series(candles, period=14):
    """
    Returns full RSI series as a list aligned with candles.
    First `period` values are None. Used for divergence detection.
    """
    if len(candles) < period + 1:
        return [None] * len(candles)
    closes = [c['close'] for c in candles]
    deltas = [closes[i] - closes[i-1] for i in range(1, len(closes))]
    gains  = [d if d > 0 else 0.0 for d in deltas]
    losses = [-d if d < 0 else 0.0 for d in deltas]
    rsi = [None] * len(closes)
    avg_gain = sum(gains[:period]) / period
    avg_loss = sum(losses[:period]) / period
    if avg_loss == 0:
        rsi[period] = 100.0
    else:
        rsi[period] = round(100 - (100 / (1 + avg_gain / avg_loss)), 2)
    for i in range(period + 1, len(closes)):
        avg_gain = (avg_gain * (period - 1) + gains[i-1]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i-1]) / period
        if avg_loss == 0:
            rsi[i] = 100.0
        else:
            rsi[i] = round(100 - (100 / (1 + avg_gain / avg_loss)), 2)
    return rsi


def _detect_rsi_divergence(candles, swing_highs, swing_lows, rsi_series, lookback=50):
    """
    Regular divergence detection using confirmed swing pivots.
    Bullish divergence:  price makes lower low,  RSI makes higher low  → potential reversal up
    Bearish divergence:  price makes higher high, RSI makes lower high  → potential reversal down
    Returns list of divergence dicts: {type, strength, price_diff, rsi_diff, candle_index}
    """
    n = len(candles)
    cutoff = n - lookback
    results = []
    recent_sh = [p for p in swing_highs if p['index'] >= cutoff]
    recent_sl = [p for p in swing_lows  if p['index'] >= cutoff]
    if len(recent_sh) >= 2:
        p1 = recent_sh[-2]
        p2 = recent_sh[-1]
        rsi1 = rsi_series[p1['index']] if p1['index'] < len(rsi_series) else None
        rsi2 = rsi_series[p2['index']] if p2['index'] < len(rsi_series) else None
        if rsi1 is not None and rsi2 is not None:
            if p2['price'] > p1['price'] and rsi2 < rsi1:
                rsi_diff = round(rsi1 - rsi2, 1)
                results.append({'type': 'bearish', 'strength': 'strong' if rsi_diff > 10 else 'moderate' if rsi_diff > 5 else 'weak',
                                 'price_diff': round(p2['price'] - p1['price'], 4), 'rsi_diff': rsi_diff, 'candle_index': p2['index']})
    if len(recent_sl) >= 2:
        p1 = recent_sl[-2]
        p2 = recent_sl[-1]
        rsi1 = rsi_series[p1['index']] if p1['index'] < len(rsi_series) else None
        rsi2 = rsi_series[p2['index']] if p2['index'] < len(rsi_series) else None
        if rsi1 is not None and rsi2 is not None:
            if p2['price'] < p1['price'] and rsi2 > rsi1:
                rsi_diff = round(rsi2 - rsi1, 1)
                results.append({'type': 'bullish', 'strength': 'strong' if rsi_diff > 10 else 'moderate' if rsi_diff > 5 else 'weak',
                                 'price_diff': round(p1['price'] - p2['price'], 4), 'rsi_diff': rsi_diff, 'candle_index': p2['index']})
    return results


def _format_rsi_status(stoch_k, stoch_d, rsi_val, divergences, ob=80, os_=20):
    """Format Stoch RSI OB/OS status and RSI divergence into commentary lines."""
    lines = []
    if stoch_k is not None and stoch_d is not None:
        either_ob = stoch_k >= ob or stoch_d >= ob
        either_os = stoch_k <= os_ or stoch_d <= os_
        if either_ob:
            lines.append(f"Stoch RSI overbought (K={stoch_k}, D={stoch_d}) — potential local top signal")
        elif either_os:
            lines.append(f"Stoch RSI oversold (K={stoch_k}, D={stoch_d}) — potential local bottom signal")
        else:
            dist_ob = min(ob - stoch_k, ob - stoch_d)
            dist_os = min(stoch_k - os_, stoch_d - os_)
            if dist_ob <= 10:
                lines.append(f"Stoch RSI approaching overbought (K={stoch_k}, D={stoch_d}, {round(dist_ob,1)} pts away)")
            elif dist_os <= 10:
                lines.append(f"Stoch RSI approaching oversold (K={stoch_k}, D={stoch_d}, {round(dist_os,1)} pts away)")
            else:
                lines.append(f"Stoch RSI neutral (K={stoch_k}, D={stoch_d})")
    for div in divergences:
        lines.append(f"RSI {div['type']} divergence detected ({div['strength']}, RSI diff={div['rsi_diff']} pts) — reversal signal, not yet confirmed")
    return lines


def _ict_swing_points(candles, lookback=100, left_bars=2, right_bars=2, **kwargs):
    """
    Pivot-based swing detection matching Pine Script ta.pivothigh/pivotlow.
    A pivot high at index i requires left_bars bars of lower highs to the left
    and right_bars bars of lower highs to the right.
    A pivot is confirmed right_bars candles AFTER it forms — matching Pine Script
    confirmation delay. Returns list of dicts with price, index, time.

    'index' is FULL-ARRAY basis — an index into the `candles` argument as
    passed, NOT into the internal lookback slice. Consumers index the full
    array with it (candles[pivot['index']]['time'] == pivot['time'] always).
    """
    c = candles[-lookback:] if len(candles) > lookback else candles
    offset = len(candles) - len(c)   # slice → full-array index correction
    highs, lows = [], []
    for i in range(left_bars, len(c) - right_bars):
        if (all(c[i]['high'] > c[i - j]['high'] for j in range(1, left_bars + 1)) and
                all(c[i]['high'] > c[i + j]['high'] for j in range(1, right_bars + 1))):
            highs.append({
                'price': round(c[i]['high'], 6),
                'index': i + offset,
                'time': c[i].get('time', c[i].get('timestamp', None)),
            })
        if (all(c[i]['low'] < c[i - j]['low'] for j in range(1, left_bars + 1)) and
                all(c[i]['low'] < c[i + j]['low'] for j in range(1, right_bars + 1))):
            lows.append({
                'price': round(c[i]['low'], 6),
                'index': i + offset,
                'time': c[i].get('time', c[i].get('timestamp', None)),
            })
    return highs, lows


def _ict_market_structure(swing_highs, swing_lows):
    if len(swing_highs) >= 2 and len(swing_lows) >= 2:
        if (swing_highs[-1]['price'] > swing_highs[-2]['price']
                and swing_lows[-1]['price'] > swing_lows[-2]['price']):
            return 'bullish'
        if (swing_highs[-1]['price'] < swing_highs[-2]['price']
                and swing_lows[-1]['price'] < swing_lows[-2]['price']):
            return 'bearish'
    return 'ranging'



def _ict_dealing_range(swing_highs, swing_lows, current_close, candles=None,
                       right_bars=2, trace_events=None, trace_collector=None,
                       pivot_min_prominence_atr=0.0):
    """
    BREAK-ONLY dealing-range state machine (D2.5 spec — replaces the D2.1–D2.4
    "break + pivot-driven extension" version, whose confirmed pivots could move
    a boundary with no close-break, pulling the range low to a swept wick on
    pivot confirmation alone).

    Rules (locked, D2.5):
      1. A boundary (range_high / range_low) moves ONLY when a candle CLOSES
         beyond it. No buffer. Closed bars only (the final array element is the
         still-forming live candle — excluded from the walk).
      2. On such a break, the BROKEN boundary jumps to the most extreme wick
         reached so far in the break direction — always, regardless of whether
         the previous level was wick-made or close-made. "So far" = a per-side
         running extreme, reset when that side's own boundary is assigned
         (down-break → running MIN of lows since range_low was last assigned;
         mirror for up). The break bar's own wick is INCLUDED (inclusive).
      3. Wicks alone NEVER move a boundary. Confirmed pivots alone NEVER move a
         boundary. These are sweeps. The pivot-driven extension machinery is
         DELETED (not disabled).
      4. On a break, the OPPOSITE boundary is reassigned to the LEG ORIGIN =
         the most recent usable confirmed opposite pivot occurring before the
         break WITH prominence score >= pivot_min_prominence_atr (D2.6). A
         sub-threshold (noise) pivot is invisible to selection — the walk keeps
         looking further back. If there is no eligible opposite pivot, HOLD the
         existing anchor — the break still fires and the broken side still moves.

      D2.6 SIGNIFICANCE FILTER (pivot_min_prominence_atr, default 0.0 = off):
         prominence = min(prior_leg, following_leg) / ATR14_at_pivot_bar, where
         prior_leg / following_leg are the distances to the ADJACENT confirmed
         OPPOSITE pivots before / after the pivot, and ATR14 is _ict_atr at the
         pivot's bar (SMA-of-TR). No-lookahead: the following leg is simply the
         first opposite pivot already in the usable list at the decision bar, so
         it counts only once that pivot has confirmed; until then the provisional
         score uses prior_leg alone. A pivot with neither leg is ineligible; a
         pivot whose ATR14 is undefined (too few bars before it) PASSES (never
         over-filtered). The filter constrains ONLY selection (origin + seed) —
         never break detection, running-extreme wick tracking, or the ledger.
      5. No last_break_index / origin-window machinery. "Most recent" needs no
         window.
      6. The pivot detector and pivot ledger are unchanged; origins are still
         sourced from confirmed pivots only.

      - Seed: the first usable confirmed pivot high + low pair forms the initial
        range (independent of any extension path); if none ever confirms, return
        None. The per-side running extremes initialise to the seed bar's wicks.
      - A candle closing beyond BOTH anchors (possible only while the range is
        transiently inverted) tie-breaks by the close's side of the prior
        equilibrium.

    Fib convention preserved (0.0 = range HIGH, 1.0 = range LOW):
      price_at_fib = drHigh + fib_val * (drLow - drHigh)

    trace_events (D2.3 / updated D2.5, optional): a caller-supplied list. When
    provided, plain-dict walk events are appended (pivot ledger, seed, break,
    tiebreak, per-bar tail). D2.5 break events carry prior_high/prior_low (old
    boundaries), new_high/new_low, origin (price or None), origin_held (bool),
    and run_hi/run_lo (the running extremes at the break). Extension events no
    longer exist. When None (both live call sites) behavior is unchanged.

    trace_collector (D2.4c, optional): a second caller-supplied list for the
    diagnose per-bar walk window (UTC date strings + raw broke_up/broke_dn).
    Every append is behind a cheap `if _tc is not None` guard and builds only
    plain dicts from already-local variables — when None (all live call sites)
    behavior and allocations are byte-for-byte unchanged. Record types: 'bar'
    (once per break-tested bar), 'break_up', 'break_dn', 'seed'; break records
    carry 'origin' and 'origin_held'. (D2.5 removed the 'extend_*' records.)

    Returns the same contract as before:
      {'high','low','eq','level_786','level_618','zone',
       'anchor_high_time','anchor_low_time'} — or None (no range).
    """
    if not swing_highs or not swing_lows:
        return None

    events = sorted(
        [{'kind': 'high', 'price': p['price'], 'index': p['index'],
          'time': p.get('time')} for p in swing_highs] +
        [{'kind': 'low', 'price': p['price'], 'index': p['index'],
          'time': p.get('time')} for p in swing_lows],
        key=lambda p: p['index'])

    _tr = trace_events
    _tc = trace_collector   # D2.4c per-bar walk-window collector (diagnose only)
    if _tr is not None:
        # D2.6 — the last CLOSED bar the walk processes is len(candles)-2 (the
        # final element is the still-forming live candle, excluded). A pivot is
        # 'pending' (not yet usable/selectable) if its confirm bar exceeds that;
        # its confirm_index is prospective, not proof of confirmation.
        _last_closed = (len(candles) - 2) if candles else None
        _tr.append({'type': 'pivots', 'pivots': [
            {'side': p['kind'], 'price': p['price'], 'time': p['time'],
             'index': p['index'], 'confirm_index': p['index'] + right_bars,
             'pending': (_last_closed is None
                         or (p['index'] + right_bars) > _last_closed)}
            for p in events]})

    rng_h = None   # {'kind','price','index','time'} (or {'price','time','index'} after a jump)
    rng_l = None

    if not candles:
        # No candle clock → no closes to break on: seed from the first
        # pivot high + low pair (in confirmation order) and hold.
        first_h = next((p for p in events if p['kind'] == 'high'), None)
        first_l = next((p for p in events if p['kind'] == 'low'), None)
        if first_h is None or first_l is None:
            return None
        rng_h, rng_l = first_h, first_l
    else:
        usable_highs = []   # confirmed pivots so far, chronological
        usable_lows = []
        ei = 0
        run_hi = None   # D2.5 per-side running extreme since that boundary was
        run_lo = None   # last assigned; {'price','time','index'}. Tracks wicks
                        # every bar; a boundary NEVER moves on the running
                        # extreme alone — only a CLOSE-break consumes it.
        _tail_buf = [] if _tr is not None else None   # D2.4: rolling per-bar log

        def _mk(price, idx):
            return {'price': price, 'time': candles[idx].get('time'), 'index': idx}

        def _prominence(p, opp_usable):
            """D2.6 significance of pivot p vs the confirmed OPPOSITE pivots
            usable so far (opp_usable, chronological). Returns a dict:
            {score, prior_leg, following_leg, atr, filtered_out}. No-lookahead
            is automatic: opp_usable holds only pivots confirmed by the current
            bar, so the 'following' opposite is the first one already usable.
            ATR undefined → passes (never over-filtered); no leg → ineligible.
            Threshold <= 0 short-circuits to a pure no-op (exact D2.5 behaviour,
            never touches ATR) so the default keeps every pivot selectable."""
            if pivot_min_prominence_atr <= 0:
                return {'score': None, 'prior_leg': None, 'following_leg': None,
                        'atr': None, 'filtered_out': False}
            before = None
            after = None
            for o in opp_usable:
                if o['index'] < p['index']:
                    before = o                     # last opposite before p
                elif o['index'] > p['index'] and after is None:
                    after = o                      # first opposite after p (confirmed)
            prior_leg = abs(p['price'] - before['price']) if before else None
            following_leg = abs(p['price'] - after['price']) if after else None
            legs = [x for x in (prior_leg, following_leg) if x is not None]
            if not legs:
                return {'score': None, 'prior_leg': prior_leg,
                        'following_leg': following_leg, 'atr': None,
                        'filtered_out': True}            # ineligible (no leg)
            a = _ict_atr(candles, 14, at_index=p['index'])
            if not a:
                return {'score': None, 'prior_leg': prior_leg,
                        'following_leg': following_leg, 'atr': a,
                        'filtered_out': False}           # ATR undefined → pass
            score = min(legs) / a
            return {'score': score, 'prior_leg': prior_leg,
                    'following_leg': following_leg, 'atr': a,
                    'filtered_out': score < pivot_min_prominence_atr}

        def _select_origin(opp_usable, self_usable):
            """Most recent opposite pivot passing the prominence filter, plus a
            capped candidate trace (chronological, last 6) each annotated with
            score + filtered_out. (self_usable unused; kept for symmetry.)"""
            picked = None
            for p in reversed(opp_usable):
                info = _prominence(p, self_usable)
                if picked is None and not info['filtered_out']:
                    picked = p
            cands = []
            for p in opp_usable[-6:]:
                info = _prominence(p, self_usable)
                cands.append({'price': p['price'], 'time': p['time'],
                              'index': p['index'],
                              'confirm_index': p['index'] + right_bars,
                              'score': (round(info['score'], 4)
                                        if info['score'] is not None else None),
                              'filtered_out': info['filtered_out']})
            return picked, cands

        # CLOSED BARS ONLY (D2.1): the final array element is the still-forming
        # live candle — excluded (no confirmation, no break on its live close).
        for k in range(len(candles) - 1):
            if rng_h is not None and rng_l is not None:
                hi = candles[k]['high']
                lo = candles[k]['low']
                close = candles[k]['close']
                # D2.5 rule 2 — maintain the per-side running extremes INCLUSIVE
                # of this bar's wick BEFORE the break test, so a break bar's own
                # wick is eligible as the jump target.
                if run_hi is None or hi > run_hi['price']:
                    run_hi = _mk(hi, k)
                if run_lo is None or lo < run_lo['price']:
                    run_lo = _mk(lo, k)
                _ph, _pl = rng_h['price'], rng_l['price']
                _rhi, _rlo = run_hi['price'], run_lo['price']   # extremes AT the break (pre-reset)
                if _tr is not None:
                    _tail_buf.append({'k': k, 'candle_time': candles[k].get('time'),
                                       'close': close, 'range_low': _pl, 'range_high': _ph})
                broke_up = close > rng_h['price']
                broke_dn = close < rng_l['price']
                if _tc is not None:
                    _tc.append({'type': 'bar', 'k': k,
                                'date': time.strftime('%Y-%m-%d',
                                    time.gmtime(candles[k].get('time') or 0)),
                                'close': close, 'rng_h': _ph, 'rng_l': _pl,
                                'broke_up': broke_up, 'broke_dn': broke_dn})
                if broke_up and broke_dn:
                    # Transiently inverted range (rare under break-only) —
                    # tie-break by the close's side of the prior equilibrium.
                    _eq_prior = (rng_h['price'] + rng_l['price']) / 2.0
                    if close >= _eq_prior:
                        broke_dn = False
                    else:
                        broke_up = False
                    if _tr is not None:
                        _tr.append({'type': 'tiebreak', 'k': k,
                                    'resolved_direction': 'up' if broke_up else 'down'})

                if broke_dn:
                    # D2.5 — broken side (low) jumps to the running-min wick
                    # reached so far (inclusive of this bar). Opposite side
                    # (high) re-derives to the leg origin = most recent usable
                    # confirmed HIGH pivot before this bar PASSING the D2.6
                    # prominence filter; HOLD if none eligible.
                    rng_l = dict(run_lo)
                    origin, _origin_cands = _select_origin(usable_highs, usable_lows)
                    origin_held = origin is None
                    if origin is not None:
                        rng_h = dict(origin)
                        run_hi = _mk(hi, k)   # opposite side reassigned → reset
                    run_lo = _mk(lo, k)       # broken side assigned → reset
                    if _tr is not None:
                        _tr.append({'type': 'break', 'k': k,
                                    'candle_time': candles[k].get('time'),
                                    'direction': 'down', 'close': close,
                                    'prior_high': _ph, 'prior_low': _pl,
                                    'new_high': rng_h['price'], 'new_low': rng_l['price'],
                                    'origin_candidates': _origin_cands,
                                    'origin': (origin['price'] if origin is not None else None),
                                    'origin_held': origin_held,
                                    'run_hi': _rhi, 'run_lo': _rlo})
                    if _tc is not None:
                        _tc.append({'type': 'break_dn', 'k': k,
                                    'date': time.strftime('%Y-%m-%d',
                                        time.gmtime(candles[k].get('time') or 0)),
                                    'close': close, 'rng_h': rng_h['price'],
                                    'rng_l': rng_l['price'], 'origin_held': origin_held,
                                    'origin': (origin['price'] if origin is not None else None)})
                elif broke_up:
                    rng_h = dict(run_hi)
                    origin, _origin_cands = _select_origin(usable_lows, usable_highs)
                    origin_held = origin is None
                    if origin is not None:
                        rng_l = dict(origin)
                        run_lo = _mk(lo, k)   # opposite side reassigned → reset
                    run_hi = _mk(hi, k)       # broken side assigned → reset
                    if _tr is not None:
                        _tr.append({'type': 'break', 'k': k,
                                    'candle_time': candles[k].get('time'),
                                    'direction': 'up', 'close': close,
                                    'prior_high': _ph, 'prior_low': _pl,
                                    'new_high': rng_h['price'], 'new_low': rng_l['price'],
                                    'origin_candidates': _origin_cands,
                                    'origin': (origin['price'] if origin is not None else None),
                                    'origin_held': origin_held,
                                    'run_hi': _rhi, 'run_lo': _rlo})
                    if _tc is not None:
                        _tc.append({'type': 'break_up', 'k': k,
                                    'date': time.strftime('%Y-%m-%d',
                                        time.gmtime(candles[k].get('time') or 0)),
                                    'close': close, 'rng_h': rng_h['price'],
                                    'rng_l': rng_l['price'], 'origin_held': origin_held,
                                    'origin': (origin['price'] if origin is not None else None)})

            # Absorb pivots that become usable at candle k — a pivot at index i
            # is usable only once ALL right_bars bars after it have closed.
            # D2.5: pivots feed the usable lists (origin sourcing) and the seed
            # ONLY — they never move a boundary directly (extension path deleted).
            while ei < len(events) and events[ei]['index'] + right_bars <= k:
                p = events[ei]
                ei += 1
                if p['kind'] == 'high':
                    usable_highs.append(p)
                else:
                    usable_lows.append(p)

            # Seed once the first ELIGIBLE usable high + low pair exists. D2.6:
            # the seed selects from pivots, so the prominence filter applies —
            # a sub-threshold pivot is skipped and the walk waits for the next
            # eligible one (early pivots pass by ATR-undefined policy, so this
            # is a no-op until a pivot can be positively scored below threshold).
            if rng_h is None or rng_l is None:
                _seed_h = next((p for p in usable_highs
                                if not _prominence(p, usable_lows)['filtered_out']), None)
                _seed_l = next((p for p in usable_lows
                                if not _prominence(p, usable_highs)['filtered_out']), None)
                if _seed_h is not None and _seed_l is not None:
                    rng_h = dict(_seed_h)
                    rng_l = dict(_seed_l)
                    run_hi = _mk(candles[k]['high'], k)   # running extremes start at seed
                    run_lo = _mk(candles[k]['low'], k)
                    if _tr is not None:
                        _tr.append({'type': 'seed',
                                    'high': {'price': rng_h['price'], 'time': rng_h['time']},
                                    'low': {'price': rng_l['price'], 'time': rng_l['time']}})
                    if _tc is not None:
                        _tc.append({'type': 'seed', 'k': k,
                                    'date': time.strftime('%Y-%m-%d',
                                        time.gmtime(candles[k].get('time') or 0)),
                                    'rng_h': rng_h['price'], 'rng_l': rng_l['price']})

        if _tr is not None:
            # D2.4 — per-candle tail log (last ~40 closed bars): candle time,
            # close, and the range low/high THAT WERE IN EFFECT for that
            # bar's break test (i.e. the same prior_high/prior_low a 'break'
            # event would show, logged for every bar, not just fired ones —
            # this is how a MISSED break becomes visible in the live trace).
            for t in _tail_buf[-40:]:
                _tr.append({'type': 'tail', 'k': t['k'], 'candle_time': t['candle_time'],
                            'close': t['close'], 'range_low': t['range_low'],
                            'range_high': t['range_high']})

        if rng_h is None or rng_l is None:
            return None

    dr_high, dr_high_time = rng_h['price'], rng_h['time']
    dr_low, dr_low_time = rng_l['price'], rng_l['time']

    # Safety inversion guard (a break can leave the range transiently
    # inverted until the developing side confirms — never return it inverted).
    if dr_low > dr_high:
        dr_high, dr_low = dr_low, dr_high
        dr_high_time, dr_low_time = dr_low_time, dr_high_time

    eq = (dr_high + dr_low) / 2.0

    # Fib levels matching Pine Script convention (0.0=high, 1.0=low)
    # price = drHigh + fib * (drLow - drHigh)
    range_size = dr_high - dr_low
    level_786 = dr_high - 0.786 * range_size
    level_618 = dr_high - 0.618 * range_size

    return {
        'high':             round(dr_high,    6),
        'low':              round(dr_low,     6),
        'eq':               round(eq,         6),
        'level_786':        round(level_786,  6),
        'level_618':        round(level_618,  6),
        'zone':             'discount' if current_close < eq else 'premium',
        'anchor_high_time': dr_high_time,
        'anchor_low_time':  dr_low_time,
    }


def _ict_fvg_in_zone(candles, direction, zone_low, zone_high,
                     span_start_idx, span_end_idx, atr, min_atr_frac):
    """
    STANDALONE in-zone FVG detection (cascade phase 1a) — FVGs as first-class
    POIs, independent of any order block. Additive capability: nothing on the
    live gate path calls this yet; the diagnose worksheet previews it.

    Scans 3-candle triples ONLY within [span_start_idx, span_end_idx] (the DR
    leg span — same convention as _ict_order_block). All indices FULL-ARRAY
    basis: candles[start_idx]['time'] == gap_start_time always.

    Gap math (same proven formulas as the displacement-FVG detector):
      bearish → c0.low  > c2.high → gap top=c0.low,  bottom=c2.high
      bullish → c0.high < c2.low  → gap top=c2.low,  bottom=c0.high
    ALL qualifying gaps in the span are detected (not first-only).

    Filters, in order (each exclusion is counted):
      1. size:  (top − bottom) >= min_atr_frac * atr. Falsy/zero atr → size
         filter skipped entirely (never divides, never errors).
      2. zone:  the gap's [bottom, top] must intersect [zone_low, zone_high]
         (delegated to _poi_in_ote, which accepts any {top, bottom} dict).
      3. fill:  gaps fully traded through after formation (state 'full') are
         spent POIs — excluded from candidates but counted.

    Fill state mirrors the displacement-FVG fill logic: extreme of price
    action from formation+1 (gap index + 3) to the end of the FULL array;
    untouched / partial / full + integer pct.

    Returns:
      {'candidates': [ {top, bottom, start_idx, end_idx, gap_start_time,
                        gap_end_time, size, size_atr_frac, fill: {state, pct},
                        direction} ... ],           # chronological order
       'counts': {'found', 'size_filtered', 'out_of_zone', 'fully_filled',
                  'kept'}}
    """
    counts = {'found': 0, 'size_filtered': 0, 'out_of_zone': 0,
              'fully_filled': 0, 'kept': 0}
    if (not candles or direction not in ('bullish', 'bearish')
            or span_start_idx is None or span_end_idx is None):
        return {'candidates': [], 'counts': counts}

    bear = (direction == 'bearish')
    s0 = max(0, int(span_start_idx))
    s1 = min(int(span_end_idx), len(candles) - 1)

    candidates = []
    for g in range(s0, s1 - 1):          # triple g, g+1, g+2 within the span
        c0, c2 = candles[g], candles[g + 2]
        if bear and c0['low'] > c2['high']:
            top, bottom = c0['low'], c2['high']
        elif (not bear) and c0['high'] < c2['low']:
            top, bottom = c2['low'], c0['high']
        else:
            continue
        counts['found'] += 1
        size = top - bottom

        # 1. size filter (skipped entirely when atr is falsy/zero)
        if atr and size < min_atr_frac * atr:
            counts['size_filtered'] += 1
            continue

        # 2. zone filter — reuse the generic interval-overlap helper
        if not _poi_in_ote([{'top': top, 'bottom': bottom}], zone_low, zone_high):
            counts['out_of_zone'] += 1
            continue

        # 3. fill state from price action after formation (full array)
        if bear:
            _ext = max((candles[k]['high'] for k in range(g + 3, len(candles))),
                       default=None)
            _touched = _ext is not None and _ext > bottom
            _depth = (_ext - bottom) if _touched else 0.0
        else:
            _ext = min((candles[k]['low'] for k in range(g + 3, len(candles))),
                       default=None)
            _touched = _ext is not None and _ext < top
            _depth = (top - _ext) if _touched else 0.0
        fill_pct = int(round(max(0.0, min(1.0,
            (_depth / size) if size > 0 else 0.0)) * 100))
        if not _touched:
            fill = {'state': 'untouched', 'pct': fill_pct}
        elif fill_pct >= 100:
            counts['fully_filled'] += 1
            continue                     # spent POI — counted, not returned
        else:
            fill = {'state': 'partial', 'pct': fill_pct}

        counts['kept'] += 1
        candidates.append({
            'top': round(top, 6),
            'bottom': round(bottom, 6),
            'start_idx': g,
            'end_idx': g + 2,
            'gap_start_time': c0.get('time'),
            'gap_end_time': c2.get('time'),
            'size': round(size, 6),
            'size_atr_frac': (round(size / atr, 3) if atr else None),
            'fill': fill,
            'direction': direction,
        })
    return {'candidates': candidates, 'counts': counts}


def _leg_span_indices(candles, anchor_high_time, anchor_low_time, direction):
    """Resolve the DR leg's searchable span from the anchor timestamps.

    Extracted verbatim from the closure formerly inside _ict_order_block so
    the OB detector, the phase_fvg_zone diagnostics, and the cascade snapshot
    builder share ONE implementation. Semantics identical to the original,
    including the 3-candle backward margin at the leg origin:
      bearish (down leg)  → [high_idx − 3, low_idx]
      bullish (up leg)    → [low_idx − 3, high_idx]
    FULL-ARRAY index convention. Returns (span_start, span_end), or None when
    either anchor time cannot be resolved or the leg is degenerate/inverted.
    """
    if not candles or direction not in ('bullish', 'bearish'):
        return None

    def _idx_of_time(t):
        if t is None:
            return None
        for i in range(len(candles) - 1, -1, -1):
            if candles[i].get('time') == t:
                return i
        return None

    hi_idx = _idx_of_time(anchor_high_time)
    lo_idx = _idx_of_time(anchor_low_time)
    if hi_idx is None or lo_idx is None:
        return None

    bear = (direction == 'bearish')
    origin_idx = hi_idx if bear else lo_idx
    span_end = lo_idx if bear else hi_idx
    span_start = max(0, origin_idx - 3)
    if span_end <= span_start:
        return None   # degenerate/inverted leg — no searchable span
    return span_start, span_end


def _ict_order_block(candles, direction, dr):
    """
    LEG-ANCHORED order-block detection (July spec — replaces the BOS-window
    approach, which anchored candidates to stale swings via the newest swing
    level the current close was beyond plus a hardcoded 30-candle window).

    Search span = the DR leg itself, resolved from the DR anchor timestamps:
      bearish (down leg)  → [high_anchor_idx − 3, low_anchor_idx]
      bullish (up leg)    → [low_anchor_idx − 3, high_anchor_idx]
    OBs outside the leg are impossible by construction.

    Candidates = MERGED CLUSTERS of consecutive opposing candles (bearish leg
    → up candles close>open; bullish mirror). Cluster bounds are wick-to-wick:
    top = max(high), bottom = min(low) across the cluster. A single opposing
    candle is a valid one-candle cluster.

    Per-candidate facts (the candle-dimension check itself lives in the gate,
    judged on the DISPLACEMENT candle at cluster_end_idx + 1):
      - displacement_fvg: FIRST same-direction 3-candle gap at/after the
        cluster end within the leg span (the triple may start on the cluster
        end candle — displacement leaves from it). Existence gates.
      - fvg_fill: untouched/partial/full + pct from later price action over
        the full array. Annotation only, never gates.
      - htf_close_through: any later HTF candle CLOSED through the cluster's
        far side (bearish: close > cluster top; bullish: close < cluster
        bottom). ANNOTATION ONLY — invalidation gating is LTF-sourced in the
        pipeline; no gate reads this key.
      - swept: a wick beyond the cluster extreme within the 5 candles before
        cluster start (bearish: high > top; bullish: low < bottom).
      - tier/tier_reason: display annotation only (strict when swept and/or
        FVG exists), kept for the signals API / chart overlay.

    Direction-parameterized — one implementation, no bull/bear duplication.

    Returns list of candidate dicts, newest-first (by cluster end):
      {
        'type': 'bullish'|'bearish',
        'top': float, 'bottom': float,          # merged wick-to-wick bounds
        'cluster_start_idx': int, 'cluster_end_idx': int,
        'cluster_start_time': int|None, 'cluster_end_time': int|None,
        'index': int, 'time': int|None,         # legacy single-candle fields:
                                                 # == cluster END candle (the
                                                 # candle displacement leaves
                                                 # from)
        'tier': 'strict'|'standard', 'tier_reason': str,
        'swept': bool,
        'displacement_fvg': {'top','bottom','gap_start_time','gap_end_time'}|None,
        'fvg_fill': {'state','pct'}|None,
        'htf_close_through': bool,
      }
    """
    if not candles or not dr or direction not in ('bullish', 'bearish'):
        return []

    _span = _leg_span_indices(candles, dr.get('anchor_high_time'),
                              dr.get('anchor_low_time'), direction)
    if _span is None:
        return []
    span_start, span_end = _span
    bear = (direction == 'bearish')

    def _opposing(c):
        return (c['close'] > c['open']) if bear else (c['close'] < c['open'])

    # ── merge consecutive opposing candles into clusters ────────────────────
    clusters = []
    i = span_start
    while i <= span_end:
        if _opposing(candles[i]):
            j = i
            while j + 1 <= span_end and _opposing(candles[j + 1]):
                j += 1
            clusters.append((i, j))
            i = j + 1
        else:
            i += 1

    out = []
    for (s, e) in clusters:
        top = round(max(candles[k]['high'] for k in range(s, e + 1)), 6)
        bottom = round(min(candles[k]['low'] for k in range(s, e + 1)), 6)

        # Sweep: wick beyond the cluster extreme in the 5 candles before start.
        if bear:
            swept = any(candles[k]['high'] > top for k in range(max(0, s - 5), s))
        else:
            swept = any(candles[k]['low'] < bottom for k in range(max(0, s - 5), s))

        # First same-direction 3-candle gap at/after cluster end within the leg.
        disp_fvg = None
        fvg_fill = None
        for g in range(e, span_end - 1):
            c0, c2 = candles[g], candles[g + 2]
            if bear and c0['low'] > c2['high']:
                disp_fvg = {'top': round(c0['low'], 6),
                            'bottom': round(c2['high'], 6),
                            'gap_start_time': c0.get('time'),
                            'gap_end_time': c2.get('time')}
            elif (not bear) and c0['high'] < c2['low']:
                disp_fvg = {'top': round(c2['low'], 6),
                            'bottom': round(c0['high'], 6),
                            'gap_start_time': c0.get('time'),
                            'gap_end_time': c2.get('time')}
            if disp_fvg is None:
                continue
            # Fill state (annotation only) from price action after formation,
            # over the FULL array — fill is about later price, not the leg.
            _gap = disp_fvg['top'] - disp_fvg['bottom']
            if bear:
                _ext = max((candles[k]['high'] for k in range(g + 3, len(candles))),
                           default=None)
                _touched = _ext is not None and _ext > disp_fvg['bottom']
                _depth = (_ext - disp_fvg['bottom']) if _touched else 0.0
            else:
                _ext = min((candles[k]['low'] for k in range(g + 3, len(candles))),
                           default=None)
                _touched = _ext is not None and _ext < disp_fvg['top']
                _depth = (disp_fvg['top'] - _ext) if _touched else 0.0
            fill_pct = int(round(max(0.0, min(1.0,
                (_depth / _gap) if _gap > 0 else 0.0)) * 100))
            if not _touched:
                _fill_state = 'untouched'
            elif fill_pct >= 100:
                _fill_state = 'full'
            else:
                _fill_state = 'partial'
            fvg_fill = {'state': _fill_state, 'pct': fill_pct}
            break   # FIRST qualifying gap only

        # HTF close-through (annotation only — invalidation gating is LTF-side).
        if bear:
            htf_close_through = any(candles[k]['close'] > top
                                    for k in range(e + 1, len(candles)))
        else:
            htf_close_through = any(candles[k]['close'] < bottom
                                    for k in range(e + 1, len(candles)))

        # Tier annotation (display only, never a gate).
        if swept and disp_fvg is not None:
            tier, tier_reason = 'strict', 'liquidity swept + FVG created'
        elif swept:
            tier, tier_reason = 'strict', 'liquidity swept'
        elif disp_fvg is not None:
            tier, tier_reason = 'strict', 'FVG created'
        else:
            tier, tier_reason = 'standard', 'opposing cluster in leg'

        out.append({
            'type': direction,
            'top': top,
            'bottom': bottom,
            'cluster_start_idx': s,
            'cluster_end_idx': e,
            'cluster_start_time': candles[s].get('time'),
            'cluster_end_time': candles[e].get('time'),
            'index': e,
            'time': candles[e].get('time'),
            'tier': tier,
            'tier_reason': tier_reason,
            'swept': swept,
            'displacement_fvg': disp_fvg,
            'fvg_fill': fvg_fill,
            'htf_close_through': htf_close_through,
        })

    out.reverse()   # newest-first (by cluster end), matching prior convention
    return out


def _detect_choch(candles, swing_highs, swing_lows, direction):
    """
    Detect a close-confirmed Change of Character (CHoCH) in the given direction.

    For bullish CHoCH: price must close ABOVE the most recent swing high
    (a wick that closes back does not count — close must be >= swing high price).

    For bearish CHoCH: price must close BELOW the most recent swing low.

    Uses the second-most-recent swing point as the CHoCH level to detect
    (the most recent opposing swing is the one being broken).

    Returns dict:
      {
        'fired': bool,
        'level': float,        # the swing level that was broken
        'candle_index': int,   # index of the candle whose close confirmed
        'close': float,        # confirming close price
      }
    """
    if not candles:
        return {'fired': False, 'level': None, 'candle_index': None, 'close': None}

    if direction == 'bullish':
        # Most recent swing high is the CHoCH level to break
        if not swing_highs:
            return {'fired': False, 'level': None, 'candle_index': None, 'close': None}
        choch_level = swing_highs[-1]['price']
        # Scan from the swing high index forward for a close above it
        sh_idx = swing_highs[-1]['index']
        for i in range(sh_idx + 1, len(candles)):
            if candles[i]['close'] >= choch_level:
                return {
                    'fired': True,
                    'level': round(choch_level, 6),
                    'candle_index': i,
                    'close': round(candles[i]['close'], 6),
                }
        return {'fired': False, 'level': round(choch_level, 6), 'candle_index': None, 'close': None}

    elif direction == 'bearish':
        if not swing_lows:
            return {'fired': False, 'level': None, 'candle_index': None, 'close': None}
        choch_level = swing_lows[-1]['price']
        sl_idx = swing_lows[-1]['index']
        for i in range(sl_idx + 1, len(candles)):
            if candles[i]['close'] <= choch_level:
                return {
                    'fired': True,
                    'level': round(choch_level, 6),
                    'candle_index': i,
                    'close': round(candles[i]['close'], 6),
                }
        return {'fired': False, 'level': round(choch_level, 6), 'candle_index': None, 'close': None}

    return {'fired': False, 'level': None, 'candle_index': None, 'close': None}


def _poi_in_ote(poi_list, ote_low, ote_high):
    """
    Filter a list of POI dicts to those whose zone overlaps the OTE band.

    Each POI dict must have 'top' and 'bottom' keys (as returned by
    _ict_order_block and _ict_fvg_in_zone).

    OTE band: level_618 (ote_low) to level_786 (ote_high) — discount for bull,
    premium for bear. The band is directional but this helper just checks overlap;
    the caller is responsible for passing the correct ote_low/ote_high.

    Returns list of matching POIs (may be empty).
    """
    if not poi_list or ote_low is None or ote_high is None:
        return []
    band_lo = min(ote_low, ote_high)
    band_hi = max(ote_low, ote_high)
    result = []
    for poi in poi_list:
        if poi is None:
            continue
        top = poi.get('top')
        bottom = poi.get('bottom')
        if top is None or bottom is None:
            continue
        poi_lo = min(top, bottom)
        poi_hi = max(top, bottom)
        # Overlap check: poi range intersects band range
        if poi_lo <= band_hi and poi_hi >= band_lo:
            result.append(poi)
    return result


# ── ATR + scanner tunables + triage ranking/builders ───────────────────────

_LTF_DUR_MIN = {
    '5m': 5, '15m': 15, '1h': 60, '4h': 240,
    '12h': 720, '1d': 1440, '1w': 10080,
}

_SCANNER_SETTINGS_PATH = os.path.join('data', 'scanner_settings.json')
_SCANNER_SETTINGS_DEFAULTS = {
    'ob_body_atr_min':         0.7,    # OB body must be >= this * ATR(14)
    'ob_body_range_ratio_min': 0.35,   # OB body / candle_range must be >= this
    'fvg_min_atr_frac':        0.10,   # standalone FVG gap size must be >= this * ATR(14)
    'dr_pivot_min_prominence_atr': 1.49,  # D2.6: DR origin/seed pivot must score
                                          # min(prior_leg,following_leg)/ATR14 >= this
                                          # to be selectable (derived from the H12
                                          # 60935-vs-65587 discriminating gap)
    'scan_min_volume':         100000, # USD 24h notional floor for the scheduled wide-scan universe
    'scan_max_tickers':        250,    # safety cap on universe size
    'auto_scan_enabled':       False,  # master switch for scheduled scan
    'scan_times_utc':          ['11:00', '18:00', '21:30'],  # up to 3 HH:MM UTC; '' = slot off
    'scan_asset_type':         'all',  # universe filter: 'all' | 'crypto' | 'tradfi'
}


def _scanner_settings(updates=None):
    """Key-value scanner tunables, file-backed (same mechanism as display
    prefs). Missing/invalid file → defaults. No DB schema involved.

    No-args call behaves EXACTLY as the original read-only helper (whitelisted
    merge over defaults, corrupt-file safe). Passing `updates` merges them in,
    persists to disk, and returns the result (mirrors _telegram_settings)."""
    s = dict(_SCANNER_SETTINGS_DEFAULTS)
    try:
        if os.path.exists(_SCANNER_SETTINGS_PATH):
            with open(_SCANNER_SETTINGS_PATH, 'r') as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                for k in _SCANNER_SETTINGS_DEFAULTS:
                    if k in saved:
                        s[k] = saved[k]
    except Exception:
        pass
    if updates:
        s.update(updates)
        with open(_SCANNER_SETTINGS_PATH, 'w') as f:
            json.dump(s, f)
    return s


@app.route('/api/trading/scanner/settings', methods=['GET'])
@login_required
def api_scanner_settings_get():
    return jsonify(_scanner_settings())


@app.route('/api/trading/scanner/settings', methods=['PUT'])
@login_required
def api_scanner_settings_put():
    data = request.get_json() or {}
    updates = {}
    # scheduled-scan keys
    if 'auto_scan_enabled' in data:
        updates['auto_scan_enabled'] = bool(data['auto_scan_enabled'])
    if 'scan_min_volume' in data:
        try:
            updates['scan_min_volume'] = max(0, float(data['scan_min_volume']))
        except (TypeError, ValueError):
            return jsonify({'error': 'scan_min_volume must be a number'}), 400
    if 'scan_max_tickers' in data:
        try:
            updates['scan_max_tickers'] = max(1, int(data['scan_max_tickers']))
        except (TypeError, ValueError):
            return jsonify({'error': 'scan_max_tickers must be an integer'}), 400
    if 'scan_asset_type' in data:
        if data['scan_asset_type'] not in ('all', 'crypto', 'tradfi'):
            return jsonify({'error': 'scan_asset_type must be all, crypto, or tradfi'}), 400
        updates['scan_asset_type'] = data['scan_asset_type']
    # OB threshold keys (also editable, so the existing backend tunables
    # become UI-accessible)
    if 'ob_body_atr_min' in data:
        try:
            updates['ob_body_atr_min'] = max(0.0, float(data['ob_body_atr_min']))
        except (TypeError, ValueError):
            return jsonify({'error': 'ob_body_atr_min must be a number'}), 400
    if 'ob_body_range_ratio_min' in data:
        try:
            v = float(data['ob_body_range_ratio_min'])
            if not (0.0 <= v <= 1.0):
                return jsonify({'error': 'ob_body_range_ratio_min must be 0–1'}), 400
            updates['ob_body_range_ratio_min'] = v
        except (TypeError, ValueError):
            return jsonify({'error': 'ob_body_range_ratio_min must be a number'}), 400
    # scheduled-scan windows: up to 3 "HH:MM" UTC slots; '' = that slot off
    if 'scan_times_utc' in data:
        raw = data['scan_times_utc']
        if not isinstance(raw, list):
            return jsonify({'error': 'scan_times_utc must be a list'}), 400
        cleaned = []
        for entry in raw[:3]:   # max 3 slots
            s = (str(entry) if entry is not None else '').strip()
            if s == '':
                cleaned.append('')        # empty slot = disabled
                continue
            # validate HH:MM 24h
            import re as _re
            if not _re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', s):
                return jsonify({'error': f'invalid time "{s}" — use HH:MM 24h UTC'}), 400
            # normalize to zero-padded HH:MM
            hh, mm = s.split(':')
            cleaned.append(f"{int(hh):02d}:{int(mm):02d}")
        updates['scan_times_utc'] = cleaned
    if not updates:
        return jsonify({'error': 'no valid settings provided'}), 400
    saved = _scanner_settings(updates=updates)
    return jsonify(saved)


def _telegram_settings(updates=None):
    """File-backed KV for the triage Telegram digest (separate from the
    daily-notification telegram config). Mirrors _scanner_settings()."""
    path = os.path.join('data', 'telegram_settings.json')
    defaults = {
        'bot_token': '',
        'chat_id': '',
        'enabled': False,
        'send_times_utc': ['12:00', '19:00', '22:30'],
        # Custom recurring reminders, each fired at its own times_utc slots.
        # Surfaced on read via the {**defaults, **stored} merge below (this helper
        # spreads the full stored dict — no whitelist — so the key isn't dropped).
        'reminders': [
            {
                'id': 'camel_finance',
                'label': 'Camel Finance Posts',
                'message': '👀 Check Camel Finance Posts',
                'enabled': False,
                'times_utc': []
            }
        ],
    }
    if not os.path.exists(path):
        if updates:
            with open(path, 'w') as f:
                json.dump({**defaults, **updates}, f)
            return {**defaults, **updates}
        return defaults
    with open(path, 'r') as f:
        stored = json.load(f)
    merged = {**defaults, **stored}
    if updates:
        merged.update(updates)
        with open(path, 'w') as f:
            json.dump(merged, f)
    return merged


# Startup migration: loosen the OB candle-dimension thresholds in the live
# scanner_settings.json on the volume. Overwrite only values still at the OLD
# defaults (1.2 / 0.55) so a genuine manual override is respected. No-op if the
# file is absent (the in-code defaults already apply) or already migrated.
try:
    if os.path.exists(_SCANNER_SETTINGS_PATH):
        with open(_SCANNER_SETTINGS_PATH, 'r') as _ssf:
            _ss_data = json.load(_ssf)
        if isinstance(_ss_data, dict):
            _ss_changed = False
            if _ss_data.get('ob_body_atr_min') == 1.2:
                _ss_data['ob_body_atr_min'] = 0.7
                _ss_changed = True
            if _ss_data.get('ob_body_range_ratio_min') == 0.55:
                _ss_data['ob_body_range_ratio_min'] = 0.35
                _ss_changed = True
            if _ss_changed:
                with open(_SCANNER_SETTINGS_PATH, 'w') as _ssf:
                    json.dump(_ss_data, _ssf, indent=2)
                print("[startup] scanner_settings.json OB thresholds migrated "
                      "to 0.7 / 0.35", flush=True)
except Exception as _ss_err:
    print(f"[startup] scanner_settings migration skipped: "
          f"{type(_ss_err).__name__}: {_ss_err}", flush=True)


def _ict_atr(candles, period=14, at_index=None):
    """Average True Range over `period` true ranges, evaluated at `at_index`
    (default: last candle). Returns None when there is insufficient data."""
    if not candles or len(candles) < period + 1:
        return None
    end = (len(candles) - 1) if at_index is None else at_index
    if end < period:
        return None
    trs = []
    for i in range(1, end + 1):
        h = candles[i]['high']
        l = candles[i]['low']
        pc = candles[i - 1]['close']
        trs.append(max(h - l, abs(h - pc), abs(l - pc)))
    window = trs[-period:]
    if not window:
        return None
    return sum(window) / len(window)


def _freshness_tier(triggered_mins, ltf):
    """fresh (<1x dur) / aging (1-3x dur) / stale (>3x dur), scaled to the
    LTF candle duration so freshness is not timeframe-blind."""
    dur = _LTF_DUR_MIN.get((ltf or '').lower(), 240)
    if triggered_mins is None:
        return 'stale'
    if triggered_mins < dur:
        return 'fresh'
    if triggered_mins <= 3 * dur:
        return 'aging'
    return 'stale'


def _rank_and_cap_setups(setups, cap=3):
    """Rank A setups by rr-to-T1 desc, then OTE depth (closer to 0.705
    equilibrium = better), then freshness (fresher first); cap PER TIER.
    Groups by tier, ranks within each group, caps each group at `cap`, and
    returns the combined list with swing setups first, then intraday.
    Pure + reusable so the Telegram digest can call it later."""
    def _key(s):
        rk = s.get('_rank') or {}
        rr = rk.get('rr_t1')
        rr = rr if isinstance(rr, (int, float)) else -1e9
        depth = rk.get('ote_depth')
        depth = depth if isinstance(depth, (int, float)) else 1e9
        fresh = rk.get('triggered_mins')
        fresh = fresh if isinstance(fresh, (int, float)) else 1e18
        return (-rr, depth, fresh)

    # Group by tier, preserving first-seen order for the grouping pass.
    groups = {}
    order = []
    for s in setups:
        tier = s.get('tier')
        if tier not in groups:
            groups[tier] = []
            order.append(tier)
        groups[tier].append(s)

    # Tier display order: swing first, then intraday, then any others
    # (stable sort keeps first-seen order among equal priorities).
    tier_priority = {'swing': 0, 'intraday': 1}
    order.sort(key=lambda t: tier_priority.get(t, 2))

    ranked = []
    for tier in order:
        ranked.extend(sorted(groups[tier], key=_key)[:cap])
    return ranked


def _setup_from_triage(ticker, triage):
    """Build a Setup contract object from a stored triage dict. Recomputes
    freshness from the CHoCH candle time so it stays current at read time."""
    ltf = triage.get('ltf')
    choch_time = triage.get('choch_time')
    tmins = None
    if choch_time:
        tmins = max(0.0, (time.time() - choch_time) / 60.0)
    rank = dict(triage.get('_rank') or {})
    rank['triggered_mins'] = tmins
    return {
        'id':            f"{ticker}-{triage.get('htf')}-{triage.get('ltf')}",
        'ticker':        ticker,
        'direction':     triage.get('direction'),
        'htf':           triage.get('htf'),
        'ltf':           triage.get('ltf'),
        'tier':          triage.get('tier'),
        'grade':         triage.get('grade', 'A'),
        'entryLow':      triage.get('entryLow'),
        'entryHigh':     triage.get('entryHigh'),
        'stop':          triage.get('stop'),
        'targets':       triage.get('targets') or [],
        'triggeredMins': (round(tmins, 1) if tmins is not None else None),
        'freshnessTier': _freshness_tier(tmins, ltf),
        'rationale':     triage.get('rationale'),
        'fvgFill':       triage.get('fvg_fill'),
        'htfCloseThrough': bool(triage.get('htf_close_through')),
        '_rank':         rank,
    }


def _watch_from_triage(ticker, triage):
    """Build a WatchItem contract object from a stored triage dict."""
    return {
        'ticker':     ticker,
        'state':      'waiting',
        'htf':        triage.get('htf'),
        'ltf':        triage.get('ltf'),
        'waitingFor': triage.get('waitingFor'),
        'eta':        triage.get('eta'),
        'fvgFill':    triage.get('fvg_fill'),
        'htfCloseThrough': bool(triage.get('htf_close_through')),
    }


def _public_setup(s):
    """Strip internal ranking keys before returning a Setup to a client."""
    return {k: v for k, v in s.items() if k != '_rank'}


# ── Per-TF candle params (module-level; shared by the scanner pipeline and
# the cascade snapshot builder). Wider lookbacks for slower TFs. Values are
# the pipeline's original locals, hoisted unchanged.
HTF_PARAMS = {
    '1w':  {'limit': 150, 'lookback': 100, 'left_bars': 2, 'right_bars': 2},
    '1d':  {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2},
    '12h': {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2},
    '4h':  {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2},
    '1h':  {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2},
}
LTF_PARAMS = {
    '1d':  {'limit': 300, 'lookback': 200, 'left_bars': 3,  'right_bars': 3},
    '4h':  {'limit': 200, 'lookback': 150, 'left_bars': 5,  'right_bars': 5},
    '1h':  {'limit': 200, 'lookback': 150, 'left_bars': 8,  'right_bars': 8},
    '15m': {'limit': 200, 'lookback': 150, 'left_bars': 10, 'right_bars': 10},
    '5m':  {'limit': 200, 'lookback': 150, 'left_bars': 10, 'right_bars': 10},
}


def _run_ict_pipeline(coin, htf, ltf, fetch_fn, diagnostics=False):
    """
    ICT Scanner v3 — locked SETUP_READY chain. Pure computation, no AI calls.

    DR + OTE are ALWAYS computed. The DR LEG DIRECTION sets the directional
    bias: high anchored before the low → down leg → bearish (premium OTE);
    low before high → up leg → bullish (discount OTE). Zone (discount/premium)
    is informational display only. There is no separate HTF-trend check.
    Then, in strict order:
      1. Valid POI in OTE — an order block inside the OTE matching bias that
         passes candle-dimension prominence AND left an UNmitigated displacement
         FVG. (No valid POI → DROPPED.)
      2. Sweep — the OB must have swept prior liquidity (a formation fact).
         (Absent → DROPPED.)
      3. LTF CHoCH — the trigger.
    Surviving outcomes:
      SETUP_READY (1+2+3) → ACT TODAY, grade 'A', with direction-aware levels.
      POI_WAITING (1+2)   → ON WATCH.
    Anything else is DROPPED (not surfaced). Never raises — returns DROPPED on
    any exception so the scan loop can continue.
    """
    _dropped = {
        'setup_state': 'DROPPED', 'direction': None,
        'htf': htf, 'ltf': ltf, 'current_price': None,
        'raw_indicators': {}, 'triage': None, 'signal_text': None,
        'drop_reason': None,
    }

    # ── Diagnostics worksheet (opt-in; ZERO effect on gates when off) ────────
    # Built alongside normal execution when diagnostics=True and attached to
    # the result as result['worksheet']. Gate order for the verdict index:
    _GATE_INDEX = {
        'insufficient HTF data':            0,
        'no dealing range':                 1,
        'no OB in OTE':                     2,
        'OB failed candle-dimension check': 3,
        'no displacement FVG':              4,
        'OB invalidated':                   5,
        'no liquidity sweep':               6,
    }
    ws = {} if diagnostics else None

    def _ret(res):
        # Identity when diagnostics is off — existing callers see no change.
        if not diagnostics:
            return res
        _state = res.get('setup_state')
        _reason = res.get('drop_reason') if _state == 'DROPPED' else None
        ws['verdict'] = {
            'state': _state,
            'drop_reason': _reason,
            'died_at_gate': _GATE_INDEX.get(_reason) if _reason else None,
        }
        res = dict(res)
        res['worksheet'] = ws
        return res

    try:
        # HTF/LTF candle params are module-level (HTF_PARAMS / LTF_PARAMS) —
        # shared with the cascade snapshot builder. Values unchanged.
        hp = HTF_PARAMS.get(htf, {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2})
        lp = LTF_PARAMS.get(ltf, {'limit': 200, 'lookback': 150, 'left_bars': 10, 'right_bars': 10})

        candles_htf = fetch_fn(coin, htf, limit=hp['limit'])
        candles_ltf = fetch_fn(coin, ltf, limit=lp['limit'])

        if diagnostics:
            ws['phase_structure'] = {
                'htf_interval': htf, 'ltf_interval': ltf,
                'candles_fetched': {'htf': len(candles_htf or []),
                                    'ltf': len(candles_ltf or [])},
            }

        if not candles_htf or len(candles_htf) < 10:
            return _ret({**_dropped, 'drop_reason': 'insufficient HTF data'})
        current_price = candles_htf[-1]['close']

        htf_sh, htf_sl = _ict_swing_points(
            candles_htf,
            lookback=hp['lookback'],
            left_bars=hp['left_bars'],
            right_bars=hp['right_bars'],
        )

        # ── DR + OTE (always) ────────────────────────────────────────────────
        dr = _ict_dealing_range(
            htf_sh, htf_sl, current_price, candles=candles_htf,
            pivot_min_prominence_atr=_scanner_settings().get(
                'dr_pivot_min_prominence_atr', 1.49))
        if not dr:
            return _ret({**_dropped, 'current_price': current_price,
                         'drop_reason': 'no dealing range'})

        # Bias = DR LEG DIRECTION (July spec — reverses the June location-bias):
        # DR high anchored BEFORE the low → the range is a down leg → bearish
        # (hunt bearish OBs in the premium OTE band); low anchored before the
        # high → up leg → bullish (discount band). Zone stays computed as
        # informational display only — it no longer sets bias.
        # Guard: missing or equal anchor timestamps → fall back to the old
        # zone-based bias; bias_source exposes the fallback in diagnose output.
        _h_t = dr.get('anchor_high_time')
        _l_t = dr.get('anchor_low_time')
        if _h_t is not None and _l_t is not None and _h_t != _l_t:
            leg_direction = 'down' if _h_t < _l_t else 'up'
            direction = 'bearish' if leg_direction == 'down' else 'bullish'
            bias_source = 'leg'
        else:
            leg_direction = None
            direction = 'bullish' if dr['zone'] == 'discount' else 'bearish'
            bias_source = 'zone_fallback'

        # Directional OTE band (61.8–78.6 retracement):
        #   bullish → discount band (high − 0.618·rng … high − 0.786·rng)
        #   bearish → premium band  (high − 0.382·rng … high − 0.214·rng)
        _rng = dr['high'] - dr['low']
        if direction == 'bullish':
            _a = dr['high'] - 0.618 * _rng
            _b = dr['high'] - 0.786 * _rng
        else:
            _a = dr['high'] - 0.382 * _rng
            _b = dr['high'] - 0.214 * _rng
        ote_low  = round(min(_a, _b), 6)
        ote_high = round(max(_a, _b), 6)

        if diagnostics:
            ws['phase_structure'].update({
                'dr_high': dr['high'], 'dr_low': dr['low'],
                'dr_anchor_high_time': dr.get('anchor_high_time'),
                'dr_anchor_low_time': dr.get('anchor_low_time'),
                'equilibrium': dr['eq'], 'zone': dr['zone'],
                'bias': direction,
                'leg_direction': leg_direction,
                'bias_source': bias_source,
            })
            ws['phase_ote'] = {
                'band_low': ote_low, 'band_high': ote_high,
                'current_price': current_price,
                'last_candle_time': candles_htf[-1].get('time'),
            }

        # ── Gate 1: valid POI = OB cluster in zone + displacement dims + FVG ─
        # EVALUATE-ALL: every in-zone candidate runs the full sequential chain
        # (dimension → FVG existence → invalidation → sweep); the pair passes
        # if ANY candidate survives all four. The first fully-surviving
        # candidate (newest-first) becomes the POI. If none survive, the drop
        # reason is the FURTHEST-ADVANCED candidate's failure. Gate order and
        # drop-reason strings unchanged.
        #
        # QUALIFICATION ZONE (July spec): an OB qualifies if it intersects
        # [0.618 retracement level → leg extreme] — wider than the OTE band,
        # which keeps its other jobs (ote_depth ranking vs 0.705, display):
        #   bearish → [band low (0.618 level), DR high]
        #   bullish → [DR low, band high (0.618 level)]
        if direction == 'bearish':
            zone_low, zone_high = ote_low, dr['high']
        else:
            zone_low, zone_high = dr['low'], ote_high
        if diagnostics:
            ws['phase_ote']['zone_low'] = zone_low
            ws['phase_ote']['zone_high'] = zone_high

        ob_candidates = _ict_order_block(candles_htf, direction, dr)
        obs_in_zone = _poi_in_ote(ob_candidates, zone_low, zone_high)

        # Invalidation is LTF-SOURCED (July spec): a candidate is invalidated
        # only by an LTF candle CLOSING through its far side within the
        # fetched LTF window (wicks never invalidate). HTF is bias + POI
        # notation only — the detector's HTF close-through is annotation.
        # Missing LTF data → NOT invalidated (never drop on missing data).
        def _ltf_invalidated(_o):
            if not candles_ltf:
                return False
            if direction == 'bearish':
                return any(_c['close'] > _o['top'] for _c in candles_ltf)
            return any(_c['close'] < _o['bottom'] for _c in candles_ltf)

        st = _scanner_settings()
        _GATE_SEQ = ('OB failed candle-dimension check', 'no displacement FVG',
                     'OB invalidated', 'no liquidity sweep')
        _PROG_LABEL = ('died: dims', 'died: FVG', 'died: invalidated',
                       'died: sweep', 'survived')

        def _disp_metrics(_o):
            # Candle-dimension metrics of the DISPLACEMENT candle — the first
            # candle after the cluster end (the candle that leaves the OB).
            # A cluster ending on the last candle has no displacement candle
            # yet → cannot pass dims.
            _dc_idx = _o['cluster_end_idx'] + 1
            if _dc_idx >= len(candles_htf):
                return 0.0, 0.0, False
            _dc = candles_htf[_dc_idx]
            _bd = abs(_dc['close'] - _dc['open'])
            _cr = _dc['high'] - _dc['low']
            _at = _ict_atr(candles_htf, 14, at_index=_dc_idx)
            _ba = (_bd / _at) if (_at and _at > 0) else 0.0
            _br = (_bd / _cr) if _cr > 0 else 0.0
            _ok = ((_at is not None and _at > 0 and _ba >= st['ob_body_atr_min'])
                   and (_cr > 0 and _br >= st['ob_body_range_ratio_min']))
            return _ba, _br, _ok

        evals = []      # per in-zone candidate: metrics + gate progress (0-4)
        poi = None
        for _ob in obs_in_zone:
            _ba, _br, _dim_ok = _disp_metrics(_ob)
            _ltf_inv = _ltf_invalidated(_ob)
            progress = 0
            if _dim_ok:
                progress = 1
                if _ob.get('displacement_fvg'):
                    progress = 2
                    if not _ltf_inv:
                        progress = 3
                        if _ob.get('swept'):
                            progress = 4    # survived every gate
            evals.append({'ob': _ob, 'body_atr': _ba, 'body_range': _br,
                          'dim_ok': _dim_ok, 'ltf_invalidated': _ltf_inv,
                          'progress': progress})
            print(
                f"[OB CHECK] {coin} {htf}: cluster "
                f"{_ob.get('cluster_start_time')}→{_ob.get('cluster_end_time')} "
                f"disp body={_ba:.2f}×ATR ratio={_br:.2f} — "
                f"{_PROG_LABEL[progress]}",
                flush=True
            )
            if progress == 4 and poi is None:
                poi = _ob   # first fully-surviving candidate = the pair's POI

        # Furthest-advanced candidate: the surviving POI when one exists, else
        # the first candidate at max progress (drop-reason attribution).
        _best_eval = None
        if evals:
            if poi is not None:
                _best_eval = next(_e for _e in evals if _e['ob'] is poi)
            else:
                _best_eval = max(evals, key=lambda _e: _e['progress'])

        if diagnostics:
            # Worksheet for EVERY candidate considered (cap: 10 nearest the OTE
            # band mid, plus the furthest-advanced candidate, always included)
            # — populated even when the pair dies at 'no OB in OTE'.
            _eval_by_id = {id(_e['ob']): _e for _e in evals}
            _best_ob = _best_eval['ob'] if _best_eval else None
            _band_mid = (ote_low + ote_high) / 2.0
            _cands = sorted(
                ob_candidates,
                key=lambda _o: abs(((_o['top'] + _o['bottom']) / 2.0) - _band_mid)
            )[:10]
            if _best_ob is not None and not any(_o is _best_ob for _o in _cands):
                _cands.append(_best_ob)
            _ob_rows = []
            for _o in _cands:
                _ev = _eval_by_id.get(id(_o))
                if _ev is not None:
                    _ba, _br, _dim = _ev['body_atr'], _ev['body_range'], _ev['dim_ok']
                    _prog = _PROG_LABEL[_ev['progress']]
                else:
                    # Not in OTE → never gate-evaluated; compute the same
                    # displacement-candle metrics for display only.
                    _ba, _br, _dim = _disp_metrics(_o)
                    _prog = None
                _ob_rows.append({
                    'time': _o.get('time'),
                    'cluster_start_time': _o.get('cluster_start_time'),
                    'cluster_end_time': _o.get('cluster_end_time'),
                    'top': _o['top'], 'bottom': _o['bottom'],
                    'body_atr_ratio': round(_ba, 3),
                    'body_range_ratio': round(_br, 3),
                    'dimension_pass': _dim,
                    'in_zone': _ev is not None,
                    'in_ote': _ev is not None,   # legacy alias for older UI
                    'gate_progress': _prog,
                    'furthest_advanced': _o is _best_ob,
                })
            ws['phase_ob'] = _ob_rows
            # phase_fvg keyed to the surviving / furthest-advanced candidate.
            _fvg_rows = []
            if _best_ob is not None:
                _dfvg = _best_ob.get('displacement_fvg')
                if _dfvg:
                    _dfvg = dict(_dfvg, formation_times=[
                        _dfvg.get('gap_start_time'), _dfvg.get('gap_end_time')])
                _best_ltf_inv = bool(_best_eval.get('ltf_invalidated')) \
                    if _best_eval and 'ltf_invalidated' in _best_eval \
                    else _ltf_invalidated(_best_ob)
                _fvg_rows.append({
                    'ob_time': _best_ob.get('time'),
                    'displacement_fvg': _dfvg,
                    'fvg_fill': _best_ob.get('fvg_fill'),
                    'ob_invalidated': _best_ltf_inv,          # LTF-sourced gate value
                    'invalidation_source': 'ltf',
                    'ltf_data_missing': not bool(candles_ltf),
                    'htf_close_through': bool(_best_ob.get('htf_close_through')),
                    'swept': bool(_best_ob.get('swept')),
                })
            ws['phase_fvg'] = _fvg_rows

            # ── phase_fvg_zone (cascade phase 1a preview — ADDITIVE) ─────────
            # Standalone in-zone FVG detection over the same leg span the OB
            # detector uses. Diagnostics-only: the live gate path never calls
            # _ict_fvg_in_zone yet. Leg-span indices come from the shared
            # _leg_span_indices helper (same one _ict_order_block uses).
            _fz_span = _leg_span_indices(candles_htf, dr.get('anchor_high_time'),
                                         dr.get('anchor_low_time'), direction)
            _fz_span_start, _fz_span_end = _fz_span if _fz_span else (None, None)
            _fz_atr = _ict_atr(candles_htf, 14)
            _fz_min_frac = st.get('fvg_min_atr_frac', 0.10)
            _fz = _ict_fvg_in_zone(
                candles_htf, direction, zone_low, zone_high,
                _fz_span_start, _fz_span_end, _fz_atr, _fz_min_frac)
            ws['phase_fvg_zone'] = {
                'zone_low': zone_low, 'zone_high': zone_high,
                'atr': (round(_fz_atr, 6) if _fz_atr else None),
                'min_atr_frac': _fz_min_frac,
                'counts': _fz['counts'],
                'candidates': _fz['candidates'],
            }

        if not obs_in_zone:
            return _ret({**_dropped, 'current_price': current_price, 'direction': direction,
                         'drop_reason': 'no OB in OTE'})

        if poi is None:
            # No candidate survived — attribute the furthest-advanced failure.
            return _ret({**_dropped, 'current_price': current_price, 'direction': direction,
                         'drop_reason': _GATE_SEQ[_best_eval['progress']]})

        # poi survived all of: dimension, FVG existence, invalidation, sweep.

        # ── Gate 3: LTF CHoCH (the trigger) ──────────────────────────────────
        ltf_sh, ltf_sl = [], []
        choch = {'fired': False, 'level': None, 'candle_index': None, 'close': None}
        if candles_ltf and len(candles_ltf) >= 10:
            ltf_sh, ltf_sl = _ict_swing_points(
                candles_ltf,
                lookback=lp['lookback'],
                left_bars=lp['left_bars'],
                right_bars=lp['right_bars'],
            )
            choch = _detect_choch(candles_ltf, ltf_sh, ltf_sl, direction)

        # Chart overlay — kept so the legacy /signals reader still resolves.
        raw_indicators = {
            'htf': {
                'structure': direction,
                'swing_highs': htf_sh,
                'swing_lows': htf_sl,
                'dr': dr,
                'fvg': None,
                'ob': poi,
                'poi_source': 'ob',
                'ote_low': ote_low,
                'ote_high': ote_high,
            },
            'ltf': {
                'structure': None,
                'swing_highs': ltf_sh,
                'swing_lows': ltf_sl,
                'dr': None,
                'choch': choch,
            },
        }

        zone_word = dr['zone']                                   # 'discount'|'premium'
        swept_label = 'SSL' if direction == 'bullish' else 'BSL'
        dir_word = 'LONG' if direction == 'bullish' else 'SHORT'

        # ── POI_WAITING — gates 1+2 hold, trigger pending ────────────────────
        if not choch.get('fired'):
            triage = {
                'kind': 'watch',
                'state': 'waiting',
                'direction': dir_word,
                'htf': htf, 'ltf': ltf,
                'waitingFor': f'LTF CHoCH on {ltf.upper()}',
                'eta': None,
                'fvg_fill': poi.get('fvg_fill'),
                'htf_close_through': bool(poi.get('htf_close_through')),
            }
            return _ret({
                'setup_state': 'POI_WAITING', 'direction': direction,
                'htf': htf, 'ltf': ltf, 'current_price': current_price,
                'raw_indicators': raw_indicators, 'triage': triage,
                'signal_text': f'{htf.upper()} {zone_word} OTE · OB+FVG · '
                               f'swept {swept_label} · awaiting CHoCH',
            })

        # ── SETUP_READY — direction-aware LTF trade levels (after CHoCH) ─────
        entryLow  = min(poi['bottom'], poi['top'])
        entryHigh = max(poi['bottom'], poi['top'])
        entry_px  = (entryLow + entryHigh) / 2.0

        ltf_atr = _ict_atr(candles_ltf, 14)
        buffer = (0.1 * ltf_atr) if (ltf_atr and ltf_atr > 0) else (0.001 * current_price)

        if direction == 'bullish':
            swing_lo = ltf_sl[-1]['price'] if ltf_sl else entryLow
            stop = round(swing_lo - buffer, 6)
            internal = [sh['price'] for sh in htf_sh if entryHigh < sh['price'] < dr['high']]
            t1 = min(internal) if internal else dr['eq']
            t2 = dr['high']
        else:
            swing_hi = ltf_sh[-1]['price'] if ltf_sh else entryHigh
            stop = round(swing_hi + buffer, 6)
            internal = [sl['price'] for sl in htf_sl if dr['low'] < sl['price'] < entryLow]
            t1 = max(internal) if internal else dr['eq']
            t2 = dr['low']

        def _rr(target):
            # Display-only; never gates. Direction-aware risk/reward.
            if direction == 'bullish':
                risk = entry_px - stop
                reward = target - entry_px
            else:
                risk = stop - entry_px
                reward = entry_px - target
            if risk <= 0:
                return None
            return round(reward / risk, 2)

        targets = [
            {'price': round(t1, 6), 'rr': _rr(t1)},
            {'price': round(t2, 6), 'rr': _rr(t2)},
        ]

        # Freshness from the confirming CHoCH candle, scaled to LTF duration.
        cidx = choch.get('candle_index')
        choch_time = None
        if cidx is not None and 0 <= cidx < len(candles_ltf):
            choch_time = candles_ltf[cidx].get('time')
        triggered_mins = None
        if choch_time:
            triggered_mins = max(0.0, (time.time() - choch_time) / 60.0)
        fresh_tier = _freshness_tier(triggered_mins, ltf)

        # OTE depth for ranking — closer to the 0.705 equilibrium = better.
        poi_mid = (entryLow + entryHigh) / 2.0
        ote_depth = abs(((dr['high'] - poi_mid) / _rng) - 0.705) if _rng else 1e9

        # ── Grade is binary today. The A→A+ upgrade hooks in HERE later:
        #    breaker detection is intentionally NOT built (see brief). Single
        #    extension point — set grade = 'A+' when a breaker confirms.
        grade = 'A'
        # if _breaker_confirms(...): grade = 'A+'

        rationale = (
            f"{htf.upper()} {zone_word} OTE · {ltf.upper()} OB+FVG · "
            f"swept {swept_label} · CHoCH ✓"
        )[:100]

        triage = {
            'kind': 'setup',
            'direction': dir_word,
            'htf': htf, 'ltf': ltf,
            'grade': grade,
            'entryLow': round(entryLow, 6), 'entryHigh': round(entryHigh, 6),
            'stop': stop, 'targets': targets,
            'triggeredMins': (round(triggered_mins, 1) if triggered_mins is not None else None),
            'freshnessTier': fresh_tier,
            'rationale': rationale,
            'choch_time': choch_time,
            'fvg_fill': poi.get('fvg_fill'),
            'htf_close_through': bool(poi.get('htf_close_through')),
            '_rank': {'rr_t1': targets[0]['rr'], 'ote_depth': ote_depth,
                      'triggered_mins': triggered_mins},
        }
        return _ret({
            'setup_state': 'SETUP_READY', 'direction': direction,
            'htf': htf, 'ltf': ltf, 'current_price': current_price,
            'raw_indicators': raw_indicators, 'triage': triage,
            'proposed_entry': entry_px, 'proposed_stop': stop,
            'proposed_target': targets[0]['price'], 'rr_ratio': targets[0]['rr'],
            'signal_text': rationale,
        })

    except Exception as _e:
        _status = ''
        # requests.exceptions.HTTPError carries .response with the status code
        resp = getattr(_e, 'response', None)
        if resp is not None and getattr(resp, 'status_code', None):
            _status = str(resp.status_code)
        _label = f"{type(_e).__name__}" + (f" {_status}" if _status else "")
        print(f'[PIPELINE] {coin} {htf}/{ltf} error: {_label}: {_e}', flush=True)
        return _ret({**_dropped,
                     'drop_reason': f'error: {type(_e).__name__}',
                     'error_status': _status,   # '' if unknown, else '429'/'500'/...
                     'is_error_drop': True})


# --- Concepts ---

@app.route('/api/trading/extract-concepts', methods=['POST'])
def api_trading_extract_concepts():
    try:
        import json as _json
        from src.engines.ai_advisor import load_ai_config
        from src.engines.llm_providers import get_provider
        from src.storage.portfolio_db import get_connection
        config = load_ai_config()
        provider = get_provider(config)
        conn = get_connection()

        # Ensure concepts_extracted_at column exists (idempotent)
        try:
            conn.execute("ALTER TABLE strategy_documents ADD COLUMN concepts_extracted_at TIMESTAMP DEFAULT NULL")
            conn.commit()
        except Exception:
            pass  # Column already exists

        data = request.json or {}
        force = data.get('force', False)

        # Fetch all docs, then split into to-process vs skipped
        all_docs = conn.execute(
            "SELECT id, filename, category, extracted_text, concepts_extracted_at "
            "FROM strategy_documents ORDER BY uploaded_at DESC"
        ).fetchall()
        if not all_docs:
            conn.close()
            return jsonify({"error": "No strategy documents uploaded yet"}), 400

        if force:
            docs_to_process = [d for d in all_docs if d['extracted_text']]
        else:
            docs_to_process = [d for d in all_docs if d['extracted_text'] and not d['concepts_extracted_at']]

        docs_skipped = len(all_docs) - len(docs_to_process)

        if not docs_to_process:
            conn.close()
            return jsonify({
                "concepts_created": 0, "docs_processed": 0,
                "docs_skipped": docs_skipped, "errors": [],
                "message": "All documents already processed. Use force=true to re-extract.",
            })

        total = 0
        errors = []
        for doc in docs_to_process:
            text = doc['extracted_text'][:8000]
            sys_prompt = "You are a DeFi and crypto trading expert. Extract key trading concepts from strategy documents as structured JSON."
            user_prompt = (
                f"Extract distinct trading concepts from this document.\n\n"
                f"For each concept return:\n"
                f"- title: short name (e.g. 'RSI Divergence Entry')\n"
                f"- summary: 2-3 sentence flashcard explanation\n"
                f"- category: one of: entry_signals, risk_management, position_sizing, "
                f"market_regimes, lp_strategy, defi_strategy, technical_analysis, macro_context, mindset\n"
                f"- tags: array of keyword strings\n\n"
                f"Return JSON: {{\"concepts\": [{{\"title\": \"...\", \"summary\": \"...\", "
                f"\"category\": \"...\", \"tags\": [...]}}]}}\n\n"
                f"Extract 5-15 concepts. Document ({doc['category']} — {doc['filename']}):\n\n{text}"
            )
            try:
                result = provider.complete(sys_prompt, user_prompt)
                for concept in result['response'].get('concepts', []):
                    if not concept.get('title') or not concept.get('summary'):
                        continue
                    conn.execute(
                        "INSERT INTO trading_concepts (title, summary, category, tags, source_doc, created_at, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
                        (concept['title'], concept.get('summary', ''), concept.get('category', 'general'),
                         _json.dumps(concept.get('tags', [])), doc['filename'])
                    )
                    total += 1
                conn.execute(
                    "UPDATE strategy_documents SET concepts_extracted_at=CURRENT_TIMESTAMP WHERE id=?",
                    (doc['id'],)
                )
            except Exception as e:
                errors.append(f"{doc['filename']}: {str(e)}")

        conn.commit()
        conn.close()
        return jsonify({
            "concepts_created": total,
            "docs_processed": len(docs_to_process),
            "docs_skipped": docs_skipped,
            "errors": errors,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/concepts')
def api_trading_concepts():
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        category = request.args.get('category', '')
        if category:
            rows = conn.execute(
                "SELECT * FROM trading_concepts WHERE category=? ORDER BY created_at DESC", (category,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM trading_concepts ORDER BY created_at DESC").fetchall()
        conn.close()
        concepts = [{
            'id': r['id'], 'title': r['title'], 'summary': r['summary'],
            'category': r['category'], 'tags': _json.loads(r['tags'] or '[]'),
            'source_doc': r['source_doc'], 'created_at': r['created_at'],
        } for r in rows]
        return jsonify({"concepts": concepts, "count": len(concepts)})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/concepts/<int:concept_id>')
def api_trading_concept_get(concept_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        r = conn.execute("SELECT * FROM trading_concepts WHERE id=?", (concept_id,)).fetchone()
        conn.close()
        if not r:
            return jsonify({"error": "Not found"}), 404
        return jsonify({
            'id': r['id'], 'title': r['title'], 'summary': r['summary'],
            'full_text': r['full_text'], 'category': r['category'],
            'tags': _json.loads(r['tags'] or '[]'), 'source_doc': r['source_doc'],
            'confidence': r['confidence'], 'created_at': r['created_at'], 'updated_at': r['updated_at'],
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/concepts/<int:concept_id>', methods=['PUT'])
def api_trading_concept_update(concept_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE trading_concepts SET "
            "title=COALESCE(?,title), summary=COALESCE(?,summary), full_text=COALESCE(?,full_text), "
            "category=COALESCE(?,category), tags=COALESCE(?,tags), updated_at=CURRENT_TIMESTAMP "
            "WHERE id=?",
            (data.get('title'), data.get('summary'), data.get('full_text'), data.get('category'),
             _json.dumps(data['tags']) if 'tags' in data else None, concept_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/concepts/<int:concept_id>', methods=['DELETE'])
def api_trading_concept_delete(concept_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM trading_concepts WHERE id=?", (concept_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- Quiz ---

@app.route('/api/trading/quiz/generate', methods=['POST'])
def api_trading_quiz_generate():
    try:
        import json as _json
        from datetime import date
        from src.engines.ai_advisor import load_ai_config
        from src.engines.llm_providers import get_provider
        from src.storage.portfolio_db import get_connection
        config = load_ai_config()
        provider = get_provider(config)
        conn = get_connection()
        today = date.today().isoformat()
        data = request.json or {}
        force = data.get('force', False)
        existing = conn.execute("SELECT id FROM daily_quizzes WHERE quiz_date=?", (today,)).fetchone()
        if existing and not force:
            conn.close()
            return jsonify({"error": "Quiz already generated today. Use force=true to regenerate."}), 409
        concepts = conn.execute(
            "SELECT id, title, summary FROM trading_concepts ORDER BY RANDOM() LIMIT 10"
        ).fetchall()
        if not concepts:
            conn.close()
            return jsonify({"error": "No trading concepts found. Run extract-concepts first."}), 400
        question_ids = []
        errors = []
        for concept in concepts:
            sys_prompt = "You are a DeFi trading educator. Generate quiz questions as valid JSON."
            user_prompt = (
                f"Generate a quiz question for this trading concept.\n\n"
                f"Concept: {concept['title']}\nExplanation: {concept['summary']}\n\n"
                f"Return JSON:\n{{\n"
                f"  \"question_text\": \"A clear question testing understanding\",\n"
                f"  \"answer_text\": \"Correct answer in 2-4 sentences\",\n"
                f"  \"distractors\": [\"wrong answer 1\", \"wrong answer 2\", \"wrong answer 3\"]\n}}"
            )
            try:
                result = provider.complete(sys_prompt, user_prompt)
                q = result['response']
                if not q.get('question_text') or not q.get('answer_text'):
                    continue
                cur = conn.execute(
                    "INSERT INTO quiz_questions (concept_id, question_text, answer_text, distractors_json, created_at) "
                    "VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
                    (concept['id'], q['question_text'], q['answer_text'], _json.dumps(q.get('distractors', [])))
                )
                question_ids.append(cur.lastrowid)
            except Exception as e:
                errors.append(f"Concept {concept['id']}: {str(e)}")
        if not question_ids:
            conn.close()
            return jsonify({"error": "Failed to generate any questions", "errors": errors}), 500
        if existing and force:
            conn.execute("DELETE FROM daily_quizzes WHERE quiz_date=?", (today,))
        conn.execute(
            "INSERT INTO daily_quizzes (quiz_date, questions_json, generated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
            (today, _json.dumps(question_ids))
        )
        conn.commit()
        conn.close()
        return jsonify({"quiz_date": today, "questions_generated": len(question_ids), "errors": errors})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/quiz/today')
def api_trading_quiz_today():
    import json as _json
    from datetime import date
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        today = date.today().isoformat()
        quiz = conn.execute("SELECT questions_json FROM daily_quizzes WHERE quiz_date=?", (today,)).fetchone()
        if not quiz:
            conn.close()
            return jsonify({"error": "No quiz generated for today", "quiz_date": today}), 404
        question_ids = _json.loads(quiz['questions_json'])
        questions = []
        for qid in question_ids:
            q = conn.execute(
                "SELECT qq.id, qq.question_text, qq.distractors_json, tc.title as concept_title "
                "FROM quiz_questions qq LEFT JOIN trading_concepts tc ON tc.id=qq.concept_id WHERE qq.id=?",
                (qid,)
            ).fetchone()
            if not q:
                continue
            questions.append({
                'id': q['id'], 'question_text': q['question_text'],
                'concept_title': q['concept_title'],
                'distractors': _json.loads(q['distractors_json'] or '[]'),
            })
        conn.close()
        return jsonify({"quiz_date": today, "questions": questions})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/quiz/today/answers')
def api_trading_quiz_today_answers():
    import json as _json
    from datetime import date
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        today = date.today().isoformat()
        quiz = conn.execute("SELECT questions_json FROM daily_quizzes WHERE quiz_date=?", (today,)).fetchone()
        if not quiz:
            conn.close()
            return jsonify({"error": "No quiz for today"}), 404
        question_ids = _json.loads(quiz['questions_json'])
        attempts = {r['question_id']: r for r in conn.execute(
            "SELECT * FROM quiz_attempts WHERE quiz_date=?", (today,)
        ).fetchall()}
        questions = []
        for qid in question_ids:
            q = conn.execute(
                "SELECT qq.*, tc.title as concept_title, tc.summary as concept_summary "
                "FROM quiz_questions qq LEFT JOIN trading_concepts tc ON tc.id=qq.concept_id WHERE qq.id=?",
                (qid,)
            ).fetchone()
            if not q:
                continue
            attempt = attempts.get(qid)
            questions.append({
                'id': q['id'], 'question_text': q['question_text'],
                'answer_text': q['answer_text'],
                'distractors': _json.loads(q['distractors_json'] or '[]'),
                'concept_title': q['concept_title'], 'concept_summary': q['concept_summary'],
                'user_answer': attempt['user_answer'] if attempt else None,
                'is_correct': bool(attempt['is_correct']) if attempt else None,
            })
        streak = conn.execute("SELECT * FROM concept_streak ORDER BY id LIMIT 1").fetchone()
        conn.close()
        return jsonify({
            "quiz_date": today, "questions": questions,
            "total_attempted": len(attempts),
            "total_correct": sum(1 for a in attempts.values() if a['is_correct']),
            "streak": dict(streak) if streak else {"current_streak": 0, "longest_streak": 0},
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/quiz/submit', methods=['POST'])
def api_trading_quiz_submit():
    import json as _json
    from datetime import date, timedelta
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    answers = data.get('answers', [])
    conn = get_connection()
    try:
        today = date.today().isoformat()
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        correct_count = 0
        total_count = 0
        for ans in answers:
            qid = ans.get('question_id')
            user_answer = (ans.get('user_answer') or '').strip()
            if not qid:
                continue
            q = conn.execute("SELECT answer_text FROM quiz_questions WHERE id=?", (qid,)).fetchone()
            if not q:
                continue
            is_correct = int(user_answer.lower() == q['answer_text'].strip().lower())
            conn.execute(
                "INSERT OR IGNORE INTO quiz_attempts (quiz_date, question_id, user_answer, is_correct, answered_at) "
                "VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
                (today, qid, user_answer, is_correct)
            )
            conn.execute(
                "UPDATE quiz_questions SET times_asked=times_asked+1, times_correct=times_correct+? WHERE id=?",
                (is_correct, qid)
            )
            if is_correct:
                correct_count += 1
            total_count += 1
        # Update streak
        streak = conn.execute("SELECT * FROM concept_streak ORDER BY id LIMIT 1").fetchone()
        if streak:
            new_streak = streak['current_streak']
            if streak['last_quiz_date'] == yesterday:
                new_streak += 1
            elif streak['last_quiz_date'] != today:
                new_streak = 1
            longest = max(streak['longest_streak'], new_streak)
            conn.execute(
                "UPDATE concept_streak SET current_streak=?, longest_streak=?, last_quiz_date=?, "
                "total_correct=total_correct+?, total_attempted=total_attempted+?, updated_at=CURRENT_TIMESTAMP "
                "WHERE id=?",
                (new_streak, longest, today, correct_count, total_count, streak['id'])
            )
        else:
            conn.execute(
                "INSERT INTO concept_streak (current_streak, longest_streak, last_quiz_date, "
                "total_correct, total_attempted, updated_at) VALUES (1, 1, ?, ?, ?, CURRENT_TIMESTAMP)",
                (today, correct_count, total_count)
            )
        conn.commit()
        conn.close()
        return jsonify({
            "submitted": total_count, "correct": correct_count,
            "score_pct": round(correct_count / total_count * 100, 1) if total_count else 0,
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


# --- Scanner ---

@app.route('/api/trading/scanner/watchlist')
def api_trading_scanner_watchlist():
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute("SELECT * FROM scanner_watchlist ORDER BY created_at DESC").fetchall()
        conn.close()
        return jsonify({"watchlist": [dict(r) for r in rows]})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/scanner/watchlist', methods=['POST'])
def api_trading_scanner_watchlist_add():
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    # Shared canonical form — identical to the HL-import path, so the same asset
    # can never enter the watchlist twice under two spellings (BTC vs BTCUSDT).
    symbol = _normalize_symbol(data.get('symbol') or '')
    if not symbol:
        return jsonify({"error": "symbol required"}), 400
    htf = data.get('htf_timeframe', data.get('interval', '4h')) or '4h'
    ltf = data.get('ltf_timeframe', '15m') or '15m'
    conn = get_connection()
    try:
        # symbol-only existence check: UNIQUE(symbol) means htf/ltf no longer
        # distinguish rows, and a dup symbol would otherwise raise IntegrityError.
        existing = conn.execute(
            "SELECT id FROM scanner_watchlist WHERE symbol=?",
            (symbol,)
        ).fetchone()
        if existing:
            conn.close()
            return jsonify({"error": "Already in watchlist", "symbol": symbol}), 409
        contract_address = (data.get('contract_address') or '').strip()
        # Manual adds can't be auto-classified — honor an explicit asset_type
        # from the body, else default to 'crypto'.
        asset_type = (data.get('asset_type') or 'crypto')
        # OR IGNORE is a backstop: the check above handles the normal case, this
        # keeps a race from violating UNIQUE(symbol) and 500-ing.
        cur = conn.execute(
            "INSERT OR IGNORE INTO scanner_watchlist (symbol, exchange, asset_type, htf_timeframe, ltf_timeframe, notes, contract_address) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (symbol, data.get('exchange', 'binance'), asset_type, htf, ltf, data.get('notes', ''), contract_address)
        )
        conn.commit()
        conn.close()
        if cur.rowcount == 0:   # lost a race — the row already exists
            return jsonify({"error": "Already in watchlist", "symbol": symbol}), 409
        return jsonify({"success": True, "symbol": symbol})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/scanner/watchlist/<int:item_id>', methods=['DELETE'])
def api_trading_scanner_watchlist_delete(item_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("DELETE FROM scanner_watchlist WHERE id=?", (item_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/scanner/watchlist/all', methods=['DELETE'])
def api_trading_scanner_watchlist_clear():
    from src.storage.portfolio_db import get_connection
    try:
        conn = get_connection()
        wl_count = conn.execute("SELECT COUNT(*) FROM scanner_watchlist").fetchone()[0]
        sig_count = conn.execute("SELECT COUNT(*) FROM scanner_signals").fetchone()[0]
        conn.execute("DELETE FROM scanner_watchlist")
        conn.execute("DELETE FROM scanner_signals")
        conn.commit()
        conn.close()
        return jsonify({'deleted': wl_count + sig_count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/watchlist/<int:item_id>', methods=['PUT'])
def api_trading_scanner_watchlist_update(item_id):
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        conn.execute("UPDATE scanner_watchlist SET notes=? WHERE id=?", (data.get('notes', ''), item_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/scanner/ohlcv')
def api_trading_scanner_ohlcv():
    import requests as _req
    import time as _time_mod
    symbol = (request.args.get('symbol') or '').strip().upper()
    interval = (request.args.get('interval') or '').strip()
    limit = int(request.args.get('limit', 100))
    if not symbol:
        return jsonify({'error': 'symbol required', 'candles': []}), 400

    coin = symbol
    for _sfx in ('USDT', 'USDC', 'PERP'):
        if coin.endswith(_sfx) and len(coin) > len(_sfx):
            coin = coin[:-len(_sfx)]
            break

    _interval_map = {
        '1w': '1w', '1d': '1d', '12h': '12h', '4h': '4h',
        '1h': '1h', '30m': '30m', '15m': '15m', '5m': '5m',
        '3m': '3m', '1m': '1m',
    }
    hl_interval = _interval_map.get(interval, interval or '4h')

    _ms_map = {
        '1w': 7*24*3600*1000, '1d': 24*3600*1000, '12h': 12*3600*1000,
        '4h': 4*3600*1000, '1h': 3600*1000, '30m': 30*60*1000,
        '15m': 15*60*1000, '5m': 5*60*1000, '3m': 3*60*1000, '1m': 60*1000,
    }
    ms_per_candle = _ms_map.get(hl_interval, 4*3600*1000)
    end_time = int(_time_mod.time() * 1000)
    start_time = end_time - (ms_per_candle * limit)

    try:
        raw = _hl_post({
            'type': 'candleSnapshot',
            'req': {
                'coin': _hl_resolve_coin(coin),
                'interval': hl_interval,
                'startTime': start_time,
                'endTime': end_time,
            },
        })
        candles = [
            {
                'time':   int(c['t']) // 1000,
                'open':   float(c['o']),
                'high':   float(c['h']),
                'low':    float(c['l']),
                'close':  float(c['c']),
                'volume': float(c['v']),
            }
            for c in raw
        ]
        return jsonify({'candles': candles, 'symbol': coin, 'interval': hl_interval})
    except Exception as e:
        print(f"[OHLCV] Error fetching {coin} {hl_interval}: {e}", flush=True)
        return jsonify({'error': str(e), 'candles': []}), 500


@app.route('/api/trading/scanner/signals')
def api_trading_scanner_signals():
    from src.storage.portfolio_db import get_connection
    import json as _json
    try:
        conn = get_connection()
        # NOTE: DB column is detected_at; it is exposed to the frontend as scanned_at.
        rows = conn.execute(
            "SELECT * FROM scanner_signals ORDER BY detected_at DESC"
        ).fetchall()
        conn.close()

        def _safe(v):
            if v is None:
                return v
            try:
                return _json.loads(v)
            except Exception:
                return v

        # Group by symbol, keeping latest scan per symbol+pair_key (rows are detected_at DESC)
        grouped = {}
        for r in rows:
            sym = r['symbol']
            pk = r['pair_key'] or 'swing'
            if sym not in grouped:
                grouped[sym] = {'symbol': sym, 'pairs': {}}
            if pk not in grouped[sym]['pairs']:
                # First occurrence is latest (ORDER BY detected_at DESC)
                _raw = _safe(r['raw_indicators_json']) or {}
                _htf_raw = _raw.get('htf') or {}
                _ltf_raw = _raw.get('ltf') or {}
                _choch   = _ltf_raw.get('choch') or {}
                grouped[sym]['pairs'][pk] = {
                    'pair_key':         pk,
                    'htf_timeframe':    r['htf_timeframe'],
                    'ltf_timeframe':    r['ltf_timeframe'],
                    'signal_text':      r['signal_text'] or '',
                    'status':           r['status'] or 'quiet',
                    'setup_state':      r['signal_type'] or '',
                    'confidence_score': r['confidence_score'] or 0,
                    'why_flagged':      r['why_flagged'] or '',
                    'htf_label':        r['htf_label'] or '',
                    'ltf_label':        r['ltf_label'] or '',
                    'current_price':    r['current_price'],
                    'scanned_at':       r['detected_at'],
                    'raw_indicators_json': _raw,
                    'proposed_entry':   r['proposed_entry'],
                    'proposed_stop':    r['proposed_stop'],
                    'proposed_target':  r['proposed_target'],
                    'rr_ratio':         r['rr_ratio'],
                    # v3 structured fields extracted from raw_indicators_json
                    'ote_low':          _htf_raw.get('ote_low'),
                    'ote_high':         _htf_raw.get('ote_high'),
                    'poi_type':         (_htf_raw.get('ob') or {}).get('type'),
                    'poi_tier':         (_htf_raw.get('ob') or {}).get('tier'),
                    'poi_top':          (_htf_raw.get('ob') or {}).get('top'),
                    'poi_bottom':       (_htf_raw.get('ob') or {}).get('bottom'),
                    'choch_fired':      bool(_choch.get('fired')),
                    'choch_level':      _choch.get('level'),
                    'trend_direction':  _htf_raw.get('structure'),
                }
                # Track the latest scan time + price at the symbol level
                if 'scanned_at' not in grouped[sym] or (r['detected_at'] or '') > (grouped[sym].get('scanned_at') or ''):
                    grouped[sym]['scanned_at'] = r['detected_at']
                    grouped[sym]['current_price'] = r['current_price']

        return jsonify({'signals': list(grouped.values())})
    except Exception as e:
        return jsonify({'error': str(e), 'signals': []}), 500


def _hl_fetch_top_volume(n=20, min_volume=0, limit=None):
    """Fetch top-N HL assets by 24h notional volume — perps + spot, with asset_type tagging."""
    import requests as _req
    assets = []

    # Step 1: Discover all perp dex names (first entry is None for native dex)
    try:
        dex_list = _hl_post({'type': 'perpDexs'})  # [null, {name: "xyz", ...}, ...]
    except Exception:
        dex_list = [None]  # fallback to native only

    # Step 2: Fetch each dex — native (empty string) first, then HIP-3 dexes
    dex_names = []
    for entry in dex_list:
        if entry is None:
            dex_names.append(('', False))   # native crypto perps
        elif isinstance(entry, dict) and entry.get('name'):
            dex_names.append((entry['name'], True))  # HIP-3 TradFi

    for dex_name, is_hip3 in dex_names:
        try:
            payload = {'type': 'metaAndAssetCtxs'}
            if dex_name:
                payload['dex'] = dex_name
            meta, asset_ctxs = _hl_post(payload)
            universe = meta.get('universe', [])
            for i, asset in enumerate(universe):
                if i >= len(asset_ctxs):
                    break
                raw_name = asset.get('name', '')
                vol = float(asset_ctxs[i].get('dayNtlVlm', 0) or 0)
                if vol < min_volume:
                    continue
                # Live price sits in the same ctx object as the volume.
                px = float(asset_ctxs[i].get('markPx')
                           or asset_ctxs[i].get('midPx')
                           or asset_ctxs[i].get('oraclePx')
                           or 0) or None
                # HIP-3 names are "dex:COIN" — strip the prefix
                name = raw_name.split(':', 1)[1] if ':' in raw_name else raw_name
                asset_type = 'tradfi' if is_hip3 else 'crypto'
                symbol = name + '-USDT' if not name.upper().endswith(('-USDT', '-USDC', 'USDT', 'USDC')) else name
                assets.append({'name': name, 'symbol': symbol, 'volume_24h': vol, 'price': px, 'asset_type': asset_type})
        except Exception:
            continue  # skip failed dex, don't abort

    # Step 3: Spot / TradFi (USDC-quoted pairs)
    try:
        spot_meta, spot_ctxs = _hl_post({'type': 'spotMetaAndAssetCtxs'})
        spot_tokens = {t['index']: t for t in spot_meta.get('tokens', [])}
        for i, pair in enumerate(spot_meta.get('universe', [])):
            if i >= len(spot_ctxs):
                break
            token_indices = pair.get('tokens', [])
            if len(token_indices) < 2:
                continue
            base_token = spot_tokens.get(token_indices[0], {})
            quote_token = spot_tokens.get(token_indices[1], {})
            if quote_token.get('name', '').upper() != 'USDC':
                continue
            base_name = base_token.get('name', '')
            if not base_name:
                continue
            vol = float(spot_ctxs[i].get('dayNtlVlm', 0) or 0)
            if vol < min_volume:
                continue
            # Live price sits in the same ctx object as the volume.
            px = float(spot_ctxs[i].get('markPx')
                       or spot_ctxs[i].get('midPx')
                       or spot_ctxs[i].get('oraclePx')
                       or 0) or None
            asset_type = 'crypto' if pair.get('isCanonical', False) else 'tradfi'
            symbol = f'{base_name}-USDC'
            assets.append({'name': base_name, 'symbol': symbol, 'volume_24h': vol, 'price': px, 'asset_type': asset_type})
    except Exception:
        pass

    # Deduplicate by symbol — keep highest volume entry per symbol
    seen = {}
    for a in assets:
        key = a['symbol'].upper()
        if key not in seen or a['volume_24h'] > seen[key]['volume_24h']:
            seen[key] = a
    deduped = sorted(seen.values(), key=lambda x: x['volume_24h'], reverse=True)
    return deduped if limit is None else deduped[:limit]


def _fmt_vol(v):
    if v >= 1e9:   return f"${v/1e9:.1f}B"
    if v >= 1e6:   return f"${v/1e6:.0f}M"
    if v >= 1e3:   return f"${v/1e3:.0f}K"
    return f"${v:.0f}"


@app.route('/api/trading/scanner/hl-top-volume')
def api_trading_scanner_hl_top_volume():
    from src.storage.portfolio_db import get_connection
    try:
        n = min(int(request.args.get('n', 20)), 200)
        min_volume = float(request.args.get('min_volume', 0) or 0)

        assets = _hl_fetch_top_volume(n=None, min_volume=min_volume, limit=None)

        conn = get_connection()
        # Canonical forms so the in_watchlist flag matches what import would store.
        existing = {_normalize_symbol(r['symbol'])
                    for r in conn.execute("SELECT symbol FROM scanner_watchlist").fetchall()}
        conn.close()

        type_filter = request.args.get('asset_type', 'all')  # 'all' | 'crypto' | 'tradfi'
        result = []
        for a in assets:
            if type_filter != 'all' and a.get('asset_type', 'crypto') != type_filter:
                continue
            result.append({
                'symbol': a['symbol'],
                'name': a['name'],
                'volume_24h': a['volume_24h'],
                'volume_display': _fmt_vol(a['volume_24h']),
                'asset_type': a.get('asset_type', 'crypto'),
                'in_watchlist': _normalize_symbol(a['symbol']) in existing,
            })
        # Slice to N AFTER the type filter so a type request yields up to N of THAT
        # type (was previously sliced across the combined list before filtering).
        result = result[:n]
        return jsonify({'assets': result, 'total_found': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/hl-volumes')
def api_trading_scanner_hl_volumes():
    """Return a symbol->{volume_24h, asset_type} map for all HL assets."""
    try:
        assets = _hl_fetch_top_volume(n=99999, min_volume=0, limit=99999)
        vol_map = {}
        for a in assets:
            # Store under both dash and no-dash keys so frontend lookup always hits
            sym = a['symbol'].upper()
            entry = {'volume_24h': a['volume_24h'], 'price': a.get('price'), 'asset_type': a['asset_type']}
            vol_map[sym] = entry
            vol_map[sym.replace('-', '')] = entry
        return jsonify({'volumes': vol_map})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/debug-hl-dexs')
def api_debug_hl_dexs():
    try:
        return jsonify({'raw': _hl_post({'type': 'perpDexs'})})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/hl-import', methods=['POST'])
def api_trading_scanner_hl_import():
    from src.storage.portfolio_db import get_connection
    try:
        data = request.json or {}
        n = min(int(data.get('n', 20)), 200)
        min_volume = float(data.get('min_volume', 0) or 0)
        asset_type_filter = data.get('asset_type', 'all')  # 'all' | 'crypto' | 'tradfi'

        assets = _hl_fetch_top_volume(n=None, min_volume=min_volume, limit=None)
        if asset_type_filter != 'all':
            assets = [a for a in assets if a.get('asset_type', 'crypto') == asset_type_filter]
        # Slice to N AFTER the type filter so an import yields up to N of THAT type
        # (was previously sliced across the combined list before filtering, which
        # made crypto counts non-deterministic run-to-run). Mirrors the preview route.
        assets = assets[:n]

        symbols_filter = set(data.get('symbols', None) or [])
        if symbols_filter:
            assets = [a for a in assets if a['symbol'] in symbols_filter]

        conn = get_connection()
        existing_rows = conn.execute("SELECT id, symbol FROM scanner_watchlist").fetchall()
        existing = {r['symbol'].upper(): r['id'] for r in existing_rows}

        # Remove quiet entries: delete watchlist rows whose most recent signal is 'quiet'
        quiet_ids = []
        for r in existing_rows:
            sig = conn.execute(
                "SELECT status FROM scanner_signals WHERE symbol=? ORDER BY detected_at DESC LIMIT 1",
                (r['symbol'],)
            ).fetchone()
            if sig and sig['status'] == 'quiet':
                quiet_ids.append(r['id'])
        removed = 0
        for qid in quiet_ids:
            conn.execute("DELETE FROM scanner_watchlist WHERE id=?", (qid,))
            removed += 1

        # Refresh existing after removals. Compare in CANONICAL form on both sides
        # so a stored "BTCUSDT" correctly matches an incoming "BTC-USDT".
        existing_now = {_normalize_symbol(r['symbol'])
                        for r in conn.execute("SELECT symbol FROM scanner_watchlist").fetchall()}

        added = 0
        skipped = 0
        for a in assets:
            sym_canon = _normalize_symbol(a['symbol'])   # already uppercased/canonical
            if sym_canon in existing_now:
                skipped += 1
                continue
            # OR IGNORE backstop against UNIQUE(symbol); rowcount tells us if it landed.
            cur = conn.execute(
                "INSERT OR IGNORE INTO scanner_watchlist (symbol, asset_type, htf_timeframe, ltf_timeframe) VALUES (?,?,?,?)",
                (sym_canon, a.get('asset_type') or 'crypto', '4h', '1h')
            )
            if cur.rowcount:
                existing_now.add(sym_canon)   # guard against dup incoming rows in the same batch
                added += 1
            else:
                skipped += 1

        conn.commit()
        conn.close()
        return jsonify({'added': added, 'skipped': skipped, 'removed': removed})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Fixed HTF→LTF pairs for v3, tagged by tier ───────────────────────
# swing pairs are always scanned; intra/scalp only when the request asks.
# Module-level so the scan engine and the diagnose route share one source.
V3_PAIRS = [
    ('1W',  '1w',  '4h',  'swing'),
    ('1D',  '1d',  '1h',  'swing'),
    ('12H', '12h', '1h',  'swing'),
    ('4H',  '4h',  '15m', 'intraday'),
]


def _hl_fetch_candles(coin, interval, limit=200):
    """Fetch HL candles for the scanner pipeline. Extracted verbatim from the
    closure formerly nested in _run_scanner_scan so the diagnose route shares
    the exact live fetch path (incl. the Monday weekly realignment)."""
    _hl_ms = {
        '1w': 7*24*3600*1000, '1d': 24*3600*1000, '12h': 12*3600*1000,
        '4h': 4*3600*1000,    '1h': 3600*1000,    '30m': 30*60*1000,
        '15m': 15*60*1000,    '5m': 5*60*1000,
    }

    def _fetch_raw(_interval, _limit):
        ms = _hl_ms.get(_interval, 4*3600*1000)
        end_time   = int(time.time() * 1000)
        start_time = end_time - (ms * _limit)
        # _hl_post handles the global rate gate + retry-on-429/5xx backoff.
        raw = _hl_post({'type': 'candleSnapshot', 'req': {
            'coin': _hl_resolve_coin(coin), 'interval': _interval,
            'startTime': start_time, 'endTime': end_time,
        }})
        return [{'open': float(c['o']), 'high': float(c['h']),
                 'low':  float(c['l']), 'close': float(c['c']),
                 'volume': float(c['v']), 'time': int(c['t']) // 1000} for c in raw]

    if interval != '1w':
        return _fetch_raw(interval, limit)

    # 1w: build Monday-anchored weeklies from a dailies fetch (see the
    # _weekly_from_dailies transform). limit weeks → limit*7 dailies.
    dailies = _fetch_raw('1d', min(limit * 7, 1000))
    return _weekly_from_dailies(dailies, limit)


def _weekly_from_dailies(dailies, limit):
    """Monday-anchored weekly aggregation (dailies → realigned weeklies).

    Extracted verbatim from _hl_fetch_candles' 1w branch so the cascade
    snapshot cache can derive weeklies from an already-fetched daily array.
    HL's native '1w' candles are Thursday-anchored, which mismatches standard
    crypto charting (weeks open Monday 00:00 UTC)."""
    if not dailies:
        return []
    weeks = {}   # monday_ts -> candle dict
    for d in dailies:
        # Days since epoch; epoch (1970-01-01) was a Thursday.
        # (days + 3) % 7 maps Monday→0 ... Sunday→6.
        days = d['time'] // 86400
        dow = (days + 3) % 7          # 0=Monday ... 6=Sunday
        monday_ts = (days - dow) * 86400
        w = weeks.get(monday_ts)
        if w is None:
            weeks[monday_ts] = {
                'time': monday_ts,
                'open': d['open'], 'high': d['high'],
                'low': d['low'],   'close': d['close'],
                'volume': d['volume'],
            }
        else:
            w['high']   = max(w['high'], d['high'])
            w['low']    = min(w['low'],  d['low'])
            w['close']  = d['close']
            w['volume'] += d['volume']
    out = [weeks[k] for k in sorted(weeks.keys())]
    return out[-limit:]


# ── Cascade phase 2: per-TF snapshot builder + per-run compute-once cache ──
# ADDITIVE ONLY: nothing on the manual/scheduled scan path constructs or calls
# SnapshotRun/_compute_tf_snapshot. Stage/nesting/persistence = Phase 3.

def _compute_tf_snapshot(coin, interval, candles):
    """Single-TF-pure snapshot for the cascade scanner (Phase 2).

    Computes, for ONE timeframe in isolation (no cross-TF logic — LTF
    invalidation etc. is a composer concern, Phase 3):
      structure  — DR (via _ict_dealing_range, unmodified) + leg-direction
                   bias from anchor times (existing convention: high anchor
                   first = down leg = bearish; mirror = bullish; missing/equal
                   anchor times → zone fallback, flagged via bias_source)
      levels     — OTE band + qualification zone (exact pipeline formulas),
                   equilibrium and fib levels as returned by the DR
      obs        — OB cluster candidates via _ict_order_block (unmodified),
                   every per-candidate annotation preserved, plus dims
                   metrics / in_zone computed per candidate. NO cross-TF
                   annotations here.
      fvgs       — standalone FVGs via _ict_fvg_in_zone with this TF's zone,
                   leg span (shared _leg_span_indices), ATR(14) and the
                   fvg_min_atr_frac setting
      swings     — _ict_swing_points at the TF's HTF_PARAMS (2/2 bars).
                   NOTE: MSS-specific pivot params are deliberately deferred
                   to Phase 4 hand-charting; no MSS logic here.
      meta       — current price, last candle time, bar count, interval,
                   computed_at

    Plain-dict, JSON-serializable; all indices FULL-ARRAY basis with the
    candles[idx]['time'] invariant."""
    snap = {
        'coin': coin, 'interval': interval,
        'bar_count': len(candles or []),
        'computed_at': int(time.time()),
        'error': None,
    }
    if not candles or len(candles) < 10:
        snap['error'] = 'insufficient data'
        return snap

    hp = HTF_PARAMS.get(interval,
                        {'limit': 300, 'lookback': 200, 'left_bars': 2, 'right_bars': 2})
    sh, sl = _ict_swing_points(candles, lookback=hp['lookback'],
                               left_bars=hp['left_bars'], right_bars=hp['right_bars'])
    current_price = candles[-1]['close']
    snap['current_price'] = current_price
    snap['last_candle_time'] = candles[-1].get('time')
    snap['swings'] = {
        'high_count': len(sh), 'low_count': len(sl),
        'recent_highs': [{'price': p['price'], 'time': p.get('time')} for p in sh[-3:]],
        'recent_lows':  [{'price': p['price'], 'time': p.get('time')} for p in sl[-3:]],
    }

    _dr_tr = []
    _dr_tc = []   # D2.4c per-bar walk-window collector (diagnose only)
    dr = _ict_dealing_range(sh, sl, current_price, candles=candles,
                            trace_events=_dr_tr, trace_collector=_dr_tc,
                            pivot_min_prominence_atr=_scanner_settings().get(
                                'dr_pivot_min_prominence_atr', 1.49))
    # D2.3: DR walk trace (pivot ledger + events), attached even on a partial
    # walk so a no-range outcome is still readable in the diagnose view.
    snap['dr_trace'] = {
        'pivots': next((e['pivots'] for e in _dr_tr if e.get('type') == 'pivots'), []),
        'events': [e for e in _dr_tr if e.get('type') != 'pivots'],
    }
    # D2.4c: compact per-bar walk window (last 50 'bar' records + ALL event
    # records) plus, for 1w only, the aggregated 2026 weekly OHLC the walk
    # actually consumed — so the live diagnose response can answer whether a
    # June close truly broke the range low, and settle Monday-vs-Thursday
    # weekly bucketing. Additive keys; nothing existing is touched.
    _tc_bars = [r for r in _dr_tc if r.get('type') == 'bar']
    _tc_events = [r for r in _dr_tc if r.get('type') != 'bar']
    snap['walk_window'] = _tc_bars[-50:] + _tc_events
    if interval == '1w':
        snap['weekly_ohlc_2026'] = [
            {'date': time.strftime('%Y-%m-%d', time.gmtime(c.get('time') or 0)),
             'open': c['open'], 'high': c['high'], 'low': c['low'], 'close': c['close']}
            for c in candles
            if (c.get('time') or 0) >= 1767225600   # 2026-01-01 00:00 UTC
        ]
    if not dr:
        snap['error'] = 'no dealing range'
        return snap

    # Leg-direction bias — same convention as the pipeline.
    _h_t = dr.get('anchor_high_time')
    _l_t = dr.get('anchor_low_time')
    if _h_t is not None and _l_t is not None and _h_t != _l_t:
        leg_direction = 'down' if _h_t < _l_t else 'up'
        direction = 'bearish' if leg_direction == 'down' else 'bullish'
        bias_source = 'leg'
    else:
        leg_direction = None
        direction = 'bullish' if dr['zone'] == 'discount' else 'bearish'
        bias_source = 'zone_fallback'

    snap['structure'] = {
        'dr_high': dr['high'], 'dr_low': dr['low'],
        'dr_anchor_high_time': _h_t, 'dr_anchor_low_time': _l_t,
        'equilibrium': dr['eq'],
        'level_786': dr.get('level_786'), 'level_618': dr.get('level_618'),
        'zone': dr['zone'],                # informational display only
        'leg_direction': leg_direction,
        'bias': direction,
        'bias_source': bias_source,
    }

    # OTE band + qualification zone — exact pipeline formulas.
    _rng = dr['high'] - dr['low']
    if direction == 'bullish':
        _a = dr['high'] - 0.618 * _rng
        _b = dr['high'] - 0.786 * _rng
    else:
        _a = dr['high'] - 0.382 * _rng
        _b = dr['high'] - 0.214 * _rng
    ote_low = round(min(_a, _b), 6)
    ote_high = round(max(_a, _b), 6)
    if direction == 'bearish':
        zone_low, zone_high = ote_low, dr['high']
    else:
        zone_low, zone_high = dr['low'], ote_high
    snap['levels'] = {
        'ote_low': ote_low, 'ote_high': ote_high,
        'zone_low': zone_low, 'zone_high': zone_high,
    }

    st = _scanner_settings()
    atr = _ict_atr(candles, 14)
    snap['atr'] = round(atr, 6) if atr else None

    # OB clusters — detector output preserved + per-candidate dims/in_zone.
    ob_candidates = _ict_order_block(candles, direction, dr)
    in_zone_ids = {id(o) for o in _poi_in_ote(ob_candidates, zone_low, zone_high)}
    obs_out = []
    for o in ob_candidates:
        row = dict(o)
        dc_idx = o['cluster_end_idx'] + 1
        ba = br = 0.0
        dim_ok = False
        if dc_idx < len(candles):
            dc = candles[dc_idx]
            bd = abs(dc['close'] - dc['open'])
            cr = dc['high'] - dc['low']
            at = _ict_atr(candles, 14, at_index=dc_idx)
            ba = (bd / at) if (at and at > 0) else 0.0
            br = (bd / cr) if cr > 0 else 0.0
            dim_ok = ((at is not None and at > 0 and ba >= st['ob_body_atr_min'])
                      and (cr > 0 and br >= st['ob_body_range_ratio_min']))
        row['body_atr_ratio'] = round(ba, 3)
        row['body_range_ratio'] = round(br, 3)
        row['dimension_pass'] = dim_ok
        row['in_zone'] = id(o) in in_zone_ids
        obs_out.append(row)
    snap['obs'] = obs_out

    # Standalone FVGs — shared leg span, this TF's zone/ATR/tunable.
    span = _leg_span_indices(candles, _h_t, _l_t, direction)
    span_start, span_end = span if span else (None, None)
    fz = _ict_fvg_in_zone(candles, direction, zone_low, zone_high,
                          span_start, span_end, atr,
                          st.get('fvg_min_atr_frac', 0.10))
    snap['fvgs'] = {'counts': fz['counts'], 'candidates': fz['candidates'],
                    'min_atr_frac': st.get('fvg_min_atr_frac', 0.10)}
    return snap


class SnapshotRun:
    """Per-run compute-once cache for cascade snapshots (Phase 2).

    Two memo layers per run, both keyed (coin, interval):
      1. raw candle arrays   2. computed snapshots
    In-memory only — no disk, no SQLite. reset() starts a new run.

    Fetch plan for the five cascade intervals: dailies fetched ONCE per coin
    at limit 1000; the 1w candles are DERIVED from that same daily array via
    _weekly_from_dailies (no second network fetch). 12h/4h at limit 300;
    1h at limit 600 (no per-request cap evidence found in code — stayed
    conservative per spec)."""

    CASCADE_INTERVALS = ('1w', '1d', '12h', '4h', '1h')
    FETCH_LIMITS = {'1d': 1000, '12h': 300, '4h': 300, '1h': 600}

    def __init__(self, fetch_fn=None):
        # fetch_fn injectable for tests; defaults to the live HL fetch.
        self._fetch_fn = fetch_fn
        self.reset()

    def reset(self):
        self._candles = {}     # (coin, interval) -> list[candle]
        self._snapshots = {}   # (coin, interval) -> snapshot dict

    def candles(self, coin, interval):
        key = (coin, interval)
        if key in self._candles:
            return self._candles[key]
        if interval == '1w':
            # Derive weeklies from the shared daily fetch — zero extra fetches.
            out = _weekly_from_dailies(self.candles(coin, '1d'),
                                       HTF_PARAMS['1w']['limit'])
        else:
            fetch = self._fetch_fn or _hl_fetch_candles
            out = fetch(coin, interval, limit=self.FETCH_LIMITS.get(interval, 300))
        self._candles[key] = out
        return out

    def snapshot(self, coin, interval):
        key = (coin, interval)
        if key in self._snapshots:
            return self._snapshots[key]
        snap = _compute_tf_snapshot(coin, interval, self.candles(coin, interval))
        self._snapshots[key] = snap
        return snap


def _build_snapshot_diagnose(coin, run):
    """All five cascade TF snapshots for one coin, partial-failure safe:
    one TF failing must not kill the response — its slot is None and the
    error string is recorded. Pure helper so the route stays thin and the
    harness can drive it directly."""
    snapshots = {}
    errors = {}
    for iv in SnapshotRun.CASCADE_INTERVALS:
        try:
            snapshots[iv] = run.snapshot(coin, iv)
        except Exception as e:
            snapshots[iv] = None
            errors[iv] = f'{type(e).__name__}: {e}'
    return snapshots, errors


def _scanner_symbol_to_coin(symbol):
    """Canonical watchlist-symbol → HL coin normalization, identical to the
    scan loop: strip a USDT/USDC/PERP quote suffix (dashed or bare), keep any
    leading namespace, drop a trailing dash."""
    coin = symbol
    for _sfx in ('USDT', 'USDC', 'PERP'):
        for _form in (f'-{_sfx}', _sfx):
            if coin.endswith(_form) and len(coin) > len(_form):
                coin = coin[:-len(_form)]
                break
        else:
            continue
        break
    return coin.rstrip('-')   # belt-and-suspenders: no trailing dash


def _run_scanner_scan(symbols=None, combos=None, tiers=None, items=None, kind='manual',
                      stale_minutes=None):
    """Run the parallel ICT scan and persist survivors. Returns the same payload
    the route builds today PLUS error/attempt counts. No HTTP coupling.

    Item resolution:
      - if `items` is provided (list of dicts with 'symbol'), scan exactly those
        (used by the scheduled wide-scan)
      - elif combos/symbols provided, resolve from the watchlist as the route does
      - else scan the full watchlist
    """
    import json as _json
    import concurrent.futures
    from src.storage.portfolio_db import get_connection

    SCAN_POOL_WORKERS = 6   # parallel HL-fetch workers (compute only; DB stays serial)

    _filter_symbols = symbols
    _filter_combos  = combos

    # V3_PAIRS is module-level (shared with the diagnose route).
    # pair_key is the uppercase HTF label stored in DB and used as the
    # grouping key in the signals API and frontend.
    _all_tiers = {p[3] for p in V3_PAIRS}
    _req_tiers = tiers
    if not isinstance(_req_tiers, list) or not _req_tiers:
        _req_tiers = set(_all_tiers)   # no tiers ("scan all") → every pair present
    else:
        _req_tiers = {t for t in _req_tiers if t in _all_tiers}  # explicit list filters
    pairs_to_scan = [p for p in V3_PAIRS if p[3] in _req_tiers]

    # Status vocab mapping — only survivor states are stored.
    _STATE_TO_STATUS = {
        'SETUP_READY': 'active',
        'POI_WAITING': 'forming',
    }
    from datetime import datetime as _dt
    scanned_at_iso = _dt.utcnow().isoformat() + 'Z'

    conn = get_connection()
    _scan_started = False   # True once we've marked _SCAN_STATE running=True
    _final = None           # the result dict (distinct from the per-pair `result`)
    try:
        # Resolve items to scan
        if items is not None:
            # Scheduled path: scan exactly these symbols (may be outside the
            # watchlist). Shared dedup loop below handles duplicates.
            items = [{'symbol': it.get('symbol')} for it in items if it.get('symbol')]
        elif _filter_combos:
            seen = set()
            items = []
            for combo in _filter_combos:
                sym = combo.get('symbol', '')
                if sym and sym not in seen:
                    seen.add(sym)
                    row = conn.execute(
                        "SELECT * FROM scanner_watchlist WHERE symbol=? LIMIT 1", (sym,)
                    ).fetchone()
                    if row:
                        items.append(row)
        elif _filter_symbols:
            _ph = ','.join('?' * len(_filter_symbols))
            items = conn.execute(
                f"SELECT * FROM scanner_watchlist WHERE symbol IN ({_ph}) ORDER BY created_at",
                _filter_symbols
            ).fetchall()
        else:
            items = conn.execute(
                "SELECT * FROM scanner_watchlist ORDER BY created_at"
            ).fetchall()

        # Optional stale filter: keep only tickers never scanned, or whose
        # last_scanned_at is older than (now - stale_minutes). NULL/empty = never
        # scanned = always stale. last_scanned_at is stored UTC ISO (… 'Z').
        if stale_minutes is not None:
            try:
                _sm = int(stale_minutes)
            except (TypeError, ValueError):
                _sm = None
            if _sm is not None and _sm >= 0:
                from datetime import timedelta as _td
                _cutoff = _dt.utcnow() - _td(minutes=_sm)

                def _is_stale(_last):
                    if not _last:
                        return True
                    _s = str(_last).strip().replace(' ', 'T').rstrip('Z')
                    if not _s:
                        return True
                    try:
                        return _dt.fromisoformat(_s) < _cutoff
                    except ValueError:
                        return True   # unparseable → treat as stale (scan it)

                _orig_count = len(items)
                items = [it for it in items if _is_stale(dict(it).get('last_scanned_at'))]
                print(f"[SCAN v3] stale filter: {_sm}min → "
                      f"{len(items)} of {_orig_count} tickers stale", flush=True)

        if not items:
            conn.close()
            return {
                'setups': [], 'watchItems': [], 'scannedAt': scanned_at_iso,
                'totalScanned': 0, 'setupReadyCount': 0,
                'errorCount': 0, 'attemptedPairs': 0,
                'error429Count': 0, 'error5xxCount': 0,
                'tickerOutcomes': [],
            }

        # Deduplicate by symbol
        seen_syms = set()
        unique_items = []
        for item in items:
            sym = dict(item).get('symbol', '')
            if sym not in seen_syms:
                seen_syms.add(sym)
                unique_items.append(item)

        setups = []
        watch_items = []
        total_scanned = 0
        error_count = 0     # pairs dropped due to a worker exception (HTTPError, etc.)
        error_429_count = 0 # of those, throttling (HTTP 429)
        error_5xx_count = 0 # of those, HL server errors (HTTP 5xx)

        # ── PHASE 0 — pre-warm the HL universe cache so worker threads only
        # READ it and never trigger a concurrent (unlocked) refresh mid-pool.
        try:
            _hl_refresh_universes()
        except Exception as _pw_err:
            print(f"[SCAN v3] universe pre-warm skipped: {_pw_err}", flush=True)

        # ── PHASE 1 — parallel compute (fetch + analyze). NO DB access in
        # workers: _run_ict_pipeline is pure compute. Each worker returns
        # (task, result_or_None, error_or_None) so a single failure can't
        # crash the pool. Results are collected in task order for Phase 2.
        _tasks = []
        for item in unique_items:
            symbol = item['symbol']
            # Derive the HL coin name. Robust for BOTH the watchlist form
            # ("BTCUSDT") and the HL universe form ("BTC-USDT"): strip the quote
            # suffix whether or not a dash precedes it, then drop any trailing
            # dash. A leading namespace (e.g. "xyz:CL-USDT" → "xyz:CL") is kept.
            coin = _scanner_symbol_to_coin(symbol)
            for (pair_key, htf, ltf, tier) in pairs_to_scan:
                _tasks.append({
                    'symbol': symbol, 'coin': coin, 'pair_key': pair_key,
                    'htf': htf, 'ltf': ltf, 'tier': tier,
                })

        # Mark the scan running (for GET /scanner/status) and reset cancel.
        _SCAN_CANCEL.clear()
        _scan_state_update(running=True, kind=kind, started_at=time.time(),
                           total=len(_tasks), done=0, current=None, cancelled=False)
        _scan_started = True

        # Per-scan candle cache: the same (coin, interval, limit) candle pull is
        # shared across pairs within this one scan. Created fresh each request,
        # so it's naturally request-scoped. The 6 workers read/write it
        # concurrently → guard the dict ops with a lock; fetch happens OUTSIDE
        # the lock so distinct keys don't serialize (a rare duplicate fetch on a
        # race is harmless).
        _candle_cache = {}
        _candle_cache_lock = threading.Lock()

        def _cached_fetch(coin, interval, limit=200):
            # Key on (coin, interval) only — the same series fetched at two
            # different limits (e.g. 4h@300 HTF and 4h@200 LTF) shares one fetch.
            # Keep whichever fetch is longest; shorter requests slice the tail.
            key = (coin, interval)
            with _candle_cache_lock:
                cached = _candle_cache.get(key)
                if cached is not None and len(cached) >= limit:
                    return cached[-limit:]
            result = _hl_fetch_candles(coin, interval, limit)  # outside lock
            with _candle_cache_lock:
                existing = _candle_cache.get(key)
                if existing is None or len(result) >= len(existing):
                    _candle_cache[key] = result
                else:
                    result = existing
            return result[-limit:] if len(result) >= limit else result

        def _scan_one(_t):
            # Worker: pure compute only — must NOT touch conn or any DB.
            try:
                _res = _run_ict_pipeline(_t['coin'], _t['htf'], _t['ltf'], _cached_fetch)
                return (_t, _res, None)
            except Exception as _e:
                return (_t, None, _e)

        # Cooperative cancel: once cancel is set we stop SUBMITTING new tasks;
        # already-submitted in-flight tasks finish (never killed). done/current
        # advance as compute results arrive, so progress moves during this slow
        # (HL-fetch-bound) phase.
        _collected = {}   # task index -> (task, result, error)
        if _tasks:
            _pool = concurrent.futures.ThreadPoolExecutor(max_workers=SCAN_POOL_WORKERS)
            try:
                _fut_idx = {}
                for _idx, _t in enumerate(_tasks):
                    if _SCAN_CANCEL.is_set():
                        break   # stop queuing further work
                    _fut_idx[_pool.submit(_scan_one, _t)] = _idx
                for _fut in concurrent.futures.as_completed(_fut_idx):
                    _ci = _fut_idx[_fut]
                    _collected[_ci] = _fut.result()
                    _scan_state_update(done=len(_collected),
                                       current=(_collected[_ci][0] or {}).get('symbol'))
                    if _SCAN_CANCEL.is_set():
                        break   # stop awaiting remaining; proceed to Phase 2 now
            finally:
                # Cancel not-yet-started tasks and DON'T block on in-flight ones
                # (they're pure compute — they finish in the background, harmless).
                try:
                    _pool.shutdown(wait=False, cancel_futures=True)
                except TypeError:
                    _pool.shutdown(wait=False)   # cancel_futures added in py3.9
        # Consume results in task (submission) order for deterministic Phase 2.
        _computed = [_collected[i] for i in sorted(_collected)]

        # ── PHASE 2 — serial write on the main thread, in task order, using the
        # single request connection. Identical per-pair logic/SQL/logging as before.
        _scanned_set = set()   # symbols actually processed (for last_scanned_at)

        # Per-ticker outcome accumulator (persisted for scheduled scans so the UI
        # can show "what was scanned"). Keyed by symbol; keeps the best outcome
        # across the ticker's pairs and the most-informative (furthest-advanced)
        # reason. A "progress" score ranks how far each pair got through the gate
        # chain; the highest-scoring pair's state/reason wins.
        _ticker_outcomes = {}
        _OUTCOME_RANK = {'SETUP_READY': 100, 'POI_WAITING': 90}
        _REASON_RANK = {
            'no liquidity sweep':              60,   # got OB + FVG, OB still valid
            'OB invalidated':                  56,   # got OB + FVG; OB later invalidated
            'no displacement FVG':             54,   # got OB; no displacement FVG
            'OB failed candle-dimension check': 50,  # got an OB candidate
            'no OB in OTE':                    40,
            'no dealing range':                20,
            'insufficient HTF data':           15,
        }

        def _outcome_progress(_state, _reason):
            if _state in _OUTCOME_RANK:
                return _OUTCOME_RANK[_state]
            if isinstance(_reason, str) and _reason.startswith('error'):
                return 5
            return _REASON_RANK.get(_reason, 10)

        def _accum_outcome(_sym, _pk, _state, _reason):
            _entry = _ticker_outcomes.get(_sym)
            if _entry is None:
                _entry = {'symbol': _sym, 'outcome': _state, 'reason': _reason,
                          '_score': _outcome_progress(_state, _reason), 'pairs': []}
                _ticker_outcomes[_sym] = _entry
            _entry['pairs'].append({'pairKey': _pk, 'state': _state, 'reason': _reason})
            _sc = _outcome_progress(_state, _reason)
            if _sc > _entry['_score']:
                _entry['_score'], _entry['outcome'], _entry['reason'] = _sc, _state, _reason

        for (_t, result, error) in _computed:
            # Cooperative-cancel checkpoint: stop before processing this result.
            # Whatever was already written stays (committed below).
            if _SCAN_CANCEL.is_set():
                break
            total_scanned += 1
            symbol   = _t['symbol']
            pair_key = _t['pair_key']
            htf, ltf, tier = _t['htf'], _t['ltf'], _t['tier']
            _scanned_set.add(symbol)
            _scan_state_update(current=symbol)

            # Record this pair's outcome for the per-ticker summary (all branches).
            if error is not None:
                _p_state, _p_reason = 'DROPPED', f'error: {type(error).__name__}'
            else:
                _p_state = result.get('setup_state') or 'DROPPED'
                _p_reason = _p_state if _p_state in ('SETUP_READY', 'POI_WAITING') \
                    else (result.get('drop_reason') or 'dropped')
            _accum_outcome(symbol, pair_key, _p_state, _p_reason)

            if error is not None:
                error_count += 1
                _estr = str(error)
                if '429' in _estr:
                    error_429_count += 1
                if any(_c in _estr for _c in ('500', '502', '503', '504')):
                    error_5xx_count += 1
                print(
                    f"[SCAN v3] FAILED {symbol} {pair_key}: "
                    f"{type(error).__name__}: {error}",
                    flush=True
                )
                continue

            state         = result.get('setup_state')
            raw_ind       = result.get('raw_indicators') or {}
            triage        = result.get('triage')
            current_price = result.get('current_price')
            signal_text   = result.get('signal_text') or ''

            # Count error-drops: the pipeline swallows fetch/compute exceptions
            # into a DROPPED result (never raises), so these don't surface as a
            # worker-level `error`. Classify them here so errorCount/429/5xx are
            # accurate and _check_scan_health_and_alert can see them. Legitimate
            # no-setup drops leave is_error_drop absent → not counted.
            if result.get('is_error_drop'):
                error_count += 1
                _st = result.get('error_status', '')
                if _st == '429':
                    error_429_count += 1
                elif _st in ('500', '502', '503', '504'):
                    error_5xx_count += 1
                # falls through to the DROP logging/continue below — just counted.

            if state not in ('SETUP_READY', 'POI_WAITING') or not triage:
                # DROP: leave any prior valid row intact. A DROP can be a
                # transient miss (insufficient data / fetch error / no DR),
                # so we never destroy a stored signal without a fresh
                # survivor to replace it.
                print(f"[SCAN v3] {symbol} {pair_key}: DROPPED "
                      f"({result.get('drop_reason')})", flush=True)
                continue

            status = _STATE_TO_STATUS.get(state, 'quiet')
            # Persist the triage payload alongside the chart overlay so
            # GET /results can rebuild the contract (no schema change).
            triage_store = dict(triage)
            triage_store['tier'] = tier
            triage_store['symbol'] = symbol
            store_blob = dict(raw_ind)
            store_blob['triage'] = triage_store

            # Replace the prior row only now that we have a survivor to
            # write in its place.
            conn.execute(
                "DELETE FROM scanner_signals WHERE symbol=? AND pair_key=?",
                (symbol, pair_key)
            )
            conn.execute(
                """INSERT INTO scanner_signals
                   (symbol, interval, htf_timeframe, ltf_timeframe, pair_key,
                    signal_type, status, signal_text,
                    confidence_score, why_flagged,
                    proposed_entry, proposed_stop, proposed_target, rr_ratio,
                    concepts_triggered, raw_indicators_json,
                    htf_label, ltf_label,
                    recent_closes_htf, recent_closes_ltf,
                    current_price, strategy_id, detected_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                (
                    symbol, htf, htf, ltf, pair_key,
                    state,                           # signal_type = state
                    status,
                    signal_text,
                    0,
                    signal_text,
                    result.get('proposed_entry'), result.get('proposed_stop'),
                    result.get('proposed_target'), result.get('rr_ratio'),
                    _json.dumps([]),
                    _json.dumps(store_blob),
                    f"{htf.upper()} {result.get('direction') or ''}".strip(),
                    ltf.upper(),
                    _json.dumps([]), _json.dumps([]),
                    current_price,
                    None,
                )
            )

            if state == 'SETUP_READY':
                setups.append(_setup_from_triage(symbol, triage_store))
            else:
                watch_items.append(_watch_from_triage(symbol, triage_store))
            print(f"[SCAN v3] {symbol} {pair_key}: {state} "
                  f"({signal_text})", flush=True)

        # Stamp every ticker actually scanned this run (drop OR survivor) with the
        # scan time. On a mid-way cancel, only the processed subset gets stamped.
        _scanned_syms = list(_scanned_set)
        if _scanned_syms:
            _ph_ls = ','.join('?' * len(_scanned_syms))
            conn.execute(
                f"UPDATE scanner_watchlist SET last_scanned_at = ? WHERE symbol IN ({_ph_ls})",
                [scanned_at_iso, *_scanned_syms]
            )

        conn.commit()
        conn.close()

        # Order per-ticker outcomes: SETUP_READY → POI_WAITING → DROPPED, and
        # within each, furthest-advanced first (so error-drops sink to the end).
        _OUTCOME_ORDER = {'SETUP_READY': 0, 'POI_WAITING': 1, 'DROPPED': 2}
        _ticker_list = sorted(
            _ticker_outcomes.values(),
            key=lambda e: (_OUTCOME_ORDER.get(e['outcome'], 3), -e['_score'], e['symbol']))
        _ticker_outcomes_out = [
            {'symbol': e['symbol'], 'outcome': e['outcome'], 'reason': e['reason'],
             'pairs': e['pairs']}
            for e in _ticker_list
        ]

        ranked = _rank_and_cap_setups(setups, cap=3)
        _final = {
            "setups":          [_public_setup(s) for s in ranked],
            "watchItems":      watch_items,
            "scannedAt":       scanned_at_iso,
            "totalScanned":    total_scanned,
            "setupReadyCount": len(setups),
            "errorCount":      error_count,
            "attemptedPairs":  total_scanned,
            "error429Count":   error_429_count,
            "error5xxCount":   error_5xx_count,
            "cancelled":       _SCAN_CANCEL.is_set(),
            "tickerOutcomes":  _ticker_outcomes_out,
        }
        return _final

    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        raise
    finally:
        # The UI must never show "running" forever — always clear the flag if we
        # started, even on exception. Record a short summary of this run.
        if _scan_started:
            _was_cancelled = _SCAN_CANCEL.is_set()
            _scan_state_update(
                running=False, current=None, cancelled=_was_cancelled,
                last_result={
                    'finishedAt': time.time(),
                    'setupReadyCount': (_final or {}).get('setupReadyCount', 0),
                    'cancelled': _was_cancelled,
                    'attemptedPairs': (_final or {}).get('attemptedPairs', 0),
                })
            _SCAN_CANCEL.clear()


@app.route('/api/trading/scanner/run', methods=['POST'])
@login_required
def api_trading_scanner_run():
    data = request.json or {}
    symbols = data.get('symbols', None)
    combos  = data.get('combos', None)
    tiers   = data.get('tiers', None)
    _stale_minutes = data.get('stale_minutes')   # int or None
    # One scan at a time — don't block-wait (would risk the gunicorn timeout).
    if not _SCAN_LOCK.acquire(blocking=False):
        return jsonify({'error': 'A scan is already in progress (background or '
                                 'manual). Try again in a few minutes.'}), 409
    try:
        result = _run_scanner_scan(symbols=symbols, combos=combos, tiers=tiers,
                                   kind='manual', stale_minutes=_stale_minutes)
        # Preserve the legacy 400 UX ONLY when no stale filter is applied: an
        # explicit symbols/combos selection that resolved to nothing. With a
        # stale filter, an empty result is a valid "nothing stale enough" outcome
        # (return the empty dict with 200, not 400).
        if _stale_minutes is None and (symbols or combos) and result.get('attemptedPairs', 0) == 0:
            return jsonify({"error": "Watchlist is empty. Add symbols first."}), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        _SCAN_LOCK.release()


def _check_scan_health_and_alert(result, label=''):
    """If a scan's error rate exceeds 5%, send a Telegram alert distinguishing
    429 (our throttling) from 5xx (HL outage). No-op below threshold."""
    attempted = result.get('attemptedPairs', 0) or 0
    errors = result.get('errorCount', 0) or 0
    if attempted == 0:
        return
    rate = errors / attempted
    if rate <= 0.05:
        return
    e429 = result.get('error429Count', 0) or 0
    e5xx = result.get('error5xxCount', 0) or 0
    if e429 >= e5xx:
        cause = ("rate limiting (HTTP 429) — our requests are "
                 "being throttled; results may be incomplete")
    else:
        cause = ("Hyperliquid server errors (HTTP 5xx) — "
                 "exchange-side issue, retry later")
    msg = (f"⚠️ Scan health alert ({label})\n"
           f"Error rate {rate*100:.1f}% "
           f"({errors}/{attempted} pairs failed).\n"
           f"Likely cause: {cause}.")
    try:
        _send_telegram(msg)
    except Exception:
        pass


def _last_scheduled_scan(updates=None):
    """File-KV for the most-recent scheduled scan's per-ticker outcomes (mirrors
    the _telegram_settings file pattern). No args → read (or {} if absent);
    `updates` (a full snapshot dict) → overwrite the file and return it."""
    path = os.path.join('data', 'last_scheduled_scan.json')
    if updates is not None:
        with open(path, 'w') as f:
            json.dump(updates, f)
        return updates
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r') as f:
            return json.load(f)
    except Exception:
        return {}


def _run_scheduled_scan(min_volume_override=None):
    """Build the volume-filtered HL universe and scan it via _run_scanner_scan.
    Returns the same result dict (or None if another scan was already running).
    Intended for the scheduler and for manual testing."""
    # One scan at a time — non-blocking so a collision skips rather than stalls.
    if not _SCAN_LOCK.acquire(blocking=False):
        print("[SCHEDULED SCAN] skipped — another scan in progress", flush=True)
        return None
    try:
        settings = _scanner_settings()
        min_vol = min_volume_override if min_volume_override is not None \
            else settings.get('scan_min_volume', 100000)
        max_tickers = settings.get('scan_max_tickers', 250)
        # Pull all HL tickers above the volume floor (uncapped — slice AFTER the
        # type filter so a crypto/tradfi selection still yields up to max_tickers
        # of THAT type, not a fraction of the combined top-N).
        universe = _hl_fetch_top_volume(n=None, min_volume=min_vol, limit=None)
        scan_asset_type = settings.get('scan_asset_type', 'all')
        if scan_asset_type != 'all':
            universe = [u for u in universe if u.get('asset_type', 'crypto') == scan_asset_type]
        universe = universe[:max_tickers]
        # universe items already have 'symbol' and 'asset_type'
        items = [{'symbol': u['symbol'], 'asset_type': u.get('asset_type', 'crypto')}
                 for u in universe]
        print(f"[SCHEDULED SCAN] universe: {len(items)} tickers "
              f"(min_vol=${min_vol:,.0f})", flush=True)
        # Scan-started confirmation so you know the scheduler fired.
        try:
            _send_telegram(
                f"🔍 Scheduled scan starting — {len(items)} tickers "
                f"(min vol ${min_vol:,.0f}). Alert to follow on completion.")
        except Exception:
            pass
        result = _run_scanner_scan(items=items, tiers=None, kind='scheduled')
        print(f"[SCHEDULED SCAN] done: {result['setupReadyCount']} ready, "
              f"{result['errorCount']} errors, {result['attemptedPairs']} pairs",
              flush=True)
        # Persist this scheduled scan's per-ticker outcomes (most-recent only).
        try:
            _outcomes = result.get('tickerOutcomes') or []
            _last_scheduled_scan(updates={
                'scannedAt': result.get('scannedAt'),
                'minVolume': min_vol,
                'tickerCount': len(items),
                'setupReadyCount': result.get('setupReadyCount', 0),
                'errorCount': result.get('errorCount', 0),
                'tickers': [{'symbol': o['symbol'], 'outcome': o['outcome'], 'reason': o['reason']}
                            for o in _outcomes],
            })
        except Exception as _lss_err:
            print(f"[SCHEDULED SCAN] outcome persist skipped: {_lss_err}", flush=True)
        return result
    finally:
        _SCAN_LOCK.release()


@app.route('/api/trading/scanner/scheduled-test', methods=['POST'])
@login_required
def api_scanner_scheduled_test():
    try:
        result = _run_scheduled_scan()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/last-scheduled-scan', methods=['GET'])
@login_required
def api_last_scheduled_scan():
    data = _last_scheduled_scan()
    if not data:
        return jsonify({'scannedAt': None, 'tickers': []})
    return jsonify(data)


@app.route('/api/trading/scanner/status', methods=['GET'])
@login_required
def api_scanner_status():
    s = _scan_state_snapshot()
    # Simple ETA from observed throughput, only while running.
    eta_seconds = None
    if s.get('running') and s.get('done', 0) > 0 and s.get('total', 0) > 0:
        elapsed = time.time() - (s.get('started_at') or time.time())
        rate = s['done'] / elapsed if elapsed > 0 else 0
        remaining = s['total'] - s['done']
        if rate > 0:
            eta_seconds = int(remaining / rate)
    return jsonify({
        'running': s.get('running', False),
        'kind': s.get('kind'),
        'startedAt': s.get('started_at'),
        'total': s.get('total', 0),
        'done': s.get('done', 0),
        'current': s.get('current'),
        'cancelled': s.get('cancelled', False),
        'etaSeconds': eta_seconds,
        'lastResult': s.get('last_result'),
    })


@app.route('/api/trading/scanner/cancel', methods=['POST'])
@login_required
def api_scanner_cancel():
    s = _scan_state_snapshot()
    if not s.get('running'):
        return jsonify({'ok': False, 'message': 'No scan in progress'}), 409
    _SCAN_CANCEL.set()
    return jsonify({'ok': True, 'message':
        'Cancel requested — scan will stop at the next ticker boundary '
        '(may take a few seconds; up to ~10s if waiting on a slow fetch).'})


def _build_triage_results(tiers=None):
    """Rebuild the triage contract dict from the last stored scan. Recomputes
    freshness at read time. `tiers` may be a comma-separated string (as from the
    query param) or None for all tiers. Returns the payload dict (no jsonify).
    Raises on DB error — callers decide how to surface it."""
    import json as _json
    from src.storage.portfolio_db import get_connection

    _allowed_tiers = ('swing', 'intraday', 'intra', 'scalp')
    if tiers:
        _tiers = {t.strip() for t in tiers.split(',')
                  if t.strip() in _allowed_tiers}  # explicit list filters
    else:
        _tiers = set(_allowed_tiers)   # no filter → all tiers

    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM scanner_signals WHERE id IN "
            "(SELECT MAX(id) FROM scanner_signals GROUP BY symbol, pair_key) "
            "AND signal_type IN ('SETUP_READY', 'POI_WAITING')"
        ).fetchall()
        conn.close()

        setups = []
        watch_items = []
        latest = None
        for r in rows:
            try:
                blob = _json.loads(r['raw_indicators_json'] or '{}')
            except Exception:
                blob = {}
            triage = (blob or {}).get('triage') or {}
            if not triage or triage.get('tier') not in _tiers:
                continue
            if r['detected_at'] and (latest is None or r['detected_at'] > latest):
                latest = r['detected_at']
            if r['signal_type'] == 'SETUP_READY':
                setups.append(_setup_from_triage(r['symbol'], triage))
            else:
                watch_items.append(_watch_from_triage(r['symbol'], triage))

        ranked = _rank_and_cap_setups(setups, cap=3)
        # Normalize to exactly one trailing Z (the stored value may already
        # carry a Z, which previously produced a double "ZZ").
        scanned_at = (str(latest).replace(' ', 'T').rstrip('Z') + 'Z') if latest else None
        return {
            "setups":       [_public_setup(s) for s in ranked],
            "watchItems":   watch_items,
            "scannedAt":    scanned_at,
            "totalScanned": len(setups) + len(watch_items),
        }
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        raise


def _format_triage_digest(results, scan_time_label):
    """
    Formats a plain-text Telegram digest from a triage results dict.
    scan_time_label: e.g. 'Morning', 'Lunch', 'Afternoon'
    """
    lines = []
    lines.append(f"📋 *The Playbook — {scan_time_label} Digest*")
    lines.append("")

    setups = results.get('setups', [])
    watch  = results.get('watchItems', [])

    swing_setups     = [s for s in setups if s.get('tier') == 'swing']
    intraday_setups  = [s for s in setups if s.get('tier') == 'intraday']

    def fmt_setup_block(setup_list, label):
        if not setup_list:
            return [f"*{label}*", "No setups qualify right now.", ""]
        out = [f"*{label}*"]
        for i, s in enumerate(setup_list, 1):
            ticker   = s.get('ticker', '').replace('USDT','').replace('USDC','')
            direction = s.get('direction', '')
            htf      = s.get('htf', '')
            ltf      = s.get('ltf', '')
            entry_lo = s.get('entryLow')
            entry_hi = s.get('entryHigh')
            stop     = s.get('stop')
            targets  = s.get('targets') or []
            fresh    = s.get('freshnessTier', '')
            t_mins   = s.get('triggeredMins')
            rationale = s.get('rationale', '')

            dir_emoji = '🟢' if direction == 'LONG' else '🔴'
            fresh_emoji = {'fresh': '🟢', 'aging': '🟡', 'stale': '🔴'}.get(fresh, '⚪')

            entry_str = f"{entry_lo:.2f}–{entry_hi:.2f}" if entry_lo and entry_hi else '—'
            stop_str  = f"{stop:.2f}" if stop else '—'
            t1_str = f"{targets[0]['price']:.2f} ({targets[0].get('rr','?')}R)" if targets else '—'
            t2_str = f"{targets[1]['price']:.2f} ({targets[1].get('rr','?')}R)" if len(targets) > 1 else '—'
            age_str = f"{round(t_mins)}m ago" if t_mins is not None else '—'

            out.append(
                f"#{i} {dir_emoji} *{ticker}* {htf}→{ltf}\n"
                f"  Entry: {entry_str}\n"
                f"  Stop: {stop_str}  |  T1: {t1_str}  |  T2: {t2_str}\n"
                f"  {fresh_emoji} {fresh} · triggered {age_str}\n"
                f"  _{rationale[:100]}{'…' if len(rationale)>100 else ''}_"
            )
        out.append("")
        return out

    lines += fmt_setup_block(swing_setups, '🕐 ACT TODAY · SWING')
    if intraday_setups:
        lines += fmt_setup_block(intraday_setups, '⚡ ACT TODAY · INTRADAY')

    # ON WATCH summary
    lines.append("*👁 ON WATCH*")
    if watch:
        for w in watch:
            ticker = w.get('ticker','').replace('USDT','').replace('USDC','')
            htf    = w.get('htf','')
            ltf    = w.get('ltf','')
            waiting = w.get('waitingFor','—')
            lines.append(f"  • *{ticker}* {htf}→{ltf} — {waiting}")
    else:
        lines.append("  Nothing on watch.")
    lines.append("")
    lines.append(f"_Scanned {results.get('totalScanned', 0)} pairs · "
                 f"{results.get('scannedAt','')[:16].replace('T',' ')} UTC_")

    return "\n".join(lines)


def _send_telegram(text):
    """
    Sends text to the configured Telegram chat.
    Returns (True, None) on success, (False, error_str) on failure.
    """
    import requests as _req
    s = _telegram_settings()
    token   = s.get('bot_token', '').strip()
    chat_id = s.get('chat_id', '').strip()
    if not token or not chat_id:
        return False, 'bot_token or chat_id not configured'
    try:
        url  = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = _req.post(url, json={
            'chat_id':    chat_id,
            'text':       text,
            'parse_mode': 'Markdown',
        }, timeout=10)
        if resp.status_code == 200:
            return True, None
        return False, resp.text
    except Exception as e:
        return False, str(e)


@app.route('/api/trading/scanner/results')
def api_trading_scanner_results():
    """Rebuild the triage contract from the last stored scan. Recomputes
    freshness at read time. Same shape as POST /scanner/run."""
    try:
        return jsonify(_build_triage_results(tiers=request.args.get('tiers')))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/scanner/diagnose', methods=['POST'])
@login_required
def api_trading_scanner_diagnose():
    """On-demand per-symbol scanner worksheet. Reruns the EXACT live pipeline
    (_run_ict_pipeline with diagnostics=True) for every V3 pair using the same
    extracted candle fetch + per-request cache the scan engine uses, and
    returns the phase-by-phase values (DR, OTE band, OB candidates, FVG state,
    verdict) for hand-chart verification. Read-only: no DB writes, no
    scanner_signals rows, no Telegram."""
    from datetime import datetime as _dt
    data = request.json or {}
    symbol = (data.get('symbol') or '').strip().upper()
    if not symbol:
        return jsonify({'error': 'symbol required'}), 400

    # Same normalization as the scan loop (watchlist symbol → HL coin).
    coin = _scanner_symbol_to_coin(symbol)

    # Validate against the HL universes (the same resolver the fetch uses).
    # Only reject when a universe cache is populated — if the refresh failed
    # we can't validate here and the fetch itself will surface the error.
    try:
        _hl_refresh_universes()
    except Exception:
        pass
    _crypto = _HL_UNIVERSE_CACHE.get('crypto') or set()
    _xyz = _HL_UNIVERSE_CACHE.get('xyz') or {}
    if (_crypto or _xyz) and coin.upper() not in _crypto and coin.upper() not in _xyz:
        return jsonify({'error': f"Symbol '{symbol}' (coin '{coin}') not found "
                                 f"in the Hyperliquid universe"}), 404

    try:
        # Per-request candle cache mirroring the scan's: key on (coin,
        # interval) only, keep the longest series, slice the tail for shorter
        # requests. Single-threaded here so no lock is needed.
        _cache = {}

        def _cached_fetch(_coin, interval, limit=200):
            key = (_coin, interval)
            cached = _cache.get(key)
            if cached is not None and len(cached) >= limit:
                return cached[-limit:]
            result = _hl_fetch_candles(_coin, interval, limit)
            existing = _cache.get(key)
            if existing is None or len(result) >= len(existing):
                _cache[key] = result
            else:
                result = existing
            return result[-limit:] if len(result) >= limit else result

        pairs_out = []
        for (pair_key, htf, ltf, tier) in V3_PAIRS:
            res = _run_ict_pipeline(coin, htf, ltf, _cached_fetch, diagnostics=True)
            worksheet = res.get('worksheet') or {}
            pairs_out.append({
                'pairKey': pair_key, 'htf': htf, 'ltf': ltf, 'tier': tier,
                **worksheet,
            })
        return jsonify({
            'symbol': symbol,
            'resolvedCoin': coin,
            'generatedAt': _dt.utcnow().isoformat() + 'Z',
            'pairs': pairs_out,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trading/scanner/snapshot-diagnose', methods=['POST'])
@login_required
def api_trading_scanner_snapshot_diagnose():
    """Cascade phase 2 preview: all five per-TF snapshots for one symbol via
    a fresh SnapshotRun (per-run compute-once cache; dailies fetched once and
    shared with the 1w derivation). Read-only: no DB writes, no Telegram.
    Partial-failure safe — a failing TF returns an error string, not a 500."""
    from datetime import datetime as _dt
    data = request.json or {}
    symbol = (data.get('symbol') or '').strip().upper()
    if not symbol:
        return jsonify({'error': 'symbol required'}), 400
    coin = _scanner_symbol_to_coin(symbol)
    try:
        run = SnapshotRun()
        snapshots, errors = _build_snapshot_diagnose(coin, run)
        return jsonify({
            'symbol': symbol,
            'resolvedCoin': coin,
            'generatedAt': _dt.utcnow().isoformat() + 'Z',
            'snapshots': snapshots,
            'errors': errors,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/telegram/test', methods=['POST'])
@login_required
def api_telegram_digest_test():
    """Send the triage digest immediately (verify without a scheduled window)."""
    results = _build_triage_results(tiers=None)
    text    = _format_triage_digest(results, 'Test')
    ok, err = _send_telegram(text)
    if ok:
        return jsonify({'ok': True, 'message': 'Digest sent'})
    return jsonify({'ok': False, 'error': err}), 400


# TEMPORARY — one-time triage-contract verification seed. Remove next commit.
@app.route('/api/trading/scanner/seed-mock', methods=['POST'])
def api_scanner_seed_mock():
    import time as _time
    import json as _json
    from src.storage.portfolio_db import get_connection
    now_iso = _time.strftime('%Y-%m-%dT%H:%M:%SZ', _time.gmtime())
    choch_time = int(_time.time()) - 1800  # 30 min ago → fresh

    setup_triage = {
        "htf": "1D", "ltf": "1H", "direction": "LONG",
        "tier": "swing", "grade": "A",
        "entryLow": 104500.0, "entryHigh": 105200.0, "stop": 103800.0,
        "targets": [
            {"price": 108000.0, "rr": 2.1},
            {"price": 112000.0, "rr": 4.8}
        ],
        "choch_time": choch_time,
        "rationale": "Mock seed: 1D OB in OTE, displacement FVG unmitigated, 1H CHoCH confirmed",
        "_rank": {"rr_t1": 2.1, "ote_depth": 0.72, "triggered_mins": 30}
    }

    watch_triage = {
        "htf": "1W", "ltf": "4H", "direction": "SHORT",
        "tier": "swing",
        "waitingFor": "4H CHoCH below 2480.00",
        "eta": None
    }

    conn = get_connection()
    try:
        conn.execute("""
            INSERT OR REPLACE INTO scanner_signals
                (symbol, htf_timeframe, ltf_timeframe, interval, signal_type,
                 price, status, raw_indicators_json, pair_key, detected_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, ('BTCUSDT', '1d', '1h', '1d', 'SETUP_READY',
              104850.0, 'active',
              _json.dumps({"triage": setup_triage}),
              '1D', now_iso))
        conn.execute("""
            INSERT OR REPLACE INTO scanner_signals
                (symbol, htf_timeframe, ltf_timeframe, interval, signal_type,
                 price, status, raw_indicators_json, pair_key, detected_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, ('ETHUSDT', '1w', '4h', '1w', 'POI_WAITING',
              2510.0, 'forming',
              _json.dumps({"triage": watch_triage}),
              '1W', now_iso))
        conn.commit()
        conn.close()
        return jsonify({"seeded": True})
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({"error": str(e)}), 500


# --- Journal ---

@app.route('/api/trading/journal')
def api_trading_journal_list():
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        limit = min(int(request.args.get('limit', 20)), 100)
        offset = int(request.args.get('offset', 0))
        rows = conn.execute(
            "SELECT * FROM trading_journal ORDER BY entry_date DESC, created_at DESC LIMIT ? OFFSET ?",
            (limit, offset)
        ).fetchall()
        total = conn.execute("SELECT COUNT(*) FROM trading_journal").fetchone()[0]
        conn.close()
        def _parse_entry(r):
            d = dict(r)
            d['tags'] = _json.loads(r['tags_json'] or '[]')
            for _jf, _default in [('execution_plan_json', None), ('concepts_json', []), ('validator_answers_json', None)]:
                try:
                    d[_jf] = _json.loads(r[_jf]) if r[_jf] else _default
                except Exception:
                    d[_jf] = _default
            return d
        entries = [_parse_entry(r) for r in rows]
        return jsonify({"entries": entries, "total": total, "limit": limit, "offset": offset})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal', methods=['POST'])
def api_trading_journal_create():
    import json as _json
    from datetime import date
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    def _jdump(v):
        if v is None:
            return None
        return v if isinstance(v, str) else _json.dumps(v)
    conn = get_connection()
    try:
        cur = conn.execute(
            "INSERT INTO trading_journal "
            "(entry_date, title, body, tags_json, mood, market_regime, "
            " ticker, direction, timeframe, outcome, r_multiple, ready_score, "
            " execution_plan_json, concepts_json, validator_answers_json, "
            " created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
            (data.get('entry_date') or date.today().isoformat(), title,
             data.get('body', ''), _json.dumps(data.get('tags', [])),
             int(data.get('mood', 3)), data.get('market_regime', ''),
             data.get('ticker'), data.get('direction'), data.get('timeframe'),
             data.get('outcome', 'open'), data.get('r_multiple'),
             data.get('ready_score'),
             _jdump(data.get('execution_plan_json')),
             _jdump(data.get('concepts_json')),
             _jdump(data.get('validator_answers_json')))
        )
        entry_id = cur.lastrowid
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": entry_id})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>')
def api_trading_journal_get(entry_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        r = conn.execute("SELECT * FROM trading_journal WHERE id=?", (entry_id,)).fetchone()
        conn.close()
        if not r:
            return jsonify({"error": "Not found"}), 404
        d = dict(r)
        d['tags'] = _json.loads(r['tags_json'] or '[]')
        for _jf, _default in [('execution_plan_json', None), ('concepts_json', []), ('validator_answers_json', None)]:
            try:
                d[_jf] = _json.loads(r[_jf]) if r[_jf] else _default
            except Exception:
                d[_jf] = _default
        return jsonify(d)
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>', methods=['PUT'])
def api_trading_journal_update(entry_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        if not conn.execute("SELECT id FROM trading_journal WHERE id=?", (entry_id,)).fetchone():
            conn.close()
            return jsonify({"error": "Not found"}), 404
        def _jdump_u(v):
            if v is None:
                return None
            return v if isinstance(v, str) else _json.dumps(v)
        conn.execute(
            "UPDATE trading_journal SET "
            "title=COALESCE(?,title), body=COALESCE(?,body), "
            "tags_json=COALESCE(?,tags_json), mood=COALESCE(?,mood), "
            "market_regime=COALESCE(?,market_regime), entry_date=COALESCE(?,entry_date), "
            "ticker=COALESCE(?,ticker), direction=COALESCE(?,direction), "
            "timeframe=COALESCE(?,timeframe), outcome=COALESCE(?,outcome), "
            "r_multiple=COALESCE(?,r_multiple), ready_score=COALESCE(?,ready_score), "
            "execution_plan_json=COALESCE(?,execution_plan_json), "
            "concepts_json=COALESCE(?,concepts_json), "
            "validator_answers_json=COALESCE(?,validator_answers_json), "
            "updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (data.get('title'), data.get('body'),
             _json.dumps(data['tags']) if 'tags' in data else None,
             data.get('mood'), data.get('market_regime'), data.get('entry_date'),
             data.get('ticker'), data.get('direction'), data.get('timeframe'),
             data.get('outcome'), data.get('r_multiple'), data.get('ready_score'),
             _jdump_u(data.get('execution_plan_json')),
             _jdump_u(data.get('concepts_json')),
             _jdump_u(data.get('validator_answers_json')),
             entry_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>', methods=['DELETE'])
def api_trading_journal_delete(entry_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        # Remove child screenshots first — the FK has no ON DELETE CASCADE and
        # foreign_keys=ON, so deleting the parent while children exist would 500.
        conn.execute("DELETE FROM journal_screenshots WHERE journal_entry_id=?", (entry_id,))
        conn.execute("DELETE FROM trading_journal WHERE id=?", (entry_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>/screenshots')
def api_trading_journal_screenshots_list(entry_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT id, filename, mime_type, data FROM journal_screenshots WHERE journal_entry_id=? ORDER BY created_at ASC",
            (entry_id,)
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>/screenshots', methods=['POST'])
def api_trading_journal_screenshots_create(entry_id):
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        cur = conn.execute(
            "INSERT INTO journal_screenshots (journal_entry_id, filename, mime_type, data) VALUES (?,?,?,?)",
            (entry_id, data.get('filename', ''), data.get('mime_type', 'image/png'), data.get('data', ''))
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": cur.lastrowid})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/journal/<int:entry_id>/screenshots/<int:screenshot_id>', methods=['DELETE'])
def api_trading_journal_screenshots_delete(entry_id, screenshot_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute(
            "DELETE FROM journal_screenshots WHERE id=? AND journal_entry_id=?",
            (screenshot_id, entry_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies')
def api_trading_strategies_list():
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        rows = conn.execute("SELECT * FROM strategies WHERE is_active=1 ORDER BY id ASC").fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            try: d['flow_json'] = _json.loads(r['flow_json']) if r['flow_json'] else None
            except Exception: d['flow_json'] = None
            out.append(d)
        return jsonify(out)
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies/<int:strategy_id>')
def api_trading_strategies_get(strategy_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        r = conn.execute("SELECT * FROM strategies WHERE id=? AND is_active=1", (strategy_id,)).fetchone()
        conn.close()
        if not r:
            return jsonify({"error": "Not found"}), 404
        d = dict(r)
        try: d['flow_json'] = _json.loads(r['flow_json']) if r['flow_json'] else None
        except Exception: d['flow_json'] = None
        return jsonify(d)
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies', methods=['POST'])
def api_trading_strategies_create():
    import json as _json
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        if data.get('is_default'):
            conn.execute("UPDATE strategies SET is_default=0")
        flow = data.get('flow_json')
        flow_str = _json.dumps(flow) if isinstance(flow, dict) else (flow or _json.dumps({"base_flow": [], "adaptive_steps": {}, "risk_step": {"id": "risk", "label": "Risk / Inv.", "stepType": "LTF", "isForm": True, "question": "Define your risk parameters.", "subtext": ""}}))
        cur = conn.execute(
            "INSERT INTO strategies (name, description, ai_prompt, flow_json, is_default, is_active) VALUES (?,?,?,?,?,1)",
            (data.get('name', 'New Strategy'), data.get('description', ''), data.get('ai_prompt', ''), flow_str, 1 if data.get('is_default') else 0)
        )
        conn.commit()
        conn.close()
        return jsonify({"success": True, "id": cur.lastrowid})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies/<int:strategy_id>', methods=['PUT'])
def api_trading_strategies_update(strategy_id):
    import json as _json
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    conn = get_connection()
    try:
        if not conn.execute("SELECT id FROM strategies WHERE id=?", (strategy_id,)).fetchone():
            conn.close()
            return jsonify({"error": "Not found"}), 404
        if data.get('is_default'):
            conn.execute("UPDATE strategies SET is_default=0")
        updates, params = [], []
        for field in ('name', 'description', 'ai_prompt', 'is_default'):
            if field in data:
                updates.append(f"{field}=?"); params.append(data[field])
        if 'flow_json' in data:
            fj = data['flow_json']
            updates.append("flow_json=?"); params.append(_json.dumps(fj) if isinstance(fj, dict) else fj)
        updates.append("updated_at=CURRENT_TIMESTAMP")
        params.append(strategy_id)
        conn.execute(f"UPDATE strategies SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies/<int:strategy_id>', methods=['DELETE'])
def api_trading_strategies_delete(strategy_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("UPDATE strategies SET is_active=0, updated_at=CURRENT_TIMESTAMP WHERE id=?", (strategy_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies/<int:strategy_id>/set-default', methods=['POST'])
def api_trading_strategies_set_default(strategy_id):
    from src.storage.portfolio_db import get_connection
    conn = get_connection()
    try:
        conn.execute("UPDATE strategies SET is_default=0")
        conn.execute("UPDATE strategies SET is_default=1, updated_at=CURRENT_TIMESTAMP WHERE id=?", (strategy_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/strategies/<int:strategy_id>/update-tree', methods=['POST'])
def api_trading_strategies_update_tree(strategy_id):
    import json as _json, re as _re
    from src.engines.ai_advisor import load_ai_config
    from src.engines.llm_providers import get_provider
    from src.storage.portfolio_db import get_connection
    data = request.json or {}
    instruction = data.get('instruction', '').strip()
    if not instruction:
        return jsonify({"error": "instruction required"}), 400
    conn = get_connection()
    try:
        row = conn.execute("SELECT flow_json FROM strategies WHERE id=? AND is_active=1", (strategy_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Strategy not found"}), 404
        flow_json_str = row['flow_json'] or '{}'
        original = _json.loads(flow_json_str)
        sys_prompt = ("You are a JSON editor. You will receive a validator flow definition as JSON and a natural language instruction describing a change. "
                      "Return ONLY the updated JSON object with no markdown, no code fences, no preamble, no explanation. The response must be valid JSON only.")
        user_prompt = f"Current flow JSON:\n{flow_json_str}\n\nInstruction: {instruction}\n\nReturn the complete updated flow JSON only."
        config = load_ai_config()
        provider = get_provider(config)
        result = provider.complete(sys_prompt, user_prompt)
        raw = result.get('response', '')
        if isinstance(raw, dict):
            proposed = raw
        else:
            cleaned = _re.sub(r'^```(?:json)?\s*|\s*```$', '', str(raw).strip(), flags=_re.MULTILINE).strip()
            proposed = _json.loads(cleaned)
        if not isinstance(proposed.get('base_flow'), list) or not isinstance(proposed.get('risk_step'), dict):
            return jsonify({"error": "AI response missing required fields base_flow or risk_step"}), 400
        return jsonify({"proposed_flow_json": proposed, "original_flow_json": original})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({"error": str(e)}), 500


@app.route('/api/trading/validator/thesis', methods=['POST'])
def api_trading_validator_thesis():
    import json as _json
    from src.engines.ai_advisor import load_ai_config
    from src.engines.llm_providers import get_provider
    data = request.json or {}
    try:
        ticker    = data.get('ticker', '')
        direction = data.get('direction', '')
        timeframe = data.get('timeframe', '')
        answers   = data.get('answers', {})
        plan      = data.get('execution_plan', {})
        ready_score = data.get('ready_score', '')

        answers_text = '\n'.join(
            f"  {v.get('label', k)}: {v.get('value', '')}"
            for k, v in answers.items()
        ) if answers else '  (no answers provided)'

        user_prompt = (
            f"Trade: {ticker} {direction} on {timeframe}\n"
            f"Setup answers:\n{answers_text}\n"
            f"Execution: Entry {plan.get('entry', '—')} · Stop {plan.get('stop', '—')} · "
            f"TP1 {plan.get('tp1', '—')} · TP2 {plan.get('tp2', '—')}\n"
            f"Invalidation: {plan.get('invalidation', '—')}\n"
            f"Ready score: {ready_score}\n"
            "Write the pre-trade thesis. Respond with ONLY the thesis text. Do not wrap in JSON."
        )
        sys_prompt = (
            "You are a pre-trade reasoning assistant. Write a concise, professional "
            "pre-trade thesis in 3–5 sentences. Be specific about the setup, not generic. "
            "Reference the actual levels and structure provided. Note any warnings "
            "(e.g. R:R below target). Output plain text only — no markdown, no headers."
        )

        # Inject strategy context — prefer named strategy ai_prompt over category-based docs
        strategy_id = data.get('strategy_id')
        doc_content = None
        try:
            from src.storage.portfolio_db import get_connection as _gc_vt
            if strategy_id:
                _st_conn = _gc_vt()
                _st_row = _st_conn.execute("SELECT ai_prompt FROM strategies WHERE id=? AND is_active=1", (strategy_id,)).fetchone()
                _st_conn.close()
                if _st_row and (_st_row['ai_prompt'] or '').strip():
                    doc_content = _st_row['ai_prompt'].strip()
            if not doc_content:
                # Fallback: category-based strategy_documents
                htf_bias_val = (answers.get('htf_bias') or answers.get('setup') or {}).get('value', '')
                if direction == 'long' and 'bullish' in htf_bias_val.lower():
                    _vt_cats = ['bull', 'trading', 'trading_strategy']
                elif direction == 'short' and 'bearish' in htf_bias_val.lower():
                    _vt_cats = ['bear', 'trading', 'trading_strategy']
                else:
                    _vt_cats = ['trading', 'trading_strategy']
                _vt_conn = _gc_vt()
                _ph = ','.join('?' for _ in _vt_cats)
                _vt_rows = _vt_conn.execute(
                    f"SELECT filename, extracted_text FROM strategy_documents WHERE category IN ({_ph}) ORDER BY uploaded_at DESC",
                    _vt_cats
                ).fetchall()
                _vt_conn.close()
                _parts = [f"=== {r['filename']} ===\n{r['extracted_text']}" for r in _vt_rows if r['extracted_text']]
                _combined = '\n\n'.join(_parts)[:3000]
                if _combined.strip():
                    doc_content = _combined
        except Exception:
            pass  # Strategy context injection is best-effort — never block thesis generation

        if doc_content:
            sys_prompt += (
                "\n\nRELEVANT STRATEGY CONTEXT:\n"
                + doc_content +
                "\nUse this context to inform the thesis where relevant, but keep the thesis "
                "focused on the specific trade setup provided."
            )

        config   = load_ai_config()
        provider = get_provider(config)
        result   = provider.complete(sys_prompt, user_prompt)
        response = result.get('response', '')

        if isinstance(response, dict):
            thesis = (response.get('thesis') or response.get('text') or
                      response.get('content') or str(response))
        else:
            thesis = str(response) if response else ''

        tokens = result.get('tokens', {})
        return jsonify({"thesis": thesis, "tokens": tokens})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# DEPRECATED: scanner-prompt routes are no longer referenced by any frontend JS.
# Strategy AI prompts are now managed via /api/trading/strategies CRUD routes.
# These routes are kept for backward compatibility only — do not remove yet.
@app.route('/api/trading/scanner-prompt')
def api_trading_scanner_prompt_get():
    try:
        with open(_get_scanner_prompt_path(), 'r') as f:
            return jsonify({"prompt": f.read()})
    except FileNotFoundError:
        return jsonify({"prompt": ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# DEPRECATED: see note on GET route above.
@app.route('/api/trading/scanner-prompt', methods=['POST'])
def api_trading_scanner_prompt_save():
    data = request.json or {}
    try:
        path = _get_scanner_prompt_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, 'w') as f:
            f.write(data.get('prompt', ''))
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    start_snapshot_scheduler()
    # Debug mode is opt-in via FLASK_DEBUG=1 — Werkzeug's debugger exposes
    # an in-browser RCE console (PIN-protected, but a dangerous default).
    # Off by default so accidental binds beyond localhost stay safe.
    debug_mode = os.getenv("FLASK_DEBUG", "").strip() == "1"
    app.run(debug=debug_mode, port=5001)
else:
    # Running under gunicorn — start scheduler on import
    start_snapshot_scheduler()

